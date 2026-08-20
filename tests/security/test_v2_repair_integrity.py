from __future__ import annotations

import copy
import zipfile
from collections.abc import Callable
from pathlib import Path
from typing import Any

import pytest
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Protection

import workbooklens.repair.ooxml_patch as ooxml_patch
from workbooklens.demo.workflow import generate_demo_workbook
from workbooklens.exceptions import (
    PatchValidationError,
    StalePlanError,
    UsageError,
    WorkbookLensError,
)
from workbooklens.models import PatchKind, PatchOperation, PatchPlan
from workbooklens.ooxml.safety import PackageLimits, inspect_package
from workbooklens.repair import apply_patch_plan, build_patch_plan
from workbooklens.scanner import scan_workbook
from workbooklens.snapshot import cell_fingerprint
from workbooklens.utils import sha256_file

CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XLSM_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
VBA_CONTENT_TYPE = "application/vnd.ms-office.vbaProject"
VBA_RELATIONSHIP_TYPE = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
SCAN_CONFIG = {"keys": [{"sheet": "Sales", "range": "A2:A22", "ignore_blank": True}]}


def _plan(path: Path) -> PatchPlan:
    return build_patch_plan(scan_workbook(path, config=SCAN_CONFIG))


def _formula_patch(plan: PatchPlan) -> PatchOperation:
    return next(patch for patch in plan.patches if patch.cell == "D8")


def _mutate_after(patch: PatchOperation) -> None:
    patch.after = "=B8*C8+1"


def _mutate_safe(patch: PatchOperation) -> None:
    patch.safe = False


def _mutate_confidence(patch: PatchOperation) -> None:
    patch.confidence = 0.98


def _mutate_kind(patch: PatchOperation) -> None:
    patch.kind = PatchKind.SET_FORMULA


def _mutate_source_cell(patch: PatchOperation) -> None:
    patch.source_cell = "D11"


def _mutate_precondition(patch: PatchOperation) -> None:
    patch.precondition.expected_value = "edited"


def _mutate_before(patch: PatchOperation) -> None:
    patch.before = "edited"


def _mutate_description(patch: PatchOperation) -> None:
    patch.description = "edited repair authority"


def _mutate_sheet(patch: PatchOperation) -> None:
    patch.sheet = "Summary"


def _mutate_cell(patch: PatchOperation) -> None:
    patch.cell = "D9"


@pytest.mark.parametrize(
    ("field", "mutate", "expected_error"),
    [
        ("after", _mutate_after, PatchValidationError),
        ("safe", _mutate_safe, PatchValidationError),
        ("confidence", _mutate_confidence, PatchValidationError),
        ("kind", _mutate_kind, PatchValidationError),
        ("source_cell", _mutate_source_cell, PatchValidationError),
        ("precondition", _mutate_precondition, PatchValidationError),
        ("before", _mutate_before, PatchValidationError),
        ("description", _mutate_description, PatchValidationError),
        ("sheet", _mutate_sheet, StalePlanError),
        ("cell", _mutate_cell, StalePlanError),
    ],
)
def test_apply_rejects_every_tampered_patch_field_without_output(
    tmp_path: Path,
    field: str,
    mutate: Callable[[PatchOperation], None],
    expected_error: type[WorkbookLensError],
) -> None:
    source = tmp_path / f"source-{field}.xlsx"
    generate_demo_workbook(source)
    submitted = _plan(source)
    patch = _formula_patch(submitted)
    mutate(patch)
    output = tmp_path / f"fixed-{field}.xlsx"

    with pytest.raises(expected_error):
        apply_patch_plan(source, submitted, output, selected_ids={patch.id}, config=SCAN_CONFIG)

    assert not output.exists()


def _inject_vba_markers(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        infos = archive.infolist()
        contents = {info.filename: archive.read(info.filename) for info in infos}

    content_types = etree.fromstring(contents["[Content_Types].xml"])
    for declaration in content_types:
        if (
            etree.QName(declaration).localname == "Override"
            and declaration.get("PartName") == "/xl/workbook.xml"
        ):
            declaration.set("ContentType", XLSM_WORKBOOK_CONTENT_TYPE)
    etree.SubElement(
        content_types,
        f"{{{CONTENT_TYPES_NS}}}Override",
        PartName="/xl/vbaProject.bin",
        ContentType=VBA_CONTENT_TYPE,
    )
    contents["[Content_Types].xml"] = etree.tostring(
        content_types,
        encoding="UTF-8",
        xml_declaration=True,
        standalone=True,
    )

    relationship_name = "xl/_rels/workbook.xml.rels"
    relationships = etree.fromstring(contents[relationship_name])
    etree.SubElement(
        relationships,
        f"{{{RELATIONSHIPS_NS}}}Relationship",
        Id="rIdWorkbookLensVba",
        Type=VBA_RELATIONSHIP_TYPE,
        Target="vbaProject.bin",
    )
    contents[relationship_name] = etree.tostring(
        relationships,
        encoding="UTF-8",
        xml_declaration=True,
        standalone=True,
    )

    temporary = path.with_name(f".{path.name}.vba.tmp")
    with zipfile.ZipFile(temporary, "w", allowZip64=True) as output:
        for info in infos:
            output.writestr(copy.copy(info), contents[info.filename])
        output.writestr("xl/vbaProject.bin", b"synthetic-vba-marker")
    temporary.replace(path)


def test_disguised_macro_enabled_xlsx_is_read_only_for_repair(tmp_path: Path) -> None:
    source = tmp_path / "disguised-macro.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    _inject_vba_markers(source)

    inspection = inspect_package(source)
    assert inspection.has_vba is True
    assert inspection.content_format == "xlsm"
    assert inspection.format_mismatch is True
    assert inspection.repairable is False

    output = tmp_path / "must-not-exist.xlsx"
    with pytest.raises(UsageError, match="read-only"):
        ooxml_patch.patch_ooxml_package(
            source,
            plan,
            output,
            safe_only=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_plain_xlsx_renamed_xlsm_is_format_mismatch_and_read_only(tmp_path: Path) -> None:
    xlsx = tmp_path / "plain.xlsx"
    generate_demo_workbook(xlsx)
    xlsm = tmp_path / "plain.xlsm"
    xlsm.write_bytes(xlsx.read_bytes())

    inspection = inspect_package(xlsm)
    assert inspection.content_format == "xlsx"
    assert inspection.extension_format == "xlsm"
    assert inspection.has_vba is False
    assert inspection.format_mismatch is True
    assert inspection.repairable is False
    assert inspection.format == "xlsm"


def test_precondition_analysis_receives_original_package_limits(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "limits.xlsx"
    generate_demo_workbook(source)
    canonical = _plan(source)
    submitted = canonical.model_copy(deep=True)
    patch = submitted.patches[0]
    output = tmp_path / "limits-fixed.xlsx"
    limits = PackageLimits()
    received: list[PackageLimits | None] = []
    original_load = ooxml_patch.load_for_analysis

    def recording_load(path: Path, active_limits: PackageLimits | None) -> Any:
        received.append(active_limits)
        return original_load(path, active_limits)

    monkeypatch.setattr(ooxml_patch, "load_for_analysis", recording_load)
    ooxml_patch.patch_ooxml_package(
        source,
        submitted,
        output,
        selected_ids={patch.id},
        limits=limits,
        canonical_plan=canonical,
    )

    assert received == [limits]
    assert received[0] is limits
    assert output.exists()


@pytest.mark.parametrize("hidden_target", ["sheet", "row", "grouped_column"])
def test_low_level_repair_rejects_newly_hidden_targets(tmp_path: Path, hidden_target: str) -> None:
    source = tmp_path / f"hidden-{hidden_target}.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    patch = _formula_patch(plan)

    workbook = load_workbook(source)
    worksheet = workbook[patch.sheet]
    if hidden_target == "sheet":
        worksheet.sheet_state = "hidden"
    elif hidden_target == "row":
        worksheet.row_dimensions[worksheet[patch.cell].row].hidden = True
    else:
        worksheet.column_dimensions.group("B", "D", hidden=True)
    workbook.save(source)
    workbook.close()
    plan.source_sha256 = sha256_file(source)

    output = tmp_path / f"hidden-{hidden_target}-fixed.xlsx"
    with pytest.raises(PatchValidationError, match=r"hidden|non-visible"):
        ooxml_patch.patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            canonical_plan=plan,
        )
    assert not output.exists()


@pytest.mark.parametrize(
    ("semantic", "expected_message"),
    [
        ("protection", "protection semantics"),
        ("number_format", "number-format semantics"),
        ("quote_prefix", "quote-prefix semantics"),
        ("pivot_button", "pivot-button semantics"),
    ],
)
def test_low_level_style_copy_rejects_semantic_mismatch(
    tmp_path: Path, semantic: str, expected_message: str
) -> None:
    source = tmp_path / f"style-{semantic}.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    patch = next(item for item in plan.patches if item.kind == PatchKind.COPY_STYLE)

    workbook = load_workbook(source)
    worksheet = workbook[patch.sheet]
    target = worksheet[patch.cell]
    if semantic == "protection":
        target.protection = Protection(locked=False)
    elif semantic == "number_format":
        target.number_format = "0.00%"
    elif semantic == "quote_prefix":
        target.quotePrefix = not target.quotePrefix
    else:
        target.pivotButton = not target.pivotButton
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source)
    target = workbook[patch.sheet][patch.cell]
    plan.source_sha256 = sha256_file(source)
    patch.precondition.cell_fingerprint = cell_fingerprint(target)
    workbook.close()

    with pytest.raises(PatchValidationError, match=expected_message):
        ooxml_patch._verify_preconditions(source, plan, [patch], None)
