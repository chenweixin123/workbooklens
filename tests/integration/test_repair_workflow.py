from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from lxml import etree
from openpyxl import Workbook, load_workbook

from workbooklens.demo import run_demo
from workbooklens.demo.workflow import DemoOutput, generate_demo_workbook
from workbooklens.exceptions import PatchValidationError, StalePlanError, UsageError
from workbooklens.models import PatchPlan, PatchResult
from workbooklens.repair import apply_patch_plan, build_patch_plan
from workbooklens.repair.ooxml_patch import patch_ooxml_package
from workbooklens.repair.planning import load_patch_plan
from workbooklens.scanner import scan_workbook
from workbooklens.utils import sha256_file


@pytest.fixture(scope="module")
def completed_demo(tmp_path_factory: pytest.TempPathFactory) -> DemoOutput:
    return run_demo(tmp_path_factory.mktemp("complete-demo"))


def _plan(path: Path) -> PatchPlan:
    config = {"keys": [{"sheet": "Sales", "range": "A2:A22", "ignore_blank": True}]}
    return build_patch_plan(scan_workbook(path, config=config))


def test_demo_applies_all_five_patch_types_and_reopens(completed_demo: DemoOutput) -> None:
    plan = PatchPlan.model_validate_json(completed_demo.repair_plan.read_text(encoding="utf-8"))
    report_path = completed_demo.directory / "apply-report.json"
    result = PatchResult.model_validate_json(report_path.read_text(encoding="utf-8"))
    assert len(plan.patches) == 5
    assert plan.findings
    assert {finding.id for finding in plan.findings} == set(plan.finding_ids)
    assert all(finding.evidence.summary for finding in plan.findings)
    assert len(result.applied_patch_ids) == 5
    assert result.source_sha256 == sha256_file(completed_demo.before_workbook)
    assert result.output_sha256 == sha256_file(completed_demo.after_workbook)
    workbook = load_workbook(completed_demo.after_workbook, read_only=True, data_only=False)
    assert workbook["Sales"]["D8"].value == "=B8*C8"
    assert workbook["Sales"]["E12"].value == "=D12*0.08"
    assert workbook["Sales"]["B10"].value == 12
    assert workbook["Summary"]["B10"].value == "=SUM(B2:B9)"
    workbook.close()


def test_only_manifested_parts_change_and_chart_is_byte_identical(
    completed_demo: DemoOutput,
) -> None:
    result = PatchResult.model_validate_json(
        (completed_demo.directory / "apply-report.json").read_text(encoding="utf-8")
    )
    changed = {change.part for change in result.package_changes}
    assert changed == {
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
    }
    with (
        zipfile.ZipFile(completed_demo.before_workbook) as before_archive,
        zipfile.ZipFile(completed_demo.after_workbook) as after_archive,
    ):
        assert before_archive.namelist() == after_archive.namelist()
        chart_parts = [name for name in before_archive.namelist() if name.startswith("xl/charts/")]
        assert chart_parts
        for name in before_archive.namelist():
            if name not in changed:
                assert before_archive.read(name) == after_archive.read(name), name


def test_formula_repairs_remove_cache_and_request_full_recalculation(
    completed_demo: DemoOutput,
) -> None:
    namespaces = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(completed_demo.after_workbook) as archive:
        workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
        calc = workbook_root.find("x:calcPr", namespaces)
        assert calc is not None
        assert calc.get("calcMode") == "auto"
        assert calc.get("fullCalcOnLoad") == "1"
        assert calc.get("forceFullCalc") == "1"
        sheet = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        cell = sheet.find(".//x:c[@r='D8']", namespaces)
        assert cell is not None
        assert cell.find("x:f", namespaces) is not None
        assert cell.find("x:v", namespaces) is None


def test_stale_source_hash_fails_without_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    workbook = load_workbook(source)
    workbook["Sales"]["B2"] = 777
    workbook.save(source)
    workbook.close()
    output = tmp_path / "fixed.xlsx"
    with pytest.raises(StalePlanError, match="source hash"):
        apply_patch_plan(source, plan, output, safe_only=True)
    assert not output.exists()


def test_cell_precondition_fails_even_when_plan_hash_is_rebound(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    target_patch = plan.patches[0]
    workbook = load_workbook(source)
    workbook[target_patch.sheet][target_patch.cell] = "changed"
    workbook.save(source)
    workbook.close()
    plan.source_sha256 = sha256_file(source)
    output = tmp_path / "fixed.xlsx"
    with pytest.raises(StalePlanError, match="Cell precondition"):
        apply_patch_plan(
            source,
            plan,
            output,
            selected_ids={target_patch.id},
        )
    assert not output.exists()


def test_selection_and_output_safety_fail_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    with pytest.raises(UsageError, match="Select at least one"):
        patch_ooxml_package(source, plan, tmp_path / "none.xlsx")
    with pytest.raises(UsageError, match="Unknown patch IDs"):
        patch_ooxml_package(
            source,
            plan,
            tmp_path / "unknown.xlsx",
            selected_ids={"patch-does-not-exist"},
        )
    with pytest.raises(UsageError, match="mutually exclusive"):
        patch_ooxml_package(
            source,
            plan,
            tmp_path / "conflicting-selection.xlsx",
            selected_ids={plan.patches[0].id},
            safe_only=True,
        )
    existing = tmp_path / "existing.xlsx"
    existing.write_bytes(b"do not overwrite")
    with pytest.raises(UsageError, match="already exists"):
        patch_ooxml_package(source, plan, existing, safe_only=True)
    assert existing.read_bytes() == b"do not overwrite"
    with pytest.raises(UsageError, match="source workbook is never overwritten"):
        patch_ooxml_package(source, plan, source, safe_only=True)


def test_patch_plan_input_is_size_bounded(tmp_path: Path) -> None:
    plan_path = tmp_path / "oversized-plan.json"
    plan_path.write_bytes(b"x" * 11)
    with pytest.raises(UsageError, match="exceeds"):
        load_patch_plan(plan_path, max_bytes=10)


def _mark_formula_range(path: Path, coordinate: str, formula_type: str) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        infos = archive.infolist()
        data = {info.filename: archive.read(info.filename) for info in infos}
    root = etree.fromstring(data["xl/worksheets/sheet1.xml"])
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    formula = root.find(f".//x:c[@r='{coordinate}']/x:f", namespace)
    assert formula is not None
    formula.set("t", formula_type)
    if formula_type == "shared":
        formula.set("si", "0")
    formula.set("ref", "D2:D22")
    data["xl/worksheets/sheet1.xml"] = etree.tostring(
        root, encoding="UTF-8", xml_declaration=True, standalone=True
    )
    temporary = path.with_name("shared.tmp.xlsx")
    with zipfile.ZipFile(temporary, "w") as output:
        for info in infos:
            output.writestr(info, data[info.filename])
    temporary.replace(path)


def test_shared_formula_source_is_never_auto_patched(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet"
    for row in range(2, 23):
        worksheet[f"C{row}"] = row
        worksheet[f"D{row}"] = f"=C{row}*2"
    worksheet["D12"] = None
    path = tmp_path / "shared.xlsx"
    workbook.save(path)
    workbook.close()
    original_plan = build_patch_plan(scan_workbook(path))
    original_patch = next(patch for patch in original_plan.patches if patch.cell == "D12")
    _mark_formula_range(path, "D2", "shared")
    plan = build_patch_plan(scan_workbook(path))
    assert not any(patch.cell == "D12" for patch in plan.patches)

    original_plan.source_sha256 = sha256_file(path)
    with pytest.raises(PatchValidationError, match="intersects an unsupported shared"):
        apply_patch_plan(
            path,
            original_plan,
            tmp_path / "fixed.xlsx",
            selected_ids={original_patch.id},
        )


@pytest.mark.parametrize("formula_type", ["array", "dataTable"])
def test_array_and_data_table_ranges_are_never_auto_patched(
    tmp_path: Path, formula_type: str
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet"
    for row in range(2, 23):
        worksheet[f"C{row}"] = row
        worksheet[f"D{row}"] = f"=C{row}*2"
    worksheet["D12"] = None
    path = tmp_path / f"{formula_type}.xlsx"
    workbook.save(path)
    workbook.close()
    original_plan = build_patch_plan(scan_workbook(path))
    original_patch = next(patch for patch in original_plan.patches if patch.cell == "D12")
    _mark_formula_range(path, "D2", formula_type)

    scan = scan_workbook(path)
    assert not any(patch.cell == "D12" for patch in scan.patches)
    original_plan.source_sha256 = sha256_file(path)
    with pytest.raises(PatchValidationError, match=f"unsupported {formula_type}"):
        apply_patch_plan(
            path,
            original_plan,
            tmp_path / f"{formula_type}-fixed.xlsx",
            selected_ids={original_patch.id},
        )


def test_edited_plan_cannot_inject_structured_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generate_demo_workbook(source)
    plan = _plan(source)
    patch = next(item for item in plan.patches if item.cell == "D8")
    patch.after = "=Table1[Amount]"
    output = tmp_path / "fixed.xlsx"
    with pytest.raises(PatchValidationError, match="structured"):
        apply_patch_plan(source, plan, output, selected_ids={patch.id})
    assert not output.exists()


def test_unknown_custom_parts_remain_byte_identical(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generate_demo_workbook(source)
    custom_xml = b'<audit xmlns="urn:workbooklens:test">preserve me</audit>'
    custom_binary = bytes(range(64))
    with zipfile.ZipFile(source, "a", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("customXml/workbooklens-audit.xml", custom_xml)
        archive.writestr("xl/media/unknown-extension.bin", custom_binary)
    plan = _plan(source)
    output = tmp_path / "fixed.xlsx"
    apply_patch_plan(source, plan, output, safe_only=True)
    with zipfile.ZipFile(output, "r") as archive:
        assert archive.read("customXml/workbooklens-audit.xml") == custom_xml
        assert archive.read("xl/media/unknown-extension.bin") == custom_binary


def test_xlsm_remains_read_only(tmp_path: Path) -> None:
    xlsx = tmp_path / "source.xlsx"
    generate_demo_workbook(xlsx)
    xlsm = tmp_path / "source.xlsm"
    xlsm.write_bytes(xlsx.read_bytes())
    scan = scan_workbook(xlsm)
    assert scan.snapshot.format == "xlsm"
    assert not scan.patches
    assert not any(finding.safe_patch_available for finding in scan.findings)
    plan = build_patch_plan(scan)
    with pytest.raises(UsageError, match="read-only"):
        patch_ooxml_package(xlsm, plan, tmp_path / "fixed.xlsx", safe_only=True)


def test_repair_package_never_calls_openpyxl_save() -> None:
    repair_root = Path(__file__).parents[2] / "src" / "workbooklens" / "repair"
    source = "\n".join(path.read_text(encoding="utf-8") for path in repair_root.glob("*.py"))
    assert ".save(" not in source
