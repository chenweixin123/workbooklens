from __future__ import annotations

import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from workbooklens.repair import apply_patch_plan, build_patch_plan
from workbooklens.scanner import ScanResult, scan_workbook

LAYOUT_RULES = {
    "WL016_TEXT_DISPLAY_RISK",
    "WL017_BORDER_EDGE_INCONSISTENCY",
    "WL018_USED_RANGE_INFLATION",
    "WL020_SAVED_VIEW_OFF_CONTENT",
    "WL021_WHITESPACE_ONLY_TAIL",
}


@dataclass(frozen=True, slots=True)
class LayoutCase:
    name: str
    build: Callable[[Path], dict[str, Any]]
    expected_findings: frozenset[str]
    repaired_rules: frozenset[str]
    remaining_rules: frozenset[str] = frozenset()


def _thin() -> Side:
    return Side(style="thin", color="FF000000")


def _full_border() -> Border:
    thin = _thin()
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _populate_grid(
    worksheet: Any,
    *,
    rows: int,
    columns: int,
    width: float = 10.0,
    height: float = 18.0,
) -> None:
    border = _full_border()
    for row in range(1, rows + 1):
        worksheet.row_dimensions[row].height = height
        for column in range(1, columns + 1):
            cell = worksheet.cell(row, column, f"R{row}C{column}")
            cell.border = border
            if row == 1:
                worksheet.column_dimensions[cell.column_letter].width = width


def _remove_four_shared_edges(worksheet: Any, coordinate: str) -> None:
    cell = worksheet[coordinate]
    row = cell.row
    column = cell.column
    thin = _thin()
    worksheet.cell(row, column).border = Border()
    worksheet.cell(row, column - 1).border = Border(left=thin, top=thin, bottom=thin)
    worksheet.cell(row, column + 1).border = Border(right=thin, top=thin, bottom=thin)
    worksheet.cell(row - 1, column).border = Border(left=thin, right=thin, top=thin)
    worksheet.cell(row + 1, column).border = Border(left=thin, right=thin, bottom=thin)


def _save(workbook: Workbook, path: Path) -> dict[str, Any]:
    workbook.save(path)
    workbook.close()
    return {}


def _build_same_cell_text_border_view(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Combined"
    _populate_grid(worksheet, rows=8, columns=8, width=12.0)
    worksheet.column_dimensions["C"].width = 6.0
    worksheet["C3"] = "A long heading that must wrap while all four missing borders are restored"
    worksheet.row_dimensions[3].height = 15.0
    _remove_four_shared_edges(worksheet, "C3")
    worksheet.sheet_view.zoomScale = 160
    return _save(workbook, path)


def _build_tall_boundary_grid(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "TallGrid"
    _populate_grid(worksheet, rows=28, columns=7, width=10.0)
    worksheet.column_dimensions["G"].width = 8.0
    for row in range(2, 6):
        worksheet.cell(row, 7, "Repeated text at right edge")
        worksheet.cell(row, 7).border = _full_border()
    thin = _thin()
    worksheet["A12"].border = Border(right=thin, top=thin, bottom=thin)
    worksheet["D28"].border = Border(left=thin, right=thin, top=thin)
    worksheet.sheet_view.zoomScale = 115
    return _save(workbook, path)


def _build_frozen_pane_mixed(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Frozen"
    _populate_grid(worksheet, rows=28, columns=7, width=10.0)
    worksheet["A1"] = "Wrapped frozen heading that needs a taller first row"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    worksheet.row_dimensions[1].height = 15.0
    _remove_four_shared_edges(worksheet, "D14")
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 130
    _save(workbook, path)
    return {"pane_xml": _pane_xml(path)}


def _build_split_pane_conservative(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "SplitPane"
    _populate_grid(worksheet, rows=28, columns=7, width=10.0)
    _remove_four_shared_edges(worksheet, "D14")
    worksheet.freeze_panes = "A2"
    assert worksheet.sheet_view.pane is not None
    worksheet.sheet_view.pane.state = "split"
    worksheet.sheet_view.zoomScale = 130
    _save(workbook, path)
    return {"pane_xml": _pane_xml(path)}


def _build_formatting_tail_and_view(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "FormatTail"
    _populate_grid(worksheet, rows=20, columns=7, width=10.0)
    fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    for row in range(50, 54):
        for column in range(16, 20):
            worksheet.cell(row, column).fill = fill
    worksheet.row_dimensions[55].height = 22.0
    worksheet.sheet_view.zoomScale = 140
    return _save(workbook, path)


def _build_whitespace_tail_and_view(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "WhitespaceTail"
    _populate_grid(worksheet, rows=8, columns=13, width=7.0)
    for row in range(9, 12):
        worksheet.cell(row, 1, " ")
        worksheet.row_dimensions[row].height = 30.0
    worksheet["A10"].font = Font(name="Aptos", italic=True)
    worksheet["A11"].alignment = Alignment(horizontal="right")
    worksheet["A11"].protection = Protection(locked=False, hidden=True)
    worksheet.sheet_view.zoomScale = 140
    _save(workbook, path)
    reopened = load_workbook(path, data_only=False, keep_links=False)
    try:
        sheet = reopened["WhitespaceTail"]
        return {
            "style_ids": {coordinate: sheet[coordinate].style_id for coordinate in ("A10", "A11")}
        }
    finally:
        reopened.close()


def _build_merged_title_perimeter_view(path: Path) -> dict[str, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "MergedTitle"
    _populate_grid(worksheet, rows=12, columns=9, width=9.0)
    worksheet.merge_cells("A1:I1")
    worksheet["A1"] = (
        "A deliberately long merged report title that should remain fully visible after repair " * 3
    )
    worksheet["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 15.0
    thin = _thin()
    worksheet["A6"].border = Border(right=thin, top=thin, bottom=thin)
    worksheet["I6"].border = Border(left=thin, top=thin, bottom=thin)
    worksheet["E12"].border = Border(left=thin, right=thin, top=thin)
    worksheet.sheet_view.topLeftCell = "D5"
    worksheet.sheet_view.zoomScale = 130
    return _save(workbook, path)


def _pane_xml(path: Path) -> bytes:
    with zipfile.ZipFile(path) as archive:
        root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    panes = [element for element in root.iter() if etree.QName(element).localname == "pane"]
    assert len(panes) == 1
    return etree.tostring(panes[0])


CASES = (
    LayoutCase(
        name="same-cell-text-border-view",
        build=_build_same_cell_text_border_view,
        expected_findings=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
        repaired_rules=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
    ),
    LayoutCase(
        name="tall-boundary-grid",
        build=_build_tall_boundary_grid,
        expected_findings=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
        repaired_rules=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
    ),
    LayoutCase(
        name="frozen-pane-mixed",
        build=_build_frozen_pane_mixed,
        expected_findings=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
        repaired_rules=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
    ),
    LayoutCase(
        name="split-pane-conservative",
        build=_build_split_pane_conservative,
        expected_findings=frozenset(
            {
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
        repaired_rules=frozenset({"WL017_BORDER_EDGE_INCONSISTENCY"}),
        remaining_rules=frozenset({"WL020_SAVED_VIEW_OFF_CONTENT"}),
    ),
    LayoutCase(
        name="formatting-tail-and-view",
        build=_build_formatting_tail_and_view,
        expected_findings=frozenset({"WL018_USED_RANGE_INFLATION", "WL020_SAVED_VIEW_OFF_CONTENT"}),
        repaired_rules=frozenset({"WL018_USED_RANGE_INFLATION", "WL020_SAVED_VIEW_OFF_CONTENT"}),
    ),
    LayoutCase(
        name="whitespace-tail-and-view",
        build=_build_whitespace_tail_and_view,
        expected_findings=frozenset({"WL020_SAVED_VIEW_OFF_CONTENT", "WL021_WHITESPACE_ONLY_TAIL"}),
        repaired_rules=frozenset({"WL020_SAVED_VIEW_OFF_CONTENT", "WL021_WHITESPACE_ONLY_TAIL"}),
    ),
    LayoutCase(
        name="merged-title-perimeter-view",
        build=_build_merged_title_perimeter_view,
        expected_findings=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
        repaired_rules=frozenset(
            {
                "WL016_TEXT_DISPLAY_RISK",
                "WL017_BORDER_EDGE_INCONSISTENCY",
                "WL020_SAVED_VIEW_OFF_CONTENT",
            }
        ),
    ),
)


def _rule_ids(scan: ScanResult) -> set[str]:
    return {finding.rule_id for finding in scan.findings}


def _selected_patch_ids(scan: ScanResult, rules: frozenset[str]) -> set[str]:
    return {
        patch_id
        for finding in scan.findings
        if finding.rule_id in rules
        for patch_id in finding.patch_ids
    }


@pytest.mark.integration
@pytest.mark.parametrize("case", CASES, ids=lambda case: case.name)
def test_layout_adversarial_workbook_repairs_are_composable(
    tmp_path: Path,
    case: LayoutCase,
) -> None:
    source = tmp_path / f"{case.name}.xlsx"
    output = tmp_path / f"{case.name}-fixed.xlsx"
    state = case.build(source)

    before = scan_workbook(source)
    assert case.expected_findings <= _rule_ids(before)
    selected_ids = _selected_patch_ids(before, case.repaired_rules)
    assert selected_ids
    if "WL020_SAVED_VIEW_OFF_CONTENT" in case.remaining_rules:
        view_findings = [
            finding
            for finding in before.findings
            if finding.rule_id == "WL020_SAVED_VIEW_OFF_CONTENT"
        ]
        assert view_findings and all(not finding.patch_ids for finding in view_findings)

    result = apply_patch_plan(
        source,
        build_patch_plan(before),
        output,
        selected_ids=selected_ids,
        accept_layout_risk=True,
    )

    assert set(result.applied_patch_ids) >= selected_ids
    after = scan_workbook(output)
    after_layout_rules = _rule_ids(after) & LAYOUT_RULES
    assert not (after_layout_rules & case.repaired_rules)
    assert case.remaining_rules <= after_layout_rules
    assert not _selected_patch_ids(after, case.repaired_rules)

    if case.name in {"frozen-pane-mixed", "split-pane-conservative"}:
        assert _pane_xml(output) == state["pane_xml"]
    if case.name == "formatting-tail-and-view":
        repaired = load_workbook(output, data_only=False, keep_links=False)
        try:
            worksheet = repaired["FormatTail"]
            assert worksheet.calculate_dimension() == "A1:G20"
            assert worksheet.row_dimensions[55].height == 22.0
        finally:
            repaired.close()
    if case.name == "whitespace-tail-and-view":
        repaired = load_workbook(output, data_only=False, keep_links=False)
        try:
            worksheet = repaired["WhitespaceTail"]
            assert (9, 1) not in worksheet._cells
            assert worksheet["A10"].value is None
            assert worksheet["A10"].style_id == state["style_ids"]["A10"]
            assert worksheet["A10"].font.italic is True
            assert worksheet["A11"].value is None
            assert worksheet["A11"].style_id == state["style_ids"]["A11"]
            assert worksheet["A11"].alignment.horizontal == "right"
            assert worksheet["A11"].protection.locked is False
            assert worksheet["A11"].protection.hidden is True
            assert all(worksheet.row_dimensions[row].height == 30.0 for row in range(9, 12))
        finally:
            repaired.close()
    if case.name == "merged-title-perimeter-view":
        repaired = load_workbook(output, data_only=False, keep_links=False)
        try:
            worksheet = repaired["MergedTitle"]
            assert [str(cell_range) for cell_range in worksheet.merged_cells.ranges] == ["A1:I1"]
        finally:
            repaired.close()


def test_adversarial_case_names_are_unique_and_count_is_bounded() -> None:
    names = [case.name for case in CASES]
    assert 5 <= len(names) <= 8
    assert len(names) == len(set(names))
