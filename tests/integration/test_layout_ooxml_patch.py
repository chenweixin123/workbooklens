from __future__ import annotations

import zipfile
from collections.abc import Callable
from pathlib import Path

import pytest
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.pagebreak import Break

from workbooklens.exceptions import PatchValidationError, StalePlanError
from workbooklens.layout import (
    FormattingTail,
    WhitespaceTail,
    column_layout_fingerprint,
    row_layout_fingerprint,
    sheet_view_fingerprint,
    tail_layout_fingerprint,
    whitespace_tail_layout_fingerprint,
)
from workbooklens.models import (
    PatchKind,
    PatchOperation,
    PatchPlan,
    PatchPrecondition,
    PatchRisk,
)
from workbooklens.repair.engine import apply_patch_plan
from workbooklens.repair.ooxml_patch import patch_ooxml_package
from workbooklens.repair.planning import build_patch_plan
from workbooklens.scanner import scan_workbook
from workbooklens.snapshot import cell_fingerprint
from workbooklens.utils import sha256_file

CUSTOM_XML = b'<audit xmlns="urn:workbooklens:test">preserve me</audit>'
CUSTOM_BINARY = bytes(range(64))


def _layout_patch(
    worksheet: object,
    *,
    patch_id: str,
    kind: PatchKind,
    cell: str,
    after: object,
    source_cell: str | None = None,
    layout_fingerprint: str | None = None,
    atomic_group: str | None = None,
) -> PatchOperation:
    return PatchOperation(
        id=patch_id,
        kind=kind,
        sheet="Sheet1",
        cell=cell,
        before=None,
        after=after,
        source_cell=source_cell,
        confidence=0.99,
        safe=False,
        risk=PatchRisk.LAYOUT_REVIEW,
        description=f"integration fixture for {kind.value}",
        precondition=PatchPrecondition(
            cell_fingerprint=cell_fingerprint(worksheet[cell]),  # type: ignore[index]
            layout_fingerprint=layout_fingerprint,
        ),
        atomic_group=atomic_group,
    )


def _plan(source: Path, patches: list[PatchOperation]) -> PatchPlan:
    return PatchPlan(
        tool_version="2.1.0",
        source_name=source.name,
        source_sha256=sha256_file(source),
        patches=patches,
    )


def _column_elements(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    return [
        dict(element.attrib) for element in root.iter() if etree.QName(element).localname == "col"
    ]


def _rewrite_xml_part(
    path: Path,
    part: str,
    mutate: Callable[[etree._Element], None],
) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        infos = archive.infolist()
        data = {info.filename: archive.read(info.filename) for info in infos}
    root = etree.fromstring(data[part])
    mutate(root)
    data[part] = etree.tostring(
        root,
        encoding="UTF-8",
        xml_declaration=True,
        standalone=True,
    )
    temporary = path.with_name(f"{path.stem}.rewrite{path.suffix}")
    with zipfile.ZipFile(temporary, "w") as output:
        for info in infos:
            output.writestr(info, data[info.filename])
    temporary.replace(path)


def _formatting_tail_plan(source: Path, *, empty_rows: tuple[int, ...] = (60,)) -> PatchPlan:
    workbook = load_workbook(source, data_only=False, keep_links=False)
    worksheet = workbook["Sheet1"]
    tail = FormattingTail(
        cell_coordinates=("Z50",),
        cell_ranges=("Z50",),
        empty_rows=empty_rows,
        row_ranges=tuple(str(row) for row in empty_rows),
        styled_cell_count=1,
        content_min_row=1,
        content_min_column=1,
        content_max_row=1,
        content_max_column=1,
        observed_max_row=max((50, *empty_rows)),
        observed_max_column=26,
    )
    patch = _layout_patch(
        worksheet,
        patch_id="clear-tail",
        kind=PatchKind.CLEAR_FORMATTING_TAIL,
        cell="Z50",
        after={
            "cells": ["Z50"],
            "empty_rows": list(empty_rows),
            "expected_dimension": "A1:Z50",
            "result_dimension": "A1",
        },
        layout_fingerprint=tail_layout_fingerprint(worksheet, tail),
    )
    workbook.close()
    return _plan(source, [patch])


def _make_formatting_tail_fixture(
    path: Path,
    *,
    cross_sheet_formula: bool = False,
    page_break: str | None = None,
    create_empty_row: bool = True,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "content"
    worksheet["Z50"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    if create_empty_row:
        worksheet.row_dimensions[60].height = 22.0
    if cross_sheet_formula:
        other = workbook.create_sheet("Other")
        other["A1"] = "='Sheet1'!Z50"
    if page_break == "rowBreaks":
        worksheet.row_breaks.append(Break(id=60))
    elif page_break == "colBreaks":
        worksheet.col_breaks.append(Break(id=26))
    workbook.save(path)
    workbook.close()


def _make_layout_fixture(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet.append(["Heading", "Long heading", 123, "Target", "Source"])
    worksheet.append(["row two", "Second heading", 456, "x", "y"])

    grouped = worksheet.column_dimensions["B"]
    grouped.min = 2
    grouped.max = 4
    grouped.width = 9.0
    grouped.bestFit = True
    grouped.outlineLevel = 1
    worksheet.row_dimensions[1].height = 15.0
    worksheet.sheet_view.topLeftCell = "D4"
    worksheet.sheet_view.zoomScale = 130

    red_thin = Side(style="thin", color="FFFF0000")
    worksheet["E1"].border = Border(left=red_thin)
    worksheet["Z50"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    worksheet.row_dimensions[60].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    workbook.save(path)
    workbook.close()

    with zipfile.ZipFile(path, "a", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("customXml/workbooklens-audit.xml", CUSTOM_XML)
        archive.writestr("xl/media/unknown-extension.bin", CUSTOM_BINARY)


def test_all_layout_patch_kinds_apply_and_preserve_package_details(tmp_path: Path) -> None:
    source = tmp_path / "layout-source.xlsx"
    output = tmp_path / "layout-fixed.xlsx"
    _make_layout_fixture(source)

    workbook = load_workbook(source, data_only=False, keep_links=False)
    worksheet = workbook["Sheet1"]
    tail = FormattingTail(
        cell_coordinates=("Z50",),
        cell_ranges=("Z50",),
        empty_rows=(60,),
        row_ranges=("60",),
        styled_cell_count=1,
        content_min_row=1,
        content_min_column=1,
        content_max_row=2,
        content_max_column=5,
        observed_max_row=60,
        observed_max_column=26,
    )
    patches = [
        _layout_patch(
            worksheet,
            patch_id="column-width",
            kind=PatchKind.SET_COLUMN_WIDTH,
            cell="C1",
            after=22.5,
            layout_fingerprint=column_layout_fingerprint(worksheet, 3),
        ),
        _layout_patch(
            worksheet,
            patch_id="row-height",
            kind=PatchKind.SET_ROW_HEIGHT,
            cell="A1",
            after=30.0,
            layout_fingerprint=row_layout_fingerprint(worksheet, 1),
        ),
        _layout_patch(
            worksheet,
            patch_id="wrap-one",
            kind=PatchKind.SET_WRAP_TEXT,
            cell="B1",
            after=True,
        ),
        _layout_patch(
            worksheet,
            patch_id="wrap-two",
            kind=PatchKind.SET_WRAP_TEXT,
            cell="B2",
            after=True,
        ),
        _layout_patch(
            worksheet,
            patch_id="shrink",
            kind=PatchKind.SET_SHRINK_TO_FIT,
            cell="A1",
            after=True,
        ),
        _layout_patch(
            worksheet,
            patch_id="set-text",
            kind=PatchKind.SET_TEXT,
            cell="C1",
            after="00123",
        ),
        _layout_patch(
            worksheet,
            patch_id="sheet-view",
            kind=PatchKind.SET_SHEET_VIEW,
            cell="A1",
            after={"top_left_cell": "A1", "zoom_scale": 85},
            layout_fingerprint=sheet_view_fingerprint(worksheet),
        ),
        _layout_patch(
            worksheet,
            patch_id="copy-border",
            kind=PatchKind.COPY_BORDER,
            cell="D1",
            source_cell="E1",
            after={"target_edge": "right", "source_edge": "left"},
        ),
        _layout_patch(
            worksheet,
            patch_id="clear-tail",
            kind=PatchKind.CLEAR_FORMATTING_TAIL,
            cell="Z50",
            after={
                "cells": ["Z50"],
                "empty_rows": [60],
                "expected_dimension": "A1:Z50",
                "result_dimension": "A1:E2",
            },
            layout_fingerprint=tail_layout_fingerprint(worksheet, tail),
        ),
    ]
    workbook.close()
    plan = _plan(source, patches)

    result, selected = patch_ooxml_package(
        source,
        plan,
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
        canonical_plan=plan,
    )

    assert {patch.kind for patch in selected} == {
        PatchKind.SET_COLUMN_WIDTH,
        PatchKind.SET_ROW_HEIGHT,
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_SHRINK_TO_FIT,
        PatchKind.SET_TEXT,
        PatchKind.SET_SHEET_VIEW,
        PatchKind.COPY_BORDER,
        PatchKind.CLEAR_FORMATTING_TAIL,
    }
    assert result.formula_changed
    assert {change.part for change in result.changes} == {
        "xl/styles.xml",
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
    }

    columns = _column_elements(output)
    assert [(item["min"], item["max"]) for item in columns] == [
        ("2", "2"),
        ("3", "3"),
        ("4", "4"),
    ]
    assert columns[1]["width"] == "22.5"
    for key in ("bestFit", "outlineLevel"):
        assert columns[0][key] == columns[2][key]

    repaired = load_workbook(output, data_only=False, keep_links=False)
    repaired_sheet = repaired["Sheet1"]
    assert repaired_sheet.row_dimensions[1].height == 30.0
    assert repaired_sheet["B1"].alignment.wrap_text
    assert repaired_sheet["B2"].alignment.wrap_text
    assert repaired_sheet["B1"].style_id == repaired_sheet["B2"].style_id
    assert repaired_sheet["A1"].alignment.shrink_to_fit
    assert repaired_sheet["C1"].value == "00123"
    assert repaired_sheet["C1"].data_type == "s"
    assert repaired_sheet["C1"].number_format == "@"
    assert repaired_sheet.sheet_view.topLeftCell == "A1"
    assert repaired_sheet.sheet_view.zoomScale == 85
    assert repaired_sheet["D1"].border.right == repaired_sheet["E1"].border.left
    assert (50, 26) not in repaired_sheet._cells
    assert 60 not in repaired_sheet.row_dimensions
    assert repaired_sheet.calculate_dimension() == "A1:E2"
    repaired.close()

    with zipfile.ZipFile(output) as archive:
        assert archive.read("customXml/workbooklens-audit.xml") == CUSTOM_XML
        assert archive.read("xl/media/unknown-extension.bin") == CUSTOM_BINARY
        workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
        calc = next(
            (element for element in workbook_root if etree.QName(element).localname == "calcPr"),
            None,
        )
        assert calc is not None
        assert calc.get("calcMode") == "auto"
        assert calc.get("fullCalcOnLoad") == "1"
        assert calc.get("forceFullCalc") == "1"


def test_layout_fingerprint_change_rejects_stale_plan(tmp_path: Path) -> None:
    source = tmp_path / "fingerprint-source.xlsx"
    output = tmp_path / "fingerprint-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["B1"] = "value"
    worksheet.column_dimensions["B"].width = 10.0
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source)
    worksheet = workbook["Sheet1"]
    patch = _layout_patch(
        worksheet,
        patch_id="column-width",
        kind=PatchKind.SET_COLUMN_WIDTH,
        cell="B1",
        after=18.0,
        layout_fingerprint=column_layout_fingerprint(worksheet, 2),
    )
    workbook.close()
    plan = _plan(source, [patch])

    workbook = load_workbook(source)
    worksheet = workbook["Sheet1"]
    worksheet.column_dimensions["B"].width = 11.0
    workbook.save(source)
    workbook.close()
    plan.source_sha256 = sha256_file(source)

    with pytest.raises(StalePlanError, match="Layout precondition failed"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_zoom_only_sheet_view_patch_preserves_frozen_pane_xml(tmp_path: Path) -> None:
    source = tmp_path / "frozen-view-source.xlsx"
    output = tmp_path / "frozen-view-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "heading"
    worksheet.freeze_panes = "B2"
    worksheet.sheet_view.zoomScale = 115
    workbook.save(source)
    workbook.close()

    with zipfile.ZipFile(source) as archive:
        source_root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    source_pane = next(
        element for element in source_root.iter() if etree.QName(element).localname == "pane"
    )
    source_pane_xml = etree.tostring(source_pane)

    workbook = load_workbook(source)
    worksheet = workbook["Sheet1"]
    patch = _layout_patch(
        worksheet,
        patch_id="frozen-zoom-only",
        kind=PatchKind.SET_SHEET_VIEW,
        cell="A1",
        after={"zoom_scale": 65},
        layout_fingerprint=sheet_view_fingerprint(worksheet),
    )
    workbook.close()

    patch_ooxml_package(
        source,
        _plan(source, [patch]),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
        canonical_plan=_plan(source, [patch]),
    )

    with zipfile.ZipFile(output) as archive:
        output_root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    output_pane = next(
        element for element in output_root.iter() if etree.QName(element).localname == "pane"
    )
    assert etree.tostring(output_pane) == source_pane_xml
    repaired = load_workbook(output)
    repaired_sheet = repaired["Sheet1"]
    assert repaired_sheet.freeze_panes == "B2"
    assert repaired_sheet.sheet_view.zoomScale == 65
    repaired.close()


def test_sheet_view_origin_patch_refuses_a_frozen_pane(tmp_path: Path) -> None:
    source = tmp_path / "frozen-origin-source.xlsx"
    output = tmp_path / "frozen-origin-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "heading"
    worksheet.freeze_panes = "A2"
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source)
    worksheet = workbook["Sheet1"]
    patch = _layout_patch(
        worksheet,
        patch_id="frozen-origin",
        kind=PatchKind.SET_SHEET_VIEW,
        cell="A1",
        after={"top_left_cell": "A1"},
        layout_fingerprint=sheet_view_fingerprint(worksheet),
    )
    workbook.close()
    plan = _plan(source, [patch])

    with pytest.raises(PatchValidationError, match="origin repair refuses"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_formatting_tail_rejects_commented_style_only_cell(tmp_path: Path) -> None:
    source = tmp_path / "tail-source.xlsx"
    output = tmp_path / "tail-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "content"
    worksheet["Z50"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    worksheet["Z50"].comment = Comment("retain this note", "WorkbookLens test")
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source)
    worksheet = workbook["Sheet1"]
    tail = FormattingTail(
        cell_coordinates=("Z50",),
        cell_ranges=("Z50",),
        empty_rows=(),
        row_ranges=(),
        styled_cell_count=1,
        content_min_row=1,
        content_min_column=1,
        content_max_row=1,
        content_max_column=1,
        observed_max_row=50,
        observed_max_column=26,
    )
    patch = _layout_patch(
        worksheet,
        patch_id="clear-tail",
        kind=PatchKind.CLEAR_FORMATTING_TAIL,
        cell="Z50",
        after={
            "cells": ["Z50"],
            "empty_rows": [],
            "expected_dimension": "A1:Z50",
            "result_dimension": "A1",
        },
        layout_fingerprint=tail_layout_fingerprint(worksheet, tail),
    )
    workbook.close()
    plan = _plan(source, [patch])

    with pytest.raises(PatchValidationError, match="comment or hyperlink"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_formatting_tail_rejects_cross_sheet_formula_reference(tmp_path: Path) -> None:
    source = tmp_path / "tail-cross-sheet-formula.xlsx"
    output = tmp_path / "tail-cross-sheet-formula-output.xlsx"
    _make_formatting_tail_fixture(source, cross_sheet_formula=True)
    plan = _formatting_tail_plan(source)
    patch = plan.patches[0]

    with pytest.raises(PatchValidationError, match="referenced by a worksheet formula"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


@pytest.mark.parametrize("collection", ["rowBreaks", "colBreaks"])
def test_formatting_tail_rejects_intersecting_page_break(
    tmp_path: Path,
    collection: str,
) -> None:
    source = tmp_path / f"tail-{collection}.xlsx"
    output = tmp_path / f"tail-{collection}-output.xlsx"
    _make_formatting_tail_fixture(source, page_break=collection)
    plan = _formatting_tail_plan(source)
    patch = plan.patches[0]

    with pytest.raises(PatchValidationError, match=rf"intersects {collection}"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_formatting_tail_rejects_namespaced_empty_row_attribute(tmp_path: Path) -> None:
    source = tmp_path / "tail-namespaced-row-attribute.xlsx"
    output = tmp_path / "tail-namespaced-row-attribute-output.xlsx"
    _make_formatting_tail_fixture(source)

    def add_namespaced_empty_row(root: etree._Element) -> None:
        row = next(
            element
            for element in root.iter()
            if etree.QName(element).localname == "row" and element.get("r") == "60"
        )
        row.set("{urn:workbooklens:test}futureFlag", "retain")

    _rewrite_xml_part(source, "xl/worksheets/sheet1.xml", add_namespaced_empty_row)
    plan = _formatting_tail_plan(source)
    patch = plan.patches[0]

    with pytest.raises(PatchValidationError, match="row 60 has unsupported attributes"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_formatting_tail_rejects_custom_height_empty_row(tmp_path: Path) -> None:
    source = tmp_path / "tail-custom-height.xlsx"
    output = tmp_path / "tail-custom-height-output.xlsx"
    _make_formatting_tail_fixture(source)
    plan = _formatting_tail_plan(source)
    patch = plan.patches[0]

    with pytest.raises(PatchValidationError, match="custom-height rows"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()


def test_formatting_tail_cleanup_preserves_custom_height_and_single_cell_dimension(
    tmp_path: Path,
) -> None:
    source = tmp_path / "tail-with-custom-height.xlsx"
    output = tmp_path / "tail-with-custom-height-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "content"
    fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    for row in range(50, 54):
        for column in range(16, 20):
            worksheet.cell(row, column).fill = fill
    worksheet.row_dimensions[55].height = 22.0
    workbook.save(source)
    workbook.close()

    def include_custom_height_row_in_declared_dimension(root: etree._Element) -> None:
        dimension = next(
            element for element in root if etree.QName(element).localname == "dimension"
        )
        dimension.set("ref", "A1:S55")

    _rewrite_xml_part(
        source,
        "xl/worksheets/sheet1.xml",
        include_custom_height_row_in_declared_dimension,
    )

    scan = scan_workbook(source)
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL)
    assert patch.after["empty_rows"] == []
    assert patch.after["result_dimension"] == "A1"

    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids=[patch.id],
        accept_layout_risk=True,
    )

    repaired = load_workbook(output, data_only=False, keep_links=False)
    repaired_sheet = repaired["Sheet1"]
    assert repaired_sheet.row_dimensions[55].height == 22.0
    assert repaired_sheet.calculate_dimension() == "A1:A1"
    repaired.close()
    with zipfile.ZipFile(output) as archive:
        root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    dimension = next(element for element in root if etree.QName(element).localname == "dimension")
    assert dimension.get("ref") == "A1"


def test_formatting_tail_cleanup_preserves_large_merged_dimension(tmp_path: Path) -> None:
    source = tmp_path / "tail-with-large-merge.xlsx"
    output = tmp_path / "tail-with-large-merge-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Merged report title"
    worksheet.merge_cells("A1:H20")
    fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    for row in range(50, 54):
        for column in range(16, 20):
            worksheet.cell(row, column).fill = fill
    workbook.save(source)
    workbook.close()

    scan = scan_workbook(source)
    findings = [
        finding for finding in scan.findings if finding.rule_id == "WL018_USED_RANGE_INFLATION"
    ]
    assert len(findings) == 1
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL)
    assert patch.after["result_dimension"] == "A1:H20"

    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids=[patch.id],
        accept_layout_risk=True,
    )

    repaired = load_workbook(output, data_only=False, keep_links=False)
    repaired_sheet = repaired["Sheet1"]
    assert repaired_sheet.calculate_dimension() == "A1:H20"
    assert [str(cell_range) for cell_range in repaired_sheet.merged_cells.ranges] == ["A1:H20"]
    repaired.close()


def test_whitespace_tail_cleanup_preserves_nondefault_cell_semantics(tmp_path: Path) -> None:
    source = tmp_path / "whitespace-style-source.xlsx"
    output = tmp_path / "whitespace-style-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "content"
    worksheet["A9"] = " "
    worksheet["A10"] = " "
    worksheet["A10"].font = Font(bold=True, name="Aptos")
    worksheet["A11"] = " "
    worksheet["A11"].alignment = Alignment(horizontal="center")
    worksheet["A11"].protection = Protection(locked=False, hidden=True)
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source, data_only=False, keep_links=False)
    worksheet = workbook["Sheet1"]
    preserve_style_ids = {
        "A10": worksheet["A10"].style_id,
        "A11": worksheet["A11"].style_id,
    }
    tail = WhitespaceTail(
        cell_coordinates=("A9", "A10", "A11"),
        cell_ranges=("A9:A11",),
        row_ranges=(),
        content_min_row=1,
        content_min_column=1,
        content_max_row=1,
        content_max_column=1,
        observed_dimension="A1:A11",
        result_dimension="A1:A11",
        preserve_style_ids=tuple(sorted(preserve_style_ids.items())),
    )
    patch = _layout_patch(
        worksheet,
        patch_id="clear-whitespace-preserve-styles",
        kind=PatchKind.REMOVE_WHITESPACE_TAIL_CELLS,
        cell="A1",
        after={
            "cells": ["A9", "A10", "A11"],
            "expected_dimension": "A1:A11",
            "preserve_style_ids": preserve_style_ids,
            "result_dimension": "A1:A11",
            "preserve_row_dimensions": True,
        },
        layout_fingerprint=whitespace_tail_layout_fingerprint(worksheet, tail),
    )
    workbook.close()
    plan = _plan(source, [patch])

    patch_ooxml_package(
        source,
        plan,
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
        canonical_plan=plan,
    )

    repaired = load_workbook(output, data_only=False, keep_links=False)
    repaired_sheet = repaired["Sheet1"]
    assert (9, 1) not in repaired_sheet._cells
    assert repaired_sheet["A10"].value is None
    assert repaired_sheet["A10"].style_id == preserve_style_ids["A10"]
    assert repaired_sheet["A10"].font.bold is True
    assert repaired_sheet["A11"].value is None
    assert repaired_sheet["A11"].style_id == preserve_style_ids["A11"]
    assert repaired_sheet["A11"].alignment.horizontal == "center"
    assert repaired_sheet["A11"].protection.locked is False
    assert repaired_sheet["A11"].protection.hidden is True
    assert repaired_sheet.calculate_dimension() == "A1:A11"
    repaired.close()

    with zipfile.ZipFile(output) as archive:
        root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    cells = {
        element.get("r"): element
        for element in root.iter()
        if etree.QName(element).localname == "c"
    }
    assert "A9" not in cells
    for coordinate in ("A10", "A11"):
        assert cells[coordinate].get("s") == str(preserve_style_ids[coordinate])
        assert cells[coordinate].get("t") is None
        assert len(cells[coordinate]) == 0


def test_scanned_whitespace_tail_preserves_style_semantics_end_to_end(
    tmp_path: Path,
) -> None:
    source = tmp_path / "whitespace-style-scan-source.xlsx"
    output = tmp_path / "whitespace-style-scan-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet["A9"] = " "
    worksheet["A10"] = " "
    worksheet["A10"].font = Font(italic=True)
    worksheet["A11"] = " "
    worksheet["A11"].alignment = Alignment(horizontal="right")
    worksheet["A11"].protection = Protection(locked=False)
    workbook.save(source)
    workbook.close()

    scan = scan_workbook(source)
    patch = next(
        patch for patch in scan.patches if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS
    )
    preserved_style_ids = patch.after["preserve_style_ids"]
    assert set(preserved_style_ids) == {"A10", "A11"}
    assert patch.after["expected_dimension"] == "A1:A11"
    assert patch.after["result_dimension"] == "A1:A11"

    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
    )

    repaired = load_workbook(output, data_only=False, keep_links=False)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    assert (9, 1) not in repaired_sheet._cells
    assert repaired_sheet["A10"].value is None
    assert repaired_sheet["A10"].font.italic is True
    assert repaired_sheet["A11"].value is None
    assert repaired_sheet["A11"].alignment.horizontal == "right"
    assert repaired_sheet["A11"].protection.locked is False
    assert repaired_sheet.calculate_dimension() == "A1:A11"
    repaired.close()
    assert not any(
        finding.rule_id == "WL021_WHITESPACE_ONLY_TAIL"
        for finding in scan_workbook(output).findings
    )


def test_whitespace_tail_cleanup_rejects_unrecognized_nondefault_style(
    tmp_path: Path,
) -> None:
    source = tmp_path / "whitespace-unrecognized-style.xlsx"
    output = tmp_path / "whitespace-unrecognized-style-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A1"] = "content"
    for row in range(9, 12):
        worksheet.cell(row, 1, " ").protection = Protection(locked=False)
    workbook.save(source)
    workbook.close()

    workbook = load_workbook(source, data_only=False, keep_links=False)
    worksheet = workbook["Sheet1"]
    tail = WhitespaceTail(
        cell_coordinates=("A9", "A10", "A11"),
        cell_ranges=("A9:A11",),
        row_ranges=(),
        content_min_row=1,
        content_min_column=1,
        content_max_row=1,
        content_max_column=1,
        observed_dimension="A1:A11",
        result_dimension="A1",
    )
    patch = _layout_patch(
        worksheet,
        patch_id="unsafe-whitespace-style-removal",
        kind=PatchKind.REMOVE_WHITESPACE_TAIL_CELLS,
        cell="A1",
        after={
            "cells": ["A9", "A10", "A11"],
            "expected_dimension": "A1:A11",
            "preserve_style_ids": {},
            "result_dimension": "A1",
            "preserve_row_dimensions": True,
        },
        layout_fingerprint=whitespace_tail_layout_fingerprint(worksheet, tail),
    )
    workbook.close()
    plan = _plan(source, [patch])

    with pytest.raises(PatchValidationError, match="unauthorized non-default style"):
        patch_ooxml_package(
            source,
            plan,
            output,
            selected_ids={patch.id},
            accept_layout_risk=True,
            canonical_plan=plan,
        )
    assert not output.exists()
