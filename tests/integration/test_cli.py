from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from click import unstyle
from openpyxl import Workbook
from typer.testing import CliRunner

import workbooklens.cli as cli_module
from workbooklens.cli import app
from workbooklens.demo.workflow import generate_demo_workbook
from workbooklens.models import (
    PatchKind,
    PatchOperation,
    PatchPlan,
    PatchPrecondition,
    PatchRisk,
)
from workbooklens.utils import sha256_file

runner = CliRunner()


def test_help_version_and_required_commands() -> None:
    help_result = runner.invoke(app, ["--help"])
    assert help_result.exit_code == 0
    for command in ("scan", "plan", "apply", "diff", "test", "serve", "demo"):
        assert command in help_result.stdout
    version_result = runner.invoke(app, ["--version"])
    assert version_result.exit_code == 0
    assert "WorkbookLens 2.2.0" in version_result.stdout


def test_apply_help_explains_layout_review_opt_in() -> None:
    result = runner.invoke(
        app,
        ["apply", "--help"],
        color=False,
        terminal_width=160,
    )
    help_text = unstyle(result.stdout)

    assert result.exit_code == 0
    assert "--accept-layout-risk" in help_text
    assert "--safe-only" in help_text
    assert "layout-review" in help_text


def test_plan_reports_safe_and_layout_review_counts(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    safe_patch = PatchOperation(
        id="safe-patch",
        kind=PatchKind.SET_FORMULA,
        sheet="Sheet1",
        cell="A1",
        after="=1+1",
        confidence=0.99,
        safe=True,
        description="safe test patch",
        precondition=PatchPrecondition(cell_fingerprint="safe-fingerprint"),
    )
    review_patch = PatchOperation(
        id="review-patch",
        kind=PatchKind.SET_COLUMN_WIDTH,
        sheet="Sheet1",
        cell="A1",
        before=8.43,
        after=16.0,
        confidence=0.99,
        safe=False,
        risk=PatchRisk.LAYOUT_REVIEW,
        description="layout test patch",
        precondition=PatchPrecondition(cell_fingerprint="review-fingerprint"),
    )
    patch_plan = PatchPlan(
        tool_version="2.1.0",
        source_name="input.xlsx",
        source_sha256="abc123",
        patches=[safe_patch, review_patch],
    )
    monkeypatch.setattr(cli_module, "scan_workbook", lambda *args, **kwargs: object())
    monkeypatch.setattr(cli_module, "build_patch_plan", lambda _scan: patch_plan)

    result = runner.invoke(
        app,
        ["plan", str(tmp_path / "input.xlsx"), "--out", str(tmp_path / "plan.json")],
    )

    assert result.exit_code == 0, result.stdout
    assert "(1 safe, 1 layout review)" in result.stdout


def test_scan_outputs_and_fail_on_exit_code(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "=#REF!+1"
    path = tmp_path / "broken.xlsx"
    workbook.save(path)
    workbook.close()
    output = tmp_path / "report"
    result = runner.invoke(app, ["scan", str(path), "--out", str(output)])
    assert result.exit_code == 0, result.stdout
    assert {item.name for item in output.iterdir()} == {
        "findings.json",
        "report.html",
        "results.sarif",
        "snapshot.json",
    }
    failing = runner.invoke(
        app,
        ["scan", str(path), "--out", str(tmp_path / "second"), "--fail-on", "error"],
    )
    assert failing.exit_code == 1


def test_plan_apply_and_diff_cli_preserve_source(tmp_path: Path) -> None:
    source = tmp_path / "before.xlsx"
    generate_demo_workbook(source)
    source_hash = sha256_file(source)
    plan = tmp_path / "repair-plan.json"
    planned = runner.invoke(app, ["plan", str(source), "--out", str(plan)])
    assert planned.exit_code == 0, planned.stdout
    fixed = tmp_path / "fixed.xlsx"
    applied = runner.invoke(
        app,
        ["apply", str(source), str(plan), "--out", str(fixed), "--safe-only"],
    )
    assert applied.exit_code == 0, applied.stdout
    assert fixed.exists()
    assert fixed.with_suffix(".xlsx.apply.json").exists()
    assert sha256_file(source) == source_hash
    diff = tmp_path / "diff.html"
    compared = runner.invoke(
        app,
        ["diff", str(source), str(fixed), "--out", str(diff)],
    )
    assert compared.exit_code == 0, compared.stdout
    assert diff.exists() and diff.with_suffix(".json").exists()


def test_apply_requires_selection_with_usage_exit_code(tmp_path: Path) -> None:
    source = tmp_path / "before.xlsx"
    generate_demo_workbook(source)
    plan = tmp_path / "repair-plan.json"
    assert runner.invoke(app, ["plan", str(source), "--out", str(plan)]).exit_code == 0
    result = runner.invoke(
        app,
        ["apply", str(source), str(plan), "--out", str(tmp_path / "fixed.xlsx")],
    )
    assert result.exit_code == 2
    assert "Select at least one" in result.stdout


def test_test_command_returns_zero_and_one(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Data"
    worksheet["A1"] = 1
    path = tmp_path / "data.xlsx"
    workbook.save(path)
    workbook.close()
    passing_config = tmp_path / "pass.yml"
    passing_config.write_text(
        yaml.safe_dump(
            {
                "version": 1,
                "assertions": [
                    {"id": "filled", "type": "nonblank", "sheet": "Data", "range": "A1"}
                ],
            }
        ),
        encoding="utf-8",
    )
    passing = runner.invoke(app, ["test", str(path), "--config", str(passing_config)])
    assert passing.exit_code == 0
    assert "PASS" in passing.stdout
    failing_config = tmp_path / "fail.yml"
    failing_config.write_text(
        yaml.safe_dump(
            {
                "version": 1,
                "assertions": [
                    {"id": "blank", "type": "nonblank", "sheet": "Data", "range": "A1:A2"}
                ],
            }
        ),
        encoding="utf-8",
    )
    failing = runner.invoke(app, ["test", str(path), "--config", str(failing_config)])
    assert failing.exit_code == 1
    assert "FAIL" in failing.stdout


def test_test_command_serializes_suppression_audit(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Data"
    worksheet["A1"] = "=#REF!+1"
    path = tmp_path / "broken.xlsx"
    workbook.save(path)
    workbook.close()
    config = tmp_path / "policy.yml"
    config.write_text(
        yaml.safe_dump(
            {
                "version": 2,
                "workbook": {"max_error_findings": 0},
                "suppressions": [
                    {
                        "id": "accepted-fixture",
                        "reason": "Synthetic fixture is intentionally broken",
                        "rules": ["WL001_BROKEN_REFERENCE"],
                        "sheets": ["Data"],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    output = tmp_path / "test-results.json"
    result = runner.invoke(
        app,
        ["test", str(path), "--config", str(config), "--out", str(output)],
    )
    assert result.exit_code == 0, result.stdout
    assert "1 findings suppressed" in result.stdout
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert payload["finding_policy"]["summary"]["suppressed"] == 1
    assert payload["finding_policy"]["suppressed_findings"][0]["suppression"]["id"] == (
        "accepted-fixture"
    )


def test_unexpected_exception_uses_documented_internal_error_exit_code(
    tmp_path: Path, monkeypatch: object
) -> None:
    def fail_scan(*args: object, **kwargs: object) -> None:
        raise RuntimeError("synthetic failure")

    monkeypatch.setattr(cli_module, "scan_workbook", fail_scan)  # type: ignore[attr-defined]
    result = runner.invoke(
        app,
        ["scan", str(tmp_path / "unused.xlsx"), "--out", str(tmp_path / "report")],
    )
    assert result.exit_code == 10
    assert "Internal error" in result.stdout
    assert "Traceback" not in result.stdout
