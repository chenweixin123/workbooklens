from __future__ import annotations

import copy
import hashlib
import json
import subprocess
import zipfile
from datetime import UTC, datetime
from pathlib import Path

import pytest
from lxml import etree
from openpyxl import Workbook, load_workbook

import workbooklens.conversion as conversion
from workbooklens.conversion import ConversionProvider
from workbooklens.demo.workflow import generate_demo_workbook
from workbooklens.exceptions import UsageError

OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")


def _legacy_workbook(path: Path) -> Path:
    path.write_bytes(OLE_SIGNATURE + b"test payload")
    return path


def _excel_identity(
    process_id: int = 4242,
    *,
    creation_filetime: int = 134000000000000000,
    session_id: int = 7,
    executable_path: str = (r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"),
) -> conversion._ExcelProcessIdentity:
    creation_utc = conversion._filetime_to_creation_utc(creation_filetime)
    normalized_path = conversion._normalize_windows_executable_path(executable_path)
    assert creation_utc is not None
    assert normalized_path is not None
    return conversion._ExcelProcessIdentity(
        process_id=process_id,
        creation_utc=creation_utc,
        creation_filetime=creation_filetime,
        session_id=session_id,
        normalized_executable_path=normalized_path,
    )


def _write_excel_identity(
    path: Path,
    identity: conversion._ExcelProcessIdentity,
) -> None:
    path.write_text(
        json.dumps(
            {
                "process_id": identity.process_id,
                "creation_utc": identity.creation_utc,
                "creation_filetime": str(identity.creation_filetime),
                "session_id": identity.session_id,
                "normalized_executable_path": identity.normalized_executable_path,
            },
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )


def _locale_builtin_workbook(path: Path, existing_format: str | None = None) -> Path:
    workbook = Workbook()
    workbook.active["A1"] = 45992
    workbook.save(path)
    workbook.close()

    temporary = path.with_name(f"{path.stem}-temporary.xlsx")
    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(temporary, "w", allowZip64=True) as target,
    ):
        styles = etree.fromstring(source.read("xl/styles.xml"))
        sheet = etree.fromstring(source.read("xl/worksheets/sheet1.xml"))
        namespace = etree.QName(styles).namespace
        assert namespace is not None
        cell_xfs = styles.find(f"{{{namespace}}}cellXfs")
        assert cell_xfs is not None
        locale_xf = copy.deepcopy(cell_xfs[0])
        locale_xf.set("numFmtId", "57")
        locale_xf.set("applyNumberFormat", "1")
        cell_xfs.append(locale_xf)
        cell_xfs.set("count", str(len(cell_xfs)))
        if existing_format is not None:
            num_fmts = styles.find(f"{{{namespace}}}numFmts")
            assert num_fmts is not None
            etree.SubElement(
                num_fmts,
                f"{{{namespace}}}numFmt",
                numFmtId="164",
                formatCode=existing_format,
            )
            num_fmts.set("count", str(len(num_fmts)))
        cell = next(element for element in sheet.iter() if element.get("r") == "A1")
        cell.set("s", "1")
        modified = {
            "xl/styles.xml": etree.tostring(
                styles, encoding="UTF-8", xml_declaration=True, standalone=True
            ),
            "xl/worksheets/sheet1.xml": etree.tostring(
                sheet, encoding="UTF-8", xml_declaration=True, standalone=True
            ),
        }
        target.comment = source.comment
        for info in source.infolist():
            target.writestr(
                copy.copy(info), modified.get(info.filename, source.read(info.filename))
            )
    temporary.replace(path)
    return path


def test_available_providers_prefer_excel(monkeypatch: pytest.MonkeyPatch) -> None:
    powershell = Path("C:/Windows/System32/WindowsPowerShell/v1.0/powershell.exe")
    libreoffice = Path("C:/Program Files/LibreOffice/program/soffice.exe")
    monkeypatch.setattr(conversion, "_find_powershell", lambda: powershell)
    monkeypatch.setattr(conversion, "_excel_com_registered", lambda: True)
    monkeypatch.setattr(conversion, "_find_libreoffice", lambda: libreoffice)

    providers = conversion.available_conversion_providers()

    assert [(provider.kind, provider.runner) for provider in providers] == [
        ("excel", powershell),
        ("libreoffice", libreoffice),
    ]


def test_conversion_rejects_disguised_non_xls(tmp_path: Path) -> None:
    source = tmp_path / "fake.xls"
    source.write_bytes(b"not an OLE compound file")

    with pytest.raises(UsageError, match="not a recognized binary Excel"):
        conversion.convert_xls_to_xlsx(source, tmp_path / "fake.xlsx")


def test_conversion_explains_when_no_local_provider(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    monkeypatch.setattr(conversion, "available_conversion_providers", lambda: ())

    with pytest.raises(UsageError, match="Microsoft Excel or LibreOffice"):
        conversion.convert_xls_to_xlsx(source, tmp_path / "converted.xlsx")


def test_conversion_falls_back_and_validates_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    excel = ConversionProvider("excel", "Microsoft Excel", Path("powershell.exe"))
    libreoffice = ConversionProvider("libreoffice", "LibreOffice", Path("soffice.exe"))
    calls: list[str] = []
    monkeypatch.setattr(
        conversion,
        "available_conversion_providers",
        lambda: (excel, libreoffice),
    )

    def fake_run(
        provider: ConversionProvider,
        _source: Path,
        target: Path,
        _timeout_seconds: int,
    ) -> tuple[conversion._NumberFormatRecord, ...]:
        calls.append(provider.kind)
        if provider.kind == "excel":
            target.write_bytes(b"partial")
            raise conversion._ProviderFailure("Excel could not open the workbook")
        assert not target.exists()
        generate_demo_workbook(target)
        return ()

    monkeypatch.setattr(conversion, "_run_provider", fake_run)

    result = conversion.convert_xls_to_xlsx(source, output)

    assert result.provider == libreoffice
    assert calls == ["excel", "libreoffice"]
    assert output.read_bytes().startswith(b"PK")


def test_conversion_never_overwrites_or_deletes_output_created_during_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("powershell.exe"))
    monkeypatch.setattr(
        conversion,
        "available_conversion_providers",
        lambda: (provider,),
    )

    def fake_run(
        _provider: ConversionProvider,
        _source: Path,
        staged_output: Path,
        _timeout_seconds: int,
    ) -> tuple[conversion._NumberFormatRecord, ...]:
        generate_demo_workbook(staged_output)
        output.write_bytes(b"external-owner")
        return ()

    monkeypatch.setattr(conversion, "_run_provider", fake_run)

    with pytest.raises(UsageError, match="refused to overwrite or delete"):
        conversion.convert_xls_to_xlsx(source, output)

    assert output.read_bytes() == b"external-owner"


def test_conversion_fails_closed_when_atomic_publish_is_unsupported(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("powershell.exe"))
    monkeypatch.setattr(
        conversion,
        "available_conversion_providers",
        lambda: (provider,),
    )
    monkeypatch.setattr(
        conversion,
        "_run_provider",
        lambda _provider, _source, staged, _timeout: (
            generate_demo_workbook(staged),
            (),
        )[1],
    )
    monkeypatch.setattr(
        conversion.os,
        "link",
        lambda _source, _target: (_ for _ in ()).throw(OSError("not supported")),
    )

    with pytest.raises(UsageError, match="without overwriting"):
        conversion.convert_xls_to_xlsx(source, output)

    assert not output.exists()


def test_conversion_removes_invalid_provider_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("powershell.exe"))
    monkeypatch.setattr(conversion, "available_conversion_providers", lambda: (provider,))
    monkeypatch.setattr(
        conversion,
        "_run_provider",
        lambda _provider, _source, target, _timeout: target.write_bytes(b"invalid"),
    )

    with pytest.raises(UsageError, match="Every available local converter failed"):
        conversion.convert_xls_to_xlsx(source, output)

    assert not output.exists()


def test_excel_runner_uses_encoded_script_and_environment_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "中文 legacy.xls")
    output = tmp_path / "中文 converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("C:/Windows/powershell.exe"))
    observed: dict[str, object] = {}
    identity = _excel_identity()
    terminated: list[int] = []
    closed: list[int] = []

    class FakeProcess:
        pid = 9001
        returncode: int | None = None

        def __init__(self, command: list[str], **kwargs: object) -> None:
            observed["command"] = command
            observed["environment"] = kwargs["env"]
            environment = kwargs["env"]
            assert isinstance(environment, dict)
            _write_excel_identity(
                Path(str(environment["WORKBOOKLENS_EXCEL_IDENTITY"])),
                identity,
            )

        def poll(self) -> int | None:
            return None

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == 10
            environment = observed["environment"]
            assert isinstance(environment, dict)
            assert Path(str(environment["WORKBOOKLENS_EXCEL_GO"])).read_bytes() == b"go"
            generate_demo_workbook(Path(str(environment["WORKBOOKLENS_XLSX_OUTPUT"])))
            Path(str(environment["WORKBOOKLENS_FORMAT_MAP"])).write_text("[]", encoding="utf-8")
            self.returncode = 0
            return "", ""

        def kill(self) -> None:
            self.returncode = -9

    monkeypatch.setattr(conversion, "_snapshot_excel_process_ids", lambda _runner: frozenset())
    monkeypatch.setattr(
        conversion,
        "_open_verified_excel_process",
        lambda expected, *, missing_is_clean: (
            conversion._OwnedExcelProcess(expected, 88)
            if expected == identity and not missing_is_clean
            else None
        ),
    )
    monkeypatch.setattr(
        conversion,
        "_terminate_excel_process",
        lambda process: terminated.append(process.identity.process_id),
    )
    monkeypatch.setattr(
        conversion,
        "_close_excel_process",
        lambda process: closed.append(process.handle) if process is not None else None,
    )
    monkeypatch.setattr(conversion.subprocess, "Popen", FakeProcess)

    records = conversion._run_excel(provider, source, output, 10)

    command = observed["command"]
    environment = observed["environment"]
    assert isinstance(command, list)
    assert isinstance(environment, dict)
    assert "-EncodedCommand" in command
    assert "-ExecutionPolicy" not in command
    assert str(source) not in command
    assert environment["WORKBOOKLENS_XLS_INPUT"] == str(source)
    staged_output = Path(str(environment["WORKBOOKLENS_XLSX_OUTPUT"]))
    assert staged_output != output
    assert staged_output.name == "converted.xlsx"
    assert output.is_file()
    assert "WORKBOOKLENS_FORMAT_MAP" in environment
    assert "WORKBOOKLENS_EXCEL_IDENTITY" in environment
    assert "WORKBOOKLENS_EXCEL_GO" in environment
    assert records == ()
    assert (
        "$workbook.SaveAs($outputPath, 51, $null, $null, $false, $false, 1, 2, $false, "
        "$null, $null, $true)" in conversion._EXCEL_CONVERSION_SCRIPT
    )
    assert "$usedRange.SpecialCells($cellType)" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "foreach ($cellType in @(2, -4123))" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$areas = $matchingCells.Areas" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$cell = $areaCells.Item($cellIndex)" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "'General', 'G/通用格式', '通用格式'" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$cell.NumberFormatLocal = $format" not in conversion._EXCEL_CONVERSION_SCRIPT
    assert "Get-Process -Name EXCEL" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "GetWindowThreadProcessId" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$ownsExcelProcess = $true" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "if ($ownsExcelProcess)" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "[IO.File]::Exists($excelGoPath)" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$excel.Calculation = -4135" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "$excel.CalculateBeforeSave = $false" in conversion._EXCEL_CONVERSION_SCRIPT
    add_workbook = "$calculationWorkbook = $excel.Workbooks.Add()"
    set_calculation = "$excel.Calculation = -4135"
    close_workbook = "$calculationWorkbook.Close($false)"
    open_source = "$workbook = $excel.Workbooks.Open($inputPath, 0, $true)"
    assert (
        conversion._EXCEL_CONVERSION_SCRIPT.index(add_workbook)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(set_calculation)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(close_workbook)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(open_source)
    )
    write_pid = "[IO.File]::WriteAllText(\n        $excelIdentityTempPath"
    publish_pid = "[IO.File]::Move($excelIdentityTempPath, $excelIdentityPath)"
    wait_for_go = "while (-not [IO.File]::Exists($excelGoPath))"
    assert (
        conversion._EXCEL_CONVERSION_SCRIPT.index(write_pid)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(publish_pid)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(wait_for_go)
        < conversion._EXCEL_CONVERSION_SCRIPT.index(add_workbook)
    )
    assert "[IO.File]::Delete($excelIdentityTempPath)" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "creation_filetime" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "normalized_executable_path" in conversion._EXCEL_CONVERSION_SCRIPT
    assert terminated == [4242]
    assert closed == [88]
    assert "$calculationWorkbook = $null\n    $workbook =" in conversion._EXCEL_CONVERSION_SCRIPT
    assert "if ($null -ne $calculationWorkbook)" in conversion._EXCEL_CONVERSION_SCRIPT


def test_excel_runner_never_overwrites_a_preexisting_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    output.write_bytes(b"user-owned")
    provider = ConversionProvider("excel", "Microsoft Excel", Path("powershell.exe"))
    monkeypatch.setattr(
        conversion.subprocess,
        "Popen",
        lambda *_args, **_kwargs: pytest.fail("PowerShell must not start"),
    )

    with pytest.raises(conversion._ProviderFailure, match="already exists"):
        conversion._run_excel(provider, source, output, 10)

    assert output.read_bytes() == b"user-owned"


def test_excel_operation_timeout_kills_owned_excel_then_wrapper(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("C:/Windows/powershell.exe"))
    taskkill = Path("C:/Windows/System32/taskkill.exe")
    killed: list[int] = []
    terminated: list[int] = []
    closed: list[int] = []
    observed_environment: dict[str, object] = {}
    identity = _excel_identity()

    def fake_subprocess_run(
        command: list[str], **kwargs: object
    ) -> subprocess.CompletedProcess[str]:
        killed.append(int(command[command.index("/PID") + 1]))
        assert command == [str(taskkill), "/PID", str(killed[-1]), "/T", "/F"]
        return subprocess.CompletedProcess(command, 0, "", "")

    class FakeProcess:
        pid = 9001
        returncode: int | None = None

        def __init__(self, _command: list[str], **kwargs: object) -> None:
            environment = kwargs["env"]
            assert isinstance(environment, dict)
            observed_environment.update(environment)
            _write_excel_identity(
                Path(str(environment["WORKBOOKLENS_EXCEL_IDENTITY"])),
                identity,
            )
            self.communicate_calls = 0

        def poll(self) -> int | None:
            return None

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            self.communicate_calls += 1
            if self.communicate_calls == 1:
                assert timeout == 1
                Path(str(observed_environment["WORKBOOKLENS_XLSX_OUTPUT"])).write_bytes(b"partial")
                assert (
                    Path(str(observed_environment["WORKBOOKLENS_EXCEL_GO"])).read_bytes() == b"go"
                )
                raise subprocess.TimeoutExpired("powershell", timeout)
            assert timeout == conversion._PROCESS_DRAIN_TIMEOUT_SECONDS
            self.returncode = -9
            return "", ""

        def kill(self) -> None:
            self.returncode = -9

    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion, "_windows_taskkill_executable", lambda: taskkill)
    monkeypatch.setattr(conversion, "_snapshot_excel_process_ids", lambda _runner: frozenset())
    monkeypatch.setattr(
        conversion,
        "_open_verified_excel_process",
        lambda expected, *, missing_is_clean: conversion._OwnedExcelProcess(expected, 88),
    )
    monkeypatch.setattr(
        conversion,
        "_terminate_excel_process",
        lambda process: terminated.append(process.identity.process_id),
    )
    monkeypatch.setattr(
        conversion,
        "_close_excel_process",
        lambda process: closed.append(process.handle) if process is not None else None,
    )
    monkeypatch.setattr(conversion.subprocess, "run", fake_subprocess_run)
    monkeypatch.setattr(conversion.subprocess, "Popen", FakeProcess)

    with pytest.raises(conversion._ProviderFailure, match="timed out after 1 seconds"):
        conversion._run_excel(provider, source, output, 1)

    assert terminated == [4242]
    assert killed == [9001]
    assert closed == [88]
    assert not output.exists()


@pytest.mark.parametrize("payload", ["4242\n", "0", "-1", "42 42", "2147483648"])
def test_excel_invalid_pid_handshake_kills_only_wrapper(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    payload: str,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("C:/Windows/powershell.exe"))
    taskkill = Path("C:/Windows/System32/taskkill.exe")
    killed: list[int] = []
    observed_environment: dict[str, object] = {}

    def fake_subprocess_run(
        command: list[str], **_kwargs: object
    ) -> subprocess.CompletedProcess[str]:
        killed.append(int(command[command.index("/PID") + 1]))
        return subprocess.CompletedProcess(command, 0, "", "")

    class FakeProcess:
        pid = 9001
        returncode: int | None = None

        def __init__(self, _command: list[str], **kwargs: object) -> None:
            environment = kwargs["env"]
            assert isinstance(environment, dict)
            observed_environment.update(environment)
            Path(str(environment["WORKBOOKLENS_EXCEL_IDENTITY"])).write_text(
                payload,
                encoding="ascii",
            )

        def poll(self) -> int | None:
            return None

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == conversion._PROCESS_DRAIN_TIMEOUT_SECONDS
            assert not Path(str(observed_environment["WORKBOOKLENS_EXCEL_GO"])).exists()
            self.returncode = -9
            return "", ""

        def kill(self) -> None:
            self.returncode = -9

    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion, "_windows_taskkill_executable", lambda: taskkill)
    monkeypatch.setattr(conversion, "_snapshot_excel_process_ids", lambda _runner: frozenset())
    monkeypatch.setattr(conversion.subprocess, "run", fake_subprocess_run)
    monkeypatch.setattr(conversion.subprocess, "Popen", FakeProcess)

    with pytest.raises(conversion._ProviderFailure, match="invalid process-identity"):
        conversion._run_excel(provider, source, output, 1)

    assert killed == [9001]


def test_excel_sidecar_cannot_authorize_a_preexisting_user_pid(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("C:/Windows/powershell.exe"))
    taskkill = Path("C:/Windows/System32/taskkill.exe")
    killed: list[int] = []
    observed_environment: dict[str, object] = {}
    identity = _excel_identity()

    def fake_subprocess_run(
        command: list[str], **_kwargs: object
    ) -> subprocess.CompletedProcess[str]:
        killed.append(int(command[command.index("/PID") + 1]))
        return subprocess.CompletedProcess(command, 0, "", "")

    class FakeProcess:
        pid = 9001
        returncode: int | None = None

        def __init__(self, _command: list[str], **kwargs: object) -> None:
            environment = kwargs["env"]
            assert isinstance(environment, dict)
            observed_environment.update(environment)
            _write_excel_identity(
                Path(str(environment["WORKBOOKLENS_EXCEL_IDENTITY"])),
                identity,
            )

        def poll(self) -> int | None:
            return None

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == conversion._PROCESS_DRAIN_TIMEOUT_SECONDS
            assert not Path(str(observed_environment["WORKBOOKLENS_EXCEL_GO"])).exists()
            self.returncode = -9
            return "", ""

        def kill(self) -> None:
            self.returncode = -9

    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion, "_windows_taskkill_executable", lambda: taskkill)
    monkeypatch.setattr(
        conversion, "_snapshot_excel_process_ids", lambda _runner: frozenset({4242})
    )
    monkeypatch.setattr(
        conversion,
        "_open_verified_excel_process",
        lambda *_args, **_kwargs: pytest.fail("baseline PID must never be authorized"),
    )
    monkeypatch.setattr(conversion.subprocess, "run", fake_subprocess_run)
    monkeypatch.setattr(conversion.subprocess, "Popen", FakeProcess)

    with pytest.raises(conversion._ProviderFailure, match="existing user process"):
        conversion._run_excel(provider, source, output, 1)

    assert killed == [9001]


def test_excel_startup_timeout_kills_wrapper_then_unique_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("excel", "Microsoft Excel", Path("C:/Windows/powershell.exe"))
    taskkill = Path("C:/Windows/System32/taskkill.exe")
    killed: list[int] = []
    terminated: list[int] = []
    closed: list[int] = []
    candidate_identity = _excel_identity()

    def fake_subprocess_run(
        command: list[str], **_kwargs: object
    ) -> subprocess.CompletedProcess[str]:
        killed.append(int(command[command.index("/PID") + 1]))
        return subprocess.CompletedProcess(command, 0, "", "")

    class FakeProcess:
        pid = 9001
        returncode: int | None = None

        def __init__(self, _command: list[str], **_kwargs: object) -> None:
            pass

        def poll(self) -> int | None:
            return None

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == conversion._PROCESS_DRAIN_TIMEOUT_SECONDS
            self.returncode = -9
            return "", ""

        def kill(self) -> None:
            self.returncode = -9

    def fail_startup(
        _process: subprocess.Popen[str],
        _path: Path,
        _command: list[str],
    ) -> conversion._ExcelProcessIdentity:
        raise conversion._ExcelStartupTimeout

    def unique_candidate(
        runner: Path,
        baseline_process_ids: frozenset[int] | None,
        launched_at: datetime,
    ) -> conversion._ExcelProcessIdentity:
        assert runner == provider.runner
        assert baseline_process_ids == frozenset({111})
        assert launched_at.tzinfo is UTC
        return candidate_identity

    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion, "_windows_taskkill_executable", lambda: taskkill)
    monkeypatch.setattr(conversion, "_snapshot_excel_process_ids", lambda _runner: frozenset({111}))
    monkeypatch.setattr(
        conversion,
        "_wait_for_owned_excel_process_identity",
        fail_startup,
    )
    monkeypatch.setattr(conversion, "_find_unique_new_excel_automation_process", unique_candidate)
    monkeypatch.setattr(
        conversion,
        "_open_verified_excel_process",
        lambda expected, *, missing_is_clean: (
            conversion._OwnedExcelProcess(expected, 88)
            if missing_is_clean
            else pytest.fail("startup fallback must open in missing-is-clean mode")
        ),
    )
    monkeypatch.setattr(
        conversion,
        "_terminate_excel_process",
        lambda process: terminated.append(process.identity.process_id),
    )
    monkeypatch.setattr(
        conversion,
        "_close_excel_process",
        lambda process: closed.append(process.handle) if process is not None else None,
    )
    monkeypatch.setattr(conversion.subprocess, "run", fake_subprocess_run)
    monkeypatch.setattr(conversion.subprocess, "Popen", FakeProcess)

    with pytest.raises(conversion._ProviderFailure, match="startup handshake timed out"):
        conversion._run_excel(provider, source, output, 1)

    assert terminated == [4242]
    assert killed == [9001]
    assert closed == [88]


def test_excel_candidate_discovery_is_strict_and_fails_closed(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    runner = Path("C:/Windows/powershell.exe")
    launched_at = datetime(2025, 1, 1, tzinfo=UTC)
    strict_identity = _excel_identity()
    strict_creation_time = datetime.fromisoformat(
        strict_identity.creation_utc.replace("Z", "+00:00")
    )
    strict = conversion._ExcelProcessRecord(
        process_id=4242,
        identity=strict_identity,
        creation_time=strict_creation_time,
        command_line=(
            r'"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE" '
            r"/automation -Embedding"
        ),
        main_window_handle=0,
    )
    second_identity = _excel_identity(process_id=4343, creation_filetime=134000000000000010)
    second = conversion._ExcelProcessRecord(
        process_id=4343,
        identity=second_identity,
        creation_time=datetime.fromisoformat(second_identity.creation_utc.replace("Z", "+00:00")),
        command_line=(
            r'"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE" '
            r"/automation -Embedding"
        ),
        main_window_handle=0,
    )
    missing_information = conversion._ExcelProcessRecord(
        process_id=4444,
        identity=None,
        creation_time=launched_at,
        command_line=None,
        main_window_handle=0,
    )
    monkeypatch.setattr(conversion, "_EXCEL_CANDIDATE_DISCOVERY_SECONDS", 0.0)

    def set_snapshot(*processes: conversion._ExcelProcessRecord) -> None:
        snapshot = conversion._ExcelProcessSnapshot(7, processes)
        monkeypatch.setattr(conversion, "_query_excel_processes", lambda _runner: snapshot)

    set_snapshot(strict)
    assert (
        conversion._find_unique_new_excel_automation_process(runner, frozenset(), launched_at)
        == strict_identity
    )
    set_snapshot(strict, second)
    assert (
        conversion._find_unique_new_excel_automation_process(runner, frozenset(), launched_at)
        is None
    )
    set_snapshot(missing_information)
    assert (
        conversion._find_unique_new_excel_automation_process(runner, frozenset(), launched_at)
        is None
    )
    set_snapshot(strict)
    assert (
        conversion._find_unique_new_excel_automation_process(runner, frozenset({4242}), launched_at)
        is None
    )
    set_snapshot()
    assert (
        conversion._find_unique_new_excel_automation_process(runner, frozenset(), launched_at)
        is None
    )


@pytest.mark.parametrize("changed_field", ["creation", "path", "session"])
def test_reused_excel_pid_identity_mismatch_is_never_terminated(
    monkeypatch: pytest.MonkeyPatch,
    changed_field: str,
) -> None:
    expected = _excel_identity()
    if changed_field == "creation":
        actual = _excel_identity(creation_filetime=expected.creation_filetime + 10)
    elif changed_field == "path":
        actual = _excel_identity(executable_path=r"D:\Different Office\EXCEL.EXE")
    else:
        actual = _excel_identity(session_id=expected.session_id + 1)
    closed: list[int] = []
    monkeypatch.setattr(
        conversion,
        "_open_windows_process_handle",
        lambda process_id: 88 if process_id == expected.process_id else None,
    )
    monkeypatch.setattr(
        conversion,
        "_read_windows_process_handle_identity",
        lambda handle, process_id: actual,
    )
    monkeypatch.setattr(
        conversion,
        "_close_windows_process_handle",
        closed.append,
    )
    monkeypatch.setattr(
        conversion,
        "_taskkill_process_tree",
        lambda _process_id: pytest.fail("Excel PID must never be passed to taskkill"),
    )

    with pytest.raises(conversion._ProviderFailure, match="identity changed"):
        conversion._open_verified_excel_process(expected, missing_is_clean=False)

    assert closed == [88]


def test_matching_excel_handle_identity_is_terminated_via_handle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    identity = _excel_identity()
    closed: list[int] = []
    terminated: list[int] = []
    wait_results = iter([conversion._WAIT_TIMEOUT, conversion._WAIT_OBJECT_0])

    class FakeTerminateProcess:
        argtypes: object = None
        restype: object = None

        def __call__(self, handle: object, _exit_code: int) -> int:
            value = getattr(handle, "value", handle)
            terminated.append(int(value))
            return 1

    class FakeKernel32:
        TerminateProcess = FakeTerminateProcess()

    monkeypatch.setattr(
        conversion,
        "_open_windows_process_handle",
        lambda process_id: 88 if process_id == identity.process_id else None,
    )
    monkeypatch.setattr(
        conversion,
        "_read_windows_process_handle_identity",
        lambda handle, process_id: identity,
    )
    monkeypatch.setattr(conversion, "_windows_kernel32", lambda: FakeKernel32())
    monkeypatch.setattr(
        conversion,
        "_wait_windows_process_handle",
        lambda _handle, _timeout: next(wait_results),
    )
    monkeypatch.setattr(
        conversion,
        "_close_windows_process_handle",
        closed.append,
    )
    monkeypatch.setattr(
        conversion,
        "_taskkill_process_tree",
        lambda _process_id: pytest.fail("Excel PID must never be passed to taskkill"),
    )

    process = conversion._open_verified_excel_process(
        identity,
        missing_is_clean=False,
    )
    assert process is not None
    conversion._terminate_excel_process(process)
    conversion._close_excel_process(process)

    assert terminated == [88]
    assert closed == [88]


def test_windows_taskkill_path_is_anchored_to_systemroot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    system_root = tmp_path / "Windows"
    executable = system_root / "System32" / "taskkill.exe"
    executable.parent.mkdir(parents=True)
    executable.write_bytes(b"")
    monkeypatch.setenv("SYSTEMROOT", str(system_root))

    assert conversion._windows_taskkill_executable() == executable


def test_excel_format_metadata_has_a_bounded_size(tmp_path: Path) -> None:
    metadata = tmp_path / "number-formats.json"
    with metadata.open("wb") as handle:
        handle.truncate(conversion._MAX_FORMAT_MAP_BYTES + 1)

    with pytest.raises(conversion._ProviderFailure, match="excessive format metadata"):
        conversion._load_number_format_records(metadata)


@pytest.mark.parametrize(
    ("existing_format", "expected_format_id", "expected_format_count"),
    [
        (None, 164, 1),
        ('yyyy"年"m"月"', 164, 1),
        ("0.0000", 165, 2),
    ],
)
def test_excel_number_format_restore_materializes_locale_builtin_and_is_idempotent(
    tmp_path: Path,
    existing_format: str | None,
    expected_format_id: int,
    expected_format_count: int,
) -> None:
    output = _locale_builtin_workbook(tmp_path / "converted.xlsx", existing_format)
    number_format = 'yyyy"年"m"月"'
    records = (conversion._NumberFormatRecord(1, "A1", number_format),)

    conversion._restore_number_formats(output, records)
    first_hash = hashlib.sha256(output.read_bytes()).hexdigest()
    conversion._restore_number_formats(output, records)

    assert hashlib.sha256(output.read_bytes()).hexdigest() == first_hash
    workbook = load_workbook(output, data_only=False)
    assert workbook.active["A1"].value == datetime(2025, 12, 1)
    assert workbook.active["A1"].number_format == number_format
    workbook.close()
    with zipfile.ZipFile(output, "r") as archive:
        styles = etree.fromstring(archive.read("xl/styles.xml"))
        namespace = etree.QName(styles).namespace
        assert namespace is not None
        num_fmts = styles.find(f"{{{namespace}}}numFmts")
        cell_xfs = styles.find(f"{{{namespace}}}cellXfs")
        assert num_fmts is not None
        assert cell_xfs is not None
        assert len(num_fmts) == expected_format_count
        assert any(
            element.get("numFmtId") == str(expected_format_id)
            and element.get("formatCode") == number_format
            for element in num_fmts
        )
        assert cell_xfs[-1].get("numFmtId") == str(expected_format_id)


def test_libreoffice_runner_uses_isolated_profile(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("libreoffice", "LibreOffice", Path("C:/LibreOffice/soffice.exe"))
    observed: list[str] = []
    popen_options: dict[str, object] = {}

    class FakeProcess:
        pid = 2468
        returncode = 0

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == 10
            output_dir = Path(observed[observed.index("--outdir") + 1])
            staged_input = Path(observed[-1])
            generate_demo_workbook(output_dir / f"{staged_input.stem}.xlsx")
            return "", ""

        def kill(self) -> None:
            raise AssertionError("successful LibreOffice conversion must not be killed")

    def fake_popen(command: list[str], **kwargs: object) -> FakeProcess:
        observed.extend(command)
        popen_options.update(kwargs)
        return FakeProcess()

    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion.subprocess, "CREATE_NO_WINDOW", 8, raising=False)
    monkeypatch.setattr(conversion.subprocess, "CREATE_NEW_PROCESS_GROUP", 512, raising=False)
    monkeypatch.setattr(conversion.subprocess, "Popen", fake_popen)

    conversion._run_libreoffice(provider, source, output, 10)

    assert "--headless" in observed
    assert any(argument.startswith("-env:UserInstallation=file:") for argument in observed)
    assert popen_options["creationflags"] == 520
    assert "start_new_session" not in popen_options
    assert output.read_bytes().startswith(b"PK")


def test_libreoffice_runner_never_overwrites_a_preexisting_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    output.write_bytes(b"user-owned")
    provider = ConversionProvider("libreoffice", "LibreOffice", Path("soffice.exe"))
    monkeypatch.setattr(
        conversion.subprocess,
        "Popen",
        lambda *_args, **_kwargs: pytest.fail("LibreOffice must not start"),
    )

    with pytest.raises(conversion._ProviderFailure, match="already exists"):
        conversion._run_libreoffice(provider, source, output, 10)

    assert output.read_bytes() == b"user-owned"


def test_libreoffice_non_windows_starts_a_new_session(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("libreoffice", "LibreOffice", Path("/usr/bin/soffice"))
    observed: dict[str, object] = {}

    class FakeProcess:
        pid = 1357
        returncode = 0

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            assert timeout == 10
            command = observed["command"]
            assert isinstance(command, list)
            output_dir = Path(command[command.index("--outdir") + 1])
            staged_input = Path(command[-1])
            generate_demo_workbook(output_dir / f"{staged_input.stem}.xlsx")
            return "", ""

        def kill(self) -> None:
            raise AssertionError("successful LibreOffice conversion must not be killed")

    def fake_popen(command: list[str], **kwargs: object) -> FakeProcess:
        observed["command"] = command
        observed["options"] = kwargs
        return FakeProcess()

    monkeypatch.setattr(conversion.sys, "platform", "linux")
    monkeypatch.setattr(conversion.subprocess, "Popen", fake_popen)

    conversion._run_libreoffice(provider, source, output, 10)

    options = observed["options"]
    assert isinstance(options, dict)
    assert options["start_new_session"] is True
    assert "creationflags" not in options
    assert output.read_bytes().startswith(b"PK")


def test_libreoffice_timeout_terminates_the_windows_process_tree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _legacy_workbook(tmp_path / "legacy.xls")
    output = tmp_path / "converted.xlsx"
    provider = ConversionProvider("libreoffice", "LibreOffice", Path("C:/LibreOffice/soffice.exe"))
    killed: list[int] = []

    class HangingProcess:
        pid = 8642
        returncode = None
        communicate_calls = 0
        kill_calls = 0

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            self.communicate_calls += 1
            if self.communicate_calls == 1:
                raise subprocess.TimeoutExpired("soffice", timeout)
            return "", ""

        def kill(self) -> None:
            self.kill_calls += 1

    process = HangingProcess()
    monkeypatch.setattr(conversion.sys, "platform", "win32")
    monkeypatch.setattr(conversion.subprocess, "Popen", lambda *_args, **_kwargs: process)
    monkeypatch.setattr(conversion, "_taskkill_process_tree", killed.append)

    with pytest.raises(conversion._ProviderFailure, match="timed out after 1 seconds"):
        conversion._run_libreoffice(provider, source, output, 1)

    assert killed == [8642]
    assert process.communicate_calls == 2
    assert process.kill_calls == 0


def test_non_windows_process_tree_escalates_to_sigkill(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    signals: list[tuple[int, int]] = []

    class HangingProcess:
        pid = 9753
        communicate_calls = 0
        kill_calls = 0

        def communicate(self, *, timeout: int) -> tuple[str, str]:
            self.communicate_calls += 1
            if self.communicate_calls < 2:
                raise subprocess.TimeoutExpired("soffice", timeout)
            return "", ""

        def kill(self) -> None:
            self.kill_calls += 1

    process = HangingProcess()
    monkeypatch.setattr(conversion.sys, "platform", "linux")
    monkeypatch.setattr(
        conversion.os,
        "killpg",
        lambda process_id, sent_signal: signals.append((process_id, sent_signal)),
        raising=False,
    )

    conversion._terminate_process_tree(process)  # type: ignore[arg-type]

    assert signals == [
        (9753, conversion._SIGTERM),
        (9753, conversion._SIGKILL),
    ]
    assert process.communicate_calls == 2
    assert process.kill_calls == 0
