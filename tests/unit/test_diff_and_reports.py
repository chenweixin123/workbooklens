from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from workbooklens.demo.workflow import generate_demo_workbook
from workbooklens.diff import compare_snapshots, compare_workbooks, write_diff_report
from workbooklens.models import CellSnapshot, SheetSnapshot, WorkbookSnapshot
from workbooklens.reports import write_scan_report
from workbooklens.reports.scan import build_sarif
from workbooklens.scanner import scan_workbook


def _diff_workbooks(directory: Path) -> tuple[Path, Path]:
    before = Workbook()
    data = before.active
    assert data is not None
    data.title = "Data"
    data.append(["Name", "Value", "Formula"])
    data.append(["A", 1, "=B2*2"])
    data.append(["B", 2, "=B3*2"])
    old = before.create_sheet("Old Name")
    old["A1"] = "same content"
    before.defined_names.add(DefinedName("Rate", attr_text="Data!$B$2"))
    before.calculation.calcMode = "manual"
    before_path = directory / "before.xlsx"
    before.save(before_path)
    before.close()

    after = Workbook()
    data_after = after.active
    assert data_after is not None
    data_after.title = "Data"
    data_after.append(["Name", "Value", "Formula"])
    data_after.append(["A", 10, "=B2*3"])
    data_after.append(["B", 2, "=B3*2"])
    data_after["B2"].fill = PatternFill("solid", fgColor="FFFF00")
    data_after["B2"].number_format = "0.00"
    data_after.row_dimensions[3].hidden = True
    data_after.column_dimensions["A"].hidden = True
    data_after.merge_cells("D2:E2")
    validation = DataValidation(type="whole", operator="between", formula1="0", formula2="10")
    data_after.add_data_validation(validation)
    validation.add("B2:B3")
    renamed = after.create_sheet("Renamed")
    renamed["A1"] = "same content"
    after._sheets = [renamed, data_after]
    after.defined_names.add(DefinedName("Rate", attr_text="Data!$B$3"))
    after.calculation.calcMode = "auto"
    after_path = directory / "after.xlsx"
    after.save(after_path)
    after.close()
    return before_path, after_path


def test_semantic_diff_covers_cell_and_structure_changes(tmp_path: Path) -> None:
    before, after = _diff_workbooks(tmp_path)
    diff = compare_workbooks(before, after)
    cell_types = {change.change_type for change in diff.cell_changes}
    structural_types = {change.change_type for change in diff.structural_changes}
    assert {"value", "formula", "style", "number_format"} <= cell_types
    assert {
        "sheet_renamed",
        "sheet_reordered",
        "hidden_row_added",
        "hidden_column_added",
        "merged_range_added",
        "data_validation_added",
        "defined_name",
        "calculation_mode",
    } <= structural_types
    formula_change = next(change for change in diff.cell_changes if change.change_type == "formula")
    assert formula_change.before_signature
    assert formula_change.after_signature
    assert formula_change.before_signature != formula_change.after_signature


def test_semantic_diff_is_value_type_sensitive(tmp_path: Path) -> None:
    before = Workbook()
    before.active["A1"] = 1
    before_path = tmp_path / "integer.xlsx"
    before.save(before_path)
    before.close()

    after = Workbook()
    after.active["A1"] = True
    after_path = tmp_path / "boolean.xlsx"
    after.save(after_path)
    after.close()

    changes = compare_workbooks(before_path, after_path).cell_changes
    value_change = next(change for change in changes if change.change_type == "value")
    assert type(value_change.before) is int
    assert type(value_change.after) is bool


def test_semantic_diff_uses_style_fingerprint_not_workbook_style_id(tmp_path: Path) -> None:
    before = Workbook()
    before.active["A1"] = "same"
    before.active["A1"].fill = PatternFill("solid", fgColor="FFFF0000")
    before_path = tmp_path / "red.xlsx"
    before.save(before_path)
    before.close()

    after = Workbook()
    after.active["A1"] = "same"
    after.active["A1"].fill = PatternFill("solid", fgColor="FF0000FF")
    after_path = tmp_path / "blue.xlsx"
    after.save(after_path)
    after.close()

    changes = compare_workbooks(before_path, after_path).cell_changes
    style_change = next(change for change in changes if change.change_type == "style")
    assert style_change.before != style_change.after


def _snapshot_with_sheets(*sheets: SheetSnapshot, source: str) -> WorkbookSnapshot:
    return WorkbookSnapshot(
        source_name=f"{source}.xlsx",
        source_sha256=source,
        format="xlsx",
        sheets=list(sheets),
    )


def _sheet(
    name: str, index: int, *, style_id: int = 0, style_fingerprint: str = "default"
) -> SheetSnapshot:
    return SheetSnapshot(
        name=name,
        index=index,
        state="visible",
        max_row=1,
        max_column=1,
        cells={
            "A1": CellSnapshot(
                coordinate="A1",
                value="same",
                data_type="s",
                style_id=style_id,
                style_fingerprint=style_fingerprint,
            )
        },
    )


def test_style_id_difference_alone_is_not_a_semantic_change() -> None:
    before = _snapshot_with_sheets(
        _sheet("Data", 0, style_id=1, style_fingerprint="same"), source="a"
    )
    after = _snapshot_with_sheets(
        _sheet("Data", 0, style_id=99, style_fingerprint="same"), source="b"
    )

    assert not compare_snapshots(before, after).cell_changes


def test_ambiguous_sheet_rename_degrades_to_add_and_remove() -> None:
    before = _snapshot_with_sheets(_sheet("Old A", 0), _sheet("Old B", 1), source="a")
    after = _snapshot_with_sheets(_sheet("New", 0), source="b")

    changes = compare_snapshots(before, after).structural_changes
    assert not any(change.change_type == "sheet_renamed" for change in changes)
    assert [change.change_type for change in changes].count("sheet_removed") == 2
    assert [change.change_type for change in changes].count("sheet_added") == 1


def test_diff_report_is_self_contained_and_has_json_twin(tmp_path: Path) -> None:
    before, after = _diff_workbooks(tmp_path)
    diff = compare_workbooks(before, after)
    paths = write_diff_report(diff, tmp_path / "diff.html")
    html = paths["html"].read_text(encoding="utf-8")
    payload = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert "Semantic workbook diff" in html
    assert "<script src=" not in html
    assert payload["before_sha256"] == diff.before_sha256
    assert payload["cell_changes"]


def test_scan_report_and_sarif_handle_multi_cell_locations(tmp_path: Path) -> None:
    workbook = tmp_path / "demo.xlsx"
    generate_demo_workbook(workbook)
    scan = scan_workbook(
        workbook,
        config={"keys": [{"sheet": "Sales", "range": "A2:A22"}]},
    )
    sarif = build_sarif(scan, source_uri="models/demo.xlsx")
    duplicate = next(
        finding for finding in scan.findings if finding.rule_id == "WL012_DUPLICATE_CONFIGURED_KEY"
    )
    assert "," in (duplicate.location or "")
    sarif_result = next(
        result
        for result in sarif["runs"][0]["results"]
        if result["ruleId"] == "WL012_DUPLICATE_CONFIGURED_KEY"
    )
    assert "region" not in sarif_result["locations"][0]["physicalLocation"]
    assert sarif_result["locations"][0]["physicalLocation"]["artifactLocation"]["uri"] == (
        "models/demo.xlsx"
    )
    assert sarif["runs"][0]["automationDetails"]["id"] == "workbooklens/models/demo.xlsx"
    paths = write_scan_report(scan, tmp_path / "report")
    html = paths["html"].read_text(encoding="utf-8")
    findings = json.loads(paths["findings"].read_text(encoding="utf-8"))
    sarif_json = json.loads(paths["sarif"].read_text(encoding="utf-8"))
    assert "Local processing" in html
    assert "Sheet overview" in html
    assert 'id="sort"' in html
    assert "<script src=" not in html
    assert findings["source_sha256"] == scan.snapshot.source_sha256
    assert sarif_json["version"] == "2.1.0"
