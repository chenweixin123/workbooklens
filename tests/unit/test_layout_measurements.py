from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from workbooklens.layout import (
    WhitespaceTail,
    find_formatting_tail,
    find_whitespace_tail,
    measure_text_cell,
    visual_border_signature,
    whitespace_tail_layout_fingerprint,
)


def test_visual_border_signature_does_not_materialize_neighbors() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["C3"] = "value"
    before = set(worksheet._cells)

    visual_border_signature(worksheet, worksheet["C3"])

    assert set(worksheet._cells) == before
    workbook.close()


def test_measure_text_excludes_layouts_that_need_excel_rendering() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "long text"
    worksheet["A1"].alignment = Alignment(shrink_to_fit=True)
    worksheet["B1"] = "long text"
    worksheet["B1"].alignment = Alignment(text_rotation=45)
    worksheet["C1"] = "long text"
    worksheet["C1"].alignment = Alignment(indent=1)
    worksheet["D1"] = "long merged text"
    worksheet.merge_cells("D1:E2")

    assert measure_text_cell(worksheet, worksheet["A1"]) is None
    assert measure_text_cell(worksheet, worksheet["B1"]) is None
    assert measure_text_cell(worksheet, worksheet["C1"]) is None
    assert measure_text_cell(worksheet, worksheet["D1"]) is None
    workbook.close()


def test_formatting_tail_preserves_exact_sparse_coordinates() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    coordinates = {
        (30, 20),
        (30, 21),
        (30, 22),
        (30, 23),
        (30, 24),
        (31, 20),
        (31, 21),
        (31, 22),
        (31, 24),
        (31, 25),
        (32, 20),
        (32, 21),
        (32, 22),
        (32, 23),
        (32, 24),
        (32, 25),
        (33, 20),
        (33, 21),
        (33, 22),
        (33, 23),
    }
    for row, column in coordinates:
        worksheet.cell(row, column).fill = fill

    tail = find_formatting_tail(worksheet)

    assert tail is not None
    assert len(tail.cell_coordinates) == len(coordinates)
    assert "W31" not in tail.cell_coordinates
    assert "T31:V31" in tail.cell_ranges
    assert "X31:Y31" in tail.cell_ranges
    workbook.close()


def test_broad_column_dimension_style_is_not_a_formatting_tail() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    dimension = worksheet.column_dimensions["L"]
    dimension.min = 12
    dimension.max = 16_384
    dimension.width = 12.0

    assert find_formatting_tail(worksheet) is None
    workbook.close()


def test_formatting_tail_excludes_custom_height_blank_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in range(50, 54):
        for column in range(16, 20):
            worksheet.cell(row, column).fill = fill
    worksheet.row_dimensions[54].height = 24.0
    worksheet.row_dimensions[55].fill = fill

    tail = find_formatting_tail(worksheet)

    assert tail is not None
    assert 54 not in tail.empty_rows
    assert 55 in tail.empty_rows
    assert tail.observed_max_row == 55
    workbook.close()


def test_single_sided_shared_border_is_visually_continuous() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "left"
    worksheet["B1"] = "right"
    worksheet["B1"].border = Border(left=Side(style="thin"))

    signature = visual_border_signature(worksheet, worksheet["A1"])

    assert signature[1][0] == "thin"
    workbook.close()


def test_whitespace_tail_marks_nondefault_cell_styles_for_preservation() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet["A9"] = " "
    worksheet["A10"] = " "
    worksheet["A10"].font = Font(bold=True)
    worksheet["A11"] = " "
    worksheet["A11"].alignment = Alignment(horizontal="center")
    worksheet["A11"].protection = Protection(locked=False, hidden=True)

    tail = find_whitespace_tail(worksheet)

    assert tail is not None
    assert tail.cell_coordinates == ("A9", "A10", "A11")
    assert dict(tail.preserve_style_ids) == {
        "A10": worksheet["A10"].style_id,
        "A11": worksheet["A11"].style_id,
    }
    assert tail.result_dimension == "A1:A11"
    workbook.close()


def test_whitespace_tail_fingerprint_uses_coordinate_order_not_json_key_order() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    for coordinate in ("A9", "A10", "A11"):
        worksheet[coordinate] = " "
        worksheet[coordinate].font = Font(bold=True)
    tail = find_whitespace_tail(worksheet)
    assert tail is not None
    reordered = WhitespaceTail(
        cell_coordinates=tail.cell_coordinates,
        cell_ranges=tail.cell_ranges,
        row_ranges=tail.row_ranges,
        content_min_row=tail.content_min_row,
        content_min_column=tail.content_min_column,
        content_max_row=tail.content_max_row,
        content_max_column=tail.content_max_column,
        observed_dimension=tail.observed_dimension,
        result_dimension=tail.result_dimension,
        preserve_style_ids=tuple(sorted(tail.preserve_style_ids, key=lambda item: item[0])),
    )

    assert whitespace_tail_layout_fingerprint(
        worksheet, tail
    ) == whitespace_tail_layout_fingerprint(worksheet, reordered)
    workbook.close()
