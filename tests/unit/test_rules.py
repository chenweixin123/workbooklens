from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

import workbooklens.rules.builtin as builtin_rules
from workbooklens.demo.workflow import generate_demo_workbook
from workbooklens.models import PatchKind, PatchRisk
from workbooklens.rules.builtin import BUILTIN_RULES
from workbooklens.scanner import ScanResult, scan_workbook


@pytest.fixture(scope="module")
def demo_scan(tmp_path_factory: pytest.TempPathFactory) -> ScanResult:
    directory = tmp_path_factory.mktemp("rules-demo")
    path = directory / "demo.xlsx"
    generate_demo_workbook(path)
    config = {"keys": [{"sheet": "Sales", "range": "A2:A22", "ignore_blank": True}]}
    return scan_workbook(path, config=config)


def test_demo_exercises_fifteen_rules_and_generated_patch_kinds(demo_scan: ScanResult) -> None:
    rule_ids = {finding.rule_id for finding in demo_scan.findings}
    assert rule_ids == {
        "WL001_BROKEN_REFERENCE",
        "WL003_BLANK_IN_FORMULA_BAND",
        "WL004_HARDCODED_VALUE_IN_FORMULA_BAND",
        "WL005_SUSPICIOUS_SUM_BOUNDARY",
        "WL006_NUMERIC_TEXT",
        "WL007_STYLE_OUTLIER",
        "WL008_HIDDEN_NONEMPTY_DATA",
        "WL009_EXTERNAL_LINK",
        "WL010_VOLATILE_OR_FRAGILE_FUNCTION",
        "WL011_ERROR_CELL",
        "WL012_DUPLICATE_CONFIGURED_KEY",
        "WL013_BROKEN_DEFINED_NAME",
        "WL014_MERGED_CELL_IN_DATA_REGION",
        "WL015_INCONSISTENT_DATA_VALIDATION",
        "WL016_TEXT_DISPLAY_RISK",
    }
    assert {patch.kind for patch in demo_scan.patches} == {
        PatchKind.SET_FORMULA,
        PatchKind.SET_NUMERIC,
        PatchKind.COPY_STYLE,
        PatchKind.CREATE_FORMULA,
        PatchKind.SET_ROW_HEIGHT,
        PatchKind.SET_WRAP_TEXT,
    }
    assert all(
        patch.safe_only_eligible for patch in demo_scan.patches if patch.risk == PatchRisk.SAFE
    )
    assert all(
        not patch.safe and patch.risk == PatchRisk.LAYOUT_REVIEW
        for patch in demo_scan.patches
        if patch.kind in {PatchKind.SET_ROW_HEIGHT, PatchKind.SET_WRAP_TEXT}
    )


def test_formula_pattern_outlier_rule_and_safe_proposal(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Band"
    for column in range(2, 22):
        worksheet.cell(1, column, column)
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["K2"] = "=K1*3"
    path = tmp_path / "outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = [
        finding for finding in scan.findings if finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER"
    ]
    assert len(findings) == 1
    assert findings[0].location == "K2"
    patches = [patch for patch in scan.patches if patch.cell == "K2"]
    assert len(patches) == 1
    assert patches[0].kind == PatchKind.SET_FORMULA
    assert patches[0].after == "=K1*2"


@pytest.mark.parametrize(
    "formula",
    [
        "=SUM(B2:T2)",
        "=SUBTOTAL(9,B2:T2)",
        "=AGGREGATE(9,5,B2:T2)",
        "=SUM(B2:T2)+0",
        "=ROUND(SUM(B2:T2),2)",
        "=SUM(B2:T2,N(0))",
        "=SUM(B2:T2)+IFERROR(0,0)",
        "=AVERAGE(B2:T2)",
        "=MIN(B2:T2)",
        "=MAX(B2:T2)",
        "=COUNT(B2:T2)",
        "=COUNTA(B2:T2)",
        "=MEDIAN(B2:T2)",
    ],
)
def test_boundary_aggregate_is_not_treated_as_formula_outlier(tmp_path: Path, formula: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for column in range(2, 21):
        worksheet.cell(1, column, column)
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["U2"] = formula
    path = tmp_path / "boundary-total.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER" and finding.location == "U2"
        for finding in scan.findings
    )
    assert not any(patch.cell == "U2" for patch in scan.patches)


def test_multiple_isolated_formula_outliers_are_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for column in range(2, 22):
        worksheet.cell(1, column, column)
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["K2"] = "=K1*3"
    worksheet["Q2"] = "=Q1+7"
    path = tmp_path / "multiple-outliers.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = {
        finding.location
        for finding in scan.findings
        if finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER"
    }
    assert findings == {"K2", "Q2"}
    assert not any(patch.cell in findings for patch in scan.patches)


def test_merged_formula_outlier_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 22):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet["B10"] = "=A10*3"
    worksheet.merge_cells("B10:C10")
    path = tmp_path / "merged-formula-outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


@pytest.mark.parametrize(
    "label",
    [
        "Subtotal",
        "Sub-total",
        "Grand-total",
        "Average",
        "Mean",
        "Minimum",
        "Maximum",
        "Summary",
        "小计",
        "小计金额",
        "合计金额",
        "汇总金额",
        "平均金额",
        "最大金额",
        "最小金额",
        "总金额",
        "累计金额",
        "净额",
        "期末余额",
    ],
)
def test_summary_row_formula_outlier_is_findings_only(tmp_path: Path, label: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A2"] = label
    for column in range(2, 22):
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["K2"] = "=K1*3"
    path = tmp_path / f"summary-formula-outlier-{len(label)}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER" and finding.location == "K2"
        for finding in scan.findings
    )
    assert not any(patch.cell == "K2" for patch in scan.patches)


def test_hidden_formula_outlier_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 22):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet["B10"] = "=A10*3"
    worksheet.row_dimensions[10].hidden = True
    path = tmp_path / "hidden-formula-outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_formula_band_boundary_never_gets_automatic_replacement(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Band"
    for column in range(2, 21):
        worksheet.cell(1, column, column)
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["U2"] = "=AVERAGE(Band!B2:T2)"
    path = tmp_path / "explicit-sheet-boundary-summary.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(patch.cell == "U2" for patch in scan.patches)


def test_finding_identity_is_stable_when_evidence_content_changes(tmp_path: Path) -> None:
    first = Workbook()
    first_sheet = first.active
    assert first_sheet is not None
    first_sheet["A1"] = "=#REF!+1"
    first_path = tmp_path / "first.xlsx"
    first.save(first_path)
    first.close()

    second = Workbook()
    second_sheet = second.active
    assert second_sheet is not None
    second_sheet["A1"] = "=#REF!+2"
    second_path = tmp_path / "second.xlsx"
    second.save(second_path)
    second.close()

    first_finding = next(
        finding
        for finding in scan_workbook(first_path).findings
        if finding.rule_id == "WL001_BROKEN_REFERENCE"
    )
    second_finding = next(
        finding
        for finding in scan_workbook(second_path).findings
        if finding.rule_id == "WL001_BROKEN_REFERENCE"
    )
    assert first_finding.id == second_finding.id
    assert first_finding.content_fingerprint != second_finding.content_fingerprint


def test_unsupported_formula_suppresses_automatic_band_repair(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Band"
    for column in range(2, 22):
        worksheet.cell(1, column, column)
        coordinate = worksheet.cell(2, column).coordinate
        source = worksheet.cell(1, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet["K2"] = "=Table1[Amount]"
    path = tmp_path / "unsupported-band.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(finding.rule_id == "WL002_FORMULA_PATTERN_OUTLIER" for finding in scan.findings)
    assert not any(patch.cell == "K2" for patch in scan.patches)


def test_formula_rule_tokens_ignore_string_literals_but_keep_real_constructs(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = '="#REF! NOW() Table1[Amount] [Book.xlsx]S!A1"'
    worksheet["A2"] = "=#REF!+1"
    worksheet["A3"] = "=NOW()"
    worksheet["A4"] = "='[Book.xlsx]S'!A1"
    path = tmp_path / "formula-tokens.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    locations_by_rule = {
        rule_id: {finding.location for finding in scan.findings if finding.rule_id == rule_id}
        for rule_id in {
            "WL001_BROKEN_REFERENCE",
            "WL009_EXTERNAL_LINK",
            "WL010_VOLATILE_OR_FRAGILE_FUNCTION",
        }
    }
    assert locations_by_rule["WL001_BROKEN_REFERENCE"] == {"A2"}
    assert locations_by_rule["WL009_EXTERNAL_LINK"] == {"A4"}
    assert locations_by_rule["WL010_VOLATILE_OR_FRAGILE_FUNCTION"] == {"A3"}


def test_numeric_text_excludes_leading_zero_identifiers(demo_scan: ScanResult) -> None:
    locations = {
        finding.location
        for finding in demo_scan.findings
        if finding.rule_id == "WL006_NUMERIC_TEXT"
    }
    assert locations == {"B10"}


@pytest.mark.parametrize(
    "header",
    [
        "Customer ID",
        "SKU",
        "Account Number",
        "Postal Code",
        "Phone Number",
        "Mobile Number",
        "SSN",
        "ISBN",
        "手机号",
        "证件号码",
        "银行卡号",
        "客户编号",
    ],
)
def test_identifier_headers_suppress_numeric_text_patch(tmp_path: Path, header: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append([header, "Amount"])
    for row in range(2, 22):
        worksheet.append([10000 + row, row * 10])
    worksheet["A10"] = "12345"
    path = tmp_path / f"identifier-{len(header)}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
    )
    assert not finding.safe_patch_available
    assert not any(patch.cell == "A10" for patch in scan.patches)


def test_account_balance_still_allows_numeric_measure_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Account Balance", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    path = tmp_path / "account-balance.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        patch.kind == PatchKind.SET_NUMERIC and patch.cell == "A10" for patch in scan.patches
    )


def test_measure_patch_survives_plain_numeric_text_in_identifier_column(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Account ID", "Customer", "Credit Limit"])
    for row in range(2, 22):
        worksheet.append([10000 + row, f"Customer {row}", row * 1000])
    worksheet["A10"] = "12345"
    worksheet["C10"] = "12000"
    path = tmp_path / "identifier-and-measure-numeric-text.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    patches = {(patch.kind, patch.cell) for patch in scan.patches}
    assert (PatchKind.SET_NUMERIC, "A10") not in patches
    assert (PatchKind.SET_NUMERIC, "C10") in patches


def test_unknown_numeric_column_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Business Field", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    path = tmp_path / "unknown-numeric-column.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "A10" for patch in scan.patches)


def test_explicit_text_format_suppresses_numeric_text_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Measure", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    worksheet["A10"].number_format = "@"
    path = tmp_path / "explicit-text.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "A10" for patch in scan.patches)


def test_quote_prefix_suppresses_numeric_and_style_patches(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Measure", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    worksheet["A10"].quotePrefix = True
    path = tmp_path / "quote-prefix.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(
        patch.cell == "A10" and patch.kind in {PatchKind.SET_NUMERIC, PatchKind.COPY_STYLE}
        for patch in scan.patches
    )


def test_protected_worksheet_is_findings_only_for_numeric_text(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Amount", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    worksheet.protection.sheet = True
    path = tmp_path / "protected-numeric.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "A10" for patch in scan.patches)


def test_grouped_numeric_text_is_reported_without_auto_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Amount", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1,200"
    path = tmp_path / "grouped-numeric.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "A10" for patch in scan.patches)


def test_grouped_hidden_nonleading_column_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Context", "Amount", "Reviewer"])
    for row in range(2, 22):
        worksheet.append([f"R{row}", f"C{row}", row * 100, "Chen"])
    worksheet["C10"] = "1200"
    worksheet.column_dimensions.group("B", "D", hidden=True)
    path = tmp_path / "grouped-hidden-nonleading-column.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "C10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "C10" for patch in scan.patches)
    snapshot_sheet = scan.snapshot.sheets[0]
    assert snapshot_sheet.hidden_columns == ["B", "C", "D"]
    assert snapshot_sheet.cells["C10"].column_hidden


@pytest.mark.parametrize("sheet_state", ["hidden", "veryHidden"])
def test_nonvisible_worksheet_is_findings_only_for_numeric_text(
    tmp_path: Path, sheet_state: str
) -> None:
    workbook = Workbook()
    cover = workbook.active
    assert cover is not None
    cover.title = "Cover"
    worksheet = workbook.create_sheet("Data")
    worksheet.append(["Amount", "Context"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"R{row}"])
    worksheet["A10"] = "1200"
    worksheet.sheet_state = sheet_state
    path = tmp_path / f"{sheet_state}-numeric.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL006_NUMERIC_TEXT"
        and finding.sheet == "Data"
        and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(patch.sheet == "Data" for patch in scan.patches)


@pytest.mark.parametrize("sheet_state", ["hidden", "veryHidden"])
def test_nonvisible_worksheet_is_findings_only_for_formula_and_style(
    tmp_path: Path, sheet_state: str
) -> None:
    workbook = Workbook()
    cover = workbook.active
    assert cover is not None
    cover.title = "Cover"
    worksheet = workbook.create_sheet("Data")
    worksheet.append(["Amount", "Calculated"])
    for row in range(2, 22):
        worksheet.append([row * 100, f"=A{row}*2"])
    worksheet["B10"] = 999
    worksheet["A11"].font = Font(bold=True)
    worksheet.sheet_state = sheet_state
    path = tmp_path / f"{sheet_state}-formula-style.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL004_HARDCODED_VALUE_IN_FORMULA_BAND"
        and finding.sheet == "Data"
        and finding.location == "B10"
        for finding in scan.findings
    )
    assert any(
        finding.rule_id == "WL007_STYLE_OUTLIER"
        and finding.sheet == "Data"
        and finding.location == "A11"
        for finding in scan.findings
    )
    assert not any(patch.sheet == "Data" for patch in scan.patches)


@pytest.mark.parametrize("sheet_state", ["hidden", "veryHidden"])
def test_nonvisible_worksheet_sum_boundary_is_findings_only(
    tmp_path: Path, sheet_state: str
) -> None:
    workbook = Workbook()
    cover = workbook.active
    assert cover is not None
    cover.title = "Cover"
    worksheet = workbook.create_sheet("Data")
    for row in range(2, 10):
        worksheet.cell(row, 1, f"R{row}")
        worksheet.cell(row, 2, row * 100)
    worksheet["A10"] = "Total"
    worksheet["B10"] = "=SUM(B2:B8)"
    worksheet.sheet_state = sheet_state
    path = tmp_path / f"{sheet_state}-sum-boundary.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY" and finding.sheet == "Data"
    )
    assert not finding.safe_patch_available
    assert not any(patch.sheet == "Data" for patch in scan.patches)


def test_protection_only_difference_is_not_a_visual_style_outlier(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Input", "Context"])
    for row in range(2, 22):
        worksheet.append([row, f"R{row}"])
    worksheet["A10"].protection = Protection(locked=False)
    worksheet.protection.sheet = True
    path = tmp_path / "unlocked-style-outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(
        patch.kind == PatchKind.COPY_STYLE and patch.cell == "A10" for patch in scan.patches
    )


def test_style_outlier_with_different_number_format_has_no_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Customer ID", "Context"])
    for row in range(2, 22):
        worksheet.append([10000 + row, f"R{row}"])
    worksheet["A10"].number_format = "000000"
    path = tmp_path / "identifier-number-format-outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "A10"
    )
    assert not finding.safe_patch_available
    assert not any(
        patch.kind == PatchKind.COPY_STYLE and patch.cell == "A10" for patch in scan.patches
    )


@pytest.mark.parametrize("missing_side", ["left", "right"])
def test_single_sided_shared_border_is_visually_equivalent(
    tmp_path: Path, missing_side: str
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Context", "Amount", "Context"])
    side = Side(style="thin", color="336699")
    grid = Border(left=side, right=side, top=side, bottom=side)
    for row in range(2, 22):
        worksheet.append([f"L{row}", row * 100, f"R{row}"])
        for cell in worksheet[row]:
            cell.border = grid
    target = worksheet["B10"]
    target.border = Border(
        left=None if missing_side == "left" else side,
        right=None if missing_side == "right" else side,
        top=side,
        bottom=side,
    )
    path = tmp_path / f"shared-border-{missing_side}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "B10"
        for finding in scan.findings
    )


@pytest.mark.parametrize("component", ["font", "fill", "alignment", "number_format", "border"])
def test_material_visual_style_difference_remains_an_outlier(
    tmp_path: Path, component: str
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Context", "Amount"])
    side = Side(style="thin", color="336699")
    grid = Border(left=side, right=side, top=side, bottom=side)
    for row in range(2, 22):
        worksheet.append([f"R{row}", row * 100])
        for cell in worksheet[row]:
            cell.border = grid
    target = worksheet["B10"]
    if component == "font":
        target.font = Font(bold=True)
    elif component == "fill":
        target.fill = PatternFill("solid", fgColor="FFF2CC")
    elif component == "alignment":
        target.alignment = Alignment(horizontal="right")
    elif component == "number_format":
        target.number_format = "0.00"
    else:
        target.border = Border(right=side, top=side, bottom=side)
        worksheet["A10"].border = Border(left=side, top=side, bottom=side)
    path = tmp_path / f"material-style-{component}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "B10"
        for finding in scan.findings
    )


def test_row_local_sum_does_not_hide_a_style_outlier(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "A", "B", "C", "Row total"])
    for row in range(2, 22):
        worksheet.append([f"R{row}", row, row + 1, row + 2, f"=SUM(B{row}:D{row})+$A$1"])
    worksheet["C10"].font = Font(bold=True)
    path = tmp_path / "row-local-sum-style-outlier.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)

    assert any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "C10"
        for finding in scan.findings
    )


def test_unlabelled_cross_row_sum_still_marks_a_summary_row(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Amount"])
    for row in range(2, 22):
        worksheet.append([f"R{row}", row * 10])
    worksheet.append(["Result", "=SUM(B2:B21)"])
    worksheet["B22"].font = Font(bold=True)
    path = tmp_path / "unlabelled-summary-row.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)

    assert not any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "B22"
        for finding in scan.findings
    )


@pytest.mark.parametrize("error_type", [ValueError, IndexError, TokenizerError])
def test_malformed_aggregate_formula_is_handled_conservatively(
    monkeypatch: pytest.MonkeyPatch,
    error_type: type[Exception],
) -> None:
    def reject_formula(_formula: str):
        raise error_type("malformed aggregate formula")

    monkeypatch.setattr(builtin_rules, "Tokenizer", reject_formula)

    assert builtin_rules._aggregate_formula_spans_other_rows("=SUM(A1:A2)", 3)


def test_pivot_button_only_difference_is_not_a_visual_style_outlier(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Amount", "Context"])
    for row in range(2, 22):
        worksheet.append([row, f"R{row}"])
    worksheet["A10"].pivotButton = True
    path = tmp_path / "pivot-button-style.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location == "A10"
        for finding in scan.findings
    )
    assert not any(
        patch.kind == PatchKind.COPY_STYLE and patch.cell == "A10" for patch in scan.patches
    )


def test_multiple_style_outliers_are_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Amount", "Context"])
    for row in range(2, 24):
        worksheet.append([row, f"R{row}"])
    worksheet["A8"].font = Font(bold=True)
    worksheet["A18"].font = Font(italic=True)
    path = tmp_path / "multiple-style-outliers.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    locations = {
        finding.location for finding in scan.findings if finding.rule_id == "WL007_STYLE_OUTLIER"
    }
    assert locations == {"A8", "A18"}
    assert not any(
        patch.kind == PatchKind.COPY_STYLE and patch.cell in locations for patch in scan.patches
    )


@pytest.mark.parametrize("label", ["Grand Total", "Summary", "总金额", "累计金额", "期末余额"])
def test_summary_row_style_outliers_are_not_reported_or_auto_copied(
    tmp_path: Path, label: str
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Amount"])
    for row in range(2, 22):
        worksheet.append([f"R{row}", row * 10])
    worksheet.append([label, "=SUM(B2:B21)"])
    worksheet["A22"].font = Font(bold=True)
    worksheet["B22"].font = Font(bold=True)
    path = tmp_path / "summary-row-style.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL007_STYLE_OUTLIER" and finding.location in {"A22", "B22"}
        for finding in scan.findings
    )
    assert not any(
        patch.kind == PatchKind.COPY_STYLE and patch.cell in {"A22", "B22"}
        for patch in scan.patches
    )


@pytest.mark.parametrize("label", ["Total", "Summary", "总金额", "累计金额", "期末余额"])
def test_summary_row_numeric_text_is_findings_only(tmp_path: Path, label: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Amount"])
    for row in range(2, 22):
        worksheet.append([f"R{row}", row * 100])
    worksheet["A10"] = label
    worksheet["B10"] = "1200"
    path = tmp_path / f"summary-numeric-{len(label)}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "B10"
    )
    assert not finding.safe_patch_available
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_secondary_row_semantic_override_blocks_all_candidate_patches(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Row type", "Amount", "Calculated"])
    for row in range(2, 22):
        worksheet.append([f"Item {row}", "Regular", row * 100, f"=C{row}*2"])
    worksheet["B10"] = "Manual override"
    worksheet["C10"] = "1200"
    worksheet["C10"].font = Font(bold=True)
    worksheet["D10"] = 999
    path = tmp_path / "secondary-row-semantic-override.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = {(finding.rule_id, finding.location) for finding in scan.findings}
    assert ("WL004_HARDCODED_VALUE_IN_FORMULA_BAND", "D10") in findings
    assert ("WL006_NUMERIC_TEXT", "C10") in findings
    assert ("WL007_STYLE_OUTLIER", "C10") in findings
    assert not any(patch.cell in {"C10", "D10"} for patch in scan.patches)


@pytest.mark.parametrize(
    ("fault", "rule_id"),
    [
        ("formula", "WL002_FORMULA_PATTERN_OUTLIER"),
        ("blank", "WL003_BLANK_IN_FORMULA_BAND"),
    ],
)
def test_secondary_row_semantic_override_blocks_formula_candidate_patches(
    tmp_path: Path, fault: str, rule_id: str
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Row type", "Input", "Calculated"])
    for row in range(2, 22):
        worksheet.append([f"Item {row}", "Regular", row * 100, f"=C{row}*2"])
    worksheet["B10"] = "调整项"
    worksheet["D10"] = "=C10*3" if fault == "formula" else None
    path = tmp_path / f"secondary-row-{fault}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == rule_id and finding.location == "D10"
    )
    assert not finding.safe_patch_available
    assert not any(patch.cell == "D10" for patch in scan.patches)


def test_visually_marked_blank_formula_gap_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Calculated"])
    for row in range(2, 22):
        worksheet.append([f"Item {row}", f'=A{row}&"-done"'])
    worksheet["B10"] = None
    worksheet["B10"].fill = PatternFill("solid", fgColor="FFF2CC")
    path = tmp_path / "highlighted-blank-formula-gap.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL003_BLANK_IN_FORMULA_BAND" and finding.location == "B10"
    )
    assert not finding.safe_patch_available
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_summary_label_outside_inferred_region_suppresses_style_finding_and_patches(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["C1"] = "Item"
    worksheet["D1"] = "Amount"
    for row in range(2, 22):
        worksheet[f"C{row}"] = f"R{row}"
        worksheet[f"D{row}"] = row * 100
    worksheet["A10"] = "Summary"
    worksheet["D10"] = "1200"
    worksheet["D10"].font = Font(bold=True)
    path = tmp_path / "outside-region-summary.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = {(finding.rule_id, finding.location) for finding in scan.findings}
    assert ("WL006_NUMERIC_TEXT", "D10") in findings
    assert ("WL007_STYLE_OUTLIER", "D10") not in findings
    assert not any(patch.cell == "D10" for patch in scan.patches)


def test_intentionally_highlighted_row_blocks_content_and_style_patches(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Item", "Amount", "Calculated"])
    for row in range(2, 22):
        worksheet.append([f"Item {row}", row * 100, f"=B{row}*2"])
    worksheet["B10"] = "1200"
    worksheet["C10"] = 999
    highlight = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[10]:
        cell.font = Font(bold=True)
        cell.fill = highlight
    path = tmp_path / "highlighted-override-row.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = {(finding.rule_id, finding.location) for finding in scan.findings}
    assert ("WL004_HARDCODED_VALUE_IN_FORMULA_BAND", "C10") in findings
    assert ("WL006_NUMERIC_TEXT", "B10") in findings
    assert ("WL007_STYLE_OUTLIER", "B10") in findings
    assert not any(patch.cell in {"B10", "C10"} for patch in scan.patches)


def test_freeform_only_row_labels_keep_formula_and_style_findings_review_only(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    names = [
        "Alice",
        "Bob",
        "Carol",
        "Diego",
        "Eve",
        "Fatima",
        "Grace",
        "Special Case",
        "Hiro",
        "Iris",
        "Jamal",
        "Kai",
        "Lina",
        "Marta",
        "Nora",
        "Omar",
        "Pia",
        "Quinn",
        "Ravi",
        "Sara",
    ]
    worksheet.append(["Name", "Amount", "Calculated"])
    for row, name in enumerate(names, start=2):
        worksheet.append([name, row * 100, f"=B{row}*2"])
    worksheet["B9"].font = Font(bold=True)
    worksheet["C9"] = 999
    path = tmp_path / "freeform-label-special-case.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    findings = {(finding.rule_id, finding.location) for finding in scan.findings}
    assert ("WL004_HARDCODED_VALUE_IN_FORMULA_BAND", "C9") in findings
    assert ("WL007_STYLE_OUTLIER", "B9") in findings
    assert not any(patch.cell in {"B9", "C9"} for patch in scan.patches)


def test_hidden_adjacent_row_suppresses_sum_boundary(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 10):
        worksheet.cell(row, 2, row)
    worksheet["B10"] = "=SUM(B2:B8)"
    worksheet.row_dimensions[9].hidden = True
    path = tmp_path / "hidden-boundary.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    assert not any(finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY" for finding in scan.findings)


def test_sum_boundary_numeric_candidate_is_findings_only(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 10):
        worksheet.cell(row, 1, f"R{row}")
        worksheet.cell(row, 2, row * 100)
    worksheet["A10"] = "Total"
    worksheet["B10"] = "=SUM(B2:B8)"
    path = tmp_path / "numeric-boundary-findings-only.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding for finding in scan.findings if finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY"
    )
    assert finding.location == "B10"
    assert finding.evidence.expected == "=SUM(B2:B9)"
    assert not finding.safe_patch_available
    assert not any(patch.cell == "B10" for patch in scan.patches)


@pytest.mark.parametrize(
    "label",
    [
        "Subtotal",
        "Sub-total",
        "Grand-total",
        "Total adjustment",
        "Average",
        "Mean",
        "Minimum",
        "Maximum",
        "Summary",
        "小计",
        "合计金额",
        "汇总金额",
        "平均金额",
        "最大金额",
        "最小金额",
        "总金额",
        "累计金额",
        "净额",
        "期末余额",
    ],
)
def test_sum_boundary_subtotal_candidate_is_findings_only(tmp_path: Path, label: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 10):
        worksheet.cell(row, 2, row * 100)
    worksheet["A9"] = label
    worksheet["B10"] = "=SUM(B2:B8)"
    path = tmp_path / f"subtotal-boundary-{len(label)}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_sum_boundary_does_not_cross_explicit_sheet_reference(tmp_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = "Summary"
    source = workbook.create_sheet("Source")
    for row in range(2, 10):
        source.cell(row, 2, row)
    summary["B9"] = 999
    summary["B10"] = "=SUM(Source!B2:B8)"
    path = tmp_path / "cross-sheet-sum.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    assert not any(finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY" for finding in scan.findings)


def test_sum_boundary_reports_adjacent_formula_without_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 9):
        worksheet.cell(row, 1, f"R{row}")
        worksheet.cell(row, 2, row)
    worksheet["A9"] = "R9"
    worksheet["B9"] = "=1+1"
    worksheet["A10"] = "Total"
    worksheet["B10"] = "=SUM(B2:B8)"
    path = tmp_path / "formula-adjacent-sum.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    finding = next(
        finding for finding in scan.findings if finding.rule_id == "WL005_SUSPICIOUS_SUM_BOUNDARY"
    )
    assert finding.location == "B10"
    assert not finding.safe_patch_available
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_scientific_notation_text_is_not_auto_converted(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Measure", "Context"])
    for row in range(2, 10):
        worksheet.append([row, f"R{row}"])
    worksheet["A5"] = "1e3"
    path = tmp_path / "scientific-text.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL006_NUMERIC_TEXT" and finding.location == "A5"
        for finding in scan.findings
    )


def test_hidden_grouped_columns_report_full_range(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["B1"] = "hidden B"
    worksheet["C1"] = "hidden C"
    worksheet.column_dimensions.group("B", "C", hidden=True)
    path = tmp_path / "hidden-columns.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    locations = {
        finding.location
        for finding in scan.findings
        if finding.rule_id == "WL008_HIDDEN_NONEMPTY_DATA"
    }
    assert "B:C" in locations


def test_merged_header_is_not_reported_as_data_body_merge(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "Header"
    for row in range(2, 8):
        worksheet.cell(row, 1, row)
        worksheet.cell(row, 2, row * 2)
    path = tmp_path / "merged-header.xlsx"
    workbook.save(path)
    workbook.close()
    scan = scan_workbook(path)
    assert not any(
        finding.rule_id == "WL014_MERGED_CELL_IN_DATA_REGION" for finding in scan.findings
    )


def test_merged_non_anchor_formula_gap_is_never_auto_patched(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(4, 12):
        for column in range(2, 7):
            worksheet.cell(row, column, row + column)
    for column in (2, 3, 5, 6):
        coordinate = worksheet.cell(12, column).coordinate
        source = worksheet.cell(11, column).coordinate
        worksheet[coordinate] = f"={source}*2"
    worksheet.merge_cells("C12:D12")
    path = tmp_path / "merged-non-anchor.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    finding = next(
        finding
        for finding in scan.findings
        if finding.rule_id == "WL003_BLANK_IN_FORMULA_BAND" and finding.location == "D12"
    )
    assert not finding.safe_patch_available
    assert not any(patch.cell == "D12" for patch in scan.patches)


def test_merged_anchor_formula_gap_is_never_auto_patched(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in (2, 3, 5, 6):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet.merge_cells("B4:C4")
    path = tmp_path / "merged-anchor-gap.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL003_BLANK_IN_FORMULA_BAND" and finding.location == "B4"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B4" for patch in scan.patches)


def test_merged_anchor_literal_is_never_replaced(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 22):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet["B10"] = "Merged note"
    worksheet.merge_cells("B10:C10")
    path = tmp_path / "merged-anchor-literal.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL004_HARDCODED_VALUE_IN_FORMULA_BAND" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


@pytest.mark.parametrize(
    "label",
    [
        "Subtotal",
        "Sub-total",
        "Grand-total",
        "Average",
        "Mean",
        "Minimum",
        "Maximum",
        "小计",
        "小计金额",
        "合计金额",
        "汇总金额",
        "平均金额",
        "最大金额",
        "最小金额",
    ],
)
def test_summary_row_literal_is_never_replaced(tmp_path: Path, label: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 22):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet["A10"] = label
    worksheet["B10"] = 999
    path = tmp_path / f"summary-literal-{len(label)}.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL004_HARDCODED_VALUE_IN_FORMULA_BAND" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_hidden_literal_is_never_replaced(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(2, 22):
        worksheet[f"B{row}"] = f"=A{row}*2"
    worksheet["B10"] = 999
    worksheet.row_dimensions[10].hidden = True
    path = tmp_path / "hidden-literal.xlsx"
    workbook.save(path)
    workbook.close()

    scan = scan_workbook(path)
    assert any(
        finding.rule_id == "WL004_HARDCODED_VALUE_IN_FORMULA_BAND" and finding.location == "B10"
        for finding in scan.findings
    )
    assert not any(patch.cell == "B10" for patch in scan.patches)


def test_all_twenty_one_builtin_rule_ids_are_stable() -> None:
    assert {rule.rule_id for rule in BUILTIN_RULES} == {
        f"WL{number:03d}_{suffix}"
        for number, suffix in enumerate(
            [
                "BROKEN_REFERENCE",
                "FORMULA_PATTERN_OUTLIER",
                "BLANK_IN_FORMULA_BAND",
                "HARDCODED_VALUE_IN_FORMULA_BAND",
                "SUSPICIOUS_SUM_BOUNDARY",
                "NUMERIC_TEXT",
                "STYLE_OUTLIER",
                "HIDDEN_NONEMPTY_DATA",
                "EXTERNAL_LINK",
                "VOLATILE_OR_FRAGILE_FUNCTION",
                "ERROR_CELL",
                "DUPLICATE_CONFIGURED_KEY",
                "BROKEN_DEFINED_NAME",
                "MERGED_CELL_IN_DATA_REGION",
                "INCONSISTENT_DATA_VALIDATION",
                "TEXT_DISPLAY_RISK",
                "BORDER_EDGE_INCONSISTENCY",
                "USED_RANGE_INFLATION",
                "IDENTIFIER_SCIENTIFIC_NOTATION",
                "SAVED_VIEW_OFF_CONTENT",
                "WHITESPACE_ONLY_TAIL",
            ],
            start=1,
        )
    }
