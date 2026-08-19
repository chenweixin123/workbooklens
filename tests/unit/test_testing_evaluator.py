from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from workbooklens.exceptions import UsageError
from workbooklens.testing import evaluate_workbook_tests, load_test_config


def _assertion_workbook(path: Path) -> None:
    workbook = Workbook()
    data = workbook.active
    assert data is not None
    data.title = "Data"
    data.append([1, "K1", "Open"])
    data.append([2, "K2", "Paid"])
    data.append([3, "K3", "Void"])
    summary = workbook.create_sheet("Summary")
    summary["B1"] = 6
    workbook.save(path)
    workbook.close()


def _write_config(path: Path, assertions: list[dict[str, object]]) -> None:
    path.write_text(
        yaml.safe_dump(
            {
                "version": 1,
                "workbook": {"max_critical_findings": 0, "max_error_findings": 0},
                "assertions": assertions,
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )


def test_all_supported_assertion_types_pass(tmp_path: Path) -> None:
    workbook = tmp_path / "assertions.xlsx"
    config_path = tmp_path / "workbooklens.yml"
    _assertion_workbook(workbook)
    _write_config(
        config_path,
        [
            {"id": "clean", "type": "no_findings", "rules": ["WL001_BROKEN_REFERENCE"]},
            {
                "id": "unique",
                "type": "unique",
                "sheet": "Data",
                "range": "B1:B3",
            },
            {
                "id": "domain",
                "type": "allowed_values",
                "sheet": "Data",
                "range": "C1:C3",
                "values": ["Open", "Paid", "Void"],
            },
            {"id": "filled", "type": "nonblank", "sheet": "Data", "range": "A1:C3"},
            {
                "id": "bounds",
                "type": "numeric_bounds",
                "sheet": "Data",
                "range": "A1:A3",
                "minimum": 1,
                "maximum": 3,
            },
            {
                "id": "reconciles",
                "type": "equals",
                "left": "Summary!B1",
                "right": "SUM(Data!A1:A3)",
                "tolerance": 0.001,
            },
        ],
    )
    run = evaluate_workbook_tests(workbook, load_test_config(config_path))
    assert run.passed
    assert len(run.results) == 8


def test_assertion_failures_report_observed_values(tmp_path: Path) -> None:
    workbook = tmp_path / "assertions.xlsx"
    config_path = tmp_path / "workbooklens.yml"
    _assertion_workbook(workbook)
    _write_config(
        config_path,
        [
            {
                "id": "wrong-total",
                "type": "equals",
                "left": "Summary!B1",
                "right": "7",
            },
            {
                "id": "too-low",
                "type": "numeric_bounds",
                "sheet": "Data",
                "range": "A1:A3",
                "minimum": 2,
            },
            {
                "id": "unsupported",
                "type": "equals",
                "left": "AVERAGE(Data!A1:A3)",
                "right": "2",
            },
        ],
    )
    run = evaluate_workbook_tests(workbook, load_test_config(config_path))
    assert not run.passed
    by_id = {result.assertion_id: result for result in run.results}
    assert by_id["wrong-total"].observed == 6
    assert by_id["too-low"].observed == [{"cell": "A1", "value": 1}]
    assert "unsupported expression" in by_id["unsupported"].message


def test_formula_cell_without_cache_fails_instead_of_guessing(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet"
    sheet["A1"] = 2
    sheet["A2"] = 4
    sheet["B1"] = "=AVERAGE(A1:A2)"
    workbook.save(workbook_path)
    workbook.close()
    config_path = tmp_path / "workbooklens.yml"
    _write_config(
        config_path,
        [{"id": "cached", "type": "equals", "left": "Sheet!B1", "right": "3"}],
    )
    run = evaluate_workbook_tests(workbook_path, load_test_config(config_path))
    result = next(item for item in run.results if item.assertion_id == "cached")
    assert not result.passed
    assert "without a cached value" in result.message


def test_sum_range_with_uncached_formula_fails_instead_of_treating_it_as_blank(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "formula-in-range.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet"
    sheet["A1"] = 1
    sheet["A2"] = "=1+1"
    workbook.save(workbook_path)
    workbook.close()
    config_path = tmp_path / "workbooklens.yml"
    _write_config(
        config_path,
        [{"id": "sum", "type": "equals", "left": "SUM(Sheet!A1:A2)", "right": "3"}],
    )
    run = evaluate_workbook_tests(workbook_path, load_test_config(config_path))
    result = next(item for item in run.results if item.assertion_id == "sum")
    assert not result.passed
    assert "formula at Sheet!A2 has no cached value" in result.message


def test_sum_range_with_error_cell_fails_instead_of_ignoring_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "error-in-range.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet"
    sheet["A1"] = 1
    sheet["A2"] = "#DIV/0!"
    workbook.save(workbook_path)
    workbook.close()
    config_path = tmp_path / "workbooklens.yml"
    _write_config(
        config_path,
        [{"id": "sum", "type": "equals", "left": "SUM(Sheet!A1:A2)", "right": "1"}],
    )
    run = evaluate_workbook_tests(workbook_path, load_test_config(config_path))
    result = next(item for item in run.results if item.assertion_id == "sum")
    assert not result.passed
    assert "SUM range contains error" in result.message


def test_config_is_size_bounded_and_safe_loaded(tmp_path: Path) -> None:
    oversized = tmp_path / "oversized.yml"
    oversized.write_bytes(b"x" * 20)
    with pytest.raises(UsageError, match="exceeds"):
        load_test_config(oversized, max_bytes=10)
    unsafe = tmp_path / "unsafe.yml"
    unsafe.write_text("!!python/object/apply:os.system ['echo unsafe']", encoding="utf-8")
    with pytest.raises(UsageError, match="Invalid UTF-8 YAML"):
        load_test_config(unsafe)
    alias_heavy = tmp_path / "aliases.yml"
    alias_heavy.write_text(
        "version: 1\nanchor: &anchor {}\nassertions:\n" + "  - *anchor\n" * 51,
        encoding="utf-8",
    )
    with pytest.raises(UsageError, match="alias count exceeds"):
        load_test_config(alias_heavy)
    deeply_nested = tmp_path / "nested.yml"
    deeply_nested.write_text(
        "version: 1\nassertions: " + "[" * 101 + "]" * 101,
        encoding="utf-8",
    )
    with pytest.raises(UsageError, match="nesting exceeds"):
        load_test_config(deeply_nested)


def test_suppressed_findings_remain_auditable_in_test_run(tmp_path: Path) -> None:
    workbook_path = tmp_path / "broken.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    sheet["A1"] = "=#REF!+1"
    workbook.save(workbook_path)
    workbook.close()
    config_path = tmp_path / "workbooklens.yml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": 2,
                "workbook": {"max_error_findings": 0},
                "suppressions": [
                    {
                        "id": "accepted-fixture",
                        "reason": "Synthetic fixture is intentionally broken",
                        "rules": ["WL001_BROKEN_REFERENCE"],
                        "sheets": ["Data"],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    run = evaluate_workbook_tests(workbook_path, load_test_config(config_path))
    assert run.passed
    assert run.policy.summary["suppressed"] == 1
    assert run.policy.suppressed_findings[0].suppression_id == "accepted-fixture"


def test_explicit_range_limit_fails_assertion_clearly(tmp_path: Path) -> None:
    workbook = tmp_path / "assertions.xlsx"
    config_path = tmp_path / "workbooklens.yml"
    _assertion_workbook(workbook)
    _write_config(
        config_path,
        [{"id": "huge", "type": "nonblank", "sheet": "Data", "range": "A1:XFD1048576"}],
    )
    run = evaluate_workbook_tests(workbook, load_test_config(config_path))
    result = next(item for item in run.results if item.assertion_id == "huge")
    assert not result.passed
    assert "limit" in result.message
