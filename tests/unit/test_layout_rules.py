from __future__ import annotations

from copy import copy
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.views import Pane
from openpyxl.worksheet.worksheet import Worksheet

from workbooklens.layout import measure_text_cell
from workbooklens.models import PatchKind, PatchRisk
from workbooklens.repair.engine import apply_patch_plan
from workbooklens.repair.planning import build_patch_plan
from workbooklens.rules.builtin import _finite_nonnegative_pane_split
from workbooklens.scanner import ScanResult, scan_workbook


def _scan(workbook: Workbook, path: Path) -> ScanResult:
    workbook.save(path)
    workbook.close()
    return scan_workbook(path)


def _findings(scan: ScanResult, rule_id: str):
    return [finding for finding in scan.findings if finding.rule_id == rule_id]


def test_wrapped_text_with_small_explicit_height_gets_review_patch(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "A deliberately long wrapped heading that needs several lines"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    worksheet.column_dimensions["A"].width = 6
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "wrapped.xlsx")

    findings = _findings(scan, "WL016_TEXT_DISPLAY_RISK")
    assert [finding.location for finding in findings] == ["A1"]
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_ROW_HEIGHT)
    assert patch.after["height"] > 15
    assert not patch.safe
    assert patch.risk == PatchRisk.LAYOUT_REVIEW


def test_auto_height_and_natural_overflow_are_not_reported(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "A long wrapped title that Excel may auto-size"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    worksheet.column_dimensions["A"].width = 5
    worksheet["C3"] = "A long unwrapped title with blank cells to its right"
    worksheet.column_dimensions["C"].width = 5

    scan = _scan(workbook, tmp_path / "natural-overflow.xlsx")

    assert not _findings(scan, "WL016_TEXT_DISPLAY_RISK")


def test_blocked_overflow_uses_atomic_wrap_and_row_height(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Long text that cannot spill into the occupied adjacent cell"
    worksheet["B1"] = "blocker"
    worksheet.column_dimensions["A"].width = 6
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "blocked.xlsx")

    patches = [
        patch
        for patch in scan.patches
        if patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
    ]
    assert {patch.kind for patch in patches} == {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_ROW_HEIGHT,
    }
    assert len({patch.atomic_group for patch in patches}) == 1
    assert patches[0].atomic_group is not None
    assert all(not patch.safe and patch.risk == PatchRisk.LAYOUT_REVIEW for patch in patches)


def test_default_height_blocked_overflow_applies_without_stale_layout_fingerprint(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Long text in an automatic-height row that must wrap inside its border"
    worksheet["B1"] = "blocker"
    worksheet.column_dimensions["A"].width = 7
    source = tmp_path / "default-height-blocked.xlsx"
    scan = _scan(workbook, source)

    patches = [
        patch
        for patch in scan.patches
        if patch.cell == "A1" and patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
    ]
    assert {patch.kind for patch in patches} == {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_ROW_HEIGHT,
    }

    output = tmp_path / "default-height-blocked-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
    )

    repaired = load_workbook(output)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    assert repaired_sheet["A1"].alignment.wrap_text
    assert repaired_sheet.row_dimensions[1].height is not None
    repaired.close()
    assert not _findings(scan_workbook(output), "WL016_TEXT_DISPLAY_RISK")


def test_row_height_covers_existing_vertical_and_new_horizontal_issue(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Existing wrapped vertical clipping"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    worksheet["B1"] = "New horizontal overflow needs substantially more wrapped height"
    worksheet["C1"] = "blocker"
    worksheet["D1"] = "peer"
    worksheet["E1"] = "peer"
    for coordinate in ("C1", "D1", "E1"):
        worksheet[coordinate].alignment = Alignment(wrap_text=True)
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 6
    worksheet.row_dimensions[1].height = 15

    source = tmp_path / "mixed-row-display-risk.xlsx"
    scan = _scan(workbook, source)

    findings = _findings(scan, "WL016_TEXT_DISPLAY_RISK")
    assert {finding.location for finding in findings} == {"A1", "B1"}
    row_patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_ROW_HEIGHT)
    wrap_patch = next(
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.SET_WRAP_TEXT and patch.cell == "B1"
    )
    measured = load_workbook(source)
    measured_sheet = measured.active
    assert measured_sheet is not None
    wrapped_measurement = measure_text_cell(measured_sheet, measured_sheet["B1"], assume_wrap=True)
    assert wrapped_measurement is not None
    assert row_patch.after["height"] >= wrapped_measurement.required_height
    measured.close()
    assert row_patch.atomic_group == wrap_patch.atomic_group

    repaired = load_workbook(source)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    alignment = copy(repaired_sheet["B1"].alignment)
    alignment.wrap_text = True
    repaired_sheet["B1"].alignment = alignment
    repaired_sheet.row_dimensions[1].height = row_patch.after["height"]
    repaired_path = tmp_path / "mixed-row-display-risk-repaired.xlsx"
    repaired.save(repaired_path)
    repaired.close()

    assert not _findings(scan_workbook(repaired_path), "WL016_TEXT_DISPLAY_RISK")


def test_hidden_adjacent_columns_do_not_count_as_natural_overflow_space(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Hidden overflow"
    worksheet["D1"] = "visible blocker"
    worksheet.column_dimensions["A"].width = 5
    for column in ("B", "C"):
        worksheet.column_dimensions[column].width = 10
        worksheet.column_dimensions[column].hidden = True
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "hidden-overflow-space.xlsx")

    assert [finding.location for finding in _findings(scan, "WL016_TEXT_DISPLAY_RISK")] == ["A1"]
    patches = [
        patch
        for patch in scan.patches
        if patch.cell == "A1" and patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
    ]
    assert {patch.kind for patch in patches} == {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_ROW_HEIGHT,
    }


def test_repeated_boundary_overflow_uses_one_conservative_column_width(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    for row in range(1, 5):
        worksheet.cell(row, 7, "Repeated long text at the right edge")
        worksheet.cell(row, 7).border = Border(right=thin)
    worksheet.column_dimensions["G"].width = 8

    source = tmp_path / "repeated-overflow.xlsx"
    scan = _scan(workbook, source)

    width_patches = [patch for patch in scan.patches if patch.kind == PatchKind.SET_COLUMN_WIDTH]
    assert len(width_patches) == 1
    assert width_patches[0].after["column"] == "G"
    assert 8.5 <= width_patches[0].after["width"] <= 40
    assert not any(patch.kind == PatchKind.SET_WRAP_TEXT for patch in scan.patches)
    assert all(
        finding.patch_ids == [width_patches[0].id]
        for finding in _findings(scan, "WL016_TEXT_DISPLAY_RISK")
    )

    repaired = load_workbook(source)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    repaired_sheet.column_dimensions["G"].width = width_patches[0].after["width"]
    repaired_path = tmp_path / "repeated-overflow-repaired.xlsx"
    repaired.save(repaired_path)
    repaired.close()
    repaired_scan = scan_workbook(repaired_path)
    assert not _findings(repaired_scan, "WL016_TEXT_DISPLAY_RISK")


def test_repeated_overflow_at_width_cap_falls_back_to_atomic_wrap_and_height(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    for row in range(1, 4):
        worksheet.cell(row, 1, "x" * 120)
        worksheet.cell(row, 1).border = Border(right=thin)
        worksheet.row_dimensions[row].height = 15
    worksheet.column_dimensions["A"].width = 40

    scan = _scan(workbook, tmp_path / "width-cap.xlsx")

    assert not any(patch.kind == PatchKind.SET_COLUMN_WIDTH for patch in scan.patches)
    for row in range(1, 4):
        row_patches = [
            patch
            for patch in scan.patches
            if patch.cell == f"A{row}"
            and patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
        ]
        assert {patch.kind for patch in row_patches} == {
            PatchKind.SET_WRAP_TEXT,
            PatchKind.SET_ROW_HEIGHT,
        }
        assert len({patch.atomic_group for patch in row_patches}) == 1
        assert row_patches[0].atomic_group is not None


def test_extreme_blocked_overflow_is_finding_only_above_excel_row_height_limit(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "x" * 2_000
    worksheet["B1"] = "blocker"
    worksheet.column_dimensions["A"].width = 6
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "extreme-overflow.xlsx")

    findings = _findings(scan, "WL016_TEXT_DISPLAY_RISK")
    assert [finding.location for finding in findings] == ["A1"]
    assert findings[0].patch_ids == []
    assert not any(
        patch.cell == "A1" and patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
        for patch in scan.patches
    )


def test_overflow_crossing_a_later_blank_cell_border_is_reported(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Long text that crosses more than one otherwise blank cell"
    worksheet.column_dimensions["A"].width = 5
    worksheet.column_dimensions["B"].width = 5
    worksheet["B1"].border = Border(right=Side(style="thin"))
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "later-border.xlsx")

    assert [finding.location for finding in _findings(scan, "WL016_TEXT_DISPLAY_RISK")] == ["A1"]
    patches = [
        patch
        for patch in scan.patches
        if patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
    ]
    assert {patch.kind for patch in patches} == {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_ROW_HEIGHT,
    }


def test_single_row_merged_title_uses_atomic_wrap_and_height(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "A merged title that is substantially wider than the merged region"
    worksheet.column_dimensions["A"].width = 7
    worksheet.column_dimensions["B"].width = 7
    worksheet.column_dimensions["C"].width = 7
    worksheet.row_dimensions[1].height = 15
    source = tmp_path / "merged-title.xlsx"
    scan = _scan(workbook, source)

    findings = _findings(scan, "WL016_TEXT_DISPLAY_RISK")
    assert [finding.location for finding in findings] == ["A1"]
    patches = [
        patch
        for patch in scan.patches
        if patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_ROW_HEIGHT}
    ]
    assert {patch.kind for patch in patches} == {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_ROW_HEIGHT,
    }
    assert len({patch.atomic_group for patch in patches}) == 1
    assert patches[0].atomic_group is not None
    assert not any(patch.kind == PatchKind.SET_COLUMN_WIDTH for patch in scan.patches)

    output = tmp_path / "merged-title-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
    )
    repaired = load_workbook(output)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    assert [str(cell_range) for cell_range in repaired_sheet.merged_cells.ranges] == ["A1:C1"]
    assert repaired_sheet["A1"].alignment.wrap_text
    assert repaired_sheet.row_dimensions[1].height is not None
    repaired.close()
    assert not _findings(scan_workbook(output), "WL016_TEXT_DISPLAY_RISK")


def _border_grid(*, remove_both_sides: bool) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, 6):
        for column in range(1, 4):
            cell = worksheet.cell(row, column, f"R{row}C{column}")
            cell.border = full
    worksheet["B3"].border = Border(left=thin, top=thin, bottom=thin)
    if remove_both_sides:
        worksheet["C3"].border = Border(right=thin, top=thin, bottom=thin)
    return workbook


def test_single_sided_shared_border_is_not_reported(tmp_path: Path) -> None:
    scan = _scan(_border_grid(remove_both_sides=False), tmp_path / "one-sided.xlsx")

    assert not _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")


def test_border_missing_on_both_sides_gets_edge_only_patch(tmp_path: Path) -> None:
    scan = _scan(_border_grid(remove_both_sides=True), tmp_path / "missing-border.xlsx")

    findings = _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    assert len(findings) == 1
    patches = [patch for patch in scan.patches if patch.kind == PatchKind.COPY_BORDER]
    assert len(patches) == 1
    assert patches[0].after["target_edge"] in {"left", "right"}
    assert patches[0].source_cell is not None


def test_internal_blank_cell_with_both_shared_sides_missing_gets_edge_only_patch(
    tmp_path: Path,
) -> None:
    workbook = _border_grid(remove_both_sides=True)
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["B3"] = None

    scan = _scan(workbook, tmp_path / "blank-cell-missing-border.xlsx")

    findings = _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    assert [finding.location for finding in findings] == ["B3"]
    patches = [
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.COPY_BORDER and patch.cell == "B3"
    ]
    assert len(patches) == 1
    assert set(patches[0].after) == {"target_edge", "source_edge"}
    assert patches[0].after["target_edge"] in {"left", "right", "top", "bottom"}
    assert patches[0].after["source_edge"] in {"left", "right", "top", "bottom"}
    assert patches[0].source_cell is not None


def test_border_rule_does_not_box_intentional_ragged_note_cell(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, 8):
        for column in range(1, 6):
            worksheet.cell(row, column, f"R{row}C{column}").border = full
    worksheet["C8"] = "Open note"
    worksheet["C8"].border = Border(top=thin)

    scan = _scan(workbook, tmp_path / "ragged-note.xlsx")

    assert not any(
        finding.location == "C8" for finding in _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    )
    assert not any(
        patch.kind == PatchKind.COPY_BORDER and patch.cell == "C8" for patch in scan.patches
    )


def _rectangular_border_grid(rows: int = 5, columns: int = 5) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            worksheet.cell(row, column, f"R{row}C{column}").border = full
    return workbook


def test_completely_borderless_internal_hole_is_repaired_and_rescans_clean(
    tmp_path: Path,
) -> None:
    workbook = _rectangular_border_grid()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    worksheet["C3"].border = Border()
    worksheet["B3"].border = Border(left=thin, top=thin, bottom=thin)
    worksheet["D3"].border = Border(right=thin, top=thin, bottom=thin)
    worksheet["C2"].border = Border(left=thin, right=thin, top=thin)
    worksheet["C4"].border = Border(left=thin, right=thin, bottom=thin)
    source = tmp_path / "complete-border-hole.xlsx"

    scan = _scan(workbook, source)

    finding = next(
        finding
        for finding in _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
        if finding.location == "C3"
    )
    patches = [
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.COPY_BORDER and patch.cell == "C3"
    ]
    assert len(patches) == 4
    assert set(finding.patch_ids) == {patch.id for patch in patches}
    assert {patch.after["target_edge"] for patch in patches} == {
        "left",
        "right",
        "top",
        "bottom",
    }
    assert len({patch.atomic_group for patch in patches}) == 1
    assert patches[0].atomic_group is not None

    output = tmp_path / "complete-border-hole-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
    )

    assert not _findings(scan_workbook(output), "WL017_BORDER_EDGE_INCONSISTENCY")


def test_all_four_table_perimeter_edges_are_repaired_and_rescan_clean(
    tmp_path: Path,
) -> None:
    workbook = _rectangular_border_grid()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    worksheet["A3"].border = Border(right=thin, top=thin, bottom=thin)
    worksheet["E3"].border = Border(left=thin, top=thin, bottom=thin)
    worksheet["C1"].border = Border(left=thin, right=thin, bottom=thin)
    worksheet["C5"].border = Border(left=thin, right=thin, top=thin)
    source = tmp_path / "missing-table-perimeter.xlsx"

    scan = _scan(workbook, source)

    findings = _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    assert {finding.location for finding in findings} == {"A3", "C1", "C5", "E3"}
    patches = [patch for patch in scan.patches if patch.kind == PatchKind.COPY_BORDER]
    assert {(patch.cell, patch.after["target_edge"]) for patch in patches} == {
        ("A3", "left"),
        ("E3", "right"),
        ("C1", "top"),
        ("C5", "bottom"),
    }

    output = tmp_path / "missing-table-perimeter-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
    )

    assert not _findings(scan_workbook(output), "WL017_BORDER_EDGE_INCONSISTENCY")


def test_border_finding_below_point_nine_five_has_no_patch(tmp_path: Path) -> None:
    workbook = _rectangular_border_grid(rows=6, columns=4)
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    for row in (3, 6):
        worksheet.cell(row, 2).border = Border(left=thin, top=thin, bottom=thin)
        worksheet.cell(row, 3).border = Border(right=thin, top=thin, bottom=thin)

    scan = _scan(workbook, tmp_path / "low-confidence-border.xlsx")

    finding = next(
        finding
        for finding in _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
        if finding.location == "B3"
    )
    assert 0.75 <= float(finding.confidence) < 0.95
    assert finding.patch_ids == []
    assert not any(
        patch.kind == PatchKind.COPY_BORDER and patch.cell == "B3" for patch in scan.patches
    )


def test_border_rule_skips_merged_target_and_small_or_sparse_rectangles(tmp_path: Path) -> None:
    merged = _rectangular_border_grid()
    merged_sheet = merged.active
    assert merged_sheet is not None
    merged_sheet.merge_cells("B3:C3")
    merged_sheet["B3"].border = Border()
    merged_scan = _scan(merged, tmp_path / "merged-border-gap.xlsx")
    assert not any(
        finding.location in {"B3", "C3"}
        for finding in _findings(merged_scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    )

    small = _rectangular_border_grid(rows=2, columns=3)
    small_sheet = small.active
    assert small_sheet is not None
    thin = Side(style="thin")
    small_sheet["B1"].border = Border(left=thin, top=thin, bottom=thin)
    small_sheet["C1"].border = Border(right=thin, top=thin, bottom=thin)
    small_scan = _scan(small, tmp_path / "small-border-grid.xlsx")
    assert not _findings(small_scan, "WL017_BORDER_EDGE_INCONSISTENCY")

    sparse = Workbook()
    sparse_sheet = sparse.active
    assert sparse_sheet is not None
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for coordinate in ("A1", "B1", "B2", "C2", "C3", "D3", "C4", "D4"):
        sparse_sheet[coordinate] = coordinate
        sparse_sheet[coordinate].border = full
    sparse_sheet["B2"].border = Border(left=thin, top=thin, bottom=thin)
    sparse_sheet["C2"].border = Border(right=thin, top=thin, bottom=thin)
    sparse_scan = _scan(sparse, tmp_path / "sparse-border-grid.xlsx")
    assert not _findings(sparse_scan, "WL017_BORDER_EDGE_INCONSISTENCY")


def test_exact_formatting_tail_patch_and_broad_column_negative(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in range(30, 34):
        for column in range(10, 14):
            worksheet.cell(row, column).fill = fill
    scan = _scan(workbook, tmp_path / "tail.xlsx")

    patches = [patch for patch in scan.patches if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL]
    assert len(patches) == 1
    assert len(patches[0].after["cells"]) == 16
    assert patches[0].after["result_dimension"] == "A1"

    negative = Workbook()
    negative_sheet = negative.active
    assert negative_sheet is not None
    negative_sheet["A1"] = "content"
    dimension = negative_sheet.column_dimensions["L"]
    dimension.min = 12
    dimension.max = 16_384
    dimension.width = 12
    negative_scan = _scan(negative, tmp_path / "broad-column.xlsx")
    assert not _findings(negative_scan, "WL018_USED_RANGE_INFLATION")


def test_identifier_patch_is_width_only_and_preserves_numeric_semantics(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["J1"] = "联系电话"
    worksheet["J2"] = 19_885_186_321
    worksheet.column_dimensions["J"].width = 8
    worksheet["K1"] = "金额"
    worksheet["K2"] = 19_885_186_321
    worksheet.column_dimensions["K"].width = 8

    scan = _scan(workbook, tmp_path / "identifier.xlsx")

    findings = _findings(scan, "WL019_IDENTIFIER_SCIENTIFIC_NOTATION")
    assert [finding.location for finding in findings] == ["J2"]
    patches = [patch for patch in scan.patches if patch.kind == PatchKind.SET_COLUMN_WIDTH]
    assert len(patches) == 1
    assert patches[0].after["column"] == "J"
    assert patches[0].atomic_group is None
    assert not any(patch.kind == PatchKind.SET_TEXT for patch in scan.patches)
    assert findings[0].patch_ids == [patches[0].id]


def test_same_column_width_requests_are_coalesced_to_the_sufficient_maximum(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["J1"] = "联系电话"
    worksheet["J2"] = 19_885_186_321
    thin = Side(style="thin")
    for row in range(3, 6):
        worksheet.cell(row, 10, "Repeated long text at the right edge")
        worksheet.cell(row, 10).border = Border(right=thin)
    worksheet.column_dimensions["J"].width = 8

    source = tmp_path / "coalesced-width.xlsx"
    scan = _scan(workbook, source)

    width_patches = [patch for patch in scan.patches if patch.kind == PatchKind.SET_COLUMN_WIDTH]
    assert len(width_patches) == 1
    assert width_patches[0].after["column"] == "J"
    assert width_patches[0].after["width"] > 20
    related = [
        finding
        for finding in scan.findings
        if finding.rule_id in {"WL016_TEXT_DISPLAY_RISK", "WL019_IDENTIFIER_SCIENTIFIC_NOTATION"}
    ]
    assert related
    assert all(finding.patch_ids == [width_patches[0].id] for finding in related)

    output = tmp_path / "coalesced-width-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={width_patches[0].id},
        accept_layout_risk=True,
    )
    repaired_scan = scan_workbook(output)
    assert not _findings(repaired_scan, "WL016_TEXT_DISPLAY_RISK")
    assert not _findings(repaired_scan, "WL019_IDENTIFIER_SCIENTIFIC_NOTATION")


def test_styled_wrapped_phone_at_old_width_still_gets_sufficient_numeric_preserving_fix(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["J1"] = "联系电话"
    worksheet["J2"] = 19_885_186_321
    worksheet["J2"].font = Font(name="仿宋", size=14)
    worksheet["J2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.column_dimensions["J"].width = 13.6

    source = tmp_path / "styled-phone-width.xlsx"
    scan = _scan(workbook, source)

    findings = _findings(scan, "WL019_IDENTIFIER_SCIENTIFIC_NOTATION")
    assert [finding.location for finding in findings] == ["J2"]
    width_patch = next(
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.SET_COLUMN_WIDTH and patch.cell == "J2"
    )
    assert width_patch.after["width"] >= 15.0
    assert not any(patch.kind == PatchKind.SET_TEXT for patch in scan.patches)

    repaired = load_workbook(source)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    repaired_sheet.column_dimensions["J"].width = width_patch.after["width"]
    repaired_path = tmp_path / "styled-phone-width-repaired.xlsx"
    repaired.save(repaired_path)
    repaired.close()

    repaired_scan = scan_workbook(repaired_path)
    assert not _findings(repaired_scan, "WL019_IDENTIFIER_SCIENTIFIC_NOTATION")
    verified = load_workbook(repaired_path, data_only=False)
    verified_sheet = verified.active
    assert verified_sheet is not None
    assert verified_sheet["J2"].value == 19_885_186_321
    assert verified_sheet["J2"].data_type == "n"
    verified.close()


def test_long_general_identifiers_remain_findings_at_max_width_without_auto_patch(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    values = {
        "A2": 123_456_789_012,
        "B2": 123_456_789_012_345,
        "C2": 1_234_567_890_123_456,
    }
    for column, coordinate in zip(("A", "B", "C"), values, strict=True):
        worksheet[f"{column}1"] = "联系电话"
        worksheet[coordinate] = values[coordinate]
        worksheet.column_dimensions[column].width = 40

    scan = _scan(workbook, tmp_path / "long-general-identifiers.xlsx")

    findings = _findings(scan, "WL019_IDENTIFIER_SCIENTIFIC_NOTATION")
    by_location = {finding.location: finding for finding in findings}
    assert set(by_location) == set(values)
    assert all(finding.patch_ids == [] for finding in findings)
    assert by_location["A2"].evidence.details["precision_recoverable"] is True
    assert by_location["B2"].evidence.details["precision_recoverable"] is True
    assert by_location["C2"].evidence.details["precision_recoverable"] is False
    assert all(finding.evidence.details["width_only_fixable"] is False for finding in findings)
    assert not any(
        patch.kind in {PatchKind.SET_COLUMN_WIDTH, PatchKind.SET_TEXT} for patch in scan.patches
    )


def test_saved_view_resets_compact_sheet_and_ignores_sheet_origin_with_panes(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet.sheet_view.topLeftCell = "A3"
    worksheet.sheet_view.zoomScale = 115
    scan = _scan(workbook, tmp_path / "view.xlsx")

    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert patch.after == {"top_left_cell": "A1"}

    frozen = Workbook()
    frozen_sheet = frozen.active
    assert frozen_sheet is not None
    frozen_sheet["A1"] = "content"
    frozen_sheet.sheet_view.topLeftCell = "A3"
    frozen_sheet.freeze_panes = "A2"
    frozen_scan = _scan(frozen, tmp_path / "frozen-view.xlsx")
    assert not _findings(frozen_scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in frozen_scan.patches)


@pytest.mark.parametrize("value", [-1, float("nan"), float("inf"), float("-inf"), True])
def test_pane_split_parser_rejects_nonfinite_or_negative_values(value: object) -> None:
    assert _finite_nonnegative_pane_split(value) is None


@pytest.mark.parametrize(
    ("x_split", "y_split"),
    [(-1, 1), (16384, 1), (1, 1048576), (1.5, 1)],
)
def test_saved_view_fails_safe_for_invalid_frozen_split_counts(
    tmp_path: Path,
    x_split: float,
    y_split: float,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet.sheet_view.pane = Pane(
        xSplit=x_split,
        ySplit=y_split,
        topLeftCell="A2",
        activePane="bottomRight",
        state="frozen",
    )

    scan = _scan(workbook, tmp_path / f"invalid-frozen-{x_split}-{y_split}.xlsx")

    assert not _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in scan.patches)


@pytest.mark.parametrize("top_left", ["XFE1", "A1048577"])
def test_saved_view_fails_safe_for_out_of_bounds_top_left_cell(
    tmp_path: Path,
    top_left: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet.sheet_view.topLeftCell = top_left

    scan = _scan(workbook, tmp_path / f"invalid-top-left-{top_left}.xlsx")

    assert not _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in scan.patches)


@pytest.mark.parametrize("top_left", ["XFE1", "A1048577"])
def test_saved_view_fails_safe_for_out_of_bounds_pane_top_left_cell(
    tmp_path: Path,
    top_left: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    worksheet.sheet_view.pane = Pane(
        xSplit=0,
        ySplit=1,
        topLeftCell=top_left,
        activePane="bottomLeft",
        state="frozen",
    )

    scan = _scan(workbook, tmp_path / f"invalid-pane-top-left-{top_left}.xlsx")

    assert not _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in scan.patches)


def _populate_vertically_scrollable_sheet(worksheet: Worksheet) -> None:
    for column in range(1, 11):
        worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 12.5
    for row in range(1, 11):
        worksheet.row_dimensions[row].height = 100
        for column in range(1, 11):
            worksheet.cell(row, column, f"R{row}C{column}")


def test_saved_view_allows_default_zoom_for_vertically_scrollable_sheet(
    tmp_path: Path,
) -> None:
    for name, zoom_scale in (("implicit", None), ("explicit", 100)):
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        _populate_vertically_scrollable_sheet(worksheet)
        worksheet.sheet_view.zoomScale = zoom_scale

        scan = _scan(workbook, tmp_path / f"{name}-default-scroll-view.xlsx")

        assert not _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
        assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in scan.patches)


def test_saved_view_repairs_offset_on_vertically_scrollable_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    _populate_vertically_scrollable_sheet(worksheet)
    worksheet.sheet_view.topLeftCell = "A4"

    scan = _scan(workbook, tmp_path / "offset-scroll-view.xlsx")

    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert finding.evidence.details["vertical_scrolling_expected"] is True
    assert patch.after == {"top_left_cell": "A1"}
    output = tmp_path / "offset-scroll-view-fixed.xlsx"
    apply_patch_plan(
        tmp_path / "offset-scroll-view.xlsx",
        build_patch_plan(scan),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
    )
    assert not _findings(scan_workbook(output), "WL020_SAVED_VIEW_OFF_CONTENT")


def test_saved_view_reduces_high_zoom_on_vertically_scrollable_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    _populate_vertically_scrollable_sheet(worksheet)
    worksheet.sheet_view.zoomScale = 200

    scan = _scan(workbook, tmp_path / "high-zoom-scroll-view.xlsx")

    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert finding.evidence.details["vertical_scrolling_expected"] is True
    assert finding.evidence.observed["estimated_height_fit_zoom"] < 50
    assert patch.after == {"top_left_cell": "A1", "zoom_scale": 80}
    output = tmp_path / "high-zoom-scroll-view-fixed.xlsx"
    apply_patch_plan(
        tmp_path / "high-zoom-scroll-view.xlsx",
        build_patch_plan(scan),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
    )
    assert not _findings(scan_workbook(output), "WL020_SAVED_VIEW_OFF_CONTENT")


def test_saved_view_reduces_zoom_for_wide_compact_content(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for column in range(1, 16):
        worksheet.cell(1, column, f"Header {column}")
        worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 14
    worksheet.sheet_view.topLeftCell = "A1"
    worksheet.sheet_view.zoomScale = 150

    scan = _scan(workbook, tmp_path / "wide-view.xlsx")

    findings = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert len(findings) == 1
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert patch.after["top_left_cell"] == "A1"
    assert 50 <= patch.after["zoom_scale"] < 100
    assert patch.after["zoom_scale"] == max(
        50, findings[0].evidence.observed["estimated_fit_zoom"] - 5
    )


def test_saved_view_uses_full_merged_width_and_preserves_fitting_zoom(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.merge_cells("A1:O1")
    worksheet["A1"] = "Merged heading"
    for column in range(1, 16):
        worksheet.column_dimensions[worksheet.cell(2, column).column_letter].width = 7
    worksheet.sheet_view.topLeftCell = "A3"
    worksheet.sheet_view.zoomScale = 100

    scan = _scan(workbook, tmp_path / "merged-view.xlsx")

    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    assert finding.evidence.details["content_bounds"] == "A1:O1"
    assert finding.evidence.observed["estimated_content_width"] == 105.0
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert patch.after == {"top_left_cell": "A1"}


def test_saved_view_fits_compact_content_height_as_well_as_width(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(1, 29):
        worksheet.cell(row, 1, f"Row {row}")
        worksheet.row_dimensions[row].height = 18
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 52
    worksheet.row_dimensions[28].height = 68
    for column in range(1, 8):
        worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 10
    worksheet.sheet_view.topLeftCell = "A1"
    worksheet.sheet_view.zoomScale = 115

    scan = _scan(workbook, tmp_path / "tall-compact-view.xlsx")

    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert finding.evidence.observed["estimated_content_height"] == 596.0
    assert finding.evidence.observed["estimated_height_fit_zoom"] == 60
    assert finding.evidence.observed["estimated_width_fit_zoom"] > 60
    assert patch.after == {"top_left_cell": "A1", "zoom_scale": 55}


def test_saved_view_offers_zoom_only_for_an_unshifted_frozen_pane(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(1, 29):
        worksheet.cell(row, 1, f"Row {row}")
        worksheet.row_dimensions[row].height = 18
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 115

    scan = _scan(workbook, tmp_path / "frozen-zoom.xlsx")

    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert finding.evidence.details["pane_preserved"] is True
    assert finding.evidence.details["zoom_only_with_frozen_pane"] is True
    assert patch.after == {"zoom_scale": 65}


def test_saved_view_reports_shifted_or_split_panes_without_a_patch(tmp_path: Path) -> None:
    shifted = Workbook()
    shifted_sheet = shifted.active
    assert shifted_sheet is not None
    for row in range(1, 29):
        shifted_sheet.cell(row, 1, f"Row {row}")
        shifted_sheet.row_dimensions[row].height = 18
    shifted_sheet.freeze_panes = "A2"
    assert shifted_sheet.sheet_view.pane is not None
    shifted_sheet.sheet_view.pane.topLeftCell = "A20"
    shifted_sheet.sheet_view.zoomScale = 115

    shifted_scan = _scan(shifted, tmp_path / "shifted-frozen.xlsx")
    shifted_finding = _findings(shifted_scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    assert shifted_finding.evidence.details["frozen_pane_offset"] is True
    assert shifted_finding.patch_ids == []

    split = Workbook()
    split_sheet = split.active
    assert split_sheet is not None
    for row in range(1, 29):
        split_sheet.cell(row, 1, f"Row {row}")
        split_sheet.row_dimensions[row].height = 18
    split_sheet.freeze_panes = "A2"
    assert split_sheet.sheet_view.pane is not None
    split_sheet.sheet_view.pane.state = "split"
    split_sheet.sheet_view.zoomScale = 115

    split_scan = _scan(split, tmp_path / "split-pane.xlsx")
    split_finding = _findings(split_scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    assert split_finding.patch_ids == []


def test_saved_view_includes_prior_row_height_patches_in_fit_estimate(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = (
        "Long text in an automatic-height row that must wrap inside its border before view fitting"
    )
    worksheet["B1"] = "blocker"
    worksheet.column_dimensions["A"].width = 8
    for row in range(2, 29):
        worksheet.cell(row, 1, f"Row {row}")
        worksheet.row_dimensions[row].height = 18
    worksheet.sheet_view.zoomScale = 115

    scan = _scan(workbook, tmp_path / "planned-height-view.xlsx")

    row_patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_ROW_HEIGHT)
    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    view_patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert finding.evidence.details["planned_height_updates_included"] >= 1
    assert finding.evidence.observed["estimated_content_height"] >= (
        row_patch.after["height"] + 27 * 18
    )
    assert view_patch.after["zoom_scale"] <= 55


def test_saved_view_includes_prior_column_width_patches_in_fit_estimate(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(1, 29):
        worksheet.cell(row, 1, f"Row {row}")
        worksheet.row_dimensions[row].height = 12
    worksheet["J1"] = "联系电话"
    worksheet["J2"] = 19_885_186_321
    worksheet.column_dimensions["J"].width = 8
    worksheet.sheet_view.zoomScale = 115

    scan = _scan(workbook, tmp_path / "planned-width-view.xlsx")

    width_patch = next(
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.SET_COLUMN_WIDTH and patch.after["column"] == "J"
    )
    finding = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")[0]
    assert finding.evidence.details["planned_width_updates_included"] >= 1
    assert finding.evidence.observed["estimated_content_width"] >= width_patch.after["width"]


def test_saved_view_reports_extremely_wide_compact_sheet_without_unsafe_zoom_patch(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for column in range(1, 21):
        cell = worksheet.cell(1, column, f"Header {column}")
        worksheet.column_dimensions[cell.column_letter].width = 20
    worksheet.sheet_view.topLeftCell = "A1"
    worksheet.sheet_view.zoomScale = 200

    scan = _scan(workbook, tmp_path / "extremely-wide-view.xlsx")

    findings = _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    assert len(findings) == 1
    assert findings[0].patch_ids == []
    assert findings[0].evidence.details["fit_zoom_below_safe_floor"] is True
    assert not any(patch.kind == PatchKind.SET_SHEET_VIEW for patch in scan.patches)


def test_whitespace_only_tail_is_removed_without_deleting_custom_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, 9):
        for column in range(1, 14):
            cell = worksheet.cell(row, column)
            cell.border = full
            if row <= 5:
                cell.value = f"R{row}C{column}"
    for row in range(9, 12):
        worksheet.cell(row, 1, " ")
        worksheet.row_dimensions[row].height = 30
    source = tmp_path / "whitespace-tail.xlsx"
    scan = _scan(workbook, source)

    findings = _findings(scan, "WL021_WHITESPACE_ONLY_TAIL")
    assert [finding.location for finding in findings] == ["A9:A11"]
    patch = next(
        patch for patch in scan.patches if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS
    )
    assert patch.after["result_dimension"] == "A1:M8"

    output = tmp_path / "whitespace-tail-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
    )
    repaired = load_workbook(output)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    assert repaired_sheet.calculate_dimension() == "A1:M8"
    assert all((row, 1) not in repaired_sheet._cells for row in range(9, 12))
    assert all(repaired_sheet.row_dimensions[row].height == 30 for row in range(9, 12))
    repaired.close()
    assert not _findings(scan_workbook(output), "WL021_WHITESPACE_ONLY_TAIL")
