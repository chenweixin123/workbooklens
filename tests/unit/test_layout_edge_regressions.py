from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from workbooklens.layout import measure_text_cell
from workbooklens.models import PatchKind
from workbooklens.repair.engine import apply_patch_plan
from workbooklens.repair.planning import build_patch_plan
from workbooklens.scanner import ScanResult, scan_workbook


def _scan(workbook: Workbook, path: Path) -> ScanResult:
    workbook.save(path)
    workbook.close()
    return scan_workbook(path)


def _findings(scan: ScanResult, rule_id: str):
    return [finding for finding in scan.findings if finding.rule_id == rule_id]


def test_text_at_right_sheet_boundary_is_not_treated_as_natural_overflow(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["XFD1"] = "Text cannot overflow beyond the final worksheet column"
    worksheet.column_dimensions["XFD"].width = 5
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "right-sheet-boundary.xlsx")

    assert [finding.location for finding in _findings(scan, "WL016_TEXT_DISPLAY_RISK")] == ["XFD1"]


def test_right_aligned_text_at_left_sheet_boundary_is_reported(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Right-aligned text cannot overflow beyond column A"
    worksheet["A1"].alignment = Alignment(horizontal="right")
    worksheet.column_dimensions["A"].width = 5
    worksheet.row_dimensions[1].height = 15

    scan = _scan(workbook, tmp_path / "left-sheet-boundary.xlsx")

    assert [finding.location for finding in _findings(scan, "WL016_TEXT_DISPLAY_RISK")] == ["A1"]


def test_word_aware_wrapping_does_not_underestimate_row_height(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "Alpha Bravo Charlie"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    worksheet.column_dimensions["A"].width = 10
    worksheet.row_dimensions[1].height = 32

    scan = _scan(workbook, tmp_path / "word-wrap-height.xlsx")

    finding = _findings(scan, "WL016_TEXT_DISPLAY_RISK")[0]
    assert finding.location == "A1"
    assert finding.evidence.expected["required_lines"] == 3
    row_patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_ROW_HEIGHT)
    assert row_patch.after["height"] > 32


def test_single_row_merged_wrap_gets_height_even_without_an_explicit_row_height(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "A merged wrapped title that requires multiple visible lines"
    worksheet["A1"].alignment = Alignment(wrap_text=True)
    for column in ("A", "B", "C"):
        worksheet.column_dimensions[column].width = 5

    scan = _scan(workbook, tmp_path / "merged-auto-height.xlsx")

    finding = _findings(scan, "WL016_TEXT_DISPLAY_RISK")[0]
    assert finding.location == "A1"
    assert any(patch.kind == PatchKind.SET_ROW_HEIGHT for patch in scan.patches)


def test_repeated_overflow_width_patch_contains_the_measured_text(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    for row in range(1, 4):
        worksheet.cell(row, 7, "Repeated long text at the right edge")
        worksheet.cell(row, 7).border = Border(right=thin)
    worksheet.column_dimensions["G"].width = 8

    scan = _scan(workbook, tmp_path / "sufficient-shared-width.xlsx")

    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_COLUMN_WIDTH)
    measurements = [measure_text_cell(worksheet, worksheet.cell(row, 7)) for row in range(1, 4)]
    assert all(measurement is not None for measurement in measurements)
    assert patch.after["width"] >= max(
        measurement.required_width for measurement in measurements if measurement is not None
    )


def _formatting_tail_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = "content"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in range(50, 54):
        for column in range(16, 20):
            worksheet.cell(row, column).fill = fill
    return workbook


def test_formatting_tail_does_not_offer_cleanup_for_custom_height_cells(
    tmp_path: Path,
) -> None:
    workbook = _formatting_tail_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.row_dimensions[52].height = 24

    scan = _scan(workbook, tmp_path / "custom-height-formatting-tail.xlsx")

    assert not any(patch.kind == PatchKind.CLEAR_FORMATTING_TAIL for patch in scan.patches)


def test_formatting_tail_does_not_offer_an_unapplicable_hidden_row_cleanup(
    tmp_path: Path,
) -> None:
    workbook = _formatting_tail_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.row_dimensions[52].hidden = True

    scan = _scan(workbook, tmp_path / "hidden-row-formatting-tail.xlsx")

    assert not any(patch.kind == PatchKind.CLEAR_FORMATTING_TAIL for patch in scan.patches)


def test_hidden_sheet_formatting_tail_is_report_only(tmp_path: Path) -> None:
    workbook = _formatting_tail_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Hidden tail"
    worksheet.sheet_state = "hidden"
    workbook.create_sheet("Visible")

    scan = _scan(workbook, tmp_path / "hidden-sheet-formatting-tail.xlsx")

    assert _findings(scan, "WL018_USED_RANGE_INFLATION")
    assert not any(patch.kind == PatchKind.CLEAR_FORMATTING_TAIL for patch in scan.patches)


def test_absent_internal_border_hole_is_materialized_and_repaired(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin", color="FF336699")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, 6):
        for column in range(1, 6):
            if (row, column) == (3, 3):
                continue
            worksheet.cell(row, column, f"R{row}C{column}").border = full
    worksheet["B3"].border = Border(left=thin, top=thin, bottom=thin)
    worksheet["D3"].border = Border(right=thin, top=thin, bottom=thin)
    worksheet["C2"].border = Border(left=thin, right=thin, top=thin)
    worksheet["C4"].border = Border(left=thin, right=thin, bottom=thin)
    source = tmp_path / "absent-border-hole.xlsx"

    scan = _scan(workbook, source)

    findings = _findings(scan, "WL017_BORDER_EDGE_INCONSISTENCY")
    json.dumps(
        {
            "findings": [finding.model_dump(mode="json") for finding in scan.findings],
            "patches": [patch.model_dump(mode="json") for patch in scan.patches],
        },
        sort_keys=True,
    )
    assert any(finding.location == "C3" for finding in findings)
    patches = [
        patch
        for patch in scan.patches
        if patch.kind == PatchKind.COPY_BORDER and patch.cell == "C3"
    ]
    assert {patch.after["target_edge"] for patch in patches} == {
        "left",
        "right",
        "top",
        "bottom",
    }

    output = tmp_path / "absent-border-hole-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id for patch in patches},
        accept_layout_risk=True,
    )

    repaired = load_workbook(output)
    repaired_sheet = repaired.active
    assert repaired_sheet is not None
    assert repaired_sheet["C3"].value is None
    assert all(
        getattr(repaired_sheet["C3"].border, edge).style == "thin"
        for edge in ("left", "right", "top", "bottom")
    )
    repaired.close()
    assert not _findings(scan_workbook(output), "WL017_BORDER_EDGE_INCONSISTENCY")


def test_measurement_compact_twenty_one_column_sheet_gets_zoom_repair(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    thin = Side(style="thin")
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, 9):
        worksheet.row_dimensions[row].height = 15
        for column in range(1, 22):
            cell = worksheet.cell(row, column, f"R{row}C{column}")
            cell.border = full
            if row == 1:
                worksheet.column_dimensions[cell.column_letter].width = 5
    worksheet.sheet_view.zoomScale = 200
    source = tmp_path / "measurement-compact-wide-sheet.xlsx"

    scan = _scan(workbook, source)

    assert _findings(scan, "WL020_SAVED_VIEW_OFF_CONTENT")
    patch = next(patch for patch in scan.patches if patch.kind == PatchKind.SET_SHEET_VIEW)
    assert patch.after["zoom_scale"] < 200
    output = tmp_path / "measurement-compact-wide-sheet-fixed.xlsx"
    apply_patch_plan(
        source,
        build_patch_plan(scan),
        output,
        selected_ids={patch.id},
        accept_layout_risk=True,
    )
    assert not _findings(scan_workbook(output), "WL020_SAVED_VIEW_OFF_CONTENT")
