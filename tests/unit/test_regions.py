from __future__ import annotations

from openpyxl import Workbook

from workbooklens.regions import infer_data_regions, infer_formula_bands


def test_dense_table_is_data_region_but_sparse_decoration_is_not() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(1, 5):
        for column in range(1, 4):
            worksheet.cell(row, column, row * column)
    worksheet["Z40"] = "logo caption"
    regions = infer_data_regions(worksheet)
    assert len(regions) == 1
    assert regions[0].model_dump(include={"min_row", "max_row", "min_column", "max_column"}) == {
        "min_row": 1,
        "max_row": 4,
        "min_column": 1,
        "max_column": 3,
    }


def test_formula_band_allows_one_internal_gap() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in (2, 3, 5, 6):
        worksheet.cell(row, 2, f"=A{row}*2")
    bands = infer_formula_bands(worksheet)
    assert [(band.kind, band.min_row, band.max_row) for band in bands] == [("formula_column", 2, 6)]
