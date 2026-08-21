"""Conservative spreadsheet layout measurements used by first-party rules."""

from __future__ import annotations

import math
import re
import unicodedata
from collections import deque
from copy import copy
from dataclasses import dataclass
from typing import Any

from openpyxl.cell.cell import Cell
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from workbooklens.utils import sha256_bytes, stable_json_bytes

DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0


@dataclass(frozen=True, slots=True)
class TextMeasurement:
    """Static estimate of the space required to display one literal text cell."""

    cell: Cell
    available_width: float
    required_width: float
    required_lines: int
    required_height: float
    current_height: float
    merged_range: str | None = None

    @property
    def width_ratio(self) -> float:
        return self.required_width / max(self.available_width, 0.1)


@dataclass(frozen=True, slots=True)
class FormattingTail:
    """One isolated format-only worksheet tail that can be reviewed as a unit."""

    cell_coordinates: tuple[str, ...]
    cell_ranges: tuple[str, ...]
    empty_rows: tuple[int, ...]
    row_ranges: tuple[str, ...]
    styled_cell_count: int
    content_min_row: int
    content_min_column: int
    content_max_row: int
    content_max_column: int
    observed_max_row: int
    observed_max_column: int


@dataclass(frozen=True, slots=True)
class WhitespaceTail:
    """Literal whitespace cells outside the visible layout envelope."""

    cell_coordinates: tuple[str, ...]
    cell_ranges: tuple[str, ...]
    row_ranges: tuple[str, ...]
    content_min_row: int
    content_min_column: int
    content_max_row: int
    content_max_column: int
    observed_dimension: str
    result_dimension: str
    preserve_style_ids: tuple[tuple[str, int], ...] = ()


def _dimension_covering_column(worksheet: Worksheet, column: int) -> Any | None:
    matches = [
        dimension
        for dimension in worksheet.column_dimensions.values()
        if dimension.min is not None
        and dimension.max is not None
        and dimension.min <= column <= dimension.max
    ]
    return matches[-1] if matches else None


def column_width(worksheet: Worksheet, column: int) -> float:
    """Return the effective explicit/default Excel width for one column."""

    dimension = _dimension_covering_column(worksheet, column)
    if dimension is not None and dimension.width is not None:
        return float(dimension.width)
    if worksheet.sheet_format.defaultColWidth is not None:
        return float(worksheet.sheet_format.defaultColWidth)
    return DEFAULT_COLUMN_WIDTH


def row_height(worksheet: Worksheet, row: int) -> float:
    """Return the effective explicit/default row height in points."""

    dimension = worksheet.row_dimensions.get(row)
    if dimension is not None and dimension.height is not None:
        return float(dimension.height)
    if worksheet.sheet_format.defaultRowHeight is not None:
        return float(worksheet.sheet_format.defaultRowHeight)
    return DEFAULT_ROW_HEIGHT


def explicit_row_height(worksheet: Worksheet, row: int) -> float | None:
    """Return a custom row height, or ``None`` when Excel may auto-size it."""

    dimension = worksheet.row_dimensions.get(row)
    if dimension is None or dimension.height is None:
        return None
    return float(dimension.height)


def _character_units(character: str) -> float:
    if character == "\t":
        return 4.0
    if character.isspace():
        return 0.5
    east_asian_width = unicodedata.east_asian_width(character)
    if east_asian_width in {"W", "F"}:
        return 2.0
    if unicodedata.category(character).startswith("P"):
        return 0.8
    return 1.0


def text_display_units(text: str) -> float:
    """Estimate Excel column-width units without depending on a platform font API."""

    return sum(_character_units(character) for character in text)


def _font_width_factor(cell: Cell, text: str) -> float:
    """Return a conservative width multiplier for the cell's visible font."""

    font_size = float(cell.font.sz or 11.0)
    factor = max(0.7, min(2.5, font_size / 11.0))
    if any(unicodedata.east_asian_width(character) in {"W", "F"} for character in text):
        factor *= 1.08
    if cell.font.bold:
        factor *= 1.04
    return factor


def estimated_text_width(cell: Cell, text: str) -> float:
    """Estimate the Excel column width needed to display ``text`` in ``cell``."""

    return text_display_units(text) * _font_width_factor(cell, text) + 1.0


def _wrapped_line_count(text: str, capacity: float, font_factor: float) -> int:
    """Estimate word-aware wrapping while allowing overlong tokens to break."""

    tokens = re.findall(r"\s+|\S+", text)
    if not tokens:
        return 1
    lines = 1
    used = 0.0
    pending_space = 0.0
    for token in tokens:
        width = text_display_units(token) * font_factor
        if token.isspace():
            pending_space += width
            continue
        if used > 0.0 and used + pending_space + width > capacity:
            lines += 1
            used = 0.0
            pending_space = 0.0
        span = pending_space + width
        if used == 0.0 and span > capacity:
            occupied_lines = max(1, math.ceil(span / capacity))
            lines += occupied_lines - 1
            used = span - (occupied_lines - 1) * capacity
        else:
            used += span
        pending_space = 0.0
    if pending_space and used > 0.0:
        lines += max(0, math.ceil((used + pending_space) / capacity) - 1)
    return lines


def _merged_range_for_cell(worksheet: Worksheet, cell: Cell) -> CellRange | None:
    for merged in worksheet.merged_cells.ranges:
        if cell.coordinate not in merged:
            continue
        return CellRange(str(merged))
    return None


def _available_width(worksheet: Worksheet, cell: Cell) -> tuple[float, str | None]:
    merged = _merged_range_for_cell(worksheet, cell)
    if merged is None:
        return column_width(worksheet, cell.column), None
    width = sum(
        column_width(worksheet, column) for column in range(merged.min_col, merged.max_col + 1)
    )
    return width, str(merged)


def measure_text_cell(
    worksheet: Worksheet,
    cell: Cell,
    *,
    assume_wrap: bool | None = None,
) -> TextMeasurement | None:
    """Estimate literal-text clipping while treating merged non-anchors as non-cells."""

    if cell.data_type == "f" or not isinstance(cell.value, str) or not cell.value.strip():
        return None
    if cell.alignment.shrink_to_fit or cell.alignment.text_rotation or cell.alignment.indent:
        return None
    merged = _merged_range_for_cell(worksheet, cell)
    if merged is not None:
        if cell.row != merged.min_row or cell.column != merged.min_col:
            return None
        # A single row-height patch cannot safely size a merge spanning several rows.
        if merged.max_row > merged.min_row:
            return None
    available_width, merged_range = _available_width(worksheet, cell)
    font_size = float(cell.font.sz or 11.0)
    font_factor = _font_width_factor(cell, cell.value)
    logical_lines = cell.value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    segment_widths = [text_display_units(line) * font_factor for line in logical_lines]
    required_width = max(segment_widths, default=0.0) + 1.0
    wrap_text = bool(cell.alignment.wrap_text) if assume_wrap is None else assume_wrap
    if wrap_text:
        capacity = max(available_width - 1.0, 1.0)
        required_lines = sum(
            _wrapped_line_count(line, capacity, font_factor) for line in logical_lines
        )
    else:
        required_lines = max(1, len(logical_lines))
    line_height = max(12.0, font_size * 1.25)
    required_height = required_lines * line_height + 2.0
    return TextMeasurement(
        cell=cell,
        available_width=available_width,
        required_width=required_width,
        required_lines=required_lines,
        required_height=required_height,
        current_height=row_height(worksheet, cell.row),
        merged_range=merged_range,
    )


def has_visible_border(cell: Cell) -> bool:
    """Return whether any cardinal border side is visibly styled."""

    return any(
        side is not None and side.style is not None
        for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)
    )


def _side_signature(side: Side | None) -> tuple[Any, ...]:
    if side is None:
        return (None, None)
    color = side.color
    color_signature = None
    if color is not None:
        color_type = color.type if isinstance(color.type, str) else None
        rgb = color.rgb if isinstance(color.rgb, str) else None
        indexed = (
            color.indexed
            if isinstance(color.indexed, int) and not isinstance(color.indexed, bool)
            else None
        )
        theme = (
            color.theme
            if isinstance(color.theme, int) and not isinstance(color.theme, bool)
            else None
        )
        tint = (
            float(color.tint)
            if isinstance(color.tint, (int, float)) and not isinstance(color.tint, bool)
            else None
        )
        auto = color.auto if isinstance(color.auto, bool) else None
        color_signature = (
            color_type,
            rgb,
            indexed,
            theme,
            tint,
            auto,
        )
    return (side.style, color_signature)


def border_signature(border: Border) -> tuple[Any, ...]:
    """Return a deterministic cardinal-border signature."""

    return tuple(
        _side_signature(side) for side in (border.left, border.right, border.top, border.bottom)
    )


def visual_border_signature(worksheet: Worksheet, cell: Cell) -> tuple[Any, ...]:
    """Describe visible grid edges, accounting for the adjacent cell's matching side."""

    def choose(primary: Side | None, alternate: Side | None) -> tuple[Any, ...]:
        if (primary is not None and primary.style is not None) or alternate is None:
            return _side_signature(primary)
        return _side_signature(alternate)

    left_peer = worksheet._cells.get((cell.row, cell.column - 1)) if cell.column > 1 else None
    right_peer = worksheet._cells.get((cell.row, cell.column + 1))
    top_peer = worksheet._cells.get((cell.row - 1, cell.column)) if cell.row > 1 else None
    bottom_peer = worksheet._cells.get((cell.row + 1, cell.column))
    return (
        choose(
            cell.border.left,
            left_peer.border.right if isinstance(left_peer, Cell) else None,
        ),
        choose(
            cell.border.right,
            right_peer.border.left if isinstance(right_peer, Cell) else None,
        ),
        choose(
            cell.border.top,
            top_peer.border.bottom if isinstance(top_peer, Cell) else None,
        ),
        choose(
            cell.border.bottom,
            bottom_peer.border.top if isinstance(bottom_peer, Cell) else None,
        ),
    )


def border_is_missing_only(observed: tuple[Any, ...], expected: tuple[Any, ...]) -> bool:
    """Return true when every difference is an absent edge rather than a conflicting edge."""

    missing = False
    for observed_side, expected_side in zip(observed, expected, strict=True):
        observed_style = observed_side[0]
        expected_style = expected_side[0]
        if observed_side == expected_side:
            continue
        if observed_style is None and expected_style is not None:
            missing = True
            continue
        return False
    return missing


def non_border_style_key(cell: Cell) -> tuple[Any, ...]:
    """Group cells whose semantics and decoration differ, if at all, only by border."""

    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.alignment),
        cell.number_format,
        copy(cell.protection),
        bool(cell.quotePrefix),
        bool(cell.pivotButton),
    )


def _meaningful_cells(worksheet: Worksheet) -> list[Cell]:
    return [
        cell
        for cell in worksheet._cells.values()
        if isinstance(cell, Cell) and cell.value is not None
    ]


def _connected_components(coordinates: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    components: list[set[tuple[int, int]]] = []
    remaining = set(coordinates)
    while remaining:
        start = remaining.pop()
        component = {start}
        queue: deque[tuple[int, int]] = deque([start])
        while queue:
            row, column = queue.popleft()
            for neighbor in (
                (row - 1, column),
                (row + 1, column),
                (row, column - 1),
                (row, column + 1),
            ):
                if neighbor not in remaining:
                    continue
                remaining.remove(neighbor)
                component.add(neighbor)
                queue.append(neighbor)
        components.append(component)
    return components


def _ranges_from_numbers(numbers: list[int]) -> tuple[str, ...]:
    if not numbers:
        return ()
    ranges: list[str] = []
    start = previous = numbers[0]
    for number in numbers[1:]:
        if number == previous + 1:
            previous = number
            continue
        ranges.append(str(start) if start == previous else f"{start}:{previous}")
        start = previous = number
    ranges.append(str(start) if start == previous else f"{start}:{previous}")
    return tuple(ranges)


def _ranges_from_coordinates(coordinates: set[tuple[int, int]]) -> tuple[str, ...]:
    """Compress exact coordinates into lossless horizontal A1 runs."""

    by_row: dict[int, list[int]] = {}
    for row, column in coordinates:
        by_row.setdefault(row, []).append(column)
    ranges: list[str] = []
    for row in sorted(by_row):
        columns = sorted(by_row[row])
        start = previous = columns[0]
        for column in columns[1:]:
            if column == previous + 1:
                previous = column
                continue
            left = f"{get_column_letter(start)}{row}"
            right = f"{get_column_letter(previous)}{row}"
            ranges.append(left if start == previous else f"{left}:{right}")
            start = previous = column
        left = f"{get_column_letter(start)}{row}"
        right = f"{get_column_letter(previous)}{row}"
        ranges.append(left if start == previous else f"{left}:{right}")
    return tuple(ranges)


def _vertical_ranges_from_coordinates(coordinates: set[tuple[int, int]]) -> tuple[str, ...]:
    by_column: dict[int, list[int]] = {}
    for row, column in coordinates:
        by_column.setdefault(column, []).append(row)
    ranges: list[str] = []
    for column in sorted(by_column):
        rows = sorted(by_column[column])
        start = previous = rows[0]
        for row in rows[1:]:
            if row == previous + 1:
                previous = row
                continue
            top = f"{get_column_letter(column)}{start}"
            bottom = f"{get_column_letter(column)}{previous}"
            ranges.append(top if start == previous else f"{top}:{bottom}")
            start = previous = row
        top = f"{get_column_letter(column)}{start}"
        bottom = f"{get_column_letter(column)}{previous}"
        ranges.append(top if start == previous else f"{top}:{bottom}")
    return tuple(ranges)


def _is_literal_whitespace_cell(cell: Cell) -> bool:
    value = cell.value
    return bool(
        cell.data_type == "s"
        and isinstance(value, str)
        and value
        and not value.strip()
        and not any(character in "\r\n\t" for character in value)
        and not cell.quotePrefix
        and (cell.number_format or "General").strip().casefold() == "general"
        and cell.comment is None
        and cell.hyperlink is None
        and not has_visible_border(cell)
        and cell.fill.fill_type in {None, "none"}
    )


def _extend_bounds(
    row_bounds: list[int],
    column_bounds: list[int],
    cell_range: CellRange,
) -> None:
    row_bounds.extend((cell_range.min_row, cell_range.max_row))
    column_bounds.extend((cell_range.min_col, cell_range.max_col))


def _layout_envelope(worksheet: Worksheet) -> tuple[int, int, int, int] | None:
    """Return bounds of visible content and intentionally formatted structures."""

    row_bounds: list[int] = []
    column_bounds: list[int] = []
    for cell in worksheet._cells.values():
        if not isinstance(cell, Cell) or _is_literal_whitespace_cell(cell):
            continue
        visible_value = cell.value is not None
        visible_format = has_visible_border(cell) or cell.fill.fill_type not in {None, "none"}
        if (
            visible_value
            or visible_format
            or cell.comment is not None
            or cell.hyperlink is not None
        ):
            row_bounds.append(cell.row)
            column_bounds.append(cell.column)
    for merged in worksheet.merged_cells.ranges:
        _extend_bounds(row_bounds, column_bounds, CellRange(str(merged)))
    for table in worksheet.tables.values():
        _extend_bounds(row_bounds, column_bounds, CellRange(table.ref))
    if worksheet.auto_filter.ref:
        _extend_bounds(row_bounds, column_bounds, CellRange(worksheet.auto_filter.ref))
    if worksheet.data_validations is not None:
        for validation in worksheet.data_validations.dataValidation:
            for target in validation.ranges.ranges:
                _extend_bounds(row_bounds, column_bounds, CellRange(str(target)))
    for conditional_formatting in worksheet.conditional_formatting:
        for target in conditional_formatting.sqref.ranges:
            _extend_bounds(row_bounds, column_bounds, CellRange(str(target)))
    if not row_bounds or not column_bounds:
        return None
    return min(row_bounds), min(column_bounds), max(row_bounds), max(column_bounds)


def _dimension_without_cells(
    worksheet: Worksheet,
    excluded: set[tuple[int, int]],
) -> str:
    row_bounds: list[int] = []
    column_bounds: list[int] = []
    for cell in worksheet._cells.values():
        if not isinstance(cell, Cell) or (cell.row, cell.column) in excluded:
            continue
        row_bounds.append(cell.row)
        column_bounds.append(cell.column)
    for merged in worksheet.merged_cells.ranges:
        _extend_bounds(row_bounds, column_bounds, CellRange(str(merged)))
    if not row_bounds or not column_bounds:
        return "A1"
    start = f"{get_column_letter(min(column_bounds))}{min(row_bounds)}"
    end = f"{get_column_letter(max(column_bounds))}{max(row_bounds)}"
    return start if start == end else f"{start}:{end}"


def _row_is_hidden(worksheet: Worksheet, row: int) -> bool:
    dimension = worksheet.row_dimensions.get(row)
    return bool(dimension.hidden) if dimension is not None else False


def _column_is_hidden(worksheet: Worksheet, column: int) -> bool:
    dimension = _dimension_covering_column(worksheet, column)
    return bool(dimension.hidden) if dimension is not None else False


def _row_blocks_formatting_tail(worksheet: Worksheet, row: int) -> bool:
    dimension = worksheet.row_dimensions.get(row)
    return bool(
        dimension is not None
        and (
            dimension.height is not None
            or dimension.hidden
            or dimension.collapsed
            or dimension.outlineLevel
        )
    )


def _column_blocks_formatting_tail(worksheet: Worksheet, column: int) -> bool:
    dimension = _dimension_covering_column(worksheet, column)
    return bool(
        dimension is not None
        and (dimension.hidden or dimension.collapsed or dimension.outlineLevel)
    )


def find_whitespace_tail(worksheet: Worksheet) -> WhitespaceTail | None:
    """Find a connected literal-whitespace tail outside the visible layout.

    Default-style cells can be removed entirely. Cells with a non-default style
    remain part of the worksheet dimension: repair may clear only their literal
    whitespace value while preserving the complete style record.
    """

    envelope = _layout_envelope(worksheet)
    if envelope is None:
        return None
    min_row, min_column, max_row, max_column = envelope
    candidates = {
        (cell.row, cell.column)
        for cell in worksheet._cells.values()
        if isinstance(cell, Cell)
        and _is_literal_whitespace_cell(cell)
        and (cell.row > max_row or cell.column > max_column)
        and not _row_is_hidden(worksheet, cell.row)
        and not _column_is_hidden(worksheet, cell.column)
        and not any(cell.coordinate in merged for merged in worksheet.merged_cells.ranges)
    }
    if not candidates:
        return None
    observed_dimension = worksheet.calculate_dimension()
    observed_bounds = CellRange(observed_dimension)
    accepted: set[tuple[int, int]] = set()
    for component in _connected_components(candidates):
        touches_outer_edge = any(
            row == observed_bounds.max_row or column == observed_bounds.max_col
            for row, column in component
        )
        if 3 <= len(component) <= 4_096 and touches_outer_edge:
            accepted.update(component)
    if not accepted:
        return None
    removable = {
        (row, column) for row, column in accepted if worksheet._cells[(row, column)].style_id == 0
    }
    preserve_style_ids = tuple(
        (
            f"{get_column_letter(column)}{row}",
            worksheet._cells[(row, column)].style_id,
        )
        for row, column in sorted(accepted - removable)
    )
    result_dimension = _dimension_without_cells(worksheet, removable)
    if not preserve_style_ids and CellRange(result_dimension).bounds == observed_bounds.bounds:
        return None
    return WhitespaceTail(
        cell_coordinates=tuple(
            f"{get_column_letter(column)}{row}" for row, column in sorted(accepted)
        ),
        cell_ranges=_vertical_ranges_from_coordinates(accepted),
        row_ranges=(),
        content_min_row=min_row,
        content_min_column=min_column,
        content_max_row=max_row,
        content_max_column=max_column,
        observed_dimension=observed_dimension,
        result_dimension=result_dimension,
        preserve_style_ids=preserve_style_ids,
    )


def find_formatting_tail(worksheet: Worksheet) -> FormattingTail | None:
    """Find large disconnected style-only islands that inflate a sheet's used range."""

    meaningful = _meaningful_cells(worksheet)
    if not meaningful:
        return None
    row_bounds = [cell.row for cell in meaningful]
    column_bounds = [cell.column for cell in meaningful]
    for merged in worksheet.merged_cells.ranges:
        row_bounds.extend((merged.min_row, merged.max_row))
        column_bounds.extend((merged.min_col, merged.max_col))
    content_min_row = min(row_bounds)
    content_min_column = min(column_bounds)
    content_max_row = max(row_bounds)
    content_max_column = max(column_bounds)
    styled_blank_coordinates = {
        (cell.row, cell.column)
        for cell in worksheet._cells.values()
        if isinstance(cell, Cell)
        and cell.value is None
        and cell.has_style
        and (cell.column >= content_max_column + 8 or cell.row >= content_max_row + 25)
    }
    accepted: list[set[tuple[int, int]]] = []
    for component in _connected_components(styled_blank_coordinates):
        if len(component) < 16:
            continue
        if any(
            _row_blocks_formatting_tail(worksheet, row)
            or _column_blocks_formatting_tail(worksheet, column)
            for row, column in component
        ):
            continue
        rows = [row for row, _ in component]
        columns = [column for _, column in component]
        area = (max(rows) - min(rows) + 1) * (max(columns) - min(columns) + 1)
        density = len(component) / area
        separated = min(columns) >= content_max_column + 8 or min(rows) >= content_max_row + 25
        if density >= 0.5 and separated:
            accepted.append(component)
    if not accepted:
        return None
    all_coordinates: set[tuple[int, int]] = set()
    for component in accepted:
        all_coordinates.update(component)
    cell_coordinates = tuple(
        f"{get_column_letter(column)}{row}" for row, column in sorted(all_coordinates)
    )
    cell_ranges = _ranges_from_coordinates(all_coordinates)
    rows_with_cells = {
        cell.row
        for cell in worksheet._cells.values()
        if isinstance(cell, Cell) and cell.value is not None
    }
    observed_dimension_rows = sorted(
        row
        for row in worksheet.row_dimensions
        if row > content_max_row and row not in rows_with_cells
    )
    empty_dimension_rows = sorted(
        row
        for row, dimension in worksheet.row_dimensions.items()
        if row > content_max_row
        and row not in rows_with_cells
        and not dimension.hidden
        and not dimension.collapsed
        and dimension.outlineLevel == 0
        # A custom height can encode intentional whitespace or print layout. Only
        # style-only row records are safe enough to include in a formatting tail.
        and dimension.height is None
        and dimension.has_style
    )
    observed_max_row = max(
        [content_max_row, *observed_dimension_rows, *(row for row, _ in all_coordinates)]
    )
    observed_max_column = max([content_max_column, *(column for _, column in all_coordinates)])
    return FormattingTail(
        cell_coordinates=cell_coordinates,
        cell_ranges=cell_ranges,
        empty_rows=tuple(empty_dimension_rows),
        row_ranges=_ranges_from_numbers(empty_dimension_rows),
        styled_cell_count=len(all_coordinates),
        content_min_row=content_min_row,
        content_min_column=content_min_column,
        content_max_row=content_max_row,
        content_max_column=content_max_column,
        observed_max_row=observed_max_row,
        observed_max_column=observed_max_column,
    )


def column_layout_fingerprint(worksheet: Worksheet, column: int) -> str:
    """Hash every column-dimension field that a width patch is allowed to observe."""

    dimension = _dimension_covering_column(worksheet, column)
    payload = {
        "column": column,
        "effective_width": column_width(worksheet, column),
        "dimension": None
        if dimension is None
        else {
            "min": dimension.min,
            "max": dimension.max,
            "width": dimension.width,
            "hidden": bool(dimension.hidden),
            "best_fit": bool(dimension.bestFit),
            "style_id": dimension.style_id,
            "outline_level": dimension.outlineLevel,
            "collapsed": bool(dimension.collapsed),
        },
    }
    return sha256_bytes(stable_json_bytes(payload))


def row_layout_fingerprint(worksheet: Worksheet, row: int) -> str:
    """Hash every row-dimension field that a height patch is allowed to observe."""

    dimension = worksheet.row_dimensions.get(row)
    payload = {
        "row": row,
        "effective_height": row_height(worksheet, row),
        "dimension": None
        if dimension is None
        else {
            "height": dimension.height,
            "hidden": bool(dimension.hidden),
            "style_id": dimension.style_id,
            "outline_level": dimension.outlineLevel,
            "collapsed": bool(dimension.collapsed),
        },
    }
    return sha256_bytes(stable_json_bytes(payload))


def sheet_view_fingerprint(worksheet: Worksheet) -> str:
    """Hash saved view state while excluding cell selection coordinates from repair."""

    view = worksheet.sheet_view
    pane = view.pane
    payload = {
        "top_left_cell": view.topLeftCell,
        "zoom_scale": view.zoomScale,
        "zoom_scale_normal": view.zoomScaleNormal,
        "zoom_scale_page_layout": view.zoomScalePageLayoutView,
        "zoom_scale_sheet_layout": view.zoomScaleSheetLayoutView,
        "view": view.view,
        "show_grid_lines": view.showGridLines,
        "pane": None
        if pane is None
        else {
            "state": pane.state,
            "top_left_cell": pane.topLeftCell,
            "x_split": pane.xSplit,
            "y_split": pane.ySplit,
            "active_pane": pane.activePane,
        },
    }
    return sha256_bytes(stable_json_bytes(payload))


def tail_layout_fingerprint(worksheet: Worksheet, tail: FormattingTail) -> str:
    """Hash the exact format-only cells and empty row dimensions proposed for removal."""

    cells: list[dict[str, Any]] = []
    for coordinate in tail.cell_coordinates:
        target = CellRange(coordinate)
        row = target.min_row
        column = target.min_col
        cell = worksheet._cells.get((row, column))
        cells.append(
            {
                "row": row,
                "column": column,
                "present": isinstance(cell, Cell),
                "value": cell.value if isinstance(cell, Cell) else None,
                "style_id": cell.style_id if isinstance(cell, Cell) else 0,
                "has_style": bool(cell.has_style) if isinstance(cell, Cell) else False,
            }
        )
    rows: list[dict[str, Any]] = []
    for row in tail.empty_rows:
        dimension = worksheet.row_dimensions.get(row)
        rows.append(
            {
                "row": row,
                "present": dimension is not None,
                "height": dimension.height if dimension is not None else None,
                "hidden": bool(dimension.hidden) if dimension is not None else False,
                "style_id": dimension.style_id if dimension is not None else 0,
                "outline_level": dimension.outlineLevel if dimension is not None else 0,
                "collapsed": bool(dimension.collapsed) if dimension is not None else False,
            }
        )
    return sha256_bytes(stable_json_bytes({"cells": cells, "rows": rows}))


def whitespace_tail_layout_fingerprint(worksheet: Worksheet, tail: WhitespaceTail) -> str:
    """Hash exact whitespace cells and their preserved row-dimension state."""

    cells: list[dict[str, Any]] = []
    rows: dict[int, dict[str, Any]] = {}
    for coordinate in tail.cell_coordinates:
        target = CellRange(coordinate)
        row = target.min_row
        column = target.min_col
        cell = worksheet._cells.get((row, column))
        cells.append(
            {
                "row": row,
                "column": column,
                "present": isinstance(cell, Cell),
                "value": cell.value if isinstance(cell, Cell) else None,
                "data_type": cell.data_type if isinstance(cell, Cell) else None,
                "style_id": cell.style_id if isinstance(cell, Cell) else 0,
                "number_format": cell.number_format if isinstance(cell, Cell) else None,
                "quote_prefix": bool(cell.quotePrefix) if isinstance(cell, Cell) else False,
            }
        )
        dimension = worksheet.row_dimensions.get(row)
        rows[row] = {
            "row": row,
            "present": dimension is not None,
            "height": dimension.height if dimension is not None else None,
            "hidden": bool(dimension.hidden) if dimension is not None else False,
            "style_id": dimension.style_id if dimension is not None else 0,
            "outline_level": dimension.outlineLevel if dimension is not None else 0,
            "collapsed": bool(dimension.collapsed) if dimension is not None else False,
        }
    return sha256_bytes(
        stable_json_bytes(
            {
                "cells": cells,
                "rows": [rows[row] for row in sorted(rows)],
                "observed_dimension": tail.observed_dimension,
                "result_dimension": tail.result_dimension,
                "preserve_style_ids": [
                    list(item)
                    for item in sorted(
                        tail.preserve_style_ids,
                        key=lambda item: coordinate_to_tuple(item[0]),
                    )
                ],
            }
        )
    )


def range_intersects(left: str, right: str) -> bool:
    """Return whether two A1 ranges overlap."""

    return not CellRange(left).isdisjoint(CellRange(right))


__all__ = [
    "FormattingTail",
    "TextMeasurement",
    "WhitespaceTail",
    "border_is_missing_only",
    "border_signature",
    "column_layout_fingerprint",
    "column_width",
    "estimated_text_width",
    "explicit_row_height",
    "find_formatting_tail",
    "find_whitespace_tail",
    "has_visible_border",
    "measure_text_cell",
    "non_border_style_key",
    "range_intersects",
    "row_height",
    "row_layout_fingerprint",
    "sheet_view_fingerprint",
    "tail_layout_fingerprint",
    "text_display_units",
    "visual_border_signature",
    "whitespace_tail_layout_fingerprint",
]
