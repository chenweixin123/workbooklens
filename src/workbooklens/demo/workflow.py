"""Generate real flaws, plan repairs, apply safe patches, and compare outputs."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import cast

import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from workbooklens.diff import compare_workbooks, write_diff_report
from workbooklens.repair import apply_patch_plan, build_patch_plan
from workbooklens.repair.planning import write_patch_plan
from workbooklens.reports import write_scan_report
from workbooklens.scanner import scan_workbook
from workbooklens.utils import write_json


@dataclass(frozen=True, slots=True)
class DemoOutput:
    """Paths emitted by a successful self-contained demo."""

    directory: Path
    before_workbook: Path
    after_workbook: Path
    repair_plan: Path
    diff_html: Path


def _prepare_output(directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    expected_files = (
        "before.xlsx",
        "after.xlsx",
        "repair-plan.json",
        "apply-report.json",
        "diff.html",
        "diff.json",
        "workbooklens.yml",
    )
    for name in expected_files:
        (directory / name).unlink(missing_ok=True)


def _style_cells(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(name="Arial", size=10, bold=cell.row == 1)


def generate_demo_workbook(path: Path) -> None:
    """Create a realistic workbook with deterministic defects; saving is fixture generation only."""

    workbook = Workbook()
    sales = cast(Worksheet, workbook.active)
    sales.title = "Sales"
    sales.append(["Order ID", "Units", "Unit Price", "Revenue", "Tax", "Status"])
    statuses = ["Open", "Paid", "Void"]
    for row in range(2, 23):
        order_id = f"{row - 1:05d}"
        if row == 6:
            order_id = "00004"  # Explicit duplicate key, still a leading-zero identifier.
        units: int | str = (row % 7) + 2
        if row == 10:
            units = "12"  # Numeric text in a numeric measure column.
        sales.append(
            [
                order_id,
                units,
                9.5 + row,
                f"=B{row}*C{row}",
                f"=D{row}*0.08",
                statuses[row % len(statuses)],
            ]
        )
    sales["D8"] = None  # One blank in an otherwise exact formula band.
    sales["E12"] = 999.0  # One hardcoded override in an exact formula band.
    for row in range(2, 23):
        sales[f"C{row}"].number_format = "$#,##0.00"
        sales[f"D{row}"].number_format = "$#,##0.00"
        sales[f"E{row}"].number_format = "$#,##0.00"
    sales.row_dimensions[20].hidden = True
    validation = DataValidation(type="list", formula1='"Open,Paid,Void"', allow_blank=False)
    sales.add_data_validation(validation)
    validation.add("F2:F10")
    validation.add("F12:F22")
    sales.conditional_formatting.add(
        "D2:D22",
        CellIsRule(  # type: ignore[no-untyped-call]
            operator="greaterThan",
            formula=["300"],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    chart = BarChart()
    chart.title = "Units by order"
    chart.add_data(Reference(sales, min_col=2, min_row=1, max_row=8), titles_from_data=True)
    sales.add_chart(chart, "H2")

    summary = workbook.create_sheet("Summary")
    summary.append(["Month", "Amount"])
    for row in range(2, 10):
        summary.append([f"M{row - 1}", row * 100])
    summary["A10"] = "Total"
    summary["B10"] = "=SUM(B2:B8)"  # Obvious exclusion of adjacent B9.
    summary["D2"] = "=#REF!+1"
    summary["D3"] = "=OFFSET(B2,1,0)"
    summary["D4"] = "=SUM(B:B)"
    summary["D5"] = "='[Budget.xlsx]Plan'!A1"
    summary["D6"] = "#DIV/0!"

    inputs = workbook.create_sheet("Inputs")
    inputs.append(["Code", "Choice", "Value"])
    for row in range(2, 9):
        inputs.append([f"C{row}", "A", row])
    inputs.merge_cells("B5:C5")  # Merge crosses the body of a dense input table.

    hidden = workbook.create_sheet("Hidden Data")
    hidden["A1"] = "Undocumented assumption"
    hidden["B1"] = 0.17
    hidden.sheet_state = "hidden"
    workbook.defined_names.add(DefinedName("BrokenName", attr_text="#REF!"))
    _style_cells(workbook)
    sales["D8"]._style = copy(sales["D7"]._style)  # Pure content gap; no visual leave-blank signal.
    sales["C17"].font = Font(
        name="Arial", size=10, italic=True, color="FF9C0006"
    )  # One visual-only style outlier; numeric semantics remain unchanged.
    # openpyxl is intentionally used only to create the demo input, never to save a repair.
    workbook.save(path)
    workbook.close()


def run_demo(directory: Path) -> DemoOutput:
    """Execute the complete scan-plan-apply-rescan-diff learning workflow."""

    directory = directory.expanduser().resolve()
    _prepare_output(directory)
    before = directory / "before.xlsx"
    after = directory / "after.xlsx"
    plan_path = directory / "repair-plan.json"
    generate_demo_workbook(before)
    config_data = {
        "version": 1,
        "workbook": {"max_critical_findings": 0, "max_error_findings": 10},
        "keys": [{"sheet": "Sales", "range": "A2:A22", "ignore_blank": True}],
        "assertions": [
            {
                "id": "status-domain",
                "type": "allowed_values",
                "sheet": "Sales",
                "range": "F2:F22",
                "values": ["Open", "Paid", "Void"],
            }
        ],
    }
    (directory / "workbooklens.yml").write_text(
        yaml.safe_dump(config_data, sort_keys=False), encoding="utf-8"
    )
    before_scan = scan_workbook(before, config=config_data)
    write_scan_report(before_scan, directory / "before-report")
    plan = build_patch_plan(before_scan)
    write_patch_plan(plan_path, plan)
    result = apply_patch_plan(
        before,
        plan,
        after,
        safe_only=True,
        config=config_data,
    )
    write_json(directory / "apply-report.json", result.model_dump(mode="json"))
    after_scan = scan_workbook(after, config=config_data)
    write_scan_report(after_scan, directory / "after-report")
    semantic_diff = compare_workbooks(before, after)
    diff_path = directory / "diff.html"
    write_diff_report(semantic_diff, diff_path)
    return DemoOutput(
        directory=directory,
        before_workbook=before,
        after_workbook=after,
        repair_plan=plan_path,
        diff_html=diff_path,
    )
