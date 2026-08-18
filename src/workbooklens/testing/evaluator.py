"""Small, explicit assertion evaluator; intentionally not a general Excel engine."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, cast

import yaml
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator
from yaml.nodes import Node

from workbooklens.exceptions import UsageError
from workbooklens.models import AssertionResult, Severity, WorkbookAssertion
from workbooklens.ooxml.safety import PackageLimits, inspect_package
from workbooklens.scanner import ScanResult, scan_workbook
from workbooklens.snapshot import load_for_analysis

CELL_EXPRESSION_RE = re.compile(
    r"^(?P<sheet>'(?:[^']|'')+'|[^'!]+)!(?P<cell>\$?[A-Z]{1,3}\$?\d+)$",
    re.IGNORECASE,
)
SUM_EXPRESSION_RE = re.compile(
    r"^SUM\((?P<sheet>'(?:[^']|'')+'|[^'!]+)!(?P<range>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)\)$",
    re.IGNORECASE,
)


class WorkbookThresholds(BaseModel):
    """Optional finding-count gates evaluated before cell assertions."""

    model_config = ConfigDict(extra="forbid")
    max_critical_findings: int | None = Field(default=None, ge=0)
    max_error_findings: int | None = Field(default=None, ge=0)


class ConfiguredKey(BaseModel):
    """A duplicate-key range consumed by the deterministic key rule."""

    model_config = ConfigDict(extra="forbid")
    sheet: str = Field(min_length=1)
    range: str = Field(min_length=1)
    ignore_blank: bool = True

    @field_validator("range")
    @classmethod
    def validate_bounded_range(cls, value: str) -> str:
        try:
            _bounded_range(value)
        except ValueError as exc:
            raise ValueError(f"configured key range is invalid: {exc}") from exc
        return value


class TestConfig(BaseModel):
    """Validated v0.1 YAML configuration."""

    model_config = ConfigDict(extra="forbid")
    version: Literal[1]
    workbook: WorkbookThresholds = Field(default_factory=WorkbookThresholds)
    assertions: list[WorkbookAssertion] = Field(default_factory=list)
    keys: list[ConfiguredKey] = Field(default_factory=list)


class LimitedSafeLoader(yaml.SafeLoader):
    """SafeLoader with explicit graph-size ceilings for untrusted configuration."""

    max_aliases = 50
    max_nodes = 20_000
    max_depth = 100

    def __init__(self, stream: Any) -> None:
        super().__init__(stream)
        self.alias_count = 0
        self.node_count = 0
        self.depth = 0

    def compose_node(self, parent: Node | None, index: Any) -> Node:
        self.node_count += 1
        if self.node_count > self.max_nodes:
            raise yaml.YAMLError(f"YAML node count exceeds the {self.max_nodes}-node limit")
        if self.check_event(yaml.AliasEvent):
            self.alias_count += 1
            if self.alias_count > self.max_aliases:
                raise yaml.YAMLError(f"YAML alias count exceeds the {self.max_aliases}-alias limit")
        self.depth += 1
        if self.depth > self.max_depth:
            raise yaml.YAMLError(f"YAML nesting exceeds the {self.max_depth}-level limit")
        try:
            return cast(Node, super().compose_node(parent, index))
        finally:
            self.depth -= 1


@dataclass(slots=True)
class TestRun:
    """Full assertion outcome plus the scan used by finding-based checks."""

    config: TestConfig
    scan: ScanResult
    results: list[AssertionResult]

    @property
    def passed(self) -> bool:
        """Return true only if every configured gate and assertion passed."""

        return all(result.passed for result in self.results)


def load_test_config(path: Path, max_bytes: int = 1024 * 1024) -> TestConfig:
    """Read a size-bounded YAML file through SafeLoader and validate version 1."""

    try:
        data = path.read_bytes()
    except OSError as exc:
        raise UsageError(f"Unable to read test configuration {path}: {exc}") from exc
    if len(data) > max_bytes:
        raise UsageError(f"Test configuration exceeds the {max_bytes}-byte limit")
    try:
        loader = LimitedSafeLoader(data.decode("utf-8"))
        try:
            raw = loader.get_single_data()
        finally:
            loader.dispose()
    except (UnicodeDecodeError, yaml.YAMLError, RecursionError) as exc:
        raise UsageError(
            f"Invalid UTF-8 YAML configuration or unsafe YAML features: {exc}"
        ) from exc
    if not isinstance(raw, dict):
        raise UsageError("Workbook test configuration must be a YAML mapping")
    try:
        return TestConfig.model_validate(raw)
    except ValidationError as exc:
        raise UsageError(f"Workbook test configuration is invalid: {exc}") from exc


def _unquote_sheet(value: str) -> str:
    if value.startswith("'") and value.endswith("'"):
        return value[1:-1].replace("''", "'")
    return value


def _bounded_range(range_text: str, maximum_cells: int = 100_000) -> tuple[int, int, int, int]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_text)
    except ValueError as exc:
        raise ValueError(f"invalid A1 range {range_text!r}") from exc
    if None in {min_col, min_row, max_col, max_row}:
        raise ValueError(f"range must have explicit row and column bounds: {range_text!r}")
    bounds = (
        cast(int, min_col),
        cast(int, min_row),
        cast(int, max_col),
        cast(int, max_row),
    )
    if bounds[0] < 1 or bounds[1] < 1 or bounds[2] > 16_384 or bounds[3] > 1_048_576:
        raise ValueError(f"range is outside Excel worksheet bounds: {range_text!r}")
    cell_count = (bounds[2] - bounds[0] + 1) * (bounds[3] - bounds[1] + 1)
    if cell_count > maximum_cells:
        raise ValueError(
            f"range {range_text!r} expands to {cell_count} cells; limit is {maximum_cells}"
        )
    return bounds


def _cells(workbook: Workbook, sheet_name: str, range_text: str) -> list[Cell]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet {sheet_name!r} does not exist")
    min_col, min_row, max_col, max_row = _bounded_range(range_text)
    worksheet = workbook[sheet_name]
    return [
        cast(Cell, worksheet.cell(row, column))
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    ]


def _sum_expression(
    formula_workbook: Workbook,
    value_workbook: Workbook,
    sheet_name: str,
    range_text: str,
) -> float:
    total = 0.0
    for formula_cell in _cells(formula_workbook, sheet_name, range_text):
        if formula_cell.data_type == "f":
            cached_cell = value_workbook[sheet_name][formula_cell.coordinate]
            value = cached_cell.value
            if value is None:
                raise ValueError(
                    f"formula at {sheet_name}!{formula_cell.coordinate} has no cached value"
                )
            if cached_cell.data_type == "e":
                raise ValueError(
                    f"formula at {sheet_name}!{formula_cell.coordinate} has cached error {value!r}"
                )
        else:
            value = formula_cell.value
            if formula_cell.data_type == "e":
                raise ValueError(
                    f"SUM range contains error at {sheet_name}!{formula_cell.coordinate}: {value!r}"
                )
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)):
            total += float(value)
        elif value is None or isinstance(value, str):
            # Excel SUM ignores text and blanks in referenced ranges.
            continue
        else:
            raise ValueError(
                f"unsupported SUM value at {sheet_name}!{formula_cell.coordinate}: {value!r}"
            )
    return total


def _resolve_expression(
    formula_workbook: Workbook, value_workbook: Workbook, expression: str
) -> Any:
    expression = expression.strip()
    sum_match = SUM_EXPRESSION_RE.fullmatch(expression)
    if sum_match:
        return _sum_expression(
            formula_workbook,
            value_workbook,
            _unquote_sheet(sum_match.group("sheet")),
            sum_match.group("range").replace("$", ""),
        )
    cell_match = CELL_EXPRESSION_RE.fullmatch(expression)
    if cell_match:
        sheet_name = _unquote_sheet(cell_match.group("sheet"))
        coordinate = cell_match.group("cell").replace("$", "")
        if sheet_name not in formula_workbook.sheetnames:
            raise ValueError(f"sheet {sheet_name!r} does not exist")
        formula_cell = formula_workbook[sheet_name][coordinate]
        if formula_cell.data_type != "f":
            return formula_cell.value
        cached_cell = value_workbook[sheet_name][coordinate]
        cached = cached_cell.value
        if cached_cell.data_type == "e":
            raise ValueError(f"{expression!r} has cached error {cached!r}")
        if cached is not None:
            return cached
        formula = formula_cell.value
        if isinstance(formula, str) and formula.startswith("="):
            nested = formula[1:]
            nested_sum = SUM_EXPRESSION_RE.fullmatch(nested)
            if nested_sum:
                return _sum_expression(
                    formula_workbook,
                    value_workbook,
                    _unquote_sheet(nested_sum.group("sheet")),
                    nested_sum.group("range").replace("$", ""),
                )
        raise ValueError(
            f"{expression!r} is a formula without a cached value and is outside the supported SUM subset"
        )
    try:
        return float(expression)
    except ValueError as exc:
        raise ValueError(
            f"unsupported expression {expression!r}; use Sheet!A1, SUM(Sheet!A1:A10), or a number"
        ) from exc


def _threshold_results(config: TestConfig, scan: ScanResult) -> list[AssertionResult]:
    counts = {
        Severity.CRITICAL: sum(finding.severity == Severity.CRITICAL for finding in scan.findings),
        Severity.ERROR: sum(finding.severity == Severity.ERROR for finding in scan.findings),
    }
    results: list[AssertionResult] = []
    for severity, maximum in (
        (Severity.CRITICAL, config.workbook.max_critical_findings),
        (Severity.ERROR, config.workbook.max_error_findings),
    ):
        if maximum is None:
            continue
        observed = counts[severity]
        results.append(
            AssertionResult(
                assertion_id=f"workbook:max_{severity.value}_findings",
                passed=observed <= maximum,
                message=f"Observed {observed} {severity.value} findings; maximum is {maximum}",
                observed=observed,
                expected={"maximum": maximum},
            )
        )
    return results


def _evaluate_assertion(
    assertion: WorkbookAssertion,
    formula_workbook: Workbook,
    value_workbook: Workbook,
    scan: ScanResult,
) -> AssertionResult:
    try:
        if assertion.type == "no_findings":
            matched = [
                finding.id
                for finding in scan.findings
                if not assertion.rules or finding.rule_id in assertion.rules
            ]
            return AssertionResult(
                assertion_id=assertion.id,
                passed=not matched,
                message=f"Matched {len(matched)} prohibited findings",
                observed=matched,
                expected=[],
            )
        if assertion.type in {"unique", "allowed_values", "nonblank", "numeric_bounds"}:
            if not assertion.sheet or not assertion.range:
                raise ValueError(f"{assertion.type} requires sheet and range")
            cells = _cells(formula_workbook, assertion.sheet, assertion.range)
            values = [cell.value for cell in cells]
            if assertion.type == "unique":
                filtered = [
                    value for value in values if not (assertion.ignore_blank and value is None)
                ]
                counts = Counter((type(value).__qualname__, repr(value)) for value in filtered)
                duplicates = sorted(
                    representation for (_, representation), count in counts.items() if count > 1
                )
                return AssertionResult(
                    assertion_id=assertion.id,
                    passed=not duplicates,
                    message=f"Found {len(duplicates)} duplicate value groups",
                    observed=duplicates,
                    expected="unique values",
                )
            if assertion.type == "allowed_values":
                invalid = [
                    {"cell": cell.coordinate, "value": cell.value}
                    for cell in cells
                    if cell.value not in assertion.values
                    and not (assertion.ignore_blank and cell.value is None)
                ]
                return AssertionResult(
                    assertion_id=assertion.id,
                    passed=not invalid,
                    message=f"Found {len(invalid)} values outside the allowed domain",
                    observed=invalid,
                    expected=assertion.values,
                )
            if assertion.type == "nonblank":
                blank = [
                    cell.coordinate for cell in cells if cell.value is None or cell.value == ""
                ]
                return AssertionResult(
                    assertion_id=assertion.id,
                    passed=not blank,
                    message=f"Found {len(blank)} blank cells",
                    observed=blank,
                    expected="all cells nonblank",
                )
            invalid_numeric: list[dict[str, Any]] = []
            for cell in cells:
                if cell.value is None and assertion.ignore_blank:
                    continue
                if isinstance(cell.value, bool) or not isinstance(cell.value, (int, float)):
                    invalid_numeric.append({"cell": cell.coordinate, "value": cell.value})
                    continue
                if (assertion.minimum is not None and cell.value < assertion.minimum) or (
                    assertion.maximum is not None and cell.value > assertion.maximum
                ):
                    invalid_numeric.append({"cell": cell.coordinate, "value": cell.value})
            return AssertionResult(
                assertion_id=assertion.id,
                passed=not invalid_numeric,
                message=f"Found {len(invalid_numeric)} nonnumeric or out-of-bounds cells",
                observed=invalid_numeric,
                expected={"minimum": assertion.minimum, "maximum": assertion.maximum},
            )
        if assertion.type == "equals":
            if assertion.left is None or assertion.right is None:
                raise ValueError("equals requires left and right expressions")
            left = _resolve_expression(formula_workbook, value_workbook, assertion.left)
            right = _resolve_expression(formula_workbook, value_workbook, assertion.right)
            if isinstance(left, (int, float)) and isinstance(right, (int, float)):
                passed = abs(float(left) - float(right)) <= assertion.tolerance
            else:
                passed = left == right
            return AssertionResult(
                assertion_id=assertion.id,
                passed=passed,
                message=f"Compared {assertion.left} with {assertion.right}",
                observed=left,
                expected={"value": right, "tolerance": assertion.tolerance},
            )
        raise ValueError(f"unsupported assertion type: {assertion.type}")
    except (ValueError, TypeError) as exc:
        return AssertionResult(
            assertion_id=assertion.id,
            passed=False,
            message=f"Assertion could not be evaluated safely: {exc}",
        )


def evaluate_workbook_tests(
    workbook_path: Path,
    config: TestConfig,
    limits: PackageLimits | None = None,
) -> TestRun:
    """Scan and evaluate every configured threshold/assertion without formula execution."""

    inspection = inspect_package(workbook_path, limits)
    config_for_rules = config.model_dump(mode="python")
    scan = scan_workbook(inspection.path, config=config_for_rules, limits=limits)
    formula_workbook = load_for_analysis(inspection.path, limits)
    value_workbook = load_workbook(
        inspection.path,
        read_only=False,
        data_only=True,
        keep_links=False,
        keep_vba=False,
    )
    try:
        results = _threshold_results(config, scan)
        results.extend(
            _evaluate_assertion(assertion, formula_workbook, value_workbook, scan)
            for assertion in config.assertions
        )
        return TestRun(config=config, scan=scan, results=results)
    finally:
        formula_workbook.close()
        value_workbook.close()
