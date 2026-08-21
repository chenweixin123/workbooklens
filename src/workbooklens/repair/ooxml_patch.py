"""Direct OOXML worksheet patching with byte-identical untouched part contents."""

from __future__ import annotations

import copy
import hashlib
import math
import os
import posixpath
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

from lxml import etree
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from workbooklens.exceptions import PatchValidationError, StalePlanError, UsageError
from workbooklens.formulas import analyze_formula
from workbooklens.models import PackageChange, PatchKind, PatchOperation, PatchPlan
from workbooklens.ooxml.safety import PackageLimits, inspect_package, parse_xml_part
from workbooklens.repair.layout_ooxml import (
    StylesEditor,
    apply_alignment,
    apply_clear_formatting_tail,
    apply_column_width,
    apply_copy_border,
    apply_remove_whitespace_tail_cells,
    apply_row_height,
    apply_set_text,
    apply_sheet_view,
    is_layout_kind,
    needs_styles,
    patch_target_key,
    tail_authorization,
    validate_layout_semantics,
    verify_layout_precondition,
    whitespace_tail_authorization,
    whitespace_tail_preserved_style_ids,
)
from workbooklens.snapshot import cell_fingerprint, load_for_analysis
from workbooklens.utils import sha256_file
from workbooklens.worksheet_state import is_column_hidden


@dataclass(frozen=True, slots=True)
class OoxmlPatchOutput:
    """Low-level validated package result before scanner comparison."""

    source_sha256: str
    output_sha256: str
    changes: list[PackageChange]
    formula_changed: bool


CANONICAL_PLAN_MISMATCH = (
    "Patch plan does not match the canonical scan: an operation intersects an unsupported "
    "shared formula range, uses an unsupported array or unsupported dataTable formula, or "
    "contains edited structured or ordinary formula, safe, confidence, kind, source-cell, "
    "description, or precondition fields"
)


def _qname(namespace: str, local_name: str) -> str:
    return f"{{{namespace}}}{local_name}"


def _element_namespace(element: etree._Element) -> str:
    value = etree.QName(element).namespace
    if value is None:
        raise PatchValidationError("OOXML root element is missing a namespace")
    return value


def _resolve_part(base_part: str, target: str) -> str:
    if "\\" in target or "\x00" in target:
        raise PatchValidationError(f"Unsafe relationship target: {target!r}")
    if target.startswith("/"):
        candidate = target.lstrip("/")
    else:
        candidate = posixpath.join(posixpath.dirname(base_part), target)
    normalized = posixpath.normpath(candidate)
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise PatchValidationError(f"Relationship target escapes package root: {target!r}")
    return normalized


def _sheet_parts(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = parse_xml_part(archive.read("xl/workbook.xml"), "xl/workbook.xml")
    rels_name = "xl/_rels/workbook.xml.rels"
    if rels_name not in archive.namelist():
        raise PatchValidationError("Workbook relationships part is missing")
    rels_root = parse_xml_part(archive.read(rels_name), rels_name)
    relationships: dict[str, tuple[str, str]] = {}
    for relationship in rels_root:
        if etree.QName(relationship).localname != "Relationship":
            continue
        relationship_id = relationship.get("Id")
        target = relationship.get("Target")
        relationship_type = relationship.get("Type", "")
        if relationship_id and target:
            relationships[relationship_id] = (relationship_type, target)
    result: dict[str, str] = {}
    for sheet in workbook_root.iter():
        if etree.QName(sheet).localname != "sheet":
            continue
        name = sheet.get("name")
        relationship_id = cast(
            str | None,
            next(
                (
                    value
                    for attribute, value in sheet.attrib.items()
                    if etree.QName(attribute).localname == "id"
                ),
                None,
            ),
        )
        if not name or not relationship_id or relationship_id not in relationships:
            raise PatchValidationError("Workbook contains a sheet with a malformed relationship")
        relationship_type, target = relationships[relationship_id]
        if relationship_type.endswith("/chartsheet"):
            continue
        if not relationship_type.endswith("/worksheet"):
            raise PatchValidationError(
                f"Workbook contains an unsupported sheet relationship type: {relationship_type!r}"
            )
        result[name] = _resolve_part("xl/workbook.xml", target)
    return result


def _find_cell(root: etree._Element, coordinate: str) -> etree._Element | None:
    for element in root.iter():
        if etree.QName(element).localname == "c" and element.get("r") == coordinate:
            return element
    return None


def _find_or_create_cell(root: etree._Element, coordinate: str) -> etree._Element:
    existing = _find_cell(root, coordinate)
    if existing is not None:
        return existing
    row_number, column_number = coordinate_to_tuple(coordinate)
    namespace = _element_namespace(root)
    sheet_data = next(
        (element for element in root if etree.QName(element).localname == "sheetData"),
        None,
    )
    if sheet_data is None:
        raise PatchValidationError("Worksheet has no sheetData element")
    row_element = next(
        (
            element
            for element in sheet_data
            if etree.QName(element).localname == "row" and int(element.get("r", "0")) == row_number
        ),
        None,
    )
    if row_element is None:
        row_element = etree.Element(_qname(namespace, "row"), r=str(row_number))
        insertion_index = len(sheet_data)
        for index, candidate in enumerate(sheet_data):
            if etree.QName(candidate).localname != "row":
                continue
            if int(candidate.get("r", "0")) > row_number:
                insertion_index = index
                break
        sheet_data.insert(insertion_index, row_element)
    cell_element = etree.Element(_qname(namespace, "c"), r=coordinate)
    insertion_index = len(row_element)
    for index, candidate in enumerate(row_element):
        if etree.QName(candidate).localname != "c":
            continue
        candidate_coordinate = candidate.get("r")
        if not candidate_coordinate:
            continue
        _, candidate_column = coordinate_to_tuple(candidate_coordinate)
        if candidate_column > column_number:
            insertion_index = index
            break
    row_element.insert(insertion_index, cell_element)
    return cell_element


def _formula_element(cell: etree._Element) -> etree._Element | None:
    return next(
        (child for child in cell if etree.QName(child).localname == "f"),
        None,
    )


def _assert_simple_formula_cell(cell: etree._Element, label: str) -> None:
    formula = _formula_element(cell)
    if formula is None:
        return
    formula_type = formula.get("t")
    if formula_type in {"shared", "array", "dataTable"} or formula.get("ref"):
        raise PatchValidationError(f"{label} uses an unsupported {formula_type or 'array'} formula")
    text = formula.text or ""
    features = analyze_formula("=" + text)
    if features.external_references or features.unsupported_reason:
        raise PatchValidationError(
            f"{label} uses an external, structured, dynamic, or advanced formula"
        )


def _assert_not_in_advanced_formula_range(root: etree._Element, coordinate: str) -> None:
    for formula in root.iter():
        if etree.QName(formula).localname != "f":
            continue
        formula_type = formula.get("t")
        reference = formula.get("ref")
        if formula_type not in {"shared", "array", "dataTable"} or not reference:
            continue
        try:
            formula_range = CellRange(reference)
        except ValueError as exc:
            raise PatchValidationError(
                f"Malformed {formula_type} formula range {reference!r}"
            ) from exc
        if coordinate in formula_range:
            raise PatchValidationError(
                f"target {coordinate} intersects an unsupported {formula_type} formula range"
            )


def _assert_not_in_merged_range(
    root: etree._Element,
    coordinate: str,
    *,
    allow_anchor: bool = False,
) -> None:
    for element in root.iter():
        if etree.QName(element).localname != "mergeCell":
            continue
        reference = element.get("ref")
        if not reference:
            continue
        try:
            merged = CellRange(reference)
        except ValueError as exc:
            raise PatchValidationError(f"Invalid merged range {reference!r}") from exc
        if coordinate in merged:
            anchor = f"{get_column_letter(merged.min_col)}{merged.min_row}"
            if allow_anchor and coordinate == anchor:
                return
            raise PatchValidationError(
                f"target {coordinate} intersects merged range {reference}; automatic patches are refused"
            )


def _remove_value_children(cell: etree._Element) -> None:
    for child in list(cell):
        if etree.QName(child).localname in {"f", "v", "is"}:
            cell.remove(child)


def _set_formula(cell: etree._Element, formula: str) -> None:
    if not formula.startswith("="):
        raise PatchValidationError("Formula patch output must begin with '='")
    features = analyze_formula(formula)
    if features.external_references or features.unsupported_reason:
        raise PatchValidationError(
            "External, structured, dynamic, spilled, or advanced formulas are not patchable"
        )
    namespace = _element_namespace(cell)
    _remove_value_children(cell)
    if "t" in cell.attrib:
        del cell.attrib["t"]
    formula_element = etree.Element(_qname(namespace, "f"))
    formula_element.text = formula[1:]
    cell.insert(0, formula_element)


def _set_numeric(cell: etree._Element, value: Any) -> None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise PatchValidationError("Numeric patch output must be an integer or finite float")
    if isinstance(value, float) and not math.isfinite(value):
        raise PatchValidationError("Numeric patch output must be finite")
    namespace = _element_namespace(cell)
    _remove_value_children(cell)
    if "t" in cell.attrib:
        del cell.attrib["t"]
    value_element = etree.Element(_qname(namespace, "v"))
    value_element.text = str(value) if isinstance(value, int) else format(value, ".15g")
    cell.append(value_element)


def _copy_style(target: etree._Element, source: etree._Element) -> None:
    style = source.get("s")
    if style is None:
        if "s" in target.attrib:
            del target.attrib["s"]
    else:
        target.set("s", style)


def _update_dimension(root: etree._Element) -> None:
    cells = [
        element.get("r")
        for element in root.iter()
        if etree.QName(element).localname == "c" and element.get("r")
    ]
    if not cells:
        return
    dimension = next(
        (element for element in root if etree.QName(element).localname == "dimension"),
        None,
    )
    if dimension is None:
        return
    existing_ref = dimension.get("ref")
    if not existing_ref:
        raise PatchValidationError("Worksheet dimension is missing its ref attribute")
    try:
        existing = CellRange(existing_ref)
    except (TypeError, ValueError) as exc:
        raise PatchValidationError("Worksheet dimension ref is malformed") from exc
    coordinates = [coordinate_to_tuple(cast(str, coordinate)) for coordinate in cells]
    min_row = min(existing.min_row, *(item[0] for item in coordinates))
    max_row = max(existing.max_row, *(item[0] for item in coordinates))
    min_column = min(existing.min_col, *(item[1] for item in coordinates))
    max_column = max(existing.max_col, *(item[1] for item in coordinates))
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    dimension.set("ref", start if start == end else f"{start}:{end}")


def _serialize_xml(root: etree._Element) -> bytes:
    return etree.tostring(
        root,
        encoding="UTF-8",
        xml_declaration=True,
        standalone=True,
    )


def _mark_full_recalculation(workbook_xml: bytes) -> bytes:
    root = parse_xml_part(workbook_xml, "xl/workbook.xml")
    namespace = _element_namespace(root)
    calc_pr = next(
        (element for element in root if etree.QName(element).localname == "calcPr"),
        None,
    )
    if calc_pr is None:
        calc_pr = etree.Element(_qname(namespace, "calcPr"))
        root.append(calc_pr)
    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    return _serialize_xml(root)


def _verify_preconditions(
    source: Path,
    plan: PatchPlan,
    selected: list[PatchOperation],
    limits: PackageLimits | None,
) -> None:
    actual_hash = sha256_file(source)
    if actual_hash != plan.source_sha256:
        raise StalePlanError(
            f"Plan source hash {plan.source_sha256} does not match workbook hash {actual_hash}"
        )
    workbook = load_for_analysis(source, limits)
    try:
        worksheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}
        for patch in selected:
            worksheet = worksheets.get(patch.sheet)
            if worksheet is None:
                raise StalePlanError(f"Patch worksheet no longer exists: {patch.sheet!r}")
            cell = worksheet[patch.cell]
            if not isinstance(cell, Cell):
                raise StalePlanError(
                    f"Patch target is not a writable cell: {patch.sheet}!{patch.cell}"
                )
            actual = cell_fingerprint(cell)
            if actual != patch.precondition.cell_fingerprint:
                raise StalePlanError(
                    f"Cell precondition failed for {patch.sheet}!{patch.cell}; plan is stale"
                )
            verify_layout_precondition(worksheet, patch)
            if worksheet.protection.sheet:
                raise PatchValidationError(
                    f"Patch target is on protected worksheet {patch.sheet!r}; unprotect it before repair"
                )
            if worksheet.sheet_state != "visible":
                raise PatchValidationError(
                    f"Patch target is on non-visible worksheet {patch.sheet!r}; unhide it before repair"
                )
            row_dimension = worksheet.row_dimensions.get(cell.row)
            if row_dimension is not None and row_dimension.hidden:
                raise PatchValidationError(
                    f"Patch target is on hidden row {patch.sheet}!{cell.row}; unhide it before repair"
                )
            if is_column_hidden(worksheet, cell.column):
                raise PatchValidationError(
                    f"Patch target is in a hidden column at {patch.sheet}!{patch.cell}; "
                    "unhide it before repair"
                )
            if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL:
                tail_cells, tail_rows = tail_authorization(patch)
                for coordinate in tail_cells:
                    tail_cell = worksheet[coordinate]
                    if not isinstance(tail_cell, Cell):
                        raise PatchValidationError(
                            f"Formatting-tail target is not a cell: {patch.sheet}!{coordinate}"
                        )
                    if tail_cell.value is not None:
                        raise PatchValidationError(
                            f"Formatting-tail target contains a value: {patch.sheet}!{coordinate}"
                        )
                    if tail_cell.comment is not None or tail_cell.hyperlink is not None:
                        raise PatchValidationError(
                            f"Formatting-tail target has a comment or hyperlink: "
                            f"{patch.sheet}!{coordinate}"
                        )
                    row_dimension = worksheet.row_dimensions.get(tail_cell.row)
                    if row_dimension is not None and row_dimension.hidden:
                        raise PatchValidationError(
                            f"Formatting-tail target is on hidden row {patch.sheet}!{tail_cell.row}"
                        )
                    if is_column_hidden(worksheet, tail_cell.column):
                        raise PatchValidationError(
                            f"Formatting-tail target is in hidden column {patch.sheet}!{coordinate}"
                        )
                for row in tail_rows:
                    dimension = worksheet.row_dimensions.get(row)
                    if dimension is None:
                        raise PatchValidationError(
                            f"Formatting-tail row is absent: {patch.sheet}!{row}"
                        )
                    if dimension.hidden or dimension.collapsed or dimension.outlineLevel:
                        raise PatchValidationError(
                            f"Formatting-tail cleanup refuses hidden, collapsed, or outlined row {row}"
                        )
            if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS:
                preserve_style_ids = whitespace_tail_preserved_style_ids(patch)
                for coordinate in whitespace_tail_authorization(patch):
                    target = worksheet[coordinate]
                    if not isinstance(target, Cell):
                        raise PatchValidationError(
                            f"Whitespace-tail target is not a cell: {patch.sheet}!{coordinate}"
                        )
                    value = target.value
                    if (
                        target.data_type != "s"
                        or not isinstance(value, str)
                        or not value
                        or value.strip()
                        or any(character in "\r\n\t" for character in value)
                        or target.quotePrefix
                        or (target.number_format or "General").strip().casefold() != "general"
                    ):
                        raise PatchValidationError(
                            f"Whitespace-tail target is no longer simple literal whitespace: "
                            f"{patch.sheet}!{coordinate}"
                        )
                    if target.comment is not None or target.hyperlink is not None:
                        raise PatchValidationError(
                            f"Whitespace-tail target has a comment or hyperlink: "
                            f"{patch.sheet}!{coordinate}"
                        )
                    expected_style_id = preserve_style_ids.get(coordinate)
                    if expected_style_id is None and target.style_id != 0:
                        raise PatchValidationError(
                            f"Whitespace-tail target has an unauthorized non-default style: "
                            f"{patch.sheet}!{coordinate}"
                        )
                    if expected_style_id is not None and target.style_id != expected_style_id:
                        raise PatchValidationError(
                            f"Whitespace-tail target style no longer matches its authorization: "
                            f"{patch.sheet}!{coordinate}"
                        )
                    row_dimension = worksheet.row_dimensions.get(target.row)
                    if row_dimension is not None and row_dimension.hidden:
                        raise PatchValidationError(
                            f"Whitespace-tail target is on hidden row {patch.sheet}!{target.row}"
                        )
                    if is_column_hidden(worksheet, target.column):
                        raise PatchValidationError(
                            f"Whitespace-tail target is in hidden column {patch.sheet}!{coordinate}"
                        )
                    if any(coordinate in merged for merged in worksheet.merged_cells.ranges):
                        raise PatchValidationError(
                            f"Whitespace-tail target intersects a merged range: "
                            f"{patch.sheet}!{coordinate}"
                        )
            if patch.kind == PatchKind.COPY_STYLE:
                if not patch.source_cell:
                    raise PatchValidationError(f"Style patch {patch.id} has no source cell")
                source_cell = worksheet[patch.source_cell]
                if not isinstance(source_cell, Cell):
                    raise PatchValidationError(
                        f"Style patch source is not a writable cell: "
                        f"{patch.sheet}!{patch.source_cell}"
                    )
                if (
                    cell.protection.locked != source_cell.protection.locked
                    or cell.protection.hidden != source_cell.protection.hidden
                ):
                    raise PatchValidationError(
                        f"Style patch {patch.id} would change cell protection semantics"
                    )
                if cell.number_format != source_cell.number_format:
                    raise PatchValidationError(
                        f"Style patch {patch.id} would change number-format semantics"
                    )
                if cell.quotePrefix != source_cell.quotePrefix:
                    raise PatchValidationError(
                        f"Style patch {patch.id} would change quote-prefix semantics"
                    )
                if cell.pivotButton != source_cell.pivotButton:
                    raise PatchValidationError(
                        f"Style patch {patch.id} would change pivot-button semantics"
                    )
            elif patch.kind == PatchKind.COPY_BORDER:
                if not patch.source_cell:
                    raise PatchValidationError(f"Border patch {patch.id} has no source cell")
                source_cell = worksheet[patch.source_cell]
                if not isinstance(source_cell, Cell):
                    raise PatchValidationError(
                        f"Border patch source is not a writable cell: "
                        f"{patch.sheet}!{patch.source_cell}"
                    )
                source_row = worksheet.row_dimensions.get(source_cell.row)
                if source_row is not None and source_row.hidden:
                    raise PatchValidationError(
                        f"Border patch source is on hidden row {patch.sheet}!{source_cell.row}"
                    )
                if is_column_hidden(worksheet, source_cell.column):
                    raise PatchValidationError(
                        f"Border patch source is in a hidden column: "
                        f"{patch.sheet}!{patch.source_cell}"
                    )
    finally:
        workbook.close()


def _select_patches(
    plan: PatchPlan,
    selected_ids: set[str] | None,
    safe_only: bool,
    *,
    enforce_safety: bool = True,
    accept_layout_risk: bool = False,
) -> list[PatchOperation]:
    by_id = {patch.id: patch for patch in plan.patches}
    if len(by_id) != len(plan.patches):
        raise UsageError("Patch plan contains duplicate patch IDs")
    if safe_only and selected_ids:
        raise UsageError("--safe-only and --patch-id are mutually exclusive")
    if safe_only:
        selected = [patch for patch in plan.patches if patch.safe_only_eligible]
        selected_set = {patch.id for patch in selected}
        incomplete_safe_groups = {
            patch.atomic_group
            for patch in selected
            if patch.atomic_group
            and any(
                member.atomic_group == patch.atomic_group and member.id not in selected_set
                for member in plan.patches
            )
        }
        if incomplete_safe_groups:
            selected = [
                patch for patch in selected if patch.atomic_group not in incomplete_safe_groups
            ]
    else:
        if not selected_ids:
            raise UsageError("Select at least one --patch-id or pass --safe-only")
        unknown = selected_ids - by_id.keys()
        if unknown:
            raise UsageError("Unknown patch IDs: " + ", ".join(sorted(unknown)))
        selected = [by_id[patch_id] for patch_id in sorted(selected_ids)]
    if not selected:
        raise UsageError("No eligible patches were selected")
    if enforce_safety:
        rejected: list[str] = []
        for patch in selected:
            risk = str(getattr(patch.risk, "value", patch.risk))
            if risk == "layout_review":
                if safe_only or not accept_layout_risk or float(patch.confidence) < 0.95:
                    rejected.append(patch.id)
                continue
            if not patch.safe or float(patch.confidence) < 0.95:
                rejected.append(patch.id)
        if rejected:
            raise UsageError(
                "WorkbookLens refuses patches outside the authorized repair risk boundary: "
                + ", ".join(rejected)
            )
    selected_set = {patch.id for patch in selected}
    incomplete_groups = sorted(
        {
            patch.atomic_group
            for patch in selected
            if patch.atomic_group
            and any(
                member.atomic_group == patch.atomic_group and member.id not in selected_set
                for member in plan.patches
            )
        }
    )
    if incomplete_groups:
        raise UsageError(
            "Atomic patch groups must be selected in full: " + ", ".join(incomplete_groups)
        )
    domains: dict[tuple[str, str, str], list[str]] = {}
    for patch in selected:
        domains.setdefault(patch_target_key(patch), []).append(patch.id)
    conflicts = [
        f"{sheet}!{cell} ({domain}: {', '.join(sorted(patch_ids))})"
        for (sheet, cell, domain), patch_ids in sorted(domains.items())
        if len(patch_ids) > 1
    ]
    if conflicts:
        raise UsageError("Conflicting patches target the same cell: " + "; ".join(conflicts))
    tails = [patch for patch in selected if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL]
    for tail in tails:
        cells, rows = tail_authorization(tail)
        for patch in selected:
            if patch.id == tail.id or patch.sheet != tail.sheet:
                continue
            patch_row, _ = coordinate_to_tuple(patch.cell)
            if patch.cell in cells or patch_row in rows:
                raise UsageError(f"Patch {patch.id} intersects formatting-tail cleanup {tail.id}")
    whitespace_tails = [
        patch for patch in selected if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS
    ]
    for tail in whitespace_tails:
        cells = whitespace_tail_authorization(tail)
        for patch in selected:
            if patch.id == tail.id or patch.sheet != tail.sheet:
                continue
            if patch.cell in cells:
                raise UsageError(f"Patch {patch.id} intersects whitespace-tail cleanup {tail.id}")
    return selected


def _validate_canonical_plan(plan: PatchPlan, canonical_plan: PatchPlan) -> None:
    """Reject any serialized repair authority not reproduced by the current scan."""

    if (
        plan.schema_version != canonical_plan.schema_version
        or plan.tool_version != canonical_plan.tool_version
        or plan.source_name != canonical_plan.source_name
        or plan.source_sha256 != canonical_plan.source_sha256
    ):
        raise PatchValidationError(CANONICAL_PLAN_MISMATCH)

    incoming = {patch.id: patch for patch in plan.patches}
    canonical = {patch.id: patch for patch in canonical_plan.patches}
    if len(incoming) != len(plan.patches) or len(canonical) != len(canonical_plan.patches):
        raise PatchValidationError(CANONICAL_PLAN_MISMATCH)
    if incoming.keys() != canonical.keys():
        raise PatchValidationError(CANONICAL_PLAN_MISMATCH)
    for patch_id, patch in incoming.items():
        if patch.model_dump(mode="json") != canonical[patch_id].model_dump(mode="json"):
            raise PatchValidationError(CANONICAL_PLAN_MISMATCH)


def _apply_to_parts(
    archive: zipfile.ZipFile,
    selected: list[PatchOperation],
) -> tuple[dict[str, bytes], bool]:
    sheet_parts = _sheet_parts(archive)
    by_part: dict[str, list[PatchOperation]] = {}
    for patch in selected:
        part = sheet_parts.get(patch.sheet)
        if part is None or part not in archive.namelist():
            raise PatchValidationError(f"Worksheet package part is missing for {patch.sheet!r}")
        by_part.setdefault(part, []).append(patch)
    modified: dict[str, bytes] = {}
    formula_changed = False
    styles: StylesEditor | None = None
    if any(needs_styles(patch.kind) for patch in selected):
        styles_name = "xl/styles.xml"
        if styles_name not in archive.namelist():
            raise PatchValidationError("Workbook package is missing xl/styles.xml")
        styles = StylesEditor.from_root(parse_xml_part(archive.read(styles_name), styles_name))
    for part, patches in by_part.items():
        root = parse_xml_part(archive.read(part), part)
        legacy_dimension_update = False
        for patch in patches:
            if patch.kind == PatchKind.SET_COLUMN_WIDTH:
                apply_column_width(root, patch)
                continue
            if patch.kind == PatchKind.SET_ROW_HEIGHT:
                apply_row_height(root, patch)
                continue
            if patch.kind == PatchKind.SET_SHEET_VIEW:
                apply_sheet_view(root, patch)
                continue
            if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL:
                apply_clear_formatting_tail(archive, part, patch.sheet, root, patch)
                continue
            if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS:
                apply_remove_whitespace_tail_cells(archive, part, patch.sheet, root, patch)
                continue
            _assert_not_in_advanced_formula_range(root, patch.cell)
            _assert_not_in_merged_range(
                root,
                patch.cell,
                allow_anchor=patch.kind in {PatchKind.SET_WRAP_TEXT, PatchKind.SET_SHRINK_TO_FIT},
            )
            if patch.kind == PatchKind.SET_WRAP_TEXT:
                if styles is None:
                    raise PatchValidationError("Styles editor was not initialized")
                apply_alignment(root, patch, styles, shrink=False)
                continue
            if patch.kind == PatchKind.SET_SHRINK_TO_FIT:
                if styles is None:
                    raise PatchValidationError("Styles editor was not initialized")
                apply_alignment(root, patch, styles, shrink=True)
                continue
            if patch.kind == PatchKind.SET_TEXT:
                if styles is None:
                    raise PatchValidationError("Styles editor was not initialized")
                apply_set_text(root, patch, styles)
                formula_changed = True
                continue
            if patch.kind == PatchKind.COPY_BORDER:
                if styles is None:
                    raise PatchValidationError("Styles editor was not initialized")
                if patch.source_cell:
                    _assert_not_in_advanced_formula_range(root, patch.source_cell)
                    _assert_not_in_merged_range(root, patch.source_cell)
                apply_copy_border(root, patch, styles)
                continue
            target = _find_or_create_cell(root, patch.cell)
            _assert_simple_formula_cell(target, f"target {patch.sheet}!{patch.cell}")
            source_element = None
            if patch.source_cell:
                source_element = _find_cell(root, patch.source_cell)
                if source_element is None:
                    raise PatchValidationError(
                        f"Patch source cell is absent: {patch.sheet}!{patch.source_cell}"
                    )
                _assert_simple_formula_cell(
                    source_element, f"source {patch.sheet}!{patch.source_cell}"
                )
            if patch.kind in {
                PatchKind.SET_FORMULA,
                PatchKind.EXTEND_SUM,
                PatchKind.CREATE_FORMULA,
            }:
                if not isinstance(patch.after, str):
                    raise PatchValidationError(f"Patch {patch.id} has a non-string formula")
                _set_formula(target, patch.after)
                formula_changed = True
                legacy_dimension_update = True
            elif patch.kind == PatchKind.SET_NUMERIC:
                _set_numeric(target, patch.after)
                legacy_dimension_update = True
            elif patch.kind == PatchKind.COPY_STYLE:
                if source_element is None:
                    raise PatchValidationError(f"Style patch {patch.id} has no source cell")
                _copy_style(target, source_element)
            else:
                raise PatchValidationError(f"Unsupported patch kind: {patch.kind}")
        if legacy_dimension_update:
            _update_dimension(root)
        modified[part] = _serialize_xml(root)
    if styles is not None and styles.dirty:
        modified["xl/styles.xml"] = _serialize_xml(styles.root)
    if formula_changed:
        modified["xl/workbook.xml"] = _mark_full_recalculation(archive.read("xl/workbook.xml"))
    return modified, formula_changed


def _write_package(
    archive: zipfile.ZipFile, modified: dict[str, bytes], temporary_path: Path
) -> None:
    with zipfile.ZipFile(temporary_path, "w", allowZip64=True) as output:
        output.comment = archive.comment
        for info in archive.infolist():
            copied_info = copy.copy(info)
            if info.filename in modified:
                output.writestr(copied_info, modified[info.filename])
                continue
            with (
                archive.open(info, "r") as source_handle,
                output.open(copied_info, "w") as output_handle,
            ):
                shutil.copyfileobj(source_handle, output_handle, length=1024 * 1024)


def _hash_zip_entry(archive: zipfile.ZipFile, name: str) -> str:
    digest = hashlib.sha256()
    with archive.open(name, "r") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _package_changes(source: Path, output: Path) -> list[PackageChange]:
    with (
        zipfile.ZipFile(source, "r") as before_archive,
        zipfile.ZipFile(output, "r") as after_archive,
    ):
        before_names = set(before_archive.namelist())
        after_names = set(after_archive.namelist())
        changes: list[PackageChange] = []
        for name in sorted(before_names | after_names):
            if name not in before_names:
                changes.append(
                    PackageChange(
                        part=name,
                        action="added",
                        after_sha256=_hash_zip_entry(after_archive, name),
                    )
                )
            elif name not in after_names:
                changes.append(
                    PackageChange(
                        part=name,
                        action="removed",
                        before_sha256=_hash_zip_entry(before_archive, name),
                    )
                )
            else:
                before_hash = _hash_zip_entry(before_archive, name)
                after_hash = _hash_zip_entry(after_archive, name)
                if before_hash != after_hash:
                    changes.append(
                        PackageChange(
                            part=name,
                            action="modified",
                            before_sha256=before_hash,
                            after_sha256=after_hash,
                        )
                    )
        return changes


def _publish_without_overwrite(temporary: Path, output: Path, expected_hash: str) -> None:
    """Publish atomically when possible and always fail if another writer won the path."""

    try:
        os.link(temporary, output)
    except FileExistsError as exc:
        raise UsageError(f"Output already exists and will not be overwritten: {output}") from exc
    except OSError:
        try:
            with temporary.open("rb") as source_handle, output.open("xb") as output_handle:
                shutil.copyfileobj(source_handle, output_handle, length=1024 * 1024)
                output_handle.flush()
                os.fsync(output_handle.fileno())
        except FileExistsError as exc:
            raise UsageError(
                f"Output already exists and will not be overwritten: {output}"
            ) from exc
        except BaseException:
            output.unlink(missing_ok=True)
            raise
        if sha256_file(output) != expected_hash:
            output.unlink(missing_ok=True)
            raise PatchValidationError(
                "Published output failed its final hash verification"
            ) from None
    temporary.unlink(missing_ok=True)


def _validate_semantics(output: Path, selected: list[PatchOperation]) -> None:
    workbook = load_workbook(output, read_only=False, data_only=False, keep_links=False)
    try:
        worksheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}
        for patch in selected:
            worksheet = worksheets.get(patch.sheet)
            if worksheet is None:
                raise PatchValidationError(
                    f"Patched worksheet is missing from output: {patch.sheet!r}"
                )
            if is_layout_kind(patch.kind):
                validate_layout_semantics(worksheet, patch)
                continue
            cell = worksheet[patch.cell]
            if (
                patch.kind
                in {
                    PatchKind.SET_FORMULA,
                    PatchKind.EXTEND_SUM,
                    PatchKind.CREATE_FORMULA,
                }
                and cell.value != patch.after
            ):
                raise PatchValidationError(f"Formula verification failed for patch {patch.id}")
            if patch.kind == PatchKind.SET_NUMERIC and cell.value != patch.after:
                raise PatchValidationError(f"Numeric verification failed for patch {patch.id}")
            if patch.kind == PatchKind.COPY_STYLE:
                if not patch.source_cell:
                    raise PatchValidationError(f"Style patch {patch.id} has no source cell")
                try:
                    source_row, source_column = coordinate_to_tuple(patch.source_cell)
                except (TypeError, ValueError) as exc:
                    raise PatchValidationError(
                        f"Style patch {patch.id} has an invalid source cell"
                    ) from exc
                source_cell = worksheet._cells.get((source_row, source_column))
                if not isinstance(cell, Cell) or not isinstance(source_cell, Cell):
                    raise PatchValidationError(
                        f"Style patch source is missing from output: "
                        f"{patch.sheet}!{patch.source_cell}"
                    )
                if cast(Any, cell)._style != cast(Any, source_cell)._style:
                    raise PatchValidationError(f"Style verification failed for patch {patch.id}")
    finally:
        workbook.close()
    read_only = load_workbook(output, read_only=True, data_only=False, keep_links=False)
    read_only.close()


def patch_ooxml_package(
    source: Path,
    plan: PatchPlan,
    output: Path,
    *,
    selected_ids: set[str] | None = None,
    safe_only: bool = False,
    accept_layout_risk: bool = False,
    limits: PackageLimits | None = None,
    canonical_plan: PatchPlan,
) -> tuple[OoxmlPatchOutput, list[PatchOperation]]:
    """Apply selected safe patches through XML only and validate the resulting package."""

    source = source.expanduser().resolve()
    output = output.expanduser().resolve()
    inspection = inspect_package(source, limits)
    if not inspection.repairable:
        raise UsageError(
            "This workbook package is read-only for repair because its extension, "
            "declared content type, or VBA parts are not a verified plain .xlsx match"
        )
    if output.suffix.lower() != ".xlsx":
        raise UsageError("Repair output must use the .xlsx extension")
    if source == output:
        raise UsageError("Output must be a new path; the source workbook is never overwritten")
    if output.exists():
        raise UsageError(f"Output already exists and will not be overwritten: {output}")
    output.parent.mkdir(parents=True, exist_ok=True)
    selected = _select_patches(
        plan,
        selected_ids,
        safe_only,
        enforce_safety=False,
        accept_layout_risk=accept_layout_risk,
    )
    _verify_preconditions(source, plan, selected, limits)
    _validate_canonical_plan(plan, canonical_plan)
    selected = _select_patches(
        canonical_plan,
        selected_ids,
        safe_only,
        accept_layout_risk=accept_layout_risk,
    )
    source_hash = sha256_file(source)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.stem}.", suffix=".xlsx", dir=output.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(source, "r") as archive:
            modified, formula_changed = _apply_to_parts(archive, selected)
            _write_package(archive, modified, temporary)
        inspect_package(temporary, limits)
        changes = _package_changes(source, temporary)
        changed_parts = {change.part for change in changes}
        expected_parts = set(modified)
        if changed_parts != expected_parts:
            unexpected = changed_parts ^ expected_parts
            raise PatchValidationError(
                "Unexpected OOXML package changes: " + ", ".join(sorted(unexpected))
            )
        if not changes:
            raise PatchValidationError("Selected patches produced no package change")
        _validate_semantics(temporary, selected)
        if sha256_file(source) != source_hash:
            raise PatchValidationError("Source workbook hash changed during apply")
        output_hash = sha256_file(temporary)
        _publish_without_overwrite(temporary, output, output_hash)
        return (
            OoxmlPatchOutput(
                source_sha256=source_hash,
                output_sha256=output_hash,
                changes=changes,
                formula_changed=formula_changed,
            ),
            selected,
        )
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise
