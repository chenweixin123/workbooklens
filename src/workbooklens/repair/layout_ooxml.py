"""Fail-closed OOXML helpers for presentation-only workbook repairs."""

from __future__ import annotations

import copy
import math
import posixpath
import re
import zipfile
from dataclasses import dataclass
from itertools import pairwise
from typing import Any, cast

from lxml import etree
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from workbooklens.exceptions import PatchValidationError, StalePlanError
from workbooklens.layout import (
    FormattingTail,
    WhitespaceTail,
    column_layout_fingerprint,
    row_layout_fingerprint,
    sheet_view_fingerprint,
    tail_layout_fingerprint,
    whitespace_tail_layout_fingerprint,
)
from workbooklens.models import PatchKind, PatchOperation

XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"
MAX_EXCEL_COLUMN = 16_384
MAX_EXCEL_ROW = 1_048_576
MAX_TAIL_CELLS = 100_000
CELL_REFERENCE_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,6}$")
LAYOUT_KINDS = frozenset(
    {
        PatchKind.SET_COLUMN_WIDTH,
        PatchKind.SET_ROW_HEIGHT,
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_SHRINK_TO_FIT,
        PatchKind.SET_TEXT,
        PatchKind.SET_SHEET_VIEW,
        PatchKind.COPY_BORDER,
        PatchKind.CLEAR_FORMATTING_TAIL,
        PatchKind.REMOVE_WHITESPACE_TAIL_CELLS,
    }
)
STYLE_LAYOUT_KINDS = frozenset(
    {
        PatchKind.SET_WRAP_TEXT,
        PatchKind.SET_SHRINK_TO_FIT,
        PatchKind.SET_TEXT,
        PatchKind.COPY_BORDER,
    }
)
ALLOWED_BORDER_EDGES = frozenset({"left", "right", "top", "bottom"})


def _qname(namespace: str, local_name: str) -> str:
    return f"{{{namespace}}}{local_name}"


def _namespace(element: etree._Element) -> str:
    namespace = etree.QName(element).namespace
    if namespace is None:
        raise PatchValidationError("OOXML element is missing a namespace")
    return namespace


def _direct_child(root: etree._Element, local_name: str) -> etree._Element | None:
    return next(
        (child for child in root if etree.QName(child).localname == local_name),
        None,
    )


def _direct_children(root: etree._Element, local_name: str) -> list[etree._Element]:
    return [child for child in root if etree.QName(child).localname == local_name]


def _parse_positive_int(value: str | None, label: str, *, maximum: int) -> int:
    try:
        parsed = int(cast(str, value))
    except (TypeError, ValueError) as exc:
        raise PatchValidationError(f"{label} is not a valid integer") from exc
    if not 1 <= parsed <= maximum:
        raise PatchValidationError(f"{label} is outside Excel bounds")
    return parsed


def _number_text(value: float) -> str:
    return format(value, ".15g")


def _as_finite_number(value: Any, label: str, *, minimum: float, maximum: float) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise PatchValidationError(f"{label} must be numeric")
    parsed = float(value)
    if not math.isfinite(parsed) or not minimum <= parsed <= maximum:
        raise PatchValidationError(f"{label} is outside the supported range")
    return parsed


def _validate_coordinate(value: Any, label: str = "cell coordinate") -> str:
    if not isinstance(value, str):
        raise PatchValidationError(f"{label} must be a string")
    normalized = value.upper()
    if not CELL_REFERENCE_RE.fullmatch(normalized):
        raise PatchValidationError(f"Invalid {label}: {value!r}")
    row, column = coordinate_to_tuple(normalized)
    if row > MAX_EXCEL_ROW or column > MAX_EXCEL_COLUMN:
        raise PatchValidationError(f"{label} is outside Excel bounds: {value!r}")
    return normalized


def _cell_element(root: etree._Element, coordinate: str) -> etree._Element | None:
    return next(
        (
            element
            for element in root.iter()
            if etree.QName(element).localname == "c" and element.get("r") == coordinate
        ),
        None,
    )


def _row_element(root: etree._Element, row_number: int) -> etree._Element | None:
    sheet_data = _direct_child(root, "sheetData")
    if sheet_data is None:
        raise PatchValidationError("Worksheet has no sheetData element")
    return next(
        (
            row
            for row in sheet_data
            if etree.QName(row).localname == "row"
            and _parse_positive_int(row.get("r"), "row r", maximum=MAX_EXCEL_ROW) == row_number
        ),
        None,
    )


def _cell_element_or_create(root: etree._Element, coordinate: str) -> etree._Element:
    existing = _cell_element(root, coordinate)
    if existing is not None:
        return existing
    row_number, column_number = coordinate_to_tuple(coordinate)
    namespace = _namespace(root)
    sheet_data = _direct_child(root, "sheetData")
    if sheet_data is None:
        raise PatchValidationError("Worksheet has no sheetData element")
    row = _row_element(root, row_number)
    if row is None:
        row = etree.Element(_qname(namespace, "row"), r=str(row_number))
        insertion_index = len(sheet_data)
        for index, candidate in enumerate(sheet_data):
            if etree.QName(candidate).localname != "row":
                continue
            candidate_row = _parse_positive_int(
                candidate.get("r"),
                "row r",
                maximum=MAX_EXCEL_ROW,
            )
            if candidate_row > row_number:
                insertion_index = index
                break
        sheet_data.insert(insertion_index, row)
    cell = etree.Element(_qname(namespace, "c"), r=coordinate)
    insertion_index = len(row)
    for index, candidate in enumerate(row):
        if etree.QName(candidate).localname != "c":
            continue
        candidate_coordinate = _validate_coordinate(
            candidate.get("r"),
            "worksheet cell coordinate",
        )
        _, candidate_column = coordinate_to_tuple(candidate_coordinate)
        if candidate_column > column_number:
            insertion_index = index
            break
    row.insert(insertion_index, cell)
    return cell


def _style_index(cell: etree._Element) -> int:
    value = cell.get("s")
    if value is None:
        return 0
    try:
        parsed = int(value)
    except ValueError as exc:
        raise PatchValidationError("Cell style index is invalid") from exc
    if not 0 <= parsed <= 2_147_483_647:
        raise PatchValidationError("Cell style index is outside supported bounds")
    return parsed


def _canonical(element: etree._Element) -> bytes:
    return etree.tostring(element, method="c14n", with_comments=True)


def _validate_count(collection: etree._Element, label: str) -> None:
    declared = collection.get("count")
    if declared is None:
        return
    try:
        parsed = int(declared)
    except ValueError as exc:
        raise PatchValidationError(f"{label} has an invalid count") from exc
    if parsed != len(collection):
        raise PatchValidationError(f"{label} count does not match its children")


def _find_equivalent(collection: etree._Element, candidate: etree._Element) -> int | None:
    signature = _canonical(candidate)
    for index, existing in enumerate(collection):
        if _canonical(existing) == signature:
            return index
    return None


def _append_unique(collection: etree._Element, candidate: etree._Element) -> tuple[int, bool]:
    existing = _find_equivalent(collection, candidate)
    if existing is not None:
        return existing, False
    collection.append(candidate)
    collection.set("count", str(len(collection)))
    return len(collection) - 1, True


def _insert_xf_child(xf: etree._Element, child: etree._Element) -> None:
    order = {"alignment": 0, "protection": 1, "extLst": 2}
    target_order = order.get(etree.QName(child).localname, 99)
    insertion_index = len(xf)
    for index, existing in enumerate(xf):
        if order.get(etree.QName(existing).localname, 99) > target_order:
            insertion_index = index
            break
    xf.insert(insertion_index, child)


def _replace_border_edge(
    border: etree._Element, edge_name: str, replacement: etree._Element
) -> None:
    order = {
        "start": 0,
        "end": 1,
        "left": 2,
        "right": 3,
        "top": 4,
        "bottom": 5,
        "diagonal": 6,
        "vertical": 7,
        "horizontal": 8,
        "extLst": 9,
    }
    existing = next(
        (child for child in border if etree.QName(child).localname == edge_name),
        None,
    )
    if existing is not None:
        border.replace(existing, replacement)
        return
    target_order = order[edge_name]
    insertion_index = len(border)
    for index, child in enumerate(border):
        if order.get(etree.QName(child).localname, 99) > target_order:
            insertion_index = index
            break
    border.insert(insertion_index, replacement)


@dataclass(slots=True)
class StylesEditor:
    """Mutable style-table view that preserves every unmodified style field."""

    root: etree._Element
    cell_xfs: etree._Element
    borders: etree._Element
    dirty: bool = False

    @classmethod
    def from_root(cls, root: etree._Element) -> StylesEditor:
        cell_xfs = _direct_child(root, "cellXfs")
        borders = _direct_child(root, "borders")
        if cell_xfs is None or borders is None:
            raise PatchValidationError("Workbook styles.xml lacks cellXfs or borders")
        _validate_count(cell_xfs, "cellXfs")
        _validate_count(borders, "borders")
        if not len(cell_xfs) or not len(borders):
            raise PatchValidationError("Workbook styles.xml has an empty style collection")
        return cls(root=root, cell_xfs=cell_xfs, borders=borders)

    def _xf(self, cell: etree._Element) -> etree._Element:
        index = _style_index(cell)
        if index >= len(self.cell_xfs):
            raise PatchValidationError(f"Cell references missing style index {index}")
        return self.cell_xfs[index]

    def _assign_xf(self, cell: etree._Element, candidate: etree._Element) -> None:
        index, appended = _append_unique(self.cell_xfs, candidate)
        cell.set("s", str(index))
        self.dirty = self.dirty or appended

    def set_alignment_flag(self, cell: etree._Element, attribute: str, enabled: bool) -> None:
        candidate = copy.deepcopy(self._xf(cell))
        alignment = _direct_child(candidate, "alignment")
        if alignment is None:
            alignment = etree.Element(_qname(_namespace(candidate), "alignment"))
            _insert_xf_child(candidate, alignment)
        if enabled:
            alignment.set(attribute, "1")
        else:
            if attribute in alignment.attrib:
                del alignment.attrib[attribute]
        candidate.set("applyAlignment", "1")
        self._assign_xf(cell, candidate)

    def set_text_number_format(self, cell: etree._Element) -> None:
        candidate = copy.deepcopy(self._xf(cell))
        candidate.set("numFmtId", "49")
        candidate.set("applyNumberFormat", "1")
        self._assign_xf(cell, candidate)

    def copy_border_edge(
        self,
        target: etree._Element,
        source: etree._Element,
        *,
        target_edge: str,
        source_edge: str,
    ) -> None:
        target_xf = self._xf(target)
        source_xf = self._xf(source)
        try:
            target_border_id = int(target_xf.get("borderId", "0"))
            source_border_id = int(source_xf.get("borderId", "0"))
        except ValueError as exc:
            raise PatchValidationError("Cell style has an invalid borderId") from exc
        if not 0 <= target_border_id < len(self.borders):
            raise PatchValidationError("Target style references a missing border")
        if not 0 <= source_border_id < len(self.borders):
            raise PatchValidationError("Source style references a missing border")
        source_border = self.borders[source_border_id]
        source_side = next(
            (child for child in source_border if etree.QName(child).localname == source_edge),
            None,
        )
        if source_side is None:
            raise PatchValidationError(f"Source border has no {source_edge} edge")
        source_style = (source_side.get("style") or "").strip()
        if not source_style or source_style.casefold() == "none":
            raise PatchValidationError(f"Source border {source_edge} edge is not visibly styled")
        candidate_border = copy.deepcopy(self.borders[target_border_id])
        replacement = copy.deepcopy(source_side)
        replacement.tag = _qname(_namespace(candidate_border), target_edge)
        _replace_border_edge(candidate_border, target_edge, replacement)
        border_id, border_appended = _append_unique(self.borders, candidate_border)
        self.dirty = self.dirty or border_appended
        candidate_xf = copy.deepcopy(target_xf)
        candidate_xf.set("borderId", str(border_id))
        candidate_xf.set("applyBorder", "1")
        self._assign_xf(target, candidate_xf)


def is_layout_kind(kind: PatchKind) -> bool:
    return kind in LAYOUT_KINDS


def needs_styles(kind: PatchKind) -> bool:
    return kind in STYLE_LAYOUT_KINDS


def _column_payload(patch: PatchOperation) -> tuple[int, float]:
    if isinstance(patch.after, dict):
        if set(patch.after) != {"column", "width"}:
            raise PatchValidationError(f"Patch {patch.id} has an invalid column-width payload")
        column_text = patch.after["column"]
        width_value = patch.after["width"]
    else:
        _, target_column = coordinate_to_tuple(_validate_coordinate(patch.cell))
        column_text = get_column_letter(target_column)
        width_value = patch.after
    if not isinstance(column_text, str):
        raise PatchValidationError("Column-width payload column must be a letter")
    try:
        column = column_index_from_string(column_text.upper())
    except ValueError as exc:
        raise PatchValidationError(f"Invalid column letter: {column_text!r}") from exc
    if not 1 <= column <= MAX_EXCEL_COLUMN:
        raise PatchValidationError("Column-width target is outside Excel bounds")
    width = _as_finite_number(width_value, "column width", minimum=0.1, maximum=255.0)
    _, anchor_column = coordinate_to_tuple(_validate_coordinate(patch.cell))
    if anchor_column != column:
        raise PatchValidationError("Column-width patch anchor does not match its target column")
    return column, width


def _row_payload(patch: PatchOperation) -> tuple[int, float]:
    if isinstance(patch.after, dict):
        if set(patch.after) != {"row", "height"}:
            raise PatchValidationError(f"Patch {patch.id} has an invalid row-height payload")
        row_value = patch.after["row"]
        height_value = patch.after["height"]
    else:
        row_value, _ = coordinate_to_tuple(_validate_coordinate(patch.cell))
        height_value = patch.after
    if isinstance(row_value, bool) or not isinstance(row_value, int):
        raise PatchValidationError("Row-height payload row must be an integer")
    if not 1 <= row_value <= MAX_EXCEL_ROW:
        raise PatchValidationError("Row-height target is outside Excel bounds")
    height = _as_finite_number(height_value, "row height", minimum=0.1, maximum=409.5)
    anchor_row, _ = coordinate_to_tuple(_validate_coordinate(patch.cell))
    if anchor_row != row_value:
        raise PatchValidationError("Row-height patch anchor does not match its target row")
    return row_value, height


def _boolean_payload(patch: PatchOperation, key: str) -> bool:
    if isinstance(patch.after, dict):
        if set(patch.after) != {key}:
            raise PatchValidationError(f"Patch {patch.id} has an invalid {key} payload")
        value = patch.after[key]
    else:
        value = patch.after
    if not isinstance(value, bool):
        raise PatchValidationError(f"Patch {patch.id} {key} value must be boolean")
    return value


def _view_payload(patch: PatchOperation) -> tuple[str | None, int | None]:
    if isinstance(patch.after, str):
        return _validate_coordinate(patch.after, "top-left cell"), None
    if not isinstance(patch.after, dict):
        raise PatchValidationError(f"Patch {patch.id} has an invalid sheet-view payload")
    allowed = {"top_left_cell", "zoom_scale"}
    if not patch.after or not set(patch.after) <= allowed:
        raise PatchValidationError(f"Patch {patch.id} has an invalid sheet-view payload")
    top_left = (
        _validate_coordinate(patch.after["top_left_cell"], "top-left cell")
        if "top_left_cell" in patch.after
        else None
    )
    if "zoom_scale" not in patch.after:
        return top_left, None
    zoom_value = patch.after["zoom_scale"]
    if isinstance(zoom_value, bool) or not isinstance(zoom_value, int):
        raise PatchValidationError("Sheet-view zoom_scale must be an integer")
    if not 10 <= zoom_value <= 400:
        raise PatchValidationError("Sheet-view zoom_scale is outside Excel bounds")
    return top_left, zoom_value


def _border_payload(patch: PatchOperation) -> tuple[str, str]:
    if isinstance(patch.after, str):
        target_edge = source_edge = patch.after
    elif isinstance(patch.after, dict):
        if set(patch.after) != {"target_edge", "source_edge"}:
            raise PatchValidationError(f"Patch {patch.id} has an invalid border payload")
        target_edge = patch.after["target_edge"]
        source_edge = patch.after["source_edge"]
    else:
        raise PatchValidationError(f"Patch {patch.id} has an invalid border payload")
    if target_edge not in ALLOWED_BORDER_EDGES or source_edge not in ALLOWED_BORDER_EDGES:
        raise PatchValidationError("Border patch uses an unsupported edge")
    return target_edge, source_edge


def _tail_payload(patch: PatchOperation) -> tuple[list[str], list[int], str, str]:
    if not isinstance(patch.after, dict):
        raise PatchValidationError(f"Patch {patch.id} has an invalid formatting-tail payload")
    required = {"cells", "empty_rows", "expected_dimension", "result_dimension"}
    if set(patch.after) != required:
        raise PatchValidationError(f"Patch {patch.id} has an invalid formatting-tail payload")
    raw_cells = patch.after["cells"]
    raw_rows = patch.after["empty_rows"]
    if not isinstance(raw_cells, list) or not isinstance(raw_rows, list):
        raise PatchValidationError("Formatting-tail cells and empty_rows must be lists")
    if len(raw_cells) > MAX_TAIL_CELLS:
        raise PatchValidationError("Formatting-tail cell authorization is too large")
    cells = [_validate_coordinate(value, "formatting-tail cell") for value in raw_cells]
    if not cells or len(set(cells)) != len(cells):
        raise PatchValidationError("Formatting-tail cells must be a non-empty unique list")
    rows: list[int] = []
    for value in raw_rows:
        if isinstance(value, bool) or not isinstance(value, int):
            raise PatchValidationError("Formatting-tail row must be an integer")
        if not 1 <= value <= MAX_EXCEL_ROW:
            raise PatchValidationError("Formatting-tail row is outside Excel bounds")
        rows.append(value)
    if len(set(rows)) != len(rows):
        raise PatchValidationError("Formatting-tail rows must be unique")
    expected = patch.after["expected_dimension"]
    result = patch.after["result_dimension"]
    for label, value in (("expected_dimension", expected), ("result_dimension", result)):
        if not isinstance(value, str):
            raise PatchValidationError(f"Formatting-tail {label} must be a string")
        try:
            CellRange(value)
        except ValueError as exc:
            raise PatchValidationError(f"Formatting-tail {label} is invalid") from exc
    return cells, rows, cast(str, expected), cast(str, result)


def _whitespace_tail_payload(
    patch: PatchOperation,
) -> tuple[list[str], dict[str, int], str, str]:
    if not isinstance(patch.after, dict):
        raise PatchValidationError(f"Patch {patch.id} has an invalid whitespace-tail payload")
    required = {
        "cells",
        "expected_dimension",
        "preserve_style_ids",
        "result_dimension",
        "preserve_row_dimensions",
    }
    if set(patch.after) != required or patch.after["preserve_row_dimensions"] is not True:
        raise PatchValidationError(f"Patch {patch.id} has an invalid whitespace-tail payload")
    raw_cells = patch.after["cells"]
    if not isinstance(raw_cells, list) or not 1 <= len(raw_cells) <= 4_096:
        raise PatchValidationError("Whitespace-tail cells must be a bounded non-empty list")
    cells = [_validate_coordinate(value, "whitespace-tail cell") for value in raw_cells]
    if len(set(cells)) != len(cells):
        raise PatchValidationError("Whitespace-tail cells must be unique")
    raw_preserve_styles = patch.after["preserve_style_ids"]
    if not isinstance(raw_preserve_styles, dict):
        raise PatchValidationError("Whitespace-tail preserve_style_ids must be an object")
    preserve_style_ids: dict[str, int] = {}
    for raw_coordinate, raw_style_id in raw_preserve_styles.items():
        coordinate = _validate_coordinate(raw_coordinate, "preserved-style whitespace cell")
        if (
            coordinate not in cells
            or isinstance(raw_style_id, bool)
            or not isinstance(raw_style_id, int)
            or raw_style_id <= 0
        ):
            raise PatchValidationError(
                "Whitespace-tail preserved styles must map authorized cells to positive style IDs"
            )
        preserve_style_ids[coordinate] = raw_style_id
    expected = patch.after["expected_dimension"]
    result = patch.after["result_dimension"]
    for label, value in (("expected_dimension", expected), ("result_dimension", result)):
        if not isinstance(value, str):
            raise PatchValidationError(f"Whitespace-tail {label} must be a string")
        try:
            CellRange(value)
        except ValueError as exc:
            raise PatchValidationError(f"Whitespace-tail {label} is invalid") from exc
    return cells, preserve_style_ids, cast(str, expected), cast(str, result)


def patch_target_key(patch: PatchOperation) -> tuple[str, str, str]:
    """Return a conflict key that reflects the actual OOXML mutation domain."""

    if patch.kind == PatchKind.SET_COLUMN_WIDTH:
        column, _ = _column_payload(patch)
        return patch.sheet, f"column:{column}", "column-width"
    if patch.kind == PatchKind.SET_ROW_HEIGHT:
        row, _ = _row_payload(patch)
        return patch.sheet, f"row:{row}", "row-height"
    if patch.kind == PatchKind.SET_SHEET_VIEW:
        _view_payload(patch)
        return patch.sheet, "sheet-view", "sheet-view"
    if patch.kind == PatchKind.CLEAR_FORMATTING_TAIL:
        _tail_payload(patch)
        return patch.sheet, "worksheet-dimension-cleanup", "worksheet-dimension-cleanup"
    if patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS:
        _whitespace_tail_payload(patch)
        return patch.sheet, "worksheet-dimension-cleanup", "worksheet-dimension-cleanup"
    if patch.kind == PatchKind.SET_WRAP_TEXT:
        _boolean_payload(patch, "wrap_text")
        return patch.sheet, patch.cell, "alignment-wrap"
    if patch.kind == PatchKind.SET_SHRINK_TO_FIT:
        _boolean_payload(patch, "shrink_to_fit")
        return patch.sheet, patch.cell, "alignment-shrink"
    if patch.kind == PatchKind.COPY_BORDER:
        target_edge, _ = _border_payload(patch)
        return patch.sheet, patch.cell, f"border-{target_edge}"
    if patch.kind == PatchKind.COPY_STYLE:
        return patch.sheet, patch.cell, "style-all"
    return patch.sheet, patch.cell, "content"


def tail_authorization(patch: PatchOperation) -> tuple[set[str], set[int]]:
    cells, rows, _, _ = _tail_payload(patch)
    return set(cells), set(rows)


def whitespace_tail_authorization(patch: PatchOperation) -> set[str]:
    cells, _, _, _ = _whitespace_tail_payload(patch)
    return set(cells)


def whitespace_tail_preserved_style_ids(patch: PatchOperation) -> dict[str, int]:
    """Return cells whose style node must survive whitespace-value cleanup."""

    _, preserve_style_ids, _, _ = _whitespace_tail_payload(patch)
    return preserve_style_ids


def verify_layout_precondition(worksheet: Worksheet, patch: PatchOperation) -> None:
    """Validate the layout fingerprint with the same source state used by the rules."""

    expected = patch.precondition.layout_fingerprint
    actual: str | None
    if patch.kind == PatchKind.SET_COLUMN_WIDTH:
        column, _ = _column_payload(patch)
        actual = column_layout_fingerprint(worksheet, column)
    elif patch.kind == PatchKind.SET_ROW_HEIGHT:
        row, _ = _row_payload(patch)
        actual = row_layout_fingerprint(worksheet, row)
    elif patch.kind == PatchKind.SET_SHEET_VIEW:
        _view_payload(patch)
        actual = sheet_view_fingerprint(worksheet)
    elif patch.kind == PatchKind.CLEAR_FORMATTING_TAIL:
        cells, rows, _, _ = _tail_payload(patch)
        tail = FormattingTail(
            cell_coordinates=tuple(cells),
            cell_ranges=tuple(cells),
            empty_rows=tuple(rows),
            row_ranges=tuple(str(row) for row in rows),
            styled_cell_count=len(cells),
            content_min_row=0,
            content_min_column=0,
            content_max_row=0,
            content_max_column=0,
            observed_max_row=0,
            observed_max_column=0,
        )
        actual = tail_layout_fingerprint(worksheet, tail)
    elif patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS:
        cells, preserve_style_ids, expected_dimension, result_dimension = _whitespace_tail_payload(
            patch
        )
        ranges = tuple(cells)
        rows = [coordinate_to_tuple(coordinate)[0] for coordinate in cells]
        columns = [coordinate_to_tuple(coordinate)[1] for coordinate in cells]
        whitespace_tail = WhitespaceTail(
            cell_coordinates=tuple(cells),
            cell_ranges=ranges,
            row_ranges=(),
            content_min_row=min(rows),
            content_min_column=min(columns),
            content_max_row=max(rows),
            content_max_column=max(columns),
            observed_dimension=expected_dimension,
            result_dimension=result_dimension,
            preserve_style_ids=tuple(
                sorted(
                    preserve_style_ids.items(),
                    key=lambda item: coordinate_to_tuple(item[0]),
                )
            ),
        )
        actual = whitespace_tail_layout_fingerprint(worksheet, whitespace_tail)
    else:
        if expected is not None:
            raise PatchValidationError(
                f"Patch {patch.id} supplies an unsupported layout fingerprint"
            )
        return
    if expected is None:
        raise PatchValidationError(f"Patch {patch.id} is missing its layout fingerprint")
    if actual != expected:
        raise StalePlanError(f"Layout precondition failed for patch {patch.id}; plan is stale")


def apply_column_width(root: etree._Element, patch: PatchOperation) -> None:
    column, width = _column_payload(patch)
    namespace = _namespace(root)
    cols_elements = _direct_children(root, "cols")
    if len(cols_elements) > 1:
        raise PatchValidationError("Worksheet contains multiple cols collections")
    if cols_elements:
        cols = cols_elements[0]
    else:
        cols = etree.Element(_qname(namespace, "cols"))
        sheet_data = _direct_child(root, "sheetData")
        if sheet_data is None:
            raise PatchValidationError("Worksheet has no sheetData element")
        root.insert(root.index(sheet_data), cols)
    definitions: list[tuple[int, int, etree._Element]] = []
    for element in cols:
        if etree.QName(element).localname != "col":
            raise PatchValidationError("cols contains an unsupported child element")
        minimum = _parse_positive_int(element.get("min"), "col min", maximum=MAX_EXCEL_COLUMN)
        maximum = _parse_positive_int(element.get("max"), "col max", maximum=MAX_EXCEL_COLUMN)
        if minimum > maximum:
            raise PatchValidationError("Column definition has min greater than max")
        definitions.append((minimum, maximum, element))
    ordered = sorted(definitions, key=lambda item: (item[0], item[1]))
    for previous, current in pairwise(ordered):
        if current[0] <= previous[1]:
            raise PatchValidationError("Worksheet contains overlapping column definitions")
    covering = [item for item in definitions if item[0] <= column <= item[1]]
    if len(covering) > 1:
        raise PatchValidationError("Target column has overlapping definitions")
    replacement_elements: list[etree._Element] = []
    if covering:
        minimum, maximum, original = covering[0]
        if minimum < column:
            left = copy.deepcopy(original)
            left.set("min", str(minimum))
            left.set("max", str(column - 1))
            replacement_elements.append(left)
        target = copy.deepcopy(original)
        target.set("min", str(column))
        target.set("max", str(column))
        target.set("width", _number_text(width))
        target.set("customWidth", "1")
        replacement_elements.append(target)
        if column < maximum:
            right = copy.deepcopy(original)
            right.set("min", str(column + 1))
            right.set("max", str(maximum))
            replacement_elements.append(right)
        index = cols.index(original)
        cols.remove(original)
        for offset, element in enumerate(replacement_elements):
            cols.insert(index + offset, element)
    else:
        target = etree.Element(
            _qname(namespace, "col"),
            min=str(column),
            max=str(column),
            width=_number_text(width),
            customWidth="1",
        )
        insertion_index = len(cols)
        for minimum, _, element in definitions:
            if minimum > column:
                insertion_index = cols.index(element)
                break
        cols.insert(insertion_index, target)


def apply_row_height(root: etree._Element, patch: PatchOperation) -> None:
    row_number, height = _row_payload(patch)
    row = _row_element(root, row_number)
    if row is None:
        raise PatchValidationError("Row-height target row is absent from worksheet XML")
    row.set("ht", _number_text(height))
    row.set("customHeight", "1")


def apply_alignment(
    root: etree._Element, patch: PatchOperation, styles: StylesEditor, *, shrink: bool
) -> None:
    coordinate = _validate_coordinate(patch.cell)
    cell = _cell_element(root, coordinate)
    if cell is None:
        raise PatchValidationError(f"Alignment target cell is absent: {coordinate}")
    key = "shrink_to_fit" if shrink else "wrap_text"
    value = _boolean_payload(patch, key)
    styles.set_alignment_flag(cell, "shrinkToFit" if shrink else "wrapText", value)


def apply_set_text(root: etree._Element, patch: PatchOperation, styles: StylesEditor) -> None:
    coordinate = _validate_coordinate(patch.cell)
    cell = _cell_element(root, coordinate)
    if cell is None:
        raise PatchValidationError(f"Text target cell is absent: {coordinate}")
    if not isinstance(patch.after, str):
        raise PatchValidationError(f"Patch {patch.id} text output must be a string")
    child_names = [etree.QName(child).localname for child in cell]
    if "f" in child_names:
        raise PatchValidationError("SET_TEXT refuses to replace a formula")
    unsupported_attributes = {str(name) for name in cell.attrib} - {"r", "s", "t"}
    if unsupported_attributes:
        raise PatchValidationError(
            "SET_TEXT refuses cell metadata attributes: "
            + ", ".join(sorted(unsupported_attributes))
        )
    unsupported_children = set(child_names) - {"v", "is", "extLst"}
    if unsupported_children:
        raise PatchValidationError(
            "SET_TEXT refuses unsupported cell children: " + ", ".join(sorted(unsupported_children))
        )
    if child_names.count("v") > 1 or child_names.count("is") > 1:
        raise PatchValidationError("SET_TEXT refuses duplicate value children")
    if child_names.count("extLst") > 1:
        raise PatchValidationError("SET_TEXT refuses duplicate extLst children")
    for child in list(cell):
        if etree.QName(child).localname in {"v", "is"}:
            cell.remove(child)
    cell.set("t", "inlineStr")
    namespace = _namespace(cell)
    inline = etree.Element(_qname(namespace, "is"))
    text = etree.SubElement(inline, _qname(namespace, "t"))
    text.text = patch.after
    if patch.after != patch.after.strip() or "\n" in patch.after or "\t" in patch.after:
        text.set(XML_SPACE, "preserve")
    extension = _direct_child(cell, "extLst")
    if extension is None:
        cell.append(inline)
    else:
        cell.insert(cell.index(extension), inline)
    styles.set_text_number_format(cell)


def apply_sheet_view(root: etree._Element, patch: PatchOperation) -> None:
    top_left, zoom = _view_payload(patch)
    sheet_views = _direct_child(root, "sheetViews")
    if sheet_views is None:
        raise PatchValidationError("Worksheet has no sheetViews element")
    views = [child for child in sheet_views if etree.QName(child).localname == "sheetView"]
    if len(views) != 1:
        raise PatchValidationError("Sheet-view repair requires exactly one sheetView")
    view = views[0]
    panes = [child for child in view if etree.QName(child).localname == "pane"]
    if len(panes) > 1:
        raise PatchValidationError("Sheet-view repair requires at most one pane")
    if panes and top_left is not None:
        raise PatchValidationError("Sheet-view origin repair refuses frozen or split panes")
    if top_left is not None:
        view.set("topLeftCell", top_left)
    if zoom is not None:
        view.set("zoomScale", str(zoom))


def apply_copy_border(root: etree._Element, patch: PatchOperation, styles: StylesEditor) -> None:
    if not patch.source_cell:
        raise PatchValidationError(f"Border patch {patch.id} has no source cell")
    target_coordinate = _validate_coordinate(patch.cell)
    source_coordinate = _validate_coordinate(patch.source_cell, "source cell coordinate")
    target = _cell_element_or_create(root, target_coordinate)
    source = _cell_element(root, source_coordinate)
    if source is None:
        raise PatchValidationError("Border patch source cell is absent")
    target_edge, source_edge = _border_payload(patch)
    styles.copy_border_edge(
        target,
        source,
        target_edge=target_edge,
        source_edge=source_edge,
    )


def _reference_intersects_targets(
    reference: str,
    points: set[tuple[int, int]],
    rows: set[int],
    *,
    include_rows: bool,
    strict: bool = False,
) -> bool:
    """Return whether one unqualified A1 reference touches authorized tail state."""

    normalized = reference.strip().lstrip("@").rstrip("#")
    try:
        min_column, min_row, max_column, max_row = range_boundaries(normalized)
    except (TypeError, ValueError) as exc:
        if strict:
            raise PatchValidationError(f"Malformed OOXML range reference {reference!r}") from exc
        return False
    if include_rows and rows:
        lower_row = 1 if min_row is None else min_row
        upper_row = MAX_EXCEL_ROW if max_row is None else max_row
        if any(lower_row <= row <= upper_row for row in rows):
            return True
    lower_column = 1 if min_column is None else min_column
    upper_column = MAX_EXCEL_COLUMN if max_column is None else max_column
    lower_row = 1 if min_row is None else min_row
    upper_row = MAX_EXCEL_ROW if max_row is None else max_row
    return any(
        lower_row <= row <= upper_row and lower_column <= column <= upper_column
        for row, column in points
    )


def _iter_direct_ranges(value: str) -> list[str]:
    return [item for item in re.split(r"[\s,]+", value.strip()) if item]


def _range_token_targets_sheet(
    token: str,
    *,
    current_is_target: bool,
    target_sheet_name: str,
) -> tuple[bool, str]:
    if "!" not in token:
        return current_is_target, token
    qualifier, reference = token.rsplit("!", 1)
    qualifier = qualifier.strip()
    if qualifier.startswith("'") and qualifier.endswith("'"):
        qualifier = qualifier[1:-1].replace("''", "'")
    if "[" in qualifier or "]" in qualifier:
        return False, reference
    if ":" in qualifier:
        # Without reconstructing workbook sheet-order semantics, conservatively
        # treat a matching 3-D cell/range reference as potentially targeting the sheet.
        return True, reference
    return qualifier.casefold() == target_sheet_name.casefold(), reference


def _formula_range_tokens(expression: str, label: str) -> list[str]:
    expression = expression[1:] if expression.startswith("=") else expression
    try:
        tokens = Tokenizer("=" + expression).items
    except (TokenizerError, ValueError, IndexError) as exc:
        raise PatchValidationError(f"Formatting-tail cleanup cannot safely parse {label}") from exc
    for token in tokens:
        if token.type != "FUNC" or token.subtype != "OPEN":
            continue
        function = token.value[:-1].upper().rsplit(".", maxsplit=1)[-1]
        if function in {"INDIRECT", "OFFSET"}:
            raise PatchValidationError(
                f"Formatting-tail cleanup refuses dynamic reference function {function}"
            )
    return [token.value for token in tokens if token.type == "OPERAND" and token.subtype == "RANGE"]


def _assert_no_formula_references(
    archive: zipfile.ZipFile,
    target_part: str,
    target_sheet_name: str,
    points: set[tuple[int, int]],
    rows: set[int],
) -> None:
    from workbooklens.ooxml.safety import parse_xml_part

    formula_parts = {
        name: name == target_part
        for name in archive.namelist()
        if (name.startswith("xl/worksheets/") or name.startswith("xl/tables/"))
        and name.endswith(".xml")
        and "/_rels/" not in name
    }
    for relationship_type, related_part in _related_parts(archive, target_part):
        if relationship_type.endswith("/table"):
            formula_parts[related_part] = True
    for formula_part, current_is_target in sorted(formula_parts.items()):
        root = parse_xml_part(archive.read(formula_part), formula_part)
        for formula in root.iter():
            if (
                etree.QName(formula).localname
                not in {
                    "f",
                    "formula",
                    "formula1",
                    "formula2",
                    "calculatedColumnFormula",
                    "totalsRowFormula",
                }
                or not formula.text
            ):
                continue
            for token in _formula_range_tokens(
                formula.text,
                f"formula in {formula_part}",
            ):
                targets_sheet, reference = _range_token_targets_sheet(
                    token,
                    current_is_target=current_is_target,
                    target_sheet_name=target_sheet_name,
                )
                if targets_sheet and _reference_intersects_targets(
                    reference,
                    points,
                    rows,
                    include_rows=True,
                ):
                    raise PatchValidationError(
                        "Formatting-tail target is referenced by a worksheet formula"
                    )


def _assert_no_direct_references(
    root: etree._Element,
    points: set[tuple[int, int]],
    rows: set[int],
) -> None:
    reference_attributes = {
        "mergeCell": "ref",
        "hyperlink": "ref",
        "conditionalFormatting": "sqref",
        "dataValidation": "sqref",
        "autoFilter": "ref",
        "protectedRange": "sqref",
        "ignoredError": "sqref",
        "sortState": "ref",
        "sortCondition": "ref",
        "f": "ref",
    }
    for element in root.iter():
        local_name = etree.QName(element).localname
        attribute = reference_attributes.get(local_name)
        if attribute is None:
            continue
        value = element.get(attribute)
        if not value:
            continue
        for reference in _iter_direct_ranges(value):
            if _reference_intersects_targets(
                reference,
                points,
                rows,
                include_rows=True,
                strict=True,
            ):
                raise PatchValidationError(f"Formatting-tail target is referenced by {local_name}")


def _assert_no_page_break_references(
    root: etree._Element,
    points: set[tuple[int, int]],
    rows: set[int],
) -> None:
    target_rows = rows | {row for row, _ in points}
    target_columns = {column for _, column in points}
    for collection_name, targets, maximum in (
        ("rowBreaks", target_rows, MAX_EXCEL_ROW),
        ("colBreaks", target_columns, MAX_EXCEL_COLUMN),
    ):
        collection = _direct_child(root, collection_name)
        if collection is None:
            continue
        for child in collection:
            if etree.QName(child).localname != "brk":
                raise PatchValidationError(
                    f"{collection_name} contains an unsupported child element"
                )
            try:
                break_id = int(cast(str, child.get("id")))
            except (TypeError, ValueError) as exc:
                raise PatchValidationError(
                    f"{collection_name} contains an invalid break id"
                ) from exc
            if not 0 <= break_id <= maximum:
                raise PatchValidationError(f"{collection_name} break id is outside Excel bounds")
            if break_id in targets:
                raise PatchValidationError(f"Formatting-tail target intersects {collection_name}")


def _relationship_part_name(part: str) -> str:
    return posixpath.join(
        posixpath.dirname(part),
        "_rels",
        posixpath.basename(part) + ".rels",
    )


def _resolve_relationship_target(part: str, target: str) -> str:
    if not target or "\\" in target or "\x00" in target:
        raise PatchValidationError(f"Unsafe relationship target: {target!r}")
    if target.startswith("/"):
        candidate = target.lstrip("/")
    else:
        candidate = posixpath.join(posixpath.dirname(part), target)
    normalized = posixpath.normpath(candidate)
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise PatchValidationError(f"Relationship target escapes package root: {target!r}")
    return normalized


def _related_parts(archive: zipfile.ZipFile, part: str) -> list[tuple[str, str]]:
    relationship_name = _relationship_part_name(part)
    if relationship_name not in archive.namelist():
        return []
    from workbooklens.ooxml.safety import parse_xml_part

    root = parse_xml_part(archive.read(relationship_name), relationship_name)
    result: list[tuple[str, str]] = []
    for relationship in root:
        if etree.QName(relationship).localname != "Relationship":
            continue
        if relationship.get("TargetMode") == "External":
            continue
        target = relationship.get("Target")
        relationship_type = relationship.get("Type", "")
        if target:
            result.append((relationship_type, _resolve_relationship_target(part, target)))
    return result


def _assert_no_related_references(
    archive: zipfile.ZipFile,
    part: str,
    points: set[tuple[int, int]],
    rows: set[int],
) -> None:
    from workbooklens.ooxml.safety import parse_xml_part

    for relationship_type, target in _related_parts(archive, part):
        if target not in archive.namelist():
            raise PatchValidationError(f"Worksheet relationship target is missing: {target}")
        if relationship_type.endswith(("/comments", "/threadedComment", "/threadedComments")):
            root = parse_xml_part(archive.read(target), target)
            for element in root.iter():
                reference = element.get("ref")
                if reference and _reference_intersects_targets(
                    reference,
                    points,
                    rows,
                    include_rows=True,
                    strict=True,
                ):
                    raise PatchValidationError("Formatting-tail target has an attached comment")
        elif relationship_type.endswith("/table"):
            root = parse_xml_part(archive.read(target), target)
            reference = root.get("ref")
            if reference and _reference_intersects_targets(
                reference,
                points,
                rows,
                include_rows=True,
                strict=True,
            ):
                raise PatchValidationError("Formatting-tail target intersects an Excel table")
        elif relationship_type.endswith("/drawing"):
            root = parse_xml_part(archive.read(target), target)
            for anchor in root:
                if etree.QName(anchor).localname not in {
                    "oneCellAnchor",
                    "twoCellAnchor",
                    "absoluteAnchor",
                }:
                    continue
                if etree.QName(anchor).localname == "absoluteAnchor":
                    raise PatchValidationError(
                        "Formatting-tail cleanup refuses worksheets with absolute drawings"
                    )
                markers = [
                    child for child in anchor if etree.QName(child).localname in {"from", "to"}
                ]
                if not markers:
                    continue
                positions: list[tuple[int, int]] = []
                for marker in markers:
                    row_node = _direct_child(marker, "row")
                    col_node = _direct_child(marker, "col")
                    if row_node is None or col_node is None:
                        raise PatchValidationError("Drawing anchor is malformed")
                    try:
                        positions.append(
                            (int(row_node.text or "") + 1, int(col_node.text or "") + 1)
                        )
                    except ValueError as exc:
                        raise PatchValidationError("Drawing anchor is malformed") from exc
                min_row = min(row for row, _ in positions)
                max_row = max(row for row, _ in positions)
                min_col = min(column for _, column in positions)
                max_col = max(column for _, column in positions)
                if _reference_intersects_targets(
                    (
                        f"{get_column_letter(min_col)}{min_row}:"
                        f"{get_column_letter(max_col)}{max_row}"
                    ),
                    points,
                    rows,
                    include_rows=True,
                ):
                    raise PatchValidationError("Formatting-tail target intersects a drawing")
        elif relationship_type.endswith("/vmlDrawing"):
            raise PatchValidationError(
                "Formatting-tail cleanup refuses worksheets with VML drawings"
            )


def _assert_no_defined_name_references(
    archive: zipfile.ZipFile,
    sheet_name: str,
    points: set[tuple[int, int]],
    rows: set[int],
) -> None:
    from workbooklens.ooxml.safety import parse_xml_part

    root = parse_xml_part(archive.read("xl/workbook.xml"), "xl/workbook.xml")
    sheets = _direct_child(root, "sheets")
    sheet_names = [child.get("name") for child in sheets] if sheets is not None else []
    try:
        target_sheet_index = sheet_names.index(sheet_name)
    except ValueError as exc:
        raise PatchValidationError(
            "Formatting-tail worksheet is missing from workbook.xml"
        ) from exc
    for element in root.iter():
        if etree.QName(element).localname != "definedName" or not element.text:
            continue
        local_sheet_id = element.get("localSheetId")
        if local_sheet_id is None:
            unqualified_targets_sheet = True
        else:
            try:
                unqualified_targets_sheet = int(local_sheet_id) == target_sheet_index
            except ValueError as exc:
                raise PatchValidationError(
                    "Workbook defined name has an invalid localSheetId"
                ) from exc
        for token in _formula_range_tokens(element.text, "workbook defined name"):
            targets_sheet, reference = _range_token_targets_sheet(
                token,
                current_is_target=unqualified_targets_sheet,
                target_sheet_name=sheet_name,
            )
            if targets_sheet and _reference_intersects_targets(
                reference,
                points,
                rows,
                include_rows=True,
            ):
                raise PatchValidationError(
                    "Formatting-tail target is referenced by a workbook defined name"
                )


def _recompute_spans(root: etree._Element) -> None:
    sheet_data = _direct_child(root, "sheetData")
    if sheet_data is None:
        raise PatchValidationError("Worksheet has no sheetData element")
    for row in sheet_data:
        if etree.QName(row).localname != "row":
            continue
        columns: list[int] = []
        for cell in row:
            if etree.QName(cell).localname != "c":
                continue
            coordinate = cell.get("r")
            if not coordinate:
                raise PatchValidationError("Worksheet cell is missing its coordinate")
            _, column = coordinate_to_tuple(coordinate)
            columns.append(column)
        if columns:
            row.set("spans", f"{min(columns)}:{max(columns)}")
        else:
            if "spans" in row.attrib:
                del row.attrib["spans"]


def _computed_dimension(root: etree._Element) -> str:
    bounds: list[tuple[int, int, int, int]] = []
    for element in root.iter():
        local_name = etree.QName(element).localname
        if local_name == "c" and element.get("r"):
            row, column = coordinate_to_tuple(cast(str, element.get("r")))
            bounds.append((column, row, column, row))
        elif local_name == "mergeCell" and element.get("ref"):
            try:
                merged_bounds = range_boundaries(cast(str, element.get("ref")))
            except ValueError as exc:
                raise PatchValidationError("Worksheet contains an invalid merged range") from exc
            if any(value is None for value in merged_bounds):
                raise PatchValidationError("Worksheet contains a non-cell merged range")
            bounds.append(cast(tuple[int, int, int, int], merged_bounds))
    if not bounds:
        return "A1"
    min_column = min(minimum for minimum, _, _, _ in bounds)
    min_row = min(minimum for _, minimum, _, _ in bounds)
    max_column = max(maximum for _, _, maximum, _ in bounds)
    max_row = max(maximum for _, _, _, maximum in bounds)
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _dimensions_equivalent(left: str, right: str) -> bool:
    """Compare worksheet dimensions by cell bounds, not lexical spelling."""

    return CellRange(left).bounds == CellRange(right).bounds


def _simple_whitespace_cell_text(
    archive: zipfile.ZipFile,
    cell: etree._Element,
) -> str:
    unsupported_attributes = {str(name) for name in cell.attrib} - {"r", "s", "t"}
    if unsupported_attributes:
        raise PatchValidationError("Whitespace-tail cell contains unsupported metadata")
    child_names = [etree.QName(child).localname for child in cell]
    if "f" in child_names or "extLst" in child_names:
        raise PatchValidationError("Whitespace-tail cell contains a formula or extension")
    cell_type = cell.get("t")
    if cell_type == "s":
        if child_names != ["v"] or "xl/sharedStrings.xml" not in archive.namelist():
            raise PatchValidationError("Whitespace-tail shared-string cell is malformed")
        try:
            index = int(cell[0].text or "")
        except ValueError as exc:
            raise PatchValidationError("Whitespace-tail shared-string index is invalid") from exc
        from workbooklens.ooxml.safety import parse_xml_part

        root = parse_xml_part(archive.read("xl/sharedStrings.xml"), "xl/sharedStrings.xml")
        strings = [child for child in root if etree.QName(child).localname == "si"]
        if not 0 <= index < len(strings):
            raise PatchValidationError("Whitespace-tail shared-string index is out of bounds")
        item = strings[index]
        item_children = [child for child in item if etree.QName(child).localname == "t"]
        if len(item_children) != 1 or len(item) != 1:
            raise PatchValidationError("Whitespace-tail cleanup refuses rich shared strings")
        return item_children[0].text or ""
    if cell_type == "inlineStr":
        if child_names != ["is"]:
            raise PatchValidationError("Whitespace-tail inline string is malformed")
        inline = cell[0]
        texts = [child for child in inline if etree.QName(child).localname == "t"]
        if len(texts) != 1 or len(inline) != 1:
            raise PatchValidationError("Whitespace-tail cleanup refuses rich inline strings")
        return texts[0].text or ""
    if cell_type == "str" and child_names == ["v"]:
        return cell[0].text or ""
    raise PatchValidationError("Whitespace-tail target is not a supported literal string")


def _cell_style_id(cell: etree._Element) -> int:
    """Return an OOXML cell style index, rejecting malformed or negative values."""

    raw_style_id = cell.get("s")
    if raw_style_id is None:
        return 0
    try:
        style_id = int(raw_style_id)
    except ValueError as exc:
        raise PatchValidationError("Whitespace-tail cell style index is invalid") from exc
    if style_id < 0:
        raise PatchValidationError("Whitespace-tail cell style index is invalid")
    return style_id


def apply_clear_formatting_tail(
    archive: zipfile.ZipFile,
    part: str,
    sheet_name: str,
    root: etree._Element,
    patch: PatchOperation,
) -> None:
    cells, empty_rows, expected_dimension, result_dimension = _tail_payload(patch)
    dimension = _direct_child(root, "dimension")
    if dimension is None or dimension.get("ref") != expected_dimension:
        raise PatchValidationError("Formatting-tail expected_dimension does not match OOXML")
    points = {coordinate_to_tuple(coordinate) for coordinate in cells}
    rows = set(empty_rows)
    _assert_no_direct_references(root, points, rows)
    _assert_no_page_break_references(root, points, rows)
    _assert_no_formula_references(archive, part, sheet_name, points, rows)
    _assert_no_related_references(archive, part, points, rows)
    _assert_no_defined_name_references(archive, sheet_name, points, rows)
    for coordinate in cells:
        cell = _cell_element(root, coordinate)
        if cell is None:
            raise PatchValidationError(f"Authorized formatting-tail cell is absent: {coordinate}")
        if cell.get("s") is None:
            raise PatchValidationError(
                f"Formatting-tail cell has no explicit style to clear: {coordinate}"
            )
        unsupported_cell_attributes = {str(name) for name in cell.attrib} - {"r", "s", "t"}
        type_value = cell.get("t")
        if (
            len(cell)
            or unsupported_cell_attributes
            or (type_value is not None and type_value != "n")
        ):
            raise PatchValidationError(
                f"Formatting-tail cell {coordinate} contains value, formula, metadata, or extension data"
            )
        parent = cell.getparent()
        if parent is None or etree.QName(parent).localname != "row":
            raise PatchValidationError("Formatting-tail cell has no worksheet row parent")
        parent.remove(cell)
    sheet_data = _direct_child(root, "sheetData")
    if sheet_data is None:
        raise PatchValidationError("Worksheet has no sheetData element")
    for row_number in empty_rows:
        row = _row_element(root, row_number)
        if row is None:
            raise PatchValidationError(f"Authorized empty tail row is absent: {row_number}")
        if len(row):
            raise PatchValidationError(f"Formatting-tail row {row_number} is not empty")
        allowed_attributes = {"r", "spans", "ht", "customHeight", "s", "customFormat"}
        unsupported_row_attributes = {str(name) for name in row.attrib} - allowed_attributes
        if unsupported_row_attributes:
            raise PatchValidationError(
                f"Formatting-tail row {row_number} has unsupported attributes: "
                + ", ".join(sorted(unsupported_row_attributes))
            )
        if row.get("ht") is not None or row.get("customHeight") in {"1", "true"}:
            raise PatchValidationError("Formatting-tail cleanup refuses custom-height rows")
        if row.get("hidden") in {"1", "true"} or row.get("collapsed") in {"1", "true"}:
            raise PatchValidationError("Formatting-tail cleanup refuses hidden or collapsed rows")
        if int(row.get("outlineLevel", "0")) != 0:
            raise PatchValidationError("Formatting-tail cleanup refuses outlined rows")
        sheet_data.remove(row)
    _recompute_spans(root)
    computed = _computed_dimension(root)
    if not _dimensions_equivalent(computed, result_dimension):
        raise PatchValidationError(
            f"Formatting-tail result_dimension mismatch: computed {computed}, authorized {result_dimension}"
        )
    dimension.set("ref", computed)


def apply_remove_whitespace_tail_cells(
    archive: zipfile.ZipFile,
    part: str,
    sheet_name: str,
    root: etree._Element,
    patch: PatchOperation,
) -> None:
    cells, preserve_style_ids, expected_dimension, result_dimension = _whitespace_tail_payload(
        patch
    )
    dimension = _direct_child(root, "dimension")
    dimension_reference = dimension.get("ref") if dimension is not None else None
    if (
        dimension is None
        or not isinstance(dimension_reference, str)
        or not _dimensions_equivalent(dimension_reference, expected_dimension)
    ):
        raise PatchValidationError("Whitespace-tail expected_dimension does not match OOXML")
    points = {coordinate_to_tuple(coordinate) for coordinate in cells}
    rows: set[int] = set()
    _assert_no_direct_references(root, points, rows)
    _assert_no_page_break_references(root, points, rows)
    _assert_no_formula_references(archive, part, sheet_name, points, rows)
    _assert_no_related_references(archive, part, points, rows)
    _assert_no_defined_name_references(archive, sheet_name, points, rows)
    for coordinate in cells:
        cell = _cell_element(root, coordinate)
        if cell is None:
            raise PatchValidationError(f"Authorized whitespace-tail cell is absent: {coordinate}")
        text = _simple_whitespace_cell_text(archive, cell)
        if not text or text.strip() or any(character in "\r\n\t" for character in text):
            raise PatchValidationError(
                f"Whitespace-tail cell is no longer simple literal whitespace: {coordinate}"
            )
        style_id = _cell_style_id(cell)
        preserved_style_id = preserve_style_ids.get(coordinate)
        if preserved_style_id is None and style_id != 0:
            raise PatchValidationError(
                f"Whitespace-tail cell has an unauthorized non-default style: {coordinate}"
            )
        if preserved_style_id is not None and style_id != preserved_style_id:
            raise PatchValidationError(
                f"Whitespace-tail cell style no longer matches its authorization: {coordinate}"
            )
        parent = cell.getparent()
        if parent is None or etree.QName(parent).localname != "row":
            raise PatchValidationError("Whitespace-tail cell has no worksheet row parent")
        if preserved_style_id is None:
            parent.remove(cell)
        else:
            for child in list(cell):
                cell.remove(child)
            if "t" in cell.attrib:
                del cell.attrib["t"]
    _recompute_spans(root)
    computed = _computed_dimension(root)
    if not _dimensions_equivalent(computed, result_dimension):
        raise PatchValidationError(
            f"Whitespace-tail result_dimension mismatch: computed {computed}, authorized {result_dimension}"
        )
    dimension.set("ref", computed)


def validate_layout_semantics(
    worksheet: Worksheet,
    patch: PatchOperation,
) -> None:
    """Verify openpyxl observes the requested layout result after package rewrite."""

    if patch.kind == PatchKind.SET_COLUMN_WIDTH:
        column, width = _column_payload(patch)
        key = get_column_letter(column)
        dimension = next(
            (
                item
                for item in worksheet.column_dimensions.values()
                if item.min is not None and item.max is not None and item.min <= column <= item.max
            ),
            None,
        )
        if (
            dimension is None
            or dimension.width is None
            or not math.isclose(float(dimension.width), width, rel_tol=0.0, abs_tol=1e-9)
        ):
            raise PatchValidationError(f"Column-width verification failed for {key}")
    elif patch.kind == PatchKind.SET_ROW_HEIGHT:
        row, height = _row_payload(patch)
        observed = worksheet.row_dimensions[row].height
        if observed is None or not math.isclose(float(observed), height, rel_tol=0.0, abs_tol=1e-9):
            raise PatchValidationError(f"Row-height verification failed for row {row}")
    elif patch.kind == PatchKind.SET_WRAP_TEXT:
        if bool(worksheet[patch.cell].alignment.wrap_text) != _boolean_payload(patch, "wrap_text"):
            raise PatchValidationError(f"Wrap-text verification failed for patch {patch.id}")
    elif patch.kind == PatchKind.SET_SHRINK_TO_FIT:
        if bool(worksheet[patch.cell].alignment.shrink_to_fit) != _boolean_payload(
            patch, "shrink_to_fit"
        ):
            raise PatchValidationError(f"Shrink-to-fit verification failed for patch {patch.id}")
    elif patch.kind == PatchKind.SET_TEXT:
        cell = worksheet[patch.cell]
        if cell.value != patch.after or cell.data_type != "s" or cell.number_format != "@":
            raise PatchValidationError(f"Text verification failed for patch {patch.id}")
    elif patch.kind == PatchKind.SET_SHEET_VIEW:
        top_left, zoom = _view_payload(patch)
        if top_left is not None and worksheet.sheet_view.topLeftCell != top_left:
            raise PatchValidationError(f"Sheet-view verification failed for patch {patch.id}")
        if zoom is not None and worksheet.sheet_view.zoomScale != zoom:
            raise PatchValidationError(f"Sheet-view zoom verification failed for patch {patch.id}")
    elif patch.kind == PatchKind.COPY_BORDER:
        if not patch.source_cell:
            raise PatchValidationError(f"Border patch {patch.id} has no source cell")
        target_edge, source_edge = _border_payload(patch)
        target_side = getattr(worksheet[patch.cell].border, target_edge)
        source_side = getattr(worksheet[patch.source_cell].border, source_edge)
        if target_side != source_side:
            raise PatchValidationError(f"Border verification failed for patch {patch.id}")
    elif patch.kind == PatchKind.CLEAR_FORMATTING_TAIL:
        cells, rows, _, result = _tail_payload(patch)
        for coordinate in cells:
            row, column = coordinate_to_tuple(coordinate)
            if (row, column) in worksheet._cells:
                raise PatchValidationError(
                    f"Formatting-tail cell remains after repair: {coordinate}"
                )
        for row in rows:
            if row in worksheet.row_dimensions:
                raise PatchValidationError(f"Formatting-tail row remains after repair: {row}")
        if not _dimensions_equivalent(worksheet.calculate_dimension(), result):
            raise PatchValidationError("Formatting-tail dimension verification failed")
    elif patch.kind == PatchKind.REMOVE_WHITESPACE_TAIL_CELLS:
        cells, preserve_style_ids, _, result = _whitespace_tail_payload(patch)
        for coordinate in cells:
            row, column = coordinate_to_tuple(coordinate)
            cell = worksheet._cells.get((row, column))
            expected_style_id = preserve_style_ids.get(coordinate)
            if expected_style_id is None and cell is not None:
                raise PatchValidationError(
                    f"Whitespace-tail cell remains after repair: {coordinate}"
                )
            if expected_style_id is not None and (
                cell is None or cell.value is not None or cell.style_id != expected_style_id
            ):
                raise PatchValidationError(
                    f"Whitespace-tail style preservation failed after repair: {coordinate}"
                )
        if not _dimensions_equivalent(worksheet.calculate_dimension(), result):
            raise PatchValidationError("Whitespace-tail dimension verification failed")


__all__ = [
    "LAYOUT_KINDS",
    "StylesEditor",
    "apply_alignment",
    "apply_clear_formatting_tail",
    "apply_column_width",
    "apply_copy_border",
    "apply_remove_whitespace_tail_cells",
    "apply_row_height",
    "apply_set_text",
    "apply_sheet_view",
    "is_layout_kind",
    "needs_styles",
    "patch_target_key",
    "tail_authorization",
    "validate_layout_semantics",
    "verify_layout_precondition",
    "whitespace_tail_authorization",
    "whitespace_tail_preserved_style_ids",
]
