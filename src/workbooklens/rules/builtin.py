"""The deterministic WorkbookLens built-in rules."""

from __future__ import annotations

import math
import re
import unicodedata
from collections import Counter, defaultdict
from collections.abc import Iterable, Sequence
from copy import copy
from decimal import Decimal, InvalidOperation
from itertools import pairwise
from typing import Any, cast

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW

from workbooklens.formulas import (
    UnsupportedFormulaError,
    analyze_formula,
    normalize_formula,
    translate_formula,
)
from workbooklens.layout import (
    FormattingTail,
    WhitespaceTail,
    column_layout_fingerprint,
    column_width,
    estimated_text_width,
    explicit_row_height,
    find_formatting_tail,
    find_whitespace_tail,
    has_visible_border,
    measure_text_cell,
    range_intersects,
    row_height,
    row_layout_fingerprint,
    sheet_view_fingerprint,
    tail_layout_fingerprint,
    visual_border_signature,
    whitespace_tail_layout_fingerprint,
)
from workbooklens.models import (
    LAYOUT_REVIEW_PATCH_KINDS,
    Confidence,
    Evidence,
    Finding,
    PatchKind,
    PatchOperation,
    PatchPrecondition,
    PatchRisk,
    Region,
    Severity,
)
from workbooklens.rules.base import RuleContext, RuleResult, WorkbookRule
from workbooklens.snapshot import cell_fingerprint
from workbooklens.utils import stable_id
from workbooklens.worksheet_state import hidden_column_spans, is_column_hidden

EXCEL_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}
CONSERVATIVE_VIEWPORT_WIDTH = 110.0
CONSERVATIVE_VIEWPORT_HEIGHT = 360.0
MIN_AUTOMATIC_VIEW_ZOOM = 50
MAX_EXCEL_ROW_HEIGHT = 409.5
SIMPLE_SUM_RE = re.compile(
    r"^=SUM\((?P<sheet>(?:'(?:[^']|'')+'|[^'!]+)!)?"
    r"(?P<col1>\$?[A-Z]{1,3})(?P<start>\$?\d+):"
    r"(?P<col2>\$?[A-Z]{1,3})(?P<end>\$?\d+)\)$",
    re.IGNORECASE,
)
AGGREGATE_FUNCTION_RE = re.compile(
    r"(?<![A-Z0-9_])(?:(?:_XLFN|_XLWS)\.)*(?:SUM|SUBTOTAL|AGGREGATE)\s*\(",
    re.IGNORECASE,
)
SUMMARY_LABEL_RE = re.compile(
    r"(?<![A-Z0-9_])(?:(?:grand|sub)[\s-]+total|subtotal|total|average|avg|mean|"
    r"minimum|maximum|min|max|median|summary|ending[\s-]+balance|closing[\s-]+balance|"
    r"net[\s-]+amount|adjustment|override|manual[\s-]+override)(?![A-Z0-9_])"
    r"|(?:合计|总计|小计|汇总|平均|均值|最大|最小|中位数|总金额|累计|净额|期末余额|"
    r"期初余额|调整项|调整|手工覆盖|人工覆盖)",
    re.IGNORECASE,
)
DIRECT_MEASURE_HEADER_TOKENS = {
    "amount",
    "balance",
    "cost",
    "count",
    "duration",
    "hours",
    "length",
    "limit",
    "margin",
    "measure",
    "price",
    "quantity",
    "rate",
    "revenue",
    "score",
    "tax",
    "total",
    "units",
    "value",
    "volume",
    "weight",
}
MEASURE_HEADER_MARKERS = (
    "金额",
    "余额",
    "价格",
    "单价",
    "成本",
    "数量",
    "数额",
    "费用",
    "税额",
    "税率",
    "收入",
    "营收",
    "时长",
    "小时",
    "长度",
    "重量",
    "比率",
    "比例",
    "分数",
    "得分",
    "上限",
    "限额",
    "额度",
    "容量",
    "体积",
)
DIRECT_IDENTIFIER_HEADER_TOKENS = {
    "code",
    "id",
    "identifier",
    "iban",
    "imei",
    "imsi",
    "isbn",
    "key",
    "ssn",
    "sku",
    "vin",
    "zip",
}
IDENTIFIER_NUMBER_PREFIXES = {
    "account",
    "bank",
    "card",
    "claim",
    "contract",
    "credit",
    "customer",
    "employee",
    "invoice",
    "license",
    "member",
    "mobile",
    "order",
    "passport",
    "phone",
    "policy",
    "product",
    "record",
    "reference",
    "routing",
    "security",
    "serial",
    "social",
    "telephone",
    "vendor",
}
IDENTIFIER_HEADER_MARKERS = (
    "标识",
    "编号",
    "编码",
    "代码",
    "号码",
    "账号",
    "帐号",
    "手机号",
    "电话",
    "身份证",
    "证件",
    "银行卡",
    "卡号",
    "社保",
    "护照",
    "单号",
    "工号",
    "学号",
    "邮编",
)


def _confidence(value: float) -> Confidence:
    return Confidence(max(0.0, min(1.0, value)))


def _make_finding(
    *,
    context: RuleContext,
    rule_id: str,
    title: str,
    explanation: str,
    severity: Severity,
    confidence: float,
    sheet: str | None,
    location: str | None,
    evidence: Evidence,
    expected: str,
    suggested_action: str,
    patches: Sequence[PatchOperation] = (),
    identity_discriminator: Any = None,
) -> Finding:
    identifier = stable_id("finding", rule_id, sheet, location, identity_discriminator)
    return Finding(
        id=identifier,
        content_fingerprint=stable_id(
            "finding-content", evidence.model_dump(mode="json"), length=24
        ),
        rule_id=rule_id,
        title=title,
        explanation=explanation,
        severity=severity,
        confidence=_confidence(confidence),
        workbook=context.path.name,
        sheet=sheet,
        location=location,
        evidence=evidence,
        expected=expected,
        suggested_action=suggested_action,
        safe_patch_available=any(patch.safe_only_eligible for patch in patches),
        patch_ids=[patch.id for patch in patches],
    )


def _make_patch(
    *,
    kind: PatchKind,
    worksheet: Worksheet,
    cell: Cell,
    before: Any,
    after: Any,
    confidence: float,
    description: str,
    source_cell: str | None = None,
    layout_fingerprint: str | None = None,
    atomic_group: str | None = None,
) -> PatchOperation:
    risk = PatchRisk.LAYOUT_REVIEW if kind in LAYOUT_REVIEW_PATCH_KINDS else PatchRisk.SAFE
    safe = confidence >= 0.95 and risk == PatchRisk.SAFE
    expected_formula = cell.value if cell.data_type == "f" and isinstance(cell.value, str) else None
    patch_id = stable_id(
        "patch", kind.value, worksheet.title, cell.coordinate, before, after, source_cell
    )
    return PatchOperation(
        id=patch_id,
        kind=kind,
        sheet=worksheet.title,
        cell=cell.coordinate,
        before=before,
        after=after,
        source_cell=source_cell,
        confidence=_confidence(confidence),
        safe=safe,
        risk=risk,
        description=description,
        precondition=PatchPrecondition(
            cell_fingerprint=cell_fingerprint(cell),
            expected_value=None if cell.data_type == "f" else cell.value,
            expected_formula=expected_formula,
            expected_style_id=cell.style_id,
            layout_fingerprint=layout_fingerprint,
        ),
        atomic_group=atomic_group,
    )


def _cell(worksheet: Worksheet, row: int, column: int) -> Cell:
    return cast(Cell, worksheet.cell(row, column))


def _band_cells(worksheet: Worksheet, band: Region) -> list[Cell]:
    if band.kind == "formula_row":
        return [
            _cell(worksheet, band.min_row, column)
            for column in range(band.min_column, band.max_column + 1)
        ]
    return [_cell(worksheet, row, band.min_column) for row in range(band.min_row, band.max_row + 1)]


def _formula_signature(cell: Cell) -> str | None:
    if cell.data_type != "f" or not isinstance(cell.value, str):
        return None
    try:
        features = analyze_formula(cell.value)
        if features.external_references or features.unsupported_reason:
            return None
        return normalize_formula(cell.value, cell.coordinate)
    except (ValueError, UnsupportedFormulaError):
        return None


def _in_unsupported_formula_range(context: RuleContext, sheet: str, coordinate: str) -> bool:
    return any(
        coordinate in formula_range
        for formula_range in context.unsupported_formula_ranges.get(sheet, ())
    )


def _band_has_unsupported_formula(context: RuleContext, sheet: str, cells: Sequence[Cell]) -> bool:
    return any(_in_unsupported_formula_range(context, sheet, cell.coordinate) for cell in cells)


def _is_aggregate_formula(value: Any) -> bool:
    if not isinstance(value, str) or not value.startswith("="):
        return False
    output: list[str] = []
    index = 0
    inside_string = False
    while index < len(value):
        character = value[index]
        if character == '"':
            if inside_string and index + 1 < len(value) and value[index + 1] == '"':
                output.extend((" ", " "))
                index += 2
                continue
            inside_string = not inside_string
            output.append(" ")
        else:
            output.append(" " if inside_string else character)
        index += 1
    return AGGREGATE_FUNCTION_RE.search("".join(output)) is not None


def _aggregate_formula_spans_other_rows(value: Any, row: int) -> bool:
    if not _is_aggregate_formula(value):
        return False
    try:
        tokens = Tokenizer(cast(str, value)).items
    except (ValueError, IndexError, TokenizerError):
        return True

    function_stack: list[bool] = []
    aggregate_depth = 0
    references: list[str] = []
    for token in tokens:
        if token.type == "FUNC" and token.subtype == "OPEN":
            is_aggregate = AGGREGATE_FUNCTION_RE.search(token.value) is not None
            function_stack.append(is_aggregate)
            aggregate_depth += int(is_aggregate)
        elif token.type == "FUNC" and token.subtype == "CLOSE":
            if not function_stack:
                return True
            aggregate_depth -= int(function_stack.pop())
        elif aggregate_depth and token.type == "OPERAND" and token.subtype == "RANGE":
            references.append(token.value)
    if function_stack:
        return True

    parsed_reference = False
    for reference in references:
        local_reference = reference.rsplit("!", 1)[-1].replace("$", "")
        try:
            _min_column, min_row, _max_column, max_row = range_boundaries(local_reference)
        except ValueError:
            if "@" in reference:
                parsed_reference = True
                continue
            return True
        parsed_reference = True
        if min_row is None or max_row is None or min_row != row or max_row != row:
            return True
    return False if parsed_reference else bool(references)


def _finite_nonnegative_pane_split(value: Any) -> float | None:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return None
    try:
        split = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return split if math.isfinite(split) and split >= 0 else None


def _bounded_view_coordinate(value: Any) -> tuple[int, int] | None:
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        row, column = coordinate_to_tuple(value.replace("$", ""))
    except (TypeError, ValueError):
        return None
    if not 1 <= row <= MAX_ROW or not 1 <= column <= MAX_COLUMN:
        return None
    return row, column


def _is_band_boundary(cell: Cell, band: Region) -> bool:
    if band.kind == "formula_row":
        return cell.column in {band.min_column, band.max_column}
    return cell.row in {band.min_row, band.max_row}


def _is_boundary_aggregate(cell: Cell, band: Region) -> bool:
    if not isinstance(cell.value, str) or not cell.value.startswith("="):
        return False
    features = analyze_formula(cell.value)
    references: list[tuple[int, int, int, int]] = []
    for reference in features.references:
        if "!" in reference:
            continue
        try:
            boundaries = range_boundaries(reference.replace("$", ""))
        except ValueError:
            continue
        if not all(isinstance(boundary, int) for boundary in boundaries):
            continue
        references.append(cast(tuple[int, int, int, int], boundaries))
    if not references:
        return False
    cell_row = int(cell.row)
    cell_column = int(cell.column)
    expected: tuple[int, int, int, int] | None
    if band.kind == "formula_row":
        expected = (
            (band.min_column, cell_row, cell_column - 1, cell_row)
            if cell_column == band.max_column
            else (cell_column + 1, cell_row, band.max_column, cell_row)
            if cell_column == band.min_column
            else None
        )
    else:
        expected = (
            (cell_column, band.min_row, cell_column, cell_row - 1)
            if cell_row == band.max_row
            else (cell_column, cell_row + 1, cell_column, band.max_row)
            if cell_row == band.min_row
            else None
        )
    return expected is not None and expected in references


def _is_merged_non_anchor(worksheet: Worksheet, coordinate: str) -> bool:
    for merged in worksheet.merged_cells.ranges:
        if coordinate not in merged:
            continue
        anchor = f"{get_column_letter(merged.min_col)}{merged.min_row}"
        return coordinate != anchor
    return False


def _is_in_merged_range(worksheet: Worksheet, coordinate: str) -> bool:
    return any(coordinate in merged for merged in worksheet.merged_cells.ranges)


def _is_hidden_cell(worksheet: Worksheet, cell: Cell | MergedCell) -> bool:
    row = cell.row
    column = cell.column
    if worksheet.sheet_state != "visible" or row is None or column is None:
        return True
    row_dimension = worksheet.row_dimensions.get(row)
    return bool(
        (row_dimension is not None and row_dimension.hidden) or is_column_hidden(worksheet, column)
    )


def _is_protected_target(worksheet: Worksheet) -> bool:
    return bool(worksheet.protection.sheet)


def _looks_like_identifier_header(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    lowered = value.strip().lower()
    if not lowered:
        return False
    if any(marker in lowered for marker in IDENTIFIER_HEADER_MARKERS):
        return True
    tokens = [token for token in re.split(r"[^a-z0-9]+", lowered) if token]
    if any(token in DIRECT_IDENTIFIER_HEADER_TOKENS for token in tokens):
        return True
    number_markers = {"number", "no", "num"}
    return bool(number_markers.intersection(tokens)) and any(
        token in IDENTIFIER_NUMBER_PREFIXES for token in tokens
    )


def _looks_like_measure_header(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    lowered = value.strip().lower()
    if not lowered:
        return False
    if any(marker in lowered for marker in MEASURE_HEADER_MARKERS):
        return True
    tokens = [token for token in re.split(r"[^a-z0-9]+", lowered) if token]
    return any(token in DIRECT_MEASURE_HEADER_TOKENS for token in tokens)


def _band_position(cell: Cell, band: Region) -> int:
    return cell.column if band.kind == "formula_row" else cell.row


def _isolated_outliers(cells: Sequence[Cell], band: Region) -> bool:
    positions = sorted(_band_position(cell, band) for cell in cells)
    return all(right - left > 1 for left, right in pairwise(positions))


def _same_data_region(context: RuleContext, worksheet: Worksheet, *cells: Cell) -> bool:
    return any(
        all(
            region.min_row <= cell.row <= region.max_row
            and region.min_column <= cell.column <= region.max_column
            for cell in cells
        )
        for region in context.data_regions.get(worksheet.title, ())
    )


def _same_protection(left: Cell, right: Cell) -> bool:
    return bool(
        left.protection.locked == right.protection.locked
        and left.protection.hidden == right.protection.hidden
    )


def _style_copy_preserves_semantics(target: Cell, source: Cell) -> bool:
    return (
        _same_protection(target, source)
        and target.number_format == source.number_format
        and target.quotePrefix == source.quotePrefix
        and target.pivotButton == source.pivotButton
    )


def _intersects_any(cell_ranges: Sequence[str], target: str) -> bool:
    try:
        return any(range_intersects(cell_range, target) for cell_range in cell_ranges)
    except ValueError:
        return False


def _unqualified_ranges(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    candidates: Iterable[Any]
    ranges = getattr(value, "ranges", None)
    if ranges is not None:
        candidates = ranges
    elif isinstance(value, (list, tuple, set)):
        candidates = value
    else:
        candidates = str(value).split(",")
    normalized: list[str] = []
    for candidate in candidates:
        text = str(candidate).strip()
        if not text:
            continue
        if "!" in text:
            text = text.rsplit("!", 1)[1]
        text = text.replace("$", "").strip().strip("'")
        row_match = re.fullmatch(r"(\d+):(\d+)", text)
        column_match = re.fullmatch(r"([A-Z]+):([A-Z]+)", text, re.IGNORECASE)
        if row_match:
            text = f"A{row_match.group(1)}:XFD{row_match.group(2)}"
        elif column_match:
            text = f"{column_match.group(1)}1:{column_match.group(2)}1048576"
        if text:
            normalized.append(text)
    return tuple(normalized)


def _drawing_anchor_range(drawing: Any) -> str | None:
    anchor = getattr(drawing, "anchor", None)
    if isinstance(anchor, str):
        return anchor
    start = getattr(anchor, "_from", None)
    if start is None:
        return None
    end = getattr(anchor, "to", None)
    min_row = int(start.row) + 1
    min_column = int(start.col) + 1
    max_row = int(end.row) + 1 if end is not None else min_row
    max_column = int(end.col) + 1 if end is not None else min_column
    left = f"{get_column_letter(min_column)}{min_row}"
    right = f"{get_column_letter(max_column)}{max_row}"
    return left if left == right else f"{left}:{right}"


def _tail_intersects_structure(
    worksheet: Worksheet,
    tail: FormattingTail | WhitespaceTail,
) -> bool:
    cell_ranges = tail.cell_ranges
    row_ranges = tuple(
        f"A{start_text}:XFD{end_text if separator else start_text}"
        for range_text in tail.row_ranges
        for start_text, separator, end_text in [range_text.partition(":")]
    )
    target_ranges = (*cell_ranges, *row_ranges)
    for cell_range in target_ranges:
        if any(
            range_intersects(cell_range, str(merged)) for merged in worksheet.merged_cells.ranges
        ):
            return True
        if any(range_intersects(cell_range, table.ref) for table in worksheet.tables.values()):
            return True
        if worksheet.auto_filter.ref and range_intersects(cell_range, worksheet.auto_filter.ref):
            return True
        if worksheet.data_validations is not None:
            for validation in worksheet.data_validations.dataValidation:
                if any(
                    range_intersects(cell_range, str(validation_range))
                    for validation_range in validation.ranges.ranges
                ):
                    return True
        for conditional_formatting in worksheet.conditional_formatting:
            if any(
                range_intersects(cell_range, str(target))
                for target in conditional_formatting.sqref.ranges
            ):
                return True
    if any(
        _intersects_any(target_ranges, print_range)
        for print_range in _unqualified_ranges(
            (worksheet.print_area, worksheet.print_title_rows, worksheet.print_title_cols)
        )
    ):
        return True
    workbook = worksheet.parent
    if workbook is not None:
        for defined_name in workbook.defined_names.values():
            try:
                destinations = list(defined_name.destinations)
            except (AttributeError, TypeError, ValueError):
                continue
            if any(
                sheet_name == worksheet.title
                and any(
                    _intersects_any(target_ranges, normalized)
                    for normalized in _unqualified_ranges(range_text)
                )
                for sheet_name, range_text in destinations
            ):
                return True
    for coordinate in tail.cell_coordinates:
        row, column = coordinate_to_tuple(coordinate)
        cell = worksheet._cells.get((row, column))
        if isinstance(cell, Cell) and (cell.comment is not None or cell.hyperlink is not None):
            return True
    if any(
        any(
            CellRange(cell_range).min_row <= item.id <= CellRange(cell_range).max_row
            for cell_range in target_ranges
        )
        for item in worksheet.row_breaks.brk
    ):
        return True
    if any(
        any(
            CellRange(cell_range).min_col <= item.id <= CellRange(cell_range).max_col
            for cell_range in target_ranges
        )
        for item in worksheet.col_breaks.brk
    ):
        return True
    for drawing in (
        *getattr(worksheet, "_charts", ()),
        *getattr(worksheet, "_images", ()),
    ):
        anchor_range = _drawing_anchor_range(drawing)
        if anchor_range is not None and _intersects_any(target_ranges, anchor_range):
            return True
    return False


def _formula_reference_target(
    reference: str,
    *,
    current_sheet: str,
    target_sheet: str,
) -> tuple[bool, str] | None:
    if "[" in reference or "]" in reference:
        return False, reference
    if "!" not in reference:
        return current_sheet.casefold() == target_sheet.casefold(), reference
    qualifier, body = reference.rsplit("!", 1)
    qualifier = qualifier.strip()
    if qualifier.startswith("'") and qualifier.endswith("'"):
        qualifier = qualifier[1:-1].replace("''", "'")
    if ":" in qualifier:
        return None
    return qualifier.casefold() == target_sheet.casefold(), body


def _tail_has_formula_blocker(
    context: RuleContext,
    worksheet: Worksheet,
    tail: WhitespaceTail,
) -> bool:
    targets = tail.cell_ranges
    for candidate_sheet in context.workbook.worksheets:
        for cell in candidate_sheet._cells.values():
            if (
                not isinstance(cell, Cell)
                or cell.data_type != "f"
                or not isinstance(cell.value, str)
            ):
                continue
            features = analyze_formula(cell.value)
            if features.unsupported_reason is not None or any(
                name in {"INDIRECT", "OFFSET"} for name in features.volatile_functions
            ):
                return True
            for reference in features.references:
                target = _formula_reference_target(
                    reference,
                    current_sheet=candidate_sheet.title,
                    target_sheet=worksheet.title,
                )
                if target is None:
                    return True
                targets_sheet, body = target
                if not targets_sheet:
                    continue
                try:
                    if any(range_intersects(body, cell_range) for cell_range in targets):
                        return True
                except ValueError:
                    return True
    return False


def _visible_text_cells(worksheet: Worksheet) -> list[Cell]:
    return sorted(
        (
            cell
            for cell in worksheet._cells.values()
            if isinstance(cell, Cell)
            and isinstance(cell.value, str)
            and cell.value.strip()
            and not _is_hidden_cell(worksheet, cell)
        ),
        key=lambda cell: (cell.row, cell.column),
    )


def _merged_range(worksheet: Worksheet, cell: Cell) -> CellRange | None:
    return next(
        (merged for merged in worksheet.merged_cells.ranges if cell.coordinate in merged),
        None,
    )


def _vertical_shared_edge_has_style(
    worksheet: Worksheet,
    row: int,
    left_column: int,
) -> bool:
    """Return whether the boundary between two columns has a visible side."""

    if not 1 <= left_column < 16_384:
        return False
    left = worksheet._cells.get((row, left_column))
    right = worksheet._cells.get((row, left_column + 1))
    left_side = left.border.right if isinstance(left, Cell) else None
    right_side = right.border.left if isinstance(right, Cell) else None
    return bool(
        (left_side is not None and left_side.style is not None)
        or (right_side is not None and right_side.style is not None)
    )


def _overflow_is_visually_blocked(
    worksheet: Worksheet,
    cell: Cell,
    required_width: float,
    available_width: float,
) -> bool:
    """Return whether an unwrapped string cannot use natural blank-cell overflow."""

    merged = _merged_range(worksheet, cell)
    if merged is not None:
        return required_width > available_width
    min_column = merged.min_col if merged is not None else cell.column
    max_column = merged.max_col if merged is not None else cell.column
    alignment = (cell.alignment.horizontal or "general").lower()
    directions: tuple[tuple[int, int], ...]
    if alignment == "right":
        directions = ((-1, min_column - 1),)
    elif alignment in {"center", "centercontinuous", "distributed"}:
        directions = ((-1, min_column - 1), (1, max_column + 1))
    else:
        directions = ((1, max_column + 1),)
    extra_width = max(0.0, required_width - available_width)
    for step, start_column in directions:
        remaining = extra_width
        column = start_column
        while remaining > 0 and 1 <= column <= 16_384:
            left_column = column - 1 if step > 0 else column
            if _vertical_shared_edge_has_style(worksheet, cell.row, left_column):
                return True
            neighbor = worksheet._cells.get((cell.row, column))
            if isinstance(neighbor, Cell) and neighbor.value is not None:
                return True
            if not is_column_hidden(worksheet, column):
                remaining -= column_width(worksheet, column)
            column += step
        if remaining > 0:
            return True
    return False


def _actual_border_source(
    worksheet: Worksheet,
    cell: Cell,
    edge: str,
) -> tuple[str, str] | None:
    side = getattr(cell.border, edge)
    if side is not None and side.style is not None:
        return cell.coordinate, edge
    peer_row = cell.row
    peer_column = cell.column
    opposite = edge
    if edge == "left":
        peer_column -= 1
        opposite = "right"
    elif edge == "right":
        peer_column += 1
        opposite = "left"
    elif edge == "top":
        peer_row -= 1
        opposite = "bottom"
    else:
        peer_row += 1
        opposite = "top"
    peer = worksheet._cells.get((peer_row, peer_column))
    if isinstance(peer, Cell):
        peer_side = getattr(peer.border, opposite)
        if peer_side is not None and peer_side.style is not None:
            return peer.coordinate, opposite
    return None


def _shared_edge_key(cell: Cell, edge: str) -> tuple[str, int, int]:
    if edge == "left":
        return "vertical", cell.row, cell.column - 1
    if edge == "right":
        return "vertical", cell.row, cell.column
    if edge == "top":
        return "horizontal", cell.row - 1, cell.column
    return "horizontal", cell.row, cell.column


def _shared_edge_has_materialized_peer(
    worksheet: Worksheet,
    cell: Cell,
    edge: str,
) -> bool:
    row = cell.row
    column = cell.column
    if edge == "left":
        column -= 1
    elif edge == "right":
        column += 1
    elif edge == "top":
        row -= 1
    else:
        row += 1
    if row < 1 or not 1 <= column <= 16_384:
        return False
    peer = worksheet._cells.get((row, column))
    return bool(
        isinstance(peer, Cell)
        and (
            (
                peer.value is not None
                and not (isinstance(peer.value, str) and not peer.value.strip())
            )
            or peer.has_style
        )
        and not _is_hidden_cell(worksheet, peer)
        and not _is_in_merged_range(worksheet, peer.coordinate)
    )


def _parallel_border_source(
    worksheet: Worksheet,
    cell: Cell,
    edge: str,
    region: Region,
) -> tuple[str, str, tuple[Any, ...], float, int, int] | None:
    """Find a parallel grid-edge consensus, retaining low-confidence evidence."""

    edge_index = ("left", "right", "top", "bottom").index(edge)
    candidates: list[tuple[str, str, tuple[Any, ...]] | None] = []
    positions = (
        ((row, cell.column) for row in range(region.min_row, region.max_row + 1))
        if edge in {"left", "right"}
        else ((cell.row, column) for column in range(region.min_column, region.max_column + 1))
    )
    for row, column in positions:
        if row == cell.row and column == cell.column:
            continue
        peer = worksheet._cells.get((row, column))
        if (
            not isinstance(peer, Cell)
            or not _is_materialized_table_cell(peer)
            or _is_hidden_cell(worksheet, peer)
            or _is_in_merged_range(worksheet, peer.coordinate)
        ):
            continue
        source = _actual_border_source(worksheet, peer, edge)
        signature = visual_border_signature(worksheet, peer)[edge_index]
        if source is not None and signature[0] is not None:
            candidates.append((source[0], source[1], signature))
        else:
            candidates.append(None)
    if len(candidates) < 2:
        return None
    visible_candidates = [candidate for candidate in candidates if candidate is not None]
    if not visible_candidates:
        return None
    signature, count = Counter(item[2] for item in visible_candidates).most_common(1)[0]
    confidence = count / len(candidates)
    if count < 2 or confidence < 0.75:
        return None
    source_cell, source_edge, _ = next(item for item in visible_candidates if item[2] == signature)
    return source_cell, source_edge, signature, confidence, count, len(candidates)


def _edge_is_internal_to_region(cell: Cell, edge: str, region: Region) -> bool:
    if edge == "left":
        return cell.column > region.min_column
    if edge == "right":
        return cell.column < region.max_column
    if edge == "top":
        return cell.row > region.min_row
    return cell.row < region.max_row


def _is_materialized_table_cell(cell: Cell) -> bool:
    return bool(
        (cell.value is not None and not (isinstance(cell.value, str) and not cell.value.strip()))
        or cell.has_style
    )


def _dense_rectangular_border_region(
    worksheet: Worksheet,
    region: Region,
) -> Region | None:
    """Trim ragged margins, then require a sizeable rectangle with a dominant grid."""

    def structured_at(row: int, column: int) -> bool:
        cell = worksheet._cells.get((row, column))
        return bool(
            isinstance(cell, Cell)
            and _is_materialized_table_cell(cell)
            and not _is_hidden_cell(worksheet, cell)
            and not _is_in_merged_range(worksheet, cell.coordinate)
        )

    min_row = region.min_row
    max_row = region.max_row
    min_column = region.min_column
    max_column = region.max_column
    while min_row <= max_row and min_column <= max_column:
        height = max_row - min_row + 1
        width = max_column - min_column + 1
        if height < 3 or width < 3:
            return None
        row_coverages = {
            row: sum(structured_at(row, column) for column in range(min_column, max_column + 1))
            / width
            for row in (min_row, max_row)
        }
        column_coverages = {
            column: sum(structured_at(row, column) for row in range(min_row, max_row + 1)) / height
            for column in (min_column, max_column)
        }
        if row_coverages[min_row] < 0.75:
            min_row += 1
            continue
        if row_coverages[max_row] < 0.75:
            max_row -= 1
            continue
        if column_coverages[min_column] < 0.75:
            min_column += 1
            continue
        if column_coverages[max_column] < 0.75:
            max_column -= 1
            continue
        break
    height = max_row - min_row + 1
    width = max_column - min_column + 1
    area = height * width
    if height < 3 or width < 3 or area < 12:
        return None
    structured = 0
    bordered = 0
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            if not structured_at(row, column):
                continue
            structured += 1
            cell = worksheet._cells.get((row, column))
            bordered += int(isinstance(cell, Cell) and has_visible_border(cell))
    density = structured / area
    if density < 0.75 or bordered / area < 0.60:
        return None
    return Region(
        sheet=region.sheet,
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
        kind=region.kind,
        confidence=Confidence(min(float(region.confidence), density)),
    )


def _visible_layout_bounds(
    worksheet: Worksheet,
    *,
    excluded_coordinates: set[str] | None = None,
) -> tuple[int, int, int, int] | None:
    """Return bounds of visible values, borders, fills, merges, and sheet structures."""

    excluded = excluded_coordinates or set()
    row_bounds: list[int] = []
    column_bounds: list[int] = []

    def extend(cell_range: CellRange) -> None:
        row_bounds.extend((cell_range.min_row, cell_range.max_row))
        column_bounds.extend((cell_range.min_col, cell_range.max_col))

    for cell in worksheet._cells.values():
        if (
            not isinstance(cell, Cell)
            or cell.coordinate in excluded
            or _is_hidden_cell(worksheet, cell)
        ):
            continue
        visible_value = cell.value is not None and not (
            isinstance(cell.value, str) and not cell.value.strip()
        )
        visible_format = has_visible_border(cell) or cell.fill.fill_type not in {None, "none"}
        if (
            visible_value
            or visible_format
            or cell.comment is not None
            or cell.hyperlink is not None
        ):
            row_bounds.append(cell.row)
            column_bounds.append(cell.column)
    for merged in worksheet.merged_cells.ranges:
        extend(CellRange(str(merged)))
    for table in worksheet.tables.values():
        extend(CellRange(table.ref))
    if worksheet.auto_filter.ref:
        extend(CellRange(worksheet.auto_filter.ref))
    if worksheet.data_validations is not None:
        for validation in worksheet.data_validations.dataValidation:
            for target in validation.ranges.ranges:
                extend(CellRange(str(target)))
    for conditional_formatting in worksheet.conditional_formatting:
        for target in conditional_formatting.sqref.ranges:
            extend(CellRange(str(target)))
    if not row_bounds or not column_bounds:
        return None
    return min(row_bounds), min(column_bounds), max(row_bounds), max(column_bounds)


def _predicted_layout_size(
    context: RuleContext,
    worksheet: Worksheet,
    visible_columns: Sequence[int],
    visible_rows: Sequence[int],
) -> tuple[float, float, int, int]:
    """Estimate layout size after earlier reviewed width and height proposals."""

    widths = {column: column_width(worksheet, column) for column in visible_columns}
    heights = {row: row_height(worksheet, row) for row in visible_rows}
    width_updates = 0
    height_updates = 0
    for patch in context.prior_patches:
        if patch.sheet != worksheet.title or not isinstance(patch.after, dict):
            continue
        if patch.kind == PatchKind.SET_COLUMN_WIDTH:
            column_value = patch.after.get("column")
            width_value = patch.after.get("width")
            if (
                isinstance(column_value, str)
                and not isinstance(width_value, bool)
                and isinstance(width_value, (int, float))
            ):
                try:
                    column = column_index_from_string(column_value)
                except ValueError:
                    continue
                if column in widths and float(width_value) > widths[column]:
                    widths[column] = float(width_value)
                    width_updates += 1
        elif patch.kind == PatchKind.SET_ROW_HEIGHT:
            row_value = patch.after.get("row")
            height_value = patch.after.get("height")
            if (
                isinstance(row_value, int)
                and not isinstance(row_value, bool)
                and not isinstance(height_value, bool)
                and isinstance(height_value, (int, float))
                and row_value in heights
                and float(height_value) > heights[row_value]
            ):
                heights[row_value] = float(height_value)
                height_updates += 1
    return sum(widths.values()), sum(heights.values()), width_updates, height_updates


def _nearest_header(worksheet: Worksheet, cell: Cell) -> Cell | None:
    for row in range(cell.row - 1, max(0, cell.row - 9), -1):
        header = worksheet._cells.get((row, cell.column))
        if isinstance(header, Cell) and isinstance(header.value, str) and header.value.strip():
            return header
    return None


def _integer_identifier_text(value: Any) -> str | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value) if value >= 0 else None
    if isinstance(value, float) and math.isfinite(value) and value >= 0 and value.is_integer():
        return format(value, ".0f")
    return None


def _sheet_snapshot(context: RuleContext, worksheet: Worksheet) -> Any | None:
    return next(
        (sheet for sheet in context.snapshot.sheets if sheet.name == worksheet.title),
        None,
    )


def _visual_style_key(worksheet: Worksheet, cell: Cell) -> tuple[Any, ...]:
    """Group visually equivalent cells without treating storage details as decoration."""

    border = cell.border
    non_cardinal_border = (
        copy(border.start),
        copy(border.end),
        copy(border.diagonal),
        copy(border.vertical),
        copy(border.horizontal),
        border.outline,
        border.diagonalUp,
        border.diagonalDown,
    )

    return (
        copy(cell.font),
        copy(cell.fill),
        visual_border_signature(worksheet, cell),
        non_cardinal_border,
        copy(cell.alignment),
        cell.number_format,
    )


def _row_has_summary_semantics(
    worksheet: Worksheet,
    row: int,
    region: Region | None = None,
    *,
    include_overrides: bool = True,
) -> bool:
    cells = (
        (
            _cell(worksheet, row, column)
            for column in range(region.min_column, region.max_column + 1)
        )
        if region is not None
        else (
            cell for cell in worksheet._cells.values() if isinstance(cell, Cell) and cell.row == row
        )
    )
    for cell in cells:
        value = cell.value
        if _aggregate_formula_spans_other_rows(value, row):
            return True
        if not isinstance(value, str) or not SUMMARY_LABEL_RE.search(value.strip()):
            continue
        normalized = unicodedata.normalize("NFKC", value).casefold()
        is_override = any(
            marker in normalized for marker in ("adjustment", "override", "调整", "覆盖")
        )
        if include_overrides or not is_override:
            return True
    return False


def _label_template(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).strip().casefold()
    normalized = re.sub(r"\d+", "<n>", normalized)
    return re.sub(r"\s+", " ", normalized)


def _containing_data_region(
    context: RuleContext, worksheet: Worksheet, cell: Cell | MergedCell
) -> Region | None:
    if cell.row is None or cell.column is None:
        return None
    candidates = [
        region
        for region in context.data_regions.get(worksheet.title, ())
        if region.min_row <= cell.row <= region.max_row
        and region.min_column <= cell.column <= region.max_column
    ]
    return min(
        candidates,
        key=lambda region: (
            (region.max_row - region.min_row + 1) * (region.max_column - region.min_column + 1),
            region.min_row,
            region.min_column,
        ),
        default=None,
    )


def _row_label_matches_peer_template(
    context: RuleContext,
    worksheet: Worksheet,
    cell: Cell | MergedCell,
    region: Region | None = None,
    *,
    include_target: bool = False,
    ignore_numeric_text: bool = False,
    require_stable: bool = True,
) -> bool:
    """Require a stable record-label template when a text row label exists."""

    if cell.row is None or cell.column is None:
        return False
    cell_row = cell.row
    cell_column = cell.column
    active_region = region or _containing_data_region(context, worksheet, cell)
    label_cells = [
        peer
        for peer in worksheet._cells.values()
        if isinstance(peer, Cell)
        and peer.row == cell_row
        and peer.column != cell_column
        and peer.data_type != "f"
        and isinstance(peer.value, str)
        and peer.value.strip()
        and not (ignore_numeric_text and _numeric_text(peer.value) is not None)
    ]
    if (
        include_target
        and isinstance(cell, Cell)
        and isinstance(cell.value, str)
        and cell.value.strip()
    ):
        label_cells.append(cell)
    if not label_cells:
        return True
    if active_region is not None:
        min_row = active_region.min_row + 1
        max_row = active_region.max_row
    else:
        min_row = max(1, cell_row - 10)
        max_row = cell_row + 10
    stable_evidence = False
    for label_cell in sorted(label_cells, key=lambda peer: peer.column):
        label = cast(str, label_cell.value).strip()
        if SUMMARY_LABEL_RE.search(label):
            return False
        template = _label_template(label)
        peers = sorted(
            (
                peer
                for peer in worksheet._cells.values()
                if isinstance(peer, Cell)
                and peer.column == label_cell.column
                and min_row <= peer.row <= max_row
                and peer.row != cell_row
                and peer.data_type != "f"
                and isinstance(peer.value, str)
                and peer.value.strip()
            ),
            key=lambda peer: (abs(peer.row - cell_row), peer.row),
        )[:8]
        if not peers:
            return False
        counts = Counter(_label_template(cast(str, peer.value)) for peer in peers)
        consensus, count = counts.most_common(1)[0]
        if count >= 3 and count / len(peers) >= 0.75:
            stable_evidence = True
            if template != consensus:
                return False
    return stable_evidence or not require_stable


def _row_visual_style_matches_peer_consensus(
    context: RuleContext,
    worksheet: Worksheet,
    cell: Cell | MergedCell,
    region: Region | None = None,
    *,
    include_target: bool = False,
) -> bool:
    """Reject auto-repair when another cell proves the row is intentionally highlighted."""

    if cell.row is None or cell.column is None:
        return False
    cell_row = cell.row
    cell_column = cell.column
    active_region = region or _containing_data_region(context, worksheet, cell)
    if active_region is not None:
        min_row = active_region.min_row + 1
        max_row = active_region.max_row
        min_column = active_region.min_column
        max_column = active_region.max_column
    else:
        min_row = max(1, cell_row - 10)
        max_row = cell_row + 10
        min_column = 1
        max_column = worksheet.max_column
    row_cells = [
        peer
        for peer in worksheet._cells.values()
        if isinstance(peer, Cell)
        and peer.row == cell_row
        and min_column <= peer.column <= max_column
        and peer.column != cell_column
        and (peer.value is not None or peer.has_style)
    ]
    if include_target and isinstance(cell, Cell) and (cell.value is not None or cell.has_style):
        row_cells.append(cell)
    for row_cell in row_cells:
        peers = sorted(
            (
                peer
                for peer in worksheet._cells.values()
                if isinstance(peer, Cell)
                and peer.column == row_cell.column
                and min_row <= peer.row <= max_row
                and peer.row != cell_row
                and (peer.value is not None or peer.has_style)
            ),
            key=lambda peer: (abs(peer.row - cell_row), peer.row),
        )[:8]
        if len(peers) < 3:
            continue
        counts = Counter(_visual_style_key(worksheet, peer) for peer in peers)
        consensus, count = counts.most_common(1)[0]
        if (
            count >= 3
            and count / len(peers) >= 0.75
            and _visual_style_key(worksheet, row_cell) != consensus
        ):
            return False
    return True


def _translated_consensus(
    target: Cell, peers: Iterable[Cell], required_signature: str | None = None
) -> tuple[str, str] | None:
    translated: list[tuple[str, str]] = []
    for peer in peers:
        if peer.data_type != "f" or not isinstance(peer.value, str):
            continue
        signature = _formula_signature(peer)
        if signature is None or (
            required_signature is not None and signature != required_signature
        ):
            continue
        try:
            value = translate_formula(peer.value, peer.coordinate, target.coordinate)
        except UnsupportedFormulaError:
            continue
        translated.append((value, peer.coordinate))
    if len(translated) < 2:
        return None
    counts = Counter(value for value, _ in translated)
    formula, count = counts.most_common(1)[0]
    if count < 2 or len(counts) != 1:
        return None
    source = min(source for value, source in translated if value == formula)
    return formula, source


class BrokenReferenceRule(WorkbookRule):
    rule_id = "WL001_BROKEN_REFERENCE"
    title = "Broken formula reference"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for cell in worksheet._cells.values():
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                features = analyze_formula(cell.value)
                if not features.broken_references:
                    continue
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="The formula contains an explicit #REF! token and cannot resolve as written.",
                        severity=Severity.ERROR,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(
                            summary="Formula contains #REF!",
                            observed=cell.value,
                            details={"references": list(features.broken_references)},
                        ),
                        expected="Every formula reference resolves to an existing cell or range.",
                        suggested_action="Review the deleted or moved source range; no automatic guess was made.",
                    )
                )
        return result


class FormulaPatternOutlierRule(WorkbookRule):
    rule_id = "WL002_FORMULA_PATTERN_OUTLIER"
    title = "Formula pattern outlier"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for band in context.formula_bands[worksheet.title]:
                cells = _band_cells(worksheet, band)
                if _band_has_unsupported_formula(context, worksheet.title, cells):
                    continue
                formula_cells = [
                    cell
                    for cell in cells
                    if cell.data_type == "f" and not _is_boundary_aggregate(cell, band)
                ]
                signature_by_cell = {
                    cell.coordinate: _formula_signature(cell) for cell in formula_cells
                }
                if any(signature is None for signature in signature_by_cell.values()):
                    continue
                signatures = [
                    cast(str, signature_by_cell[cell.coordinate]) for cell in formula_cells
                ]
                if len(signatures) < 4:
                    continue
                counts = Counter(signatures)
                consensus, consensus_count = counts.most_common(1)[0]
                ratio = consensus_count / len(signatures)
                outliers = [
                    cell
                    for cell in formula_cells
                    if signature_by_cell[cell.coordinate] != consensus
                ]
                if ratio < 0.8 or not outliers or not _isolated_outliers(outliers, band):
                    continue
                peers = [
                    item.coordinate
                    for item in formula_cells
                    if signature_by_cell[item.coordinate] == consensus
                ]
                for cell in outliers:
                    if (worksheet.title, cell.coordinate) in seen:
                        continue
                    seen.add((worksheet.title, cell.coordinate))
                    proposal = _translated_consensus(cell, formula_cells, consensus)
                    patches: list[PatchOperation] = []
                    aggregate_formula = _is_aggregate_formula(cell.value)
                    row_semantics_safe = (
                        not _row_has_summary_semantics(worksheet, cell.row)
                        and _row_label_matches_peer_template(context, worksheet, cell)
                        and _row_visual_style_matches_peer_consensus(
                            context, worksheet, cell, include_target=True
                        )
                    )
                    if (
                        proposal is not None
                        and ratio >= 0.95
                        and len(outliers) == 1
                        and not aggregate_formula
                        and not _is_band_boundary(cell, band)
                        and not _is_in_merged_range(worksheet, cell.coordinate)
                        and not _is_hidden_cell(worksheet, cell)
                        and not _is_protected_target(worksheet)
                        and row_semantics_safe
                    ):
                        formula, source = proposal
                        patch = _make_patch(
                            kind=PatchKind.SET_FORMULA,
                            worksheet=worksheet,
                            cell=cell,
                            before=cell.value,
                            after=formula,
                            confidence=ratio,
                            source_cell=source,
                            description=(
                                "Replace the one-off formula with the exact translated peer consensus."
                            ),
                        )
                        patches.append(patch)
                        result.patches.append(patch)
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation=(
                                "A formula has a different relative-reference signature from a "
                                "strong band consensus."
                            ),
                            severity=Severity.WARNING,
                            confidence=ratio,
                            sheet=worksheet.title,
                            location=cell.coordinate,
                            evidence=Evidence(
                                summary=(
                                    f"{consensus_count} of {len(signatures)} formulas share one signature"
                                ),
                                observed=cell.value,
                                expected=consensus,
                                peers=peers[:12],
                            ),
                            expected=(
                                "Copied formulas in this band have the same structural signature."
                            ),
                            suggested_action=(
                                "Multiple isolated anomalies were found, so no automatic patch is "
                                "offered; compare each cell with the listed peers."
                                if len(outliers) > 1
                                else "Aggregate formulas are never replaced automatically; review the "
                                "subtotal or total manually."
                                if aggregate_formula
                                else "The row context does not match stable detail-row semantics, so "
                                "automatic replacement is withheld."
                                if not row_semantics_safe
                                else "Compare the cell with the listed peers and review any proposed formula."
                            ),
                            patches=patches,
                        )
                    )
        return result


class BlankInFormulaBandRule(WorkbookRule):
    rule_id = "WL003_BLANK_IN_FORMULA_BAND"
    title = "Blank interrupts formula band"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for band in context.formula_bands[worksheet.title]:
                cells = _band_cells(worksheet, band)
                if _band_has_unsupported_formula(context, worksheet.title, cells):
                    continue
                blanks = [cell for cell in cells[1:-1] if cell.value is None]
                formula_cells = [cell for cell in cells if cell.data_type == "f"]
                raw_signatures = [_formula_signature(cell) for cell in formula_cells]
                if any(signature is None for signature in raw_signatures):
                    continue
                signatures = [cast(str, signature) for signature in raw_signatures]
                if len(blanks) != 1 or len(signatures) < 3:
                    continue
                consensus, count = Counter(signatures).most_common(1)[0]
                if count / len(signatures) < 0.9:
                    continue
                cell = blanks[0]
                if (worksheet.title, cell.coordinate) in seen:
                    continue
                seen.add((worksheet.title, cell.coordinate))
                proposal = _translated_consensus(cell, formula_cells, consensus)
                if proposal is None:
                    continue
                formula, source = proposal
                merged_cell = _is_in_merged_range(worksheet, cell.coordinate)
                patches: list[PatchOperation] = []
                hidden_cell = _is_hidden_cell(worksheet, cell)
                protected_target = _is_protected_target(worksheet)
                row_semantics_safe = (
                    not _row_has_summary_semantics(worksheet, cell.row)
                    and _row_label_matches_peer_template(context, worksheet, cell)
                    and _row_visual_style_matches_peer_consensus(
                        context, worksheet, cell, include_target=True
                    )
                )
                if (
                    not merged_cell
                    and not hidden_cell
                    and not protected_target
                    and row_semantics_safe
                ):
                    patch = _make_patch(
                        kind=PatchKind.CREATE_FORMULA,
                        worksheet=worksheet,
                        cell=cell,
                        before=None,
                        after=formula,
                        confidence=0.99,
                        source_cell=source,
                        description=(
                            "Create the missing cell with the exact translated formula agreed by peers."
                        ),
                    )
                    patches.append(patch)
                    result.patches.append(patch)
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A single blank lies between formulas whose translations agree exactly at this cell.",
                        severity=Severity.ERROR,
                        confidence=0.99,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(
                            summary="Independent neighboring formulas translate to the same expression",
                            observed=None,
                            expected=formula,
                            peers=[peer.coordinate for peer in formula_cells[:12]],
                        ),
                        expected="The contiguous formula band has no unexplained blank.",
                        suggested_action=(
                            "The blank is inside a merged range and cannot safely receive a formula; "
                            "review the merge manually."
                            if merged_cell
                            else "The blank is in a hidden sheet, row, or column, so automatic formula creation "
                            "is withheld."
                            if hidden_cell
                            else "The blank is locked on a protected sheet, so automatic formula creation "
                            "is withheld."
                            if protected_target
                            else "The row context does not match stable detail-row semantics, so automatic "
                            "formula creation is withheld."
                            if not row_semantics_safe
                            else "Review and select the proposed translated formula."
                        ),
                        patches=patches,
                    )
                )
        return result


class HardcodedValueInFormulaBandRule(WorkbookRule):
    rule_id = "WL004_HARDCODED_VALUE_IN_FORMULA_BAND"
    title = "Hardcoded value interrupts formula band"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for band in context.formula_bands[worksheet.title]:
                cells = _band_cells(worksheet, band)
                if _band_has_unsupported_formula(context, worksheet.title, cells):
                    continue
                literals = [
                    cell for cell in cells[1:-1] if cell.value is not None and cell.data_type != "f"
                ]
                formula_cells = [cell for cell in cells if cell.data_type == "f"]
                raw_signatures = [_formula_signature(cell) for cell in formula_cells]
                if any(signature is None for signature in raw_signatures):
                    continue
                signatures = [cast(str, signature) for signature in raw_signatures]
                if len(literals) != 1 or len(signatures) < 3:
                    continue
                consensus, count = Counter(signatures).most_common(1)[0]
                ratio = count / len(signatures)
                if ratio < 0.9:
                    continue
                cell = literals[0]
                if (worksheet.title, cell.coordinate) in seen:
                    continue
                seen.add((worksheet.title, cell.coordinate))
                proposal = _translated_consensus(cell, formula_cells, consensus)
                if proposal is None:
                    continue
                formula, source = proposal
                confidence = 0.99 if ratio >= 0.95 else 0.94
                patches: list[PatchOperation] = []
                merged_cell = _is_in_merged_range(worksheet, cell.coordinate)
                hidden_cell = _is_hidden_cell(worksheet, cell)
                protected_target = _is_protected_target(worksheet)
                row_semantics_safe = (
                    not _row_has_summary_semantics(worksheet, cell.row)
                    and _row_label_matches_peer_template(
                        context, worksheet, cell, include_target=True
                    )
                    and _row_visual_style_matches_peer_consensus(
                        context, worksheet, cell, include_target=True
                    )
                )
                if (
                    confidence >= 0.95
                    and not merged_cell
                    and not hidden_cell
                    and not protected_target
                    and row_semantics_safe
                ):
                    patch = _make_patch(
                        kind=PatchKind.SET_FORMULA,
                        worksheet=worksheet,
                        cell=cell,
                        before=cell.value,
                        after=formula,
                        confidence=confidence,
                        source_cell=source,
                        description="Replace the isolated literal with the exact translated peer formula.",
                    )
                    patches.append(patch)
                    result.patches.append(patch)
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A literal value replaces one cell in an otherwise consistent copied-formula band.",
                        severity=Severity.ERROR,
                        confidence=confidence,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(
                            summary="Peer formulas translate to one exact replacement",
                            observed=cell.value,
                            expected=formula,
                            peers=[peer.coordinate for peer in formula_cells[:12]],
                        ),
                        expected="The formula band follows its consensus structure.",
                        suggested_action=(
                            "The cell is inside a merged range, so automatic replacement is withheld."
                            if merged_cell
                            else "The cell is in a hidden sheet, row, or column, so automatic replacement is withheld."
                            if hidden_cell
                            else "The cell is locked on a protected sheet, so automatic replacement is withheld."
                            if protected_target
                            else "The row context does not match stable detail-row semantics, so automatic "
                            "replacement is withheld."
                            if not row_semantics_safe
                            else "Confirm the literal is not an intentional override before selecting the patch."
                        ),
                        patches=patches,
                    )
                )
        return result


class SuspiciousSumBoundaryRule(WorkbookRule):
    rule_id = "WL005_SUSPICIOUS_SUM_BOUNDARY"
    title = "Suspicious SUM boundary"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for cell in worksheet._cells.values():
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                match = SIMPLE_SUM_RE.fullmatch(cell.value)
                if match is None:
                    continue
                if match.group("sheet") or _in_unsupported_formula_range(
                    context, worksheet.title, cell.coordinate
                ):
                    continue
                first_column = match.group("col1").replace("$", "").upper()
                second_column = match.group("col2").replace("$", "").upper()
                if first_column != second_column or first_column != cell.column_letter:
                    continue
                start_row = int(match.group("start").replace("$", ""))
                end_row = int(match.group("end").replace("$", ""))
                candidate_row = end_row + 1
                if candidate_row != cell.row - 1 or start_row >= end_row:
                    continue
                candidate = _cell(worksheet, candidate_row, cell.column)
                if isinstance(candidate.value, bool):
                    continue
                candidate_dimension = worksheet.row_dimensions.get(candidate_row)
                if candidate_dimension is not None and candidate_dimension.hidden:
                    continue
                if any(candidate.coordinate in merged for merged in worksheet.merged_cells.ranges):
                    continue
                literal_candidate = isinstance(candidate.value, (int, float))
                candidate_formula = (
                    candidate.value
                    if candidate.data_type == "f" and isinstance(candidate.value, str)
                    else None
                )
                if not literal_candidate and candidate_formula is None:
                    continue
                candidate_summary = _row_has_summary_semantics(worksheet, candidate_row)
                if candidate_formula is not None:
                    if (
                        _in_unsupported_formula_range(
                            context, worksheet.title, candidate.coordinate
                        )
                        or _is_aggregate_formula(candidate_formula)
                        or not _same_data_region(context, worksheet, candidate, cell)
                    ):
                        continue
                    try:
                        features = analyze_formula(candidate_formula)
                    except (ValueError, UnsupportedFormulaError):
                        continue
                    if (
                        features.broken_references
                        or features.external_references
                        or features.unsupported_reason
                    ):
                        continue
                raw_end = match.group("end")
                replacement_end = (
                    "$" + str(candidate_row) if raw_end.startswith("$") else str(candidate_row)
                )
                formula = (
                    cell.value[: match.start("end")]
                    + replacement_end
                    + cell.value[match.end("end") :]
                )
                confidence = 0.96 if literal_candidate else 0.9
                patches: list[PatchOperation] = []
                target_blocked = bool(
                    _is_in_merged_range(worksheet, cell.coordinate)
                    or _is_hidden_cell(worksheet, cell)
                    or _is_protected_target(worksheet)
                )
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation=(
                            "A simple total stops one row before a directly adjacent non-hidden "
                            f"{'numeric' if literal_candidate else 'formula'} peer."
                        ),
                        severity=Severity.WARNING,
                        confidence=confidence,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(
                            summary=f"SUM ends at row {end_row}, while {candidate.coordinate} is adjacent",
                            observed=cell.value,
                            expected=formula,
                            peers=[candidate.coordinate],
                        ),
                        expected="A simple contiguous total includes its directly adjacent peer row.",
                        suggested_action=(
                            "The adjacent row has subtotal or total semantics, so automatic SUM "
                            "extension is withheld."
                            if candidate_summary
                            else "The SUM target is merged, hidden, or locked on a protected sheet, "
                            "so automatic extension is withheld."
                            if target_blocked
                            else "Review the adjacent peer manually. WorkbookLens reports the candidate "
                            "formula but does not automatically extend SUM boundaries because inclusion "
                            "semantics cannot be proven from adjacency alone."
                        ),
                        patches=patches,
                    )
                )
        return result


def _numeric_text(value: Any) -> int | float | None:
    if not isinstance(value, str):
        return None
    stripped = value.strip()
    if not stripped:
        return None
    digits = re.sub(r"[^0-9]", "", stripped)
    if len(digits) > 15 or re.search(r"[()\-/]", stripped):
        return None
    grouped = re.fullmatch(r"[+-]?\d{1,3}(?:,\d{3})+(?:\.\d+)?", stripped)
    plain = re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)", stripped)
    if grouped is None and plain is None:
        return None
    normalized = stripped.replace(",", "")
    if re.match(r"^[+-]?0\d+", normalized):
        return None
    try:
        number = Decimal(normalized)
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    if number == number.to_integral_value():
        return int(number)
    result = float(number)
    return result if math.isfinite(result) else None


class NumericTextRule(WorkbookRule):
    rule_id = "WL006_NUMERIC_TEXT"
    title = "Numeric text in numeric region"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for region in context.data_regions[worksheet.title]:
                for column in range(region.min_column, region.max_column + 1):
                    cells = [
                        _cell(worksheet, row, column)
                        for row in range(region.min_row + 1, region.max_row + 1)
                        if _cell(worksheet, row, column).value is not None
                    ]
                    numeric_count = sum(
                        isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
                        for cell in cells
                    )
                    if numeric_count < 3:
                        continue
                    header_value = _cell(worksheet, region.min_row, column).value
                    identifier_header = _looks_like_identifier_header(header_value)
                    measure_header = _looks_like_measure_header(header_value)
                    for cell in cells:
                        converted = _numeric_text(cell.value)
                        if converted is None or (worksheet.title, cell.coordinate) in seen:
                            continue
                        ratio = numeric_count / len(cells)
                        if ratio < 0.75:
                            continue
                        seen.add((worksheet.title, cell.coordinate))
                        confidence = 0.97 if ratio >= 0.85 else 0.9
                        text_formatted = cell.number_format == "@" or cell.quotePrefix
                        grouped_numeric = isinstance(cell.value, str) and "," in cell.value
                        row_semantics_safe = (
                            not _row_has_summary_semantics(worksheet, cell.row)
                            and _row_label_matches_peer_template(
                                context,
                                worksheet,
                                cell,
                                region,
                                ignore_numeric_text=True,
                                require_stable=False,
                            )
                            and _row_visual_style_matches_peer_consensus(
                                context, worksheet, cell, region, include_target=True
                            )
                        )
                        patches: list[PatchOperation] = []
                        if (
                            confidence >= 0.95
                            and not identifier_header
                            and measure_header
                            and not text_formatted
                            and not grouped_numeric
                            and not _is_in_merged_range(worksheet, cell.coordinate)
                            and not _is_hidden_cell(worksheet, cell)
                            and not _is_protected_target(worksheet)
                            and row_semantics_safe
                        ):
                            patch = _make_patch(
                                kind=PatchKind.SET_NUMERIC,
                                worksheet=worksheet,
                                cell=cell,
                                before=cell.value,
                                after=converted,
                                confidence=confidence,
                                description="Convert an unambiguous numeric string to an OOXML numeric value.",
                            )
                            patches.append(patch)
                            result.patches.append(patch)
                        result.findings.append(
                            _make_finding(
                                context=context,
                                rule_id=self.rule_id,
                                title=self.title,
                                explanation="A plain numeric string appears in a column dominated by numeric values.",
                                severity=Severity.WARNING,
                                confidence=confidence,
                                sheet=worksheet.title,
                                location=cell.coordinate,
                                evidence=Evidence(
                                    summary=f"{numeric_count} peer cells are stored as numbers",
                                    observed=cell.value,
                                    expected=converted,
                                    peers=[
                                        item.coordinate
                                        for item in cells
                                        if isinstance(item.value, (int, float))
                                    ][:12],
                                ),
                                expected="Numeric measures use numeric cell storage, while identifiers remain text.",
                                suggested_action=(
                                    "Identifier semantics or explicit text formatting prevent automatic "
                                    "numeric conversion; confirm storage intentionally."
                                    if identifier_header or text_formatted
                                    else "The header does not explicitly identify a numeric measure, so "
                                    "automatic conversion is withheld; confirm the column semantics."
                                    if not measure_header
                                    else "Grouped numeric text is reported for review but is not converted "
                                    "automatically in this release."
                                    if grouped_numeric
                                    else "The row context does not match stable detail-row semantics, so "
                                    "automatic numeric conversion is withheld."
                                    if not row_semantics_safe
                                    else "Confirm the value is a measure rather than an identifier."
                                ),
                                patches=patches,
                            )
                        )
        return result


class StyleOutlierRule(WorkbookRule):
    rule_id = "WL007_STYLE_OUTLIER"
    title = "Style outlier in homogeneous region"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for region in context.data_regions[worksheet.title]:
                detail_rows = [
                    row
                    for row in range(region.min_row + 1, region.max_row + 1)
                    if not _row_has_summary_semantics(worksheet, row, include_overrides=False)
                ]
                for column in range(region.min_column, region.max_column + 1):
                    cells = [
                        _cell(worksheet, row, column)
                        for row in detail_rows
                        if _cell(worksheet, row, column).value is not None
                    ]
                    if len(cells) < 5:
                        continue
                    style_by_cell = {
                        cell.coordinate: _visual_style_key(worksheet, cell) for cell in cells
                    }
                    counts = Counter(style_by_cell.values())
                    consensus_style, count = counts.most_common(1)[0]
                    ratio = count / len(cells)
                    outliers = [
                        cell for cell in cells if style_by_cell[cell.coordinate] != consensus_style
                    ]
                    outlier_rows = sorted(cell.row for cell in outliers)
                    if (
                        ratio < 0.8
                        or not outliers
                        or any(right - left <= 1 for left, right in pairwise(outlier_rows))
                    ):
                        continue
                    sources = [
                        peer
                        for peer in cells
                        if style_by_cell[peer.coordinate] == consensus_style
                        and not _in_unsupported_formula_range(
                            context, worksheet.title, peer.coordinate
                        )
                    ]
                    for cell in outliers:
                        if (worksheet.title, cell.coordinate) in seen:
                            continue
                        seen.add((worksheet.title, cell.coordinate))
                        source = (
                            min(
                                sources,
                                key=lambda peer: (
                                    abs(peer.row - cell.row),
                                    peer.coordinate,
                                ),
                            )
                            if sources
                            else None
                        )
                        confidence = ratio
                        patches: list[PatchOperation] = []
                        row_semantics_safe = (
                            not _row_has_summary_semantics(worksheet, cell.row)
                            and _row_label_matches_peer_template(
                                context, worksheet, cell, region, include_target=True
                            )
                            and _row_visual_style_matches_peer_consensus(
                                context, worksheet, cell, region
                            )
                        )
                        if (
                            confidence >= 0.95
                            and len(outliers) == 1
                            and source is not None
                            and _style_copy_preserves_semantics(cell, source)
                            and not worksheet.protection.sheet
                            and not _is_in_merged_range(worksheet, cell.coordinate)
                            and not _is_hidden_cell(worksheet, cell)
                            and row_semantics_safe
                            and not _in_unsupported_formula_range(
                                context, worksheet.title, cell.coordinate
                            )
                        ):
                            patch = _make_patch(
                                kind=PatchKind.COPY_STYLE,
                                worksheet=worksheet,
                                cell=cell,
                                before=cell.style_id,
                                after=source.style_id,
                                confidence=confidence,
                                source_cell=source.coordinate,
                                description="Copy the existing consensus style ID from the nearest peer.",
                            )
                            patches.append(patch)
                            result.patches.append(patch)
                        result.findings.append(
                            _make_finding(
                                context=context,
                                rule_id=self.rule_id,
                                title=self.title,
                                explanation="A populated cell has a different visual style from its column peers.",
                                severity=Severity.INFO,
                                confidence=confidence,
                                sheet=worksheet.title,
                                location=cell.coordinate,
                                evidence=Evidence(
                                    summary=(
                                        f"One visual style appears in {count} of {len(cells)} peer cells"
                                    ),
                                    observed=cell.style_id,
                                    expected=source.style_id if source is not None else None,
                                    peers=[
                                        peer.coordinate
                                        for peer in cells
                                        if style_by_cell[peer.coordinate] == consensus_style
                                    ][:12],
                                    details={"number_format": cell.number_format},
                                ),
                                expected="A homogeneous measure column uses its consensus style.",
                                suggested_action=(
                                    "Multiple isolated style anomalies were found, so no automatic patch "
                                    "is offered."
                                    if len(outliers) > 1
                                    else "The row context does not match stable detail-row semantics, so no "
                                    "automatic style patch is offered."
                                    if not row_semantics_safe
                                    else "Check whether the visual distinction is intentional."
                                ),
                                patches=patches,
                            )
                        )
        return result


class HiddenNonemptyDataRule(WorkbookRule):
    rule_id = "WL008_HIDDEN_NONEMPTY_DATA"
    title = "Hidden nonempty data"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            nonempty = [cell for cell in worksheet._cells.values() if cell.value is not None]
            if worksheet.sheet_state != "visible" and nonempty:
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A hidden worksheet contains data or formulas that may affect interpretation.",
                        severity=Severity.WARNING,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=None,
                        evidence=Evidence(
                            summary=f"{worksheet.sheet_state} sheet contains {len(nonempty)} nonempty cells"
                        ),
                        expected="Hidden content is reviewed and documented.",
                        suggested_action="Inspect the hidden sheet manually; WorkbookLens never unhides it automatically.",
                    )
                )
            for row, row_dimension in sorted(worksheet.row_dimensions.items()):
                cells = [cell for cell in nonempty if cell.row == row]
                if not row_dimension.hidden or not cells:
                    continue
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A hidden row contains values or formulas.",
                        severity=Severity.WARNING
                        if any(cell.data_type == "f" for cell in cells)
                        else Severity.INFO,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=f"{row}:{row}",
                        evidence=Evidence(
                            summary=f"Hidden row {row} contains {len(cells)} nonempty cells",
                            peers=[cell.coordinate for cell in cells[:12]],
                        ),
                        expected="Hidden rows with consequential content are intentionally documented.",
                        suggested_action="Review the row manually; no automatic unhide is offered.",
                    )
                )
            for min_column, max_column in hidden_column_spans(worksheet):
                cells = [cell for cell in nonempty if min_column <= cell.column <= max_column]
                if not cells:
                    continue
                location = f"{get_column_letter(min_column)}:{get_column_letter(max_column)}"
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A hidden column contains values or formulas.",
                        severity=Severity.WARNING
                        if any(cell.data_type == "f" for cell in cells)
                        else Severity.INFO,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=location,
                        evidence=Evidence(
                            summary=f"Hidden column range {location} contains {len(cells)} nonempty cells",
                            peers=[cell.coordinate for cell in cells[:12]],
                        ),
                        expected="Hidden columns with consequential content are intentionally documented.",
                        suggested_action="Review the column manually; no automatic unhide is offered.",
                    )
                )
        return result


class ExternalLinkRule(WorkbookRule):
    rule_id = "WL009_EXTERNAL_LINK"
    title = "External workbook link"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for cell in worksheet._cells.values():
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                features = analyze_formula(cell.value)
                if not features.external_references:
                    continue
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="The formula depends on another workbook; WorkbookLens does not fetch it.",
                        severity=Severity.WARNING,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(
                            summary="Formula contains external workbook reference",
                            observed=cell.value,
                            details={"references": list(features.external_references)},
                        ),
                        expected="External dependencies are explicit, available, and reviewed.",
                        suggested_action="Verify the linked workbook and consider replacing fragile dependencies.",
                    )
                )
        for name, target in context.snapshot.defined_names.items():
            if re.search(r"\[[^\]]+\][^!]*!", target):
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A defined name refers to another workbook.",
                        severity=Severity.WARNING,
                        confidence=1.0,
                        sheet=None,
                        location=name,
                        evidence=Evidence(
                            summary="Defined name contains external reference", observed=target
                        ),
                        expected="Defined-name dependencies remain local or are explicitly reviewed.",
                        suggested_action="Review the external target; WorkbookLens never opens it.",
                    )
                )
        return result


class VolatileOrFragileFunctionRule(WorkbookRule):
    rule_id = "WL010_VOLATILE_OR_FRAGILE_FUNCTION"
    title = "Volatile or fragile formula construct"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for cell in worksheet._cells.values():
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                features = analyze_formula(cell.value)
                if not features.volatile_functions and not features.has_whole_column_reference:
                    continue
                constructs = list(features.volatile_functions)
                if features.has_whole_column_reference:
                    constructs.append("whole-column reference")
                severity = (
                    Severity.WARNING
                    if any(
                        item in {"OFFSET", "INDIRECT", "whole-column reference"}
                        for item in constructs
                    )
                    else Severity.INFO
                )
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="The formula uses constructs that can recalculate frequently or resist static tracing.",
                        severity=severity,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(summary=", ".join(constructs), observed=cell.value),
                        expected="Performance-sensitive and auditable models avoid unnecessary fragile constructs.",
                        suggested_action="Review whether a bounded direct reference can express the same intent.",
                    )
                )
        return result


class ErrorCellRule(WorkbookRule):
    rule_id = "WL011_ERROR_CELL"
    title = "Stored Excel error value"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for cell in worksheet._cells.values():
                value = cell.value
                if cell.data_type != "e" and not (
                    isinstance(value, str) and value.upper() in EXCEL_ERRORS
                ):
                    continue
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="The cell stores a recognized Excel error value.",
                        severity=Severity.ERROR,
                        confidence=1.0,
                        sheet=worksheet.title,
                        location=cell.coordinate,
                        evidence=Evidence(summary="Stored error cell", observed=value),
                        expected="Calculated or imported values do not contain Excel error tokens.",
                        suggested_action="Trace the producing formula or upstream data; no value is fabricated.",
                    )
                )
        return result


def _configured_key_identity(value: Any) -> tuple[str, Any]:
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, (int, float)):
        return ("number", value)
    try:
        hash(value)
        normalized = value
    except TypeError:
        normalized = repr(value)
    return (type(value).__qualname__, normalized)


class DuplicateConfiguredKeyRule(WorkbookRule):
    rule_id = "WL012_DUPLICATE_CONFIGURED_KEY"
    title = "Duplicate configured key"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        keys = context.config.get("keys", [])
        if not isinstance(keys, list):
            return result
        for specification in keys:
            if not isinstance(specification, dict):
                continue
            sheet_name = specification.get("sheet")
            range_text = specification.get("range")
            if not isinstance(sheet_name, str) or not isinstance(range_text, str):
                continue
            if sheet_name not in context.workbook.sheetnames:
                continue
            worksheet = context.workbook[sheet_name]
            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_text)
            except ValueError:
                continue
            if None in {min_col, min_row, max_col, max_row}:
                continue
            min_col = cast(int, min_col)
            min_row = cast(int, min_row)
            max_col = cast(int, max_col)
            max_row = cast(int, max_row)
            if (
                min_col != max_col
                or min_col < 1
                or min_row < 1
                or max_col > 16_384
                or max_row > 1_048_576
                or max_row - min_row + 1 > 100_000
            ):
                continue
            by_value: dict[tuple[str, Any], list[str]] = defaultdict(list)
            observed_values: dict[tuple[str, Any], Any] = {}
            for row in range(min_row, max_row + 1):
                cell = _cell(worksheet, row, min_col)
                if cell.value is None and specification.get("ignore_blank", True):
                    continue
                key = _configured_key_identity(cell.value)
                observed_values.setdefault(key, cell.value)
                by_value[key].append(cell.coordinate)
            for key, locations in by_value.items():
                if len(locations) < 2:
                    continue
                result.findings.append(
                    _make_finding(
                        context=context,
                        rule_id=self.rule_id,
                        title=self.title,
                        explanation="A value repeats in a column explicitly configured as a unique key.",
                        severity=Severity.ERROR,
                        confidence=1.0,
                        sheet=sheet_name,
                        location=",".join(locations),
                        evidence=Evidence(
                            summary=f"Configured key value appears {len(locations)} times",
                            observed=observed_values[key],
                            peers=locations,
                        ),
                        expected=f"Values in {sheet_name}!{range_text} are unique.",
                        suggested_action="Resolve the duplicate records or revise the explicit key configuration.",
                    )
                )
        return result


class BrokenDefinedNameRule(WorkbookRule):
    rule_id = "WL013_BROKEN_DEFINED_NAME"
    title = "Broken defined name"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for defined_name in context.workbook.defined_names.values():
            target = defined_name.attr_text or ""
            reason = None
            formula = target if target.startswith("=") else f"={target}"
            if analyze_formula(formula).broken_references:
                reason = "target contains #REF!"
            elif defined_name.type == "RANGE":
                try:
                    destinations = list(defined_name.destinations)
                except (TypeError, ValueError, AttributeError):
                    destinations = []
                    reason = "range target could not be parsed"
                if not destinations and reason is None:
                    reason = "range target has no resolvable destination"
                for sheet_name, range_text in destinations:
                    if sheet_name not in context.workbook.sheetnames:
                        reason = f"target sheet {sheet_name!r} does not exist"
                        break
                    try:
                        range_boundaries(range_text)
                    except ValueError:
                        reason = f"target range {range_text!r} is invalid"
                        break
            if reason is None:
                continue
            result.findings.append(
                _make_finding(
                    context=context,
                    rule_id=self.rule_id,
                    title=self.title,
                    explanation="A workbook defined name cannot resolve to an existing valid range.",
                    severity=Severity.ERROR,
                    confidence=1.0,
                    sheet=None,
                    location=defined_name.name,
                    evidence=Evidence(summary=reason, observed=target),
                    expected="Defined names resolve to valid local sheets and ranges.",
                    suggested_action="Repair or remove the name in Excel after confirming downstream usage.",
                )
            )
        return result


def _ranges_intersect(left: CellRange, right: Region) -> bool:
    return not (
        left.max_row < right.min_row
        or left.min_row > right.max_row
        or left.max_col < right.min_column
        or left.min_col > right.max_column
    )


class MergedCellInDataRegionRule(WorkbookRule):
    rule_id = "WL014_MERGED_CELL_IN_DATA_REGION"
    title = "Merged cells intersect a data region"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            for merged in worksheet.merged_cells.ranges:
                for region in context.data_regions[worksheet.title]:
                    if not _ranges_intersect(merged, region):
                        continue
                    if merged.max_row == merged.min_row == region.min_row:
                        # A single merged header row is common and not itself table corruption.
                        continue
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation="A merge intersects the body of a dense table-like region.",
                            severity=Severity.WARNING,
                            confidence=0.9,
                            sheet=worksheet.title,
                            location=str(merged),
                            evidence=Evidence(
                                summary="Merged range overlaps inferred data body",
                                observed=str(merged),
                                expected={
                                    "min_row": region.min_row,
                                    "max_row": region.max_row,
                                    "min_column": region.min_column,
                                    "max_column": region.max_column,
                                },
                            ),
                            expected="Table-like data bodies use one logical value per cell.",
                            suggested_action="Review downstream sort/filter behavior; no automatic unmerge is offered.",
                        )
                    )
                    break
        return result


def _validation_signature(worksheet: Worksheet, cell: Cell) -> str | None:
    if worksheet.data_validations is None:
        return None
    signatures: list[str] = []
    for validation in worksheet.data_validations.dataValidation:
        for cell_range in validation.ranges.ranges:
            if (
                cell_range.min_row <= cell.row <= cell_range.max_row
                and cell_range.min_col <= cell.column <= cell_range.max_col
            ):
                signatures.append(
                    "|".join(
                        [
                            validation.type or "",
                            validation.operator or "",
                            validation.formula1 or "",
                            validation.formula2 or "",
                            str(bool(validation.allow_blank)),
                        ]
                    )
                )
    return ";".join(sorted(signatures)) or None


class InconsistentDataValidationRule(WorkbookRule):
    rule_id = "WL015_INCONSISTENT_DATA_VALIDATION"
    title = "Inconsistent data validation"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        seen: set[tuple[str, str]] = set()
        for worksheet in context.workbook.worksheets:
            for region in context.data_regions[worksheet.title]:
                for column in range(region.min_column, region.max_column + 1):
                    cells = [
                        _cell(worksheet, row, column)
                        for row in range(region.min_row + 1, region.max_row + 1)
                        if _cell(worksheet, row, column).value is not None
                    ]
                    if len(cells) < 4:
                        continue
                    signatures = [_validation_signature(worksheet, cell) for cell in cells]
                    nonempty = [signature for signature in signatures if signature is not None]
                    if len(nonempty) < 3:
                        continue
                    consensus, count = Counter(nonempty).most_common(1)[0]
                    if count / len(cells) < 0.75:
                        continue
                    outliers = [
                        cell
                        for cell, signature in zip(cells, signatures, strict=True)
                        if signature != consensus
                    ]
                    if len(outliers) != 1:
                        continue
                    cell = outliers[0]
                    if (worksheet.title, cell.coordinate) in seen:
                        continue
                    seen.add((worksheet.title, cell.coordinate))
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation="One populated input cell lacks or differs from the validation used by its peers.",
                            severity=Severity.WARNING,
                            confidence=count / len(cells),
                            sheet=worksheet.title,
                            location=cell.coordinate,
                            evidence=Evidence(
                                summary=f"{count} peer cells share one validation signature",
                                observed=_validation_signature(worksheet, cell),
                                expected=consensus,
                                peers=[
                                    peer.coordinate
                                    for peer in cells
                                    if _validation_signature(worksheet, peer) == consensus
                                ][:12],
                            ),
                            expected="Cells in a homogeneous input column share validation constraints.",
                            suggested_action="Review and restore the intended validation rule manually.",
                        )
                    )
        return result


class TextDisplayRiskRule(WorkbookRule):
    rule_id = "WL016_TEXT_DISPLAY_RISK"
    title = "Text may be clipped or cross a visible boundary"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            issues_by_row: dict[int, list[tuple[Cell, Any, str]]] = defaultdict(list)
            for cell in _visible_text_cells(worksheet):
                measurement = measure_text_cell(worksheet, cell)
                if measurement is None:
                    continue
                wrap_text = bool(cell.alignment.wrap_text)
                explicit_height = explicit_row_height(worksheet, cell.row)
                height_limit = explicit_height
                if height_limit is None and measurement.merged_range is not None:
                    height_limit = measurement.current_height
                forced_multiline = "\n" in cast(str, cell.value) or "\r" in cast(str, cell.value)
                if (
                    (wrap_text or forced_multiline)
                    and height_limit is not None
                    and measurement.required_lines >= 2
                    and measurement.required_height > height_limit + 2.0
                ):
                    issues_by_row[cell.row].append((cell, measurement, "vertical"))
                    continue
                if (
                    not wrap_text
                    and not forced_multiline
                    and measurement.width_ratio >= 1.10
                    and _overflow_is_visually_blocked(
                        worksheet,
                        cell,
                        measurement.required_width,
                        measurement.available_width,
                    )
                ):
                    wrapped_measurement = measure_text_cell(
                        worksheet,
                        cell,
                        assume_wrap=True,
                    )
                    if wrapped_measurement is not None:
                        issues_by_row[cell.row].append((cell, wrapped_measurement, "horizontal"))
            horizontal_by_column: dict[int, list[tuple[Cell, Any, str]]] = defaultdict(list)
            for issues in issues_by_row.values():
                for issue in issues:
                    if issue[2] == "horizontal" and issue[1].merged_range is None:
                        horizontal_by_column[issue[0].column].append(issue)
            width_columns = {
                column for column, issues in horizontal_by_column.items() if len(issues) >= 3
            }
            width_patches: dict[int, PatchOperation] = {}
            if not _is_protected_target(worksheet):
                for column in sorted(width_columns):
                    column_issues = sorted(
                        horizontal_by_column[column],
                        key=lambda issue: (issue[0].row, issue[0].column),
                    )
                    current_width = column_width(worksheet, column)
                    measured_width = max(issue[1].required_width for issue in column_issues)
                    minimum_clear_width = measured_width
                    if minimum_clear_width > 40.0:
                        continue
                    target_width = max(
                        math.ceil(minimum_clear_width * 4.0) / 4.0,
                        math.ceil((current_width + 0.5) * 4.0) / 4.0,
                    )
                    if target_width <= column_width(worksheet, column) + 0.25:
                        continue
                    anchor = column_issues[0][0]
                    width_patch = _make_patch(
                        kind=PatchKind.SET_COLUMN_WIDTH,
                        worksheet=worksheet,
                        cell=anchor,
                        before={
                            "column": get_column_letter(column),
                            "width": current_width,
                        },
                        after={
                            "column": get_column_letter(column),
                            "width": target_width,
                        },
                        confidence=0.96,
                        description=(
                            "Widen the repeatedly overflowing text column to the measured local maximum."
                        ),
                        layout_fingerprint=column_layout_fingerprint(worksheet, column),
                    )
                    result.patches.append(width_patch)
                    width_patches[column] = width_patch
            for row, issues in sorted(issues_by_row.items()):
                height_issues = [
                    issue
                    for issue in issues
                    if not (issue[2] == "horizontal" and issue[0].column in width_patches)
                ]
                horizontal_issues = [issue for issue in height_issues if issue[2] == "horizontal"]
                # Every remaining horizontal issue will receive wrap_text. Its wrapped
                # height must participate in the same atomic row-height plan, even when
                # this row already contains a separate vertical clipping issue.
                height_basis = height_issues
                atomic_group = (
                    stable_id("atomic", self.rule_id, worksheet.title, row)
                    if horizontal_issues
                    else None
                )
                required_height = max(
                    (issue[1].required_height for issue in height_basis),
                    default=row_height(worksheet, row),
                )
                proposed_height = math.ceil(required_height * 2.0) / 2.0
                current_height = explicit_row_height(worksheet, row)
                effective_height = row_height(worksheet, row)
                anchor = (
                    max(
                        height_issues,
                        key=lambda issue: (issue[1].required_height, -issue[0].column),
                    )[0]
                    if height_issues
                    else issues[0][0]
                )
                row_patch: PatchOperation | None = None
                if (
                    height_issues
                    and not _is_protected_target(worksheet)
                    and proposed_height <= MAX_EXCEL_ROW_HEIGHT
                    and proposed_height > effective_height + 0.5
                ):
                    row_patch = _make_patch(
                        kind=PatchKind.SET_ROW_HEIGHT,
                        worksheet=worksheet,
                        cell=anchor,
                        before={
                            "row": row,
                            "height": current_height,
                            "effective_height": effective_height,
                        },
                        after={"row": row, "height": proposed_height},
                        confidence=0.96,
                        description=(
                            f"Increase row {row} height to {proposed_height:g} points after review."
                        ),
                        layout_fingerprint=row_layout_fingerprint(worksheet, row),
                        atomic_group=atomic_group,
                    )
                    result.patches.append(row_patch)
                for cell, measurement, issue_kind in issues:
                    patches: list[PatchOperation] = []
                    shared_width_patch = width_patches.get(cell.column)
                    if issue_kind == "horizontal" and shared_width_patch is not None:
                        patches.append(shared_width_patch)
                    elif (
                        issue_kind == "horizontal"
                        and not _is_protected_target(worksheet)
                        and (
                            row_patch is not None
                            or measurement.required_height <= effective_height + 0.5
                        )
                    ):
                        wrap_patch = _make_patch(
                            kind=PatchKind.SET_WRAP_TEXT,
                            worksheet=worksheet,
                            cell=cell,
                            before={"wrap_text": bool(cell.alignment.wrap_text)},
                            after={"wrap_text": True},
                            confidence=0.96,
                            description="Wrap the blocked text within its existing cell after review.",
                            atomic_group=atomic_group,
                        )
                        result.patches.append(wrap_patch)
                        patches.append(wrap_patch)
                    if row_patch is not None and cell.column not in width_patches:
                        patches.append(row_patch)
                    if issue_kind == "horizontal":
                        summary = "Unwrapped text exceeds its cell and natural overflow is blocked"
                        explanation = (
                            "The text is wider than the available cell width and either an adjacent value "
                            "or a visible border prevents a clean natural overflow."
                        )
                    else:
                        summary = "Wrapped or multiline text exceeds an explicit row height"
                        explanation = (
                            "Static text measurement indicates that the saved explicit row height is too "
                            "small for all wrapped lines."
                        )
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation=explanation,
                            severity=Severity.WARNING,
                            confidence=0.96,
                            sheet=worksheet.title,
                            location=cell.coordinate,
                            evidence=Evidence(
                                summary=summary,
                                observed={
                                    "available_width": round(measurement.available_width, 2),
                                    "required_width": round(measurement.required_width, 2),
                                    "current_row_height": effective_height,
                                    "wrap_text": bool(cell.alignment.wrap_text),
                                },
                                expected=(
                                    {
                                        "estimated_column_width": shared_width_patch.after["width"],
                                        "strategy": "shared_column_width",
                                    }
                                    if shared_width_patch is not None
                                    else {
                                        "required_lines": measurement.required_lines,
                                        "estimated_row_height": proposed_height,
                                        "strategy": "local_wrap_and_row_height",
                                    }
                                ),
                                details={"merged_range": measurement.merged_range},
                            ),
                            expected="Visible text remains inside its intended cell boundary without clipping.",
                            suggested_action=(
                                "The row would exceed Excel's maximum height; widen the layout or shorten "
                                "the content manually."
                                if not patches and proposed_height > MAX_EXCEL_ROW_HEIGHT
                                else "Review the proposed local wrap/row-height change in Excel; font "
                                "rendering can vary by device."
                            ),
                            patches=patches,
                            identity_discriminator=issue_kind,
                        )
                    )
        return result


class BorderEdgeInconsistencyRule(WorkbookRule):
    rule_id = "WL017_BORDER_EDGE_INCONSISTENCY"
    title = "Likely missing shared border edge"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        edge_names = ("left", "right", "top", "bottom")
        for worksheet in context.workbook.worksheets:
            reported_edges: set[tuple[str, int, int]] = set()
            dense_regions = [
                dense_region
                for region in context.data_regions.get(worksheet.title, ())
                if (dense_region := _dense_rectangular_border_region(worksheet, region)) is not None
            ]
            for region in dense_regions:
                candidates_by_coordinate = {
                    (cell.row, cell.column): cell
                    for cell in worksheet._cells.values()
                    if isinstance(cell, Cell)
                    and region.min_row <= cell.row <= region.max_row
                    and region.min_column <= cell.column <= region.max_column
                    and _is_materialized_table_cell(cell)
                    and not _is_hidden_cell(worksheet, cell)
                    and not _is_in_merged_range(worksheet, cell.coordinate)
                }
                if float(region.confidence) >= 0.95:
                    for row in range(region.min_row + 1, region.max_row):
                        for column in range(region.min_column + 1, region.max_column):
                            existing = worksheet._cells.get((row, column))
                            if isinstance(existing, Cell) and _is_materialized_table_cell(existing):
                                continue
                            coordinate = f"{get_column_letter(column)}{row}"
                            if existing is not None and not isinstance(existing, Cell):
                                continue
                            if _is_in_merged_range(worksheet, coordinate):
                                continue
                            neighbors = [
                                worksheet._cells.get((row, column - 1)),
                                worksheet._cells.get((row, column + 1)),
                                worksheet._cells.get((row - 1, column)),
                                worksheet._cells.get((row + 1, column)),
                            ]
                            if not all(
                                isinstance(neighbor, Cell)
                                and _is_materialized_table_cell(neighbor)
                                and not _is_hidden_cell(worksheet, neighbor)
                                and not _is_in_merged_range(worksheet, neighbor.coordinate)
                                for neighbor in neighbors
                            ):
                                continue
                            candidates_by_coordinate[(row, column)] = (
                                existing
                                if isinstance(existing, Cell)
                                else Cell(
                                    worksheet,
                                    row=row,
                                    column=column,
                                )
                            )
                candidates = list(candidates_by_coordinate.values())
                candidates.sort(
                    key=lambda cell: (
                        -sum(side[0] is None for side in visual_border_signature(worksheet, cell)),
                        cell.row,
                        cell.column,
                    )
                )
                for cell in candidates:
                    observed = visual_border_signature(worksheet, cell)
                    edge_consensus: dict[
                        str, tuple[str, str, tuple[Any, ...], float, int, int]
                    ] = {}
                    for index, edge in enumerate(edge_names):
                        if observed[index][0] is not None:
                            continue
                        if _edge_is_internal_to_region(
                            cell, edge, region
                        ) and not _shared_edge_has_materialized_peer(worksheet, cell, edge):
                            continue
                        edge_key = _shared_edge_key(cell, edge)
                        if edge_key in reported_edges:
                            continue
                        consensus = _parallel_border_source(worksheet, cell, edge, region)
                        if consensus is not None:
                            edge_consensus[edge] = consensus
                    if not edge_consensus:
                        continue
                    missing_edges = list(edge_consensus)
                    finding_confidence = min(edge_consensus[edge][3] for edge in missing_edges)
                    expected = list(observed)
                    for edge in missing_edges:
                        expected[edge_names.index(edge)] = edge_consensus[edge][2]
                    patches: list[PatchOperation] = []
                    patchable = finding_confidence >= 0.95 and not _is_protected_target(worksheet)
                    atomic_group = (
                        stable_id(
                            "atomic-border",
                            self.rule_id,
                            worksheet.title,
                            cell.coordinate,
                            missing_edges,
                        )
                        if patchable and len(missing_edges) > 1
                        else None
                    )
                    if patchable:
                        for edge in missing_edges:
                            source_cell, source_edge, _, confidence, _, _ = edge_consensus[edge]
                            edge_index = edge_names.index(edge)
                            patch = _make_patch(
                                kind=PatchKind.COPY_BORDER,
                                worksheet=worksheet,
                                cell=cell,
                                before={
                                    "target_edge": edge,
                                    "signature": observed[edge_index],
                                },
                                after={"target_edge": edge, "source_edge": source_edge},
                                source_cell=source_cell,
                                confidence=confidence,
                                description=(
                                    f"Copy the parallel-consensus {edge} border edge after review."
                                ),
                                atomic_group=atomic_group,
                            )
                            result.patches.append(patch)
                            patches.append(patch)
                    reported_edges.update(_shared_edge_key(cell, edge) for edge in missing_edges)
                    edge_details = {
                        edge: {
                            "confidence": edge_consensus[edge][3],
                            "support": edge_consensus[edge][4],
                            "population": edge_consensus[edge][5],
                            "source": edge_consensus[edge][0],
                        }
                        for edge in missing_edges
                    }
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation=(
                                "Both sides of one or more shared edges are absent inside a dense "
                                "rectangular table, or a table perimeter edge is absent, while parallel "
                                "edges show a stable style. A border present on either side remains "
                                "visually continuous and is not reported."
                            ),
                            severity=Severity.WARNING,
                            confidence=finding_confidence,
                            sheet=worksheet.title,
                            location=cell.coordinate,
                            evidence=Evidence(
                                summary=(
                                    "Missing parallel-consensus edge(s): "
                                    + ", ".join(missing_edges)
                                ),
                                observed=observed,
                                expected=tuple(expected),
                                peers=sorted({edge_consensus[edge][0] for edge in missing_edges}),
                                details={"edge_consensus": edge_details},
                            ),
                            expected="Shared and perimeter table edges remain visually continuous.",
                            suggested_action=(
                                "Review the proposed edge-only border copy; no fill, font, or number "
                                "format is changed. Findings below 0.95 confidence are report-only."
                            ),
                            patches=patches,
                            identity_discriminator=missing_edges,
                        )
                    )
        return result


class UsedRangeInflationRule(WorkbookRule):
    rule_id = "WL018_USED_RANGE_INFLATION"
    title = "Format-only tail inflates the worksheet used range"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            tail = find_formatting_tail(worksheet)
            if tail is None:
                continue
            snapshot = _sheet_snapshot(context, worksheet)
            declared_dimension = getattr(snapshot, "declared_dimension", None)
            content_dimension = getattr(snapshot, "content_dimension", None)
            content_start = f"{get_column_letter(tail.content_min_column)}{tail.content_min_row}"
            content_end = f"{get_column_letter(tail.content_max_column)}{tail.content_max_row}"
            computed_content_dimension = (
                content_start if content_start == content_end else f"{content_start}:{content_end}"
            )
            try:
                declared_bounds = (
                    range_boundaries(declared_dimension)
                    if isinstance(declared_dimension, str) and declared_dimension
                    else None
                )
            except ValueError:
                declared_bounds = None
            tail_cell_max_row = max(
                coordinate_to_tuple(coordinate)[0] for coordinate in tail.cell_coordinates
            )
            minimum_declared_max_row = max(tail.content_max_row, tail_cell_max_row)
            declared_max_row = declared_bounds[3] if declared_bounds is not None else None
            declared_matches = bool(
                declared_bounds
                and declared_bounds[2] == tail.observed_max_column
                and isinstance(declared_max_row, int)
                and minimum_declared_max_row <= declared_max_row <= tail.observed_max_row
            )
            content_matches = content_dimension in {computed_content_dimension, None}
            structure_intersection = _tail_intersects_structure(worksheet, tail)
            anchor = min(
                (
                    cell
                    for cell in worksheet._cells.values()
                    if isinstance(cell, Cell) and cell.value is not None
                ),
                key=lambda cell: (cell.row, cell.column),
            )
            patches: list[PatchOperation] = []
            if (
                not _is_protected_target(worksheet)
                and not _is_hidden_cell(worksheet, anchor)
                and declared_matches
                and content_matches
                and isinstance(declared_dimension, str)
                and isinstance(content_dimension, str)
                and not structure_intersection
            ):
                patch = _make_patch(
                    kind=PatchKind.CLEAR_FORMATTING_TAIL,
                    worksheet=worksheet,
                    cell=anchor,
                    before={"dimension": declared_dimension},
                    after={
                        "cells": list(tail.cell_coordinates),
                        "empty_rows": list(tail.empty_rows),
                        "expected_dimension": declared_dimension,
                        "result_dimension": content_dimension,
                    },
                    confidence=0.99,
                    description="Clear only the exact reviewed format-only cells and empty row records.",
                    layout_fingerprint=tail_layout_fingerprint(worksheet, tail),
                )
                result.patches.append(patch)
                patches.append(patch)
            result.findings.append(
                _make_finding(
                    context=context,
                    rule_id=self.rule_id,
                    title=self.title,
                    explanation=(
                        "A large, separated set of blank styled cells or empty row records extends far "
                        "beyond the populated content. Broad column-dimension styling by itself is ignored."
                    ),
                    severity=Severity.WARNING,
                    confidence=0.99 if declared_matches else 0.9,
                    sheet=worksheet.title,
                    location=",".join((*tail.cell_ranges, *tail.row_ranges)),
                    evidence=Evidence(
                        summary=(
                            f"{tail.styled_cell_count} exact blank styled cells and "
                            f"{len(tail.empty_rows)} empty row records form a separated tail"
                        ),
                        observed={
                            "declared_dimension": declared_dimension,
                            "tail_cell_runs": list(tail.cell_ranges),
                            "empty_rows": list(tail.empty_rows),
                        },
                        expected={
                            "content_dimension": content_dimension or computed_content_dimension
                        },
                        details={
                            "structure_intersection": structure_intersection,
                            "declared_dimension_verified": declared_matches,
                        },
                    ),
                    expected="Worksheet dimensions reflect meaningful content and intentional structures.",
                    suggested_action=(
                        "Apply the exact-cell cleanup only after reviewing names, print settings, comments, "
                        "links, breaks, and drawing anchors."
                    ),
                    patches=patches,
                )
            )
        return result


class IdentifierScientificNotationRule(WorkbookRule):
    rule_id = "WL019_IDENTIFIER_SCIENTIFIC_NOTATION"
    title = "Identifier may display in scientific notation"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            by_column: dict[int, list[tuple[Cell, Cell, str, float]]] = defaultdict(list)
            for cell in sorted(
                (item for item in worksheet._cells.values() if isinstance(item, Cell)),
                key=lambda item: (item.row, item.column),
            ):
                if _is_hidden_cell(worksheet, cell) or cell.data_type == "f":
                    continue
                identifier = _integer_identifier_text(cell.value)
                if identifier is None or len(identifier) < 10:
                    continue
                if (cell.number_format or "General").strip().lower() != "general":
                    continue
                header = _nearest_header(worksheet, cell)
                if header is None or not _looks_like_identifier_header(header.value):
                    continue
                if _looks_like_measure_header(header.value):
                    continue
                measured_width = estimated_text_width(cell, identifier)
                font_aware_width = math.ceil((measured_width + 0.5) * 4.0) / 4.0
                desired_width = min(
                    40.0,
                    max(12.0, len(identifier) * 1.1 + 1.5, font_aware_width),
                )
                width_only_fixable = len(identifier) <= 11
                if (
                    width_only_fixable
                    and column_width(worksheet, cell.column) >= desired_width - 0.25
                ):
                    continue
                by_column[cell.column].append((cell, header, identifier, desired_width))
            for column, candidates in sorted(by_column.items()):
                fixable = [candidate for candidate in candidates if len(candidate[2]) <= 11]
                width_patch: PatchOperation | None = None
                if fixable and not _is_protected_target(worksheet):
                    anchor = fixable[0][0]
                    target_width = round(max(candidate[3] for candidate in fixable), 2)
                    width_patch = _make_patch(
                        kind=PatchKind.SET_COLUMN_WIDTH,
                        worksheet=worksheet,
                        cell=anchor,
                        before={
                            "column": get_column_letter(column),
                            "width": column_width(worksheet, column),
                        },
                        after={
                            "column": get_column_letter(column),
                            "width": target_width,
                        },
                        confidence=0.99,
                        description=(
                            "Widen the identifier column while preserving the stored numeric value and type."
                        ),
                        layout_fingerprint=column_layout_fingerprint(worksheet, column),
                    )
                    result.patches.append(width_patch)
                for cell, header, identifier, desired_width in candidates:
                    patches: list[PatchOperation] = []
                    if len(identifier) <= 11 and width_patch is not None:
                        patches.append(width_patch)
                    result.findings.append(
                        _make_finding(
                            context=context,
                            rule_id=self.rule_id,
                            title=self.title,
                            explanation=(
                                "A long integer under an identifier-like header uses General formatting "
                                "and is wider than its column. Excel may display it in scientific notation."
                            ),
                            severity=Severity.ERROR if len(identifier) > 15 else Severity.WARNING,
                            confidence=0.99,
                            sheet=worksheet.title,
                            location=cell.coordinate,
                            evidence=Evidence(
                                summary=(
                                    "Long numeric identifier uses General format in a narrow column"
                                    if len(identifier) <= 11
                                    else "General format forces a long numeric identifier into scientific notation"
                                ),
                                observed={
                                    "value": cell.value,
                                    "digits": len(identifier),
                                    "number_format": cell.number_format,
                                    "column_width": column_width(worksheet, column),
                                    "font_name": cell.font.name,
                                    "font_size": cell.font.sz,
                                },
                                expected={
                                    "storage": "numeric value and type preserved",
                                    "minimum_estimated_width": (
                                        round(desired_width, 2) if len(identifier) <= 11 else None
                                    ),
                                },
                                peers=[header.coordinate],
                                details={
                                    "header": header.value,
                                    "precision_recoverable": len(identifier) <= 15,
                                    "width_only_fixable": len(identifier) <= 11,
                                },
                            ),
                            expected=(
                                "Phone, ID, account, and similar identifiers display fully without changing "
                                "their stored value or type."
                            ),
                            suggested_action=(
                                "Review the font-aware width-only proposal and source value."
                                if len(identifier) <= 11
                                else (
                                    "General format still uses scientific notation for 12-15 digit integers "
                                    "even in wide columns; choose an explicit integer format or intentional "
                                    "text storage after semantic review."
                                    if len(identifier) <= 15
                                    else "The value may already have lost precision and is not auto-patched."
                                )
                            ),
                            patches=patches,
                        )
                    )
        return result


class SavedViewOffContentRule(WorkbookRule):
    rule_id = "WL020_SAVED_VIEW_OFF_CONTENT"
    title = "Saved worksheet view may hide meaningful content"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            formatting_tail = find_formatting_tail(worksheet)
            whitespace_tail = find_whitespace_tail(worksheet)
            excluded_coordinates = {
                coordinate
                for tail in (formatting_tail, whitespace_tail)
                if tail is not None
                for coordinate in tail.cell_coordinates
            }
            bounds = _visible_layout_bounds(
                worksheet,
                excluded_coordinates=excluded_coordinates,
            )
            if bounds is None:
                continue
            min_row, min_column, max_row, max_column = bounds
            layout_cells = [
                cell
                for cell in worksheet._cells.values()
                if isinstance(cell, Cell)
                and cell.coordinate not in excluded_coordinates
                and min_row <= cell.row <= max_row
                and min_column <= cell.column <= max_column
                and not _is_hidden_cell(worksheet, cell)
                and (
                    (
                        cell.value is not None
                        and not (isinstance(cell.value, str) and not cell.value.strip())
                    )
                    or has_visible_border(cell)
                    or cell.fill.fill_type not in {None, "none"}
                    or cell.comment is not None
                    or cell.hyperlink is not None
                )
            ]
            if not layout_cells:
                continue
            top_left = worksheet.sheet_view.topLeftCell or "A1"
            pane = worksheet.sheet_view.pane
            pane_state = None if pane is None or pane.state is None else str(pane.state)
            expected_origin = f"{get_column_letter(min_column)}{min_row}"
            observed_origin = top_left
            pane_offset = False
            if pane is not None:
                x_split = _finite_nonnegative_pane_split(pane.xSplit)
                y_split = _finite_nonnegative_pane_split(pane.ySplit)
                if x_split is None or y_split is None:
                    continue
                if pane.topLeftCell and _bounded_view_coordinate(pane.topLeftCell) is None:
                    continue
                if pane_state == "frozen":
                    if not x_split.is_integer() or not y_split.is_integer():
                        continue
                    frozen_columns = int(x_split)
                    frozen_rows = int(y_split)
                    if frozen_columns > MAX_COLUMN - min_column or frozen_rows > MAX_ROW - min_row:
                        continue
                    expected_origin = (
                        f"{get_column_letter(min_column + frozen_columns)}{min_row + frozen_rows}"
                    )
                    if pane.topLeftCell:
                        observed_origin = pane.topLeftCell
            observed_coordinate = _bounded_view_coordinate(observed_origin)
            expected_coordinate = _bounded_view_coordinate(expected_origin)
            if observed_coordinate is None or expected_coordinate is None:
                continue
            view_row, view_column = observed_coordinate
            expected_row, expected_column = expected_coordinate
            if pane is None:
                view_off_content = view_row > expected_row or view_column > expected_column
            elif pane_state == "frozen":
                pane_offset = view_row > expected_row or view_column > expected_column
                view_off_content = pane_offset
            else:
                view_off_content = False
            count_compact = max_row - min_row + 1 <= 40 and max_column - min_column + 1 <= 20
            visible_columns = [
                column
                for column in range(min_column, max_column + 1)
                if not is_column_hidden(worksheet, column)
            ]
            visible_rows = [
                row
                for row in range(min_row, max_row + 1)
                if not (
                    (row_dimension := worksheet.row_dimensions.get(row)) is not None
                    and row_dimension.hidden
                )
            ]
            content_width, content_height, planned_width_updates, planned_height_updates = (
                _predicted_layout_size(context, worksheet, visible_columns, visible_rows)
            )
            raw_width_fit_zoom = (
                int(math.floor(CONSERVATIVE_VIEWPORT_WIDTH / content_width * 100.0 / 5.0) * 5)
                if content_width > 0
                else 100
            )
            raw_height_fit_zoom = (
                int(math.floor(CONSERVATIVE_VIEWPORT_HEIGHT / content_height * 100.0 / 5.0) * 5)
                if content_height > 0
                else 100
            )
            raw_fit_zoom = min(raw_width_fit_zoom, raw_height_fit_zoom)
            full_fit_zoom = (
                min(400, raw_fit_zoom) if raw_fit_zoom >= MIN_AUTOMATIC_VIEW_ZOOM else None
            )
            zoom_scale = worksheet.sheet_view.zoomScale
            saved_zoom = int(zoom_scale or 100)
            vertical_scrolling_expected = (
                raw_height_fit_zoom < MIN_AUTOMATIC_VIEW_ZOOM <= raw_width_fit_zoom
            )
            width_only_fit_zoom = (
                min(400, raw_width_fit_zoom)
                if vertical_scrolling_expected and zoom_scale is not None and saved_zoom > 100
                else None
            )
            fit_zoom = full_fit_zoom if full_fit_zoom is not None else width_only_fit_zoom
            compact = count_compact or fit_zoom is not None
            zoom_too_large = compact and fit_zoom is not None and saved_zoom > fit_zoom
            fit_below_safe_floor = count_compact and raw_width_fit_zoom < MIN_AUTOMATIC_VIEW_ZOOM
            if not view_off_content and not zoom_too_large and not fit_below_safe_floor:
                continue
            content_start = f"{get_column_letter(min_column)}{min_row}"
            anchor = min(layout_cells, key=lambda cell: (cell.row, cell.column))
            has_pane = pane is not None
            can_adjust_zoom = pane is None or (pane_state == "frozen" and not pane_offset)
            after: dict[str, Any] = {}
            if not has_pane and view_off_content:
                after["top_left_cell"] = content_start
            if compact and fit_zoom is not None and can_adjust_zoom:
                # Keep one five-point step of headroom for font/rendering differences and
                # other reviewed width changes selected from the same repair plan.
                target_zoom = max(MIN_AUTOMATIC_VIEW_ZOOM, fit_zoom - 5)
                if zoom_too_large and saved_zoom > target_zoom:
                    after["zoom_scale"] = target_zoom
                    if not has_pane:
                        after.setdefault("top_left_cell", content_start)
            patches: list[PatchOperation] = []
            if after:
                patch = _make_patch(
                    kind=PatchKind.SET_SHEET_VIEW,
                    worksheet=worksheet,
                    cell=anchor,
                    before={
                        "top_left_cell": top_left,
                        "pane_top_left_cell": None if pane is None else pane.topLeftCell,
                        "zoom_scale": zoom_scale,
                    },
                    after=after,
                    confidence=0.99,
                    description=(
                        "Reset the saved viewport when safe and reduce zoom using conservative "
                        "two-dimensional layout estimates."
                    ),
                    layout_fingerprint=sheet_view_fingerprint(worksheet),
                )
                result.patches.append(patch)
                patches.append(patch)
            result.findings.append(
                _make_finding(
                    context=context,
                    rule_id=self.rule_id,
                    title=self.title,
                    explanation=(
                        "The saved viewport either starts beyond the intended visible layout or uses a "
                        "zoom level that makes the intended visible layout difficult to use."
                    ),
                    severity=Severity.WARNING,
                    confidence=0.99,
                    sheet=worksheet.title,
                    location=observed_origin,
                    evidence=Evidence(
                        summary=(
                            "Compact content is too wide to fit above the automatic zoom safety floor"
                            if fit_below_safe_floor
                            else (
                                "Saved zoom is too large for the visible width of a vertically "
                                "scrollable sheet"
                                if vertical_scrolling_expected and zoom_too_large
                                else "Saved viewport is offset or too zoomed-in for the compact visible layout"
                            )
                        ),
                        observed={
                            "top_left_cell": top_left,
                            "pane_top_left_cell": None if pane is None else pane.topLeftCell,
                            "zoom_scale": zoom_scale,
                            "estimated_content_width": round(content_width, 2),
                            "estimated_content_height": round(content_height, 2),
                            "estimated_width_fit_zoom": min(400, raw_width_fit_zoom),
                            "estimated_height_fit_zoom": min(400, raw_height_fit_zoom),
                            "estimated_fit_zoom": fit_zoom,
                            "raw_estimated_fit_zoom": raw_fit_zoom,
                            "pane": None if pane is None else pane.state,
                        },
                        expected={
                            "top_left_cell": expected_origin,
                            "zoom_scale": after.get("zoom_scale", zoom_scale),
                        },
                        details={
                            "content_bounds": (
                                f"{content_start}:{get_column_letter(max_column)}{max_row}"
                            ),
                            "compact_sheet": compact,
                            "viewport_width_assumption": CONSERVATIVE_VIEWPORT_WIDTH,
                            "viewport_height_assumption": CONSERVATIVE_VIEWPORT_HEIGHT,
                            "planned_width_updates_included": planned_width_updates,
                            "planned_height_updates_included": planned_height_updates,
                            "fit_zoom_below_safe_floor": fit_below_safe_floor,
                            "vertical_scrolling_expected": vertical_scrolling_expected,
                            "pane_preserved": has_pane,
                            "frozen_pane_offset": pane_offset,
                            "zoom_only_with_frozen_pane": (
                                has_pane and pane_state == "frozen" and not pane_offset
                            ),
                        },
                    ),
                    expected=(
                        "Opening the sheet starts at the intended content origin and uses a readable zoom "
                        "for the visible layout."
                    ),
                    suggested_action=(
                        "The estimated width fit zoom is below the automatic safety floor; review the "
                        "layout, column widths, or saved zoom manually."
                        if fit_below_safe_floor
                        else (
                            "Review the frozen-pane scroll origin manually; WorkbookLens will not "
                            "change the number of frozen rows or columns."
                            if pane_offset
                            else "Review the saved viewport. The proposed zoom is a conservative estimate "
                            "for a typical desktop window, not a cross-device guarantee."
                        )
                    ),
                    patches=patches,
                )
            )
        return result


class WhitespaceOnlyTailRule(WorkbookRule):
    rule_id = "WL021_WHITESPACE_ONLY_TAIL"
    title = "Whitespace-only cells extend beyond the visible layout"

    def run(self, context: RuleContext) -> RuleResult:
        result = RuleResult()
        for worksheet in context.workbook.worksheets:
            tail = find_whitespace_tail(worksheet)
            if tail is None:
                continue
            snapshot = _sheet_snapshot(context, worksheet)
            declared_dimension = getattr(snapshot, "declared_dimension", None)
            try:
                declared_matches = bool(
                    isinstance(declared_dimension, str)
                    and CellRange(declared_dimension).bounds
                    == CellRange(tail.observed_dimension).bounds
                )
            except ValueError:
                declared_matches = False
            structure_intersection = _tail_intersects_structure(worksheet, tail)
            formula_blocker = _tail_has_formula_blocker(context, worksheet, tail)
            formatting_tail_present = find_formatting_tail(worksheet) is not None
            target_coordinates = set(tail.cell_coordinates)
            anchor = min(
                (
                    cell
                    for cell in worksheet._cells.values()
                    if isinstance(cell, Cell)
                    and cell.coordinate not in target_coordinates
                    and cell.value is not None
                    and not (isinstance(cell.value, str) and not cell.value.strip())
                ),
                key=lambda cell: (cell.row, cell.column),
                default=worksheet[tail.cell_coordinates[0]],
            )
            patches: list[PatchOperation] = []
            if (
                worksheet.sheet_state == "visible"
                and not _is_protected_target(worksheet)
                and declared_matches
                and not structure_intersection
                and not formula_blocker
                and not formatting_tail_present
                and isinstance(anchor, Cell)
            ):
                patch = _make_patch(
                    kind=PatchKind.REMOVE_WHITESPACE_TAIL_CELLS,
                    worksheet=worksheet,
                    cell=anchor,
                    before={"dimension": declared_dimension},
                    after={
                        "cells": list(tail.cell_coordinates),
                        "expected_dimension": tail.observed_dimension,
                        "preserve_style_ids": dict(tail.preserve_style_ids),
                        "result_dimension": tail.result_dimension,
                        "preserve_row_dimensions": True,
                    },
                    confidence=0.99,
                    description=(
                        "Remove reviewed default-style whitespace nodes and clear only the values "
                        "of non-default-style cells, preserving their styles and row dimensions."
                    ),
                    layout_fingerprint=whitespace_tail_layout_fingerprint(worksheet, tail),
                )
                result.patches.append(patch)
                patches.append(patch)
            result.findings.append(
                _make_finding(
                    context=context,
                    rule_id=self.rule_id,
                    title=self.title,
                    explanation=(
                        "A connected tail of literal whitespace strings extends the worksheet cell "
                        "bounds beyond its visible content. These cells can render as stray values in "
                        "some preview tools even though Excel shows them as blank."
                    ),
                    severity=Severity.WARNING,
                    confidence=0.99 if patches else 0.9,
                    sheet=worksheet.title,
                    location=",".join(tail.cell_ranges),
                    evidence=Evidence(
                        summary=(
                            f"{len(tail.cell_coordinates)} literal-whitespace cells form an outer tail"
                        ),
                        observed={
                            "cell_runs": list(tail.cell_ranges),
                            "declared_dimension": declared_dimension,
                            "observed_dimension": tail.observed_dimension,
                            "preserved_style_cells": dict(tail.preserve_style_ids),
                        },
                        expected={"result_dimension": tail.result_dimension},
                        details={
                            "row_dimensions_preserved": True,
                            "cell_styles_preserved": True,
                            "structure_intersection": structure_intersection,
                            "formula_blocker": formula_blocker,
                            "formatting_tail_cleanup_first": formatting_tail_present,
                            "declared_dimension_verified": declared_matches,
                        },
                    ),
                    expected=(
                        "Literal whitespace values outside the intended layout are cleared; only "
                        "intentional non-default style nodes may continue to define stored bounds."
                    ),
                    suggested_action=(
                        "Apply the exact value cleanup after reviewing the preserved cell styles and "
                        "blank-row layout."
                        if patches
                        else "Review the referenced structure or formula blocker; no automatic cleanup is offered."
                    ),
                    patches=patches,
                )
            )
        return result


BUILTIN_RULES: tuple[type[WorkbookRule], ...] = (
    BrokenReferenceRule,
    FormulaPatternOutlierRule,
    BlankInFormulaBandRule,
    HardcodedValueInFormulaBandRule,
    SuspiciousSumBoundaryRule,
    NumericTextRule,
    StyleOutlierRule,
    HiddenNonemptyDataRule,
    ExternalLinkRule,
    VolatileOrFragileFunctionRule,
    ErrorCellRule,
    DuplicateConfiguredKeyRule,
    BrokenDefinedNameRule,
    MergedCellInDataRegionRule,
    InconsistentDataValidationRule,
    TextDisplayRiskRule,
    BorderEdgeInconsistencyRule,
    UsedRangeInflationRule,
    IdentifierScientificNotationRule,
    SavedViewOffContentRule,
    WhitespaceOnlyTailRule,
)
