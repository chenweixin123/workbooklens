"""Workbook-to-model snapshot extraction without formula or macro execution."""

from __future__ import annotations

import posixpath
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Any, Literal, cast

from lxml import etree
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from workbooklens.exceptions import PatchValidationError
from workbooklens.models import CellSnapshot, SheetSnapshot, WorkbookSnapshot
from workbooklens.ooxml.safety import PackageLimits, inspect_package, parse_xml_part
from workbooklens.utils import sha256_bytes, sha256_file, stable_json_bytes
from workbooklens.worksheet_state import hidden_column_labels, is_column_hidden


def _formula_payload(cell: Cell) -> str | None:
    if cell.data_type != "f":
        return None
    if isinstance(cell.value, str):
        return cell.value
    if isinstance(cell.value, ArrayFormula):
        return f"<array ref={cell.value.ref!r} text={cell.value.text!r}>"
    if isinstance(cell.value, DataTableFormula):
        attributes = ",".join(
            f"{name}={getattr(cell.value, name)!r}"
            for name in ("ref", "ca", "dt2D", "dtr", "r1", "r2", "del1", "del2")
        )
        return f"<dataTable {attributes}>"
    return f"<unsupported formula object {type(cell.value).__qualname__}>"


def cell_semantic_payload(cell: Cell) -> dict[str, Any]:
    """Return the source-bound fields used by patch preconditions."""

    formula = _formula_payload(cell)
    value = None if formula is not None else cell.value
    return {
        "coordinate": cell.coordinate,
        "value": value,
        "formula": formula,
        "data_type": cell.data_type,
        "style_id": cell.style_id,
        "number_format": cell.number_format,
        "quote_prefix": bool(cell.quotePrefix),
    }


def cell_fingerprint(cell: Cell) -> str:
    """Hash effective cell semantics without workbook-local style table IDs."""

    payload = cell_semantic_payload(cell)
    payload.pop("style_id", None)
    payload["style_fingerprint"] = style_fingerprint(cell)
    return sha256_bytes(stable_json_bytes(payload))


def _xml_component_payload(component: Any) -> dict[str, Any]:
    """Return a deterministic, workbook-local-ID-free style component tree."""

    element = component.to_tree()

    def normalize(node: Any) -> dict[str, Any]:
        return {
            "tag": str(node.tag),
            "attributes": dict(
                sorted((str(key), str(value)) for key, value in node.attrib.items())
            ),
            "text": node.text or "",
            "children": [normalize(child) for child in node],
        }

    return normalize(element)


def _style_payload(cell: Cell) -> dict[str, Any]:
    """Describe effective non-number-format styling without using ``style_id``."""

    return {
        "font": _xml_component_payload(cell.font),
        "fill": _xml_component_payload(cell.fill),
        "border": _xml_component_payload(cell.border),
        "alignment": _xml_component_payload(cell.alignment),
        "protection": _xml_component_payload(cell.protection),
        "quote_prefix": bool(cell.quotePrefix),
        "pivot_button": bool(cell.pivotButton),
    }


@lru_cache(maxsize=1)
def _default_style_payload() -> dict[str, Any]:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        if worksheet is None:  # pragma: no cover - openpyxl always creates an active sheet
            raise RuntimeError("new workbook has no active worksheet")
        return _style_payload(worksheet["A1"])
    finally:
        workbook.close()


def style_fingerprint(cell: Cell) -> str:
    """Hash effective visual/protection styling independently of workbook style-table IDs."""

    payload = _style_payload(cell)
    if payload == _default_style_payload():
        return "default"
    return sha256_bytes(stable_json_bytes(payload))


def _snapshot_cell(cell: Cell, worksheet: Worksheet) -> CellSnapshot:
    payload = cell_semantic_payload(cell)
    snapshot_payload = dict(payload)
    snapshot_payload.pop("quote_prefix", None)
    row_dimension = worksheet.row_dimensions.get(cell.row)
    return CellSnapshot(
        **snapshot_payload,
        style_fingerprint=style_fingerprint(cell),
        row_hidden=bool(row_dimension.hidden) if row_dimension is not None else False,
        column_hidden=is_column_hidden(worksheet, cell.column),
    )


def _defined_names(workbook: Workbook) -> dict[str, str]:
    names: dict[str, str] = {}
    for name in workbook.defined_names.values():
        key = name.name
        if name.localSheetId is not None:
            key = f"{key}@sheet:{name.localSheetId}"
        names[key] = name.attr_text or ""
    return dict(sorted(names.items()))


def _resolve_part(base_part: str, target: str) -> str:
    """Resolve one internal OOXML relationship without allowing package-root escape."""

    if "\\" in target or "\x00" in target:
        raise PatchValidationError(f"Unsafe relationship target: {target!r}")
    candidate = (
        target.lstrip("/")
        if target.startswith("/")
        else posixpath.join(posixpath.dirname(base_part), target)
    )
    normalized = posixpath.normpath(candidate)
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise PatchValidationError(f"Relationship target escapes package root: {target!r}")
    return normalized


def _declared_sheet_dimensions(path: Path) -> dict[str, str | None]:
    """Read raw worksheet ``dimension`` refs before openpyxl normalizes sparse state."""

    with zipfile.ZipFile(path, "r") as archive:
        workbook_part = "xl/workbook.xml"
        relationships_part = "xl/_rels/workbook.xml.rels"
        workbook_root = parse_xml_part(archive.read(workbook_part), workbook_part)
        relationships_root = parse_xml_part(archive.read(relationships_part), relationships_part)
        relationships: dict[str, tuple[str, str]] = {}
        for relationship in relationships_root:
            if etree.QName(relationship).localname != "Relationship":
                continue
            identifier = relationship.get("Id")
            target = relationship.get("Target")
            relationship_type = relationship.get("Type", "")
            if identifier and target:
                relationships[identifier] = (relationship_type, target)

        dimensions: dict[str, str | None] = {}
        for sheet in workbook_root.iter():
            if etree.QName(sheet).localname != "sheet":
                continue
            name = sheet.get("name")
            relationship_id = next(
                (
                    value
                    for attribute, value in sheet.attrib.items()
                    if etree.QName(attribute).localname == "id"
                ),
                None,
            )
            if not name or not relationship_id or relationship_id not in relationships:
                raise PatchValidationError(
                    "Workbook contains a sheet with a malformed relationship"
                )
            relationship_type, target = relationships[relationship_id]
            if relationship_type.endswith("/chartsheet"):
                continue
            if not relationship_type.endswith("/worksheet"):
                raise PatchValidationError(
                    "Workbook contains an unsupported sheet relationship type: "
                    f"{relationship_type!r}"
                )
            part = _resolve_part(workbook_part, target)
            if part not in archive.namelist():
                raise PatchValidationError(f"Worksheet package part is missing: {part}")
            worksheet_root = parse_xml_part(archive.read(part), part)
            dimension = next(
                (
                    element
                    for element in worksheet_root
                    if etree.QName(element).localname == "dimension"
                ),
                None,
            )
            dimensions[name] = dimension.get("ref") if dimension is not None else None
        return dimensions


def _content_bounds(worksheet: Worksheet, cells: list[Cell]) -> str | None:
    """Return content bounds while preserving the visible extent of merged ranges."""

    meaningful = [cell for cell in cells if cell.value is not None]
    row_bounds = [cell.row for cell in meaningful]
    column_bounds = [cell.column for cell in meaningful]
    for merged in worksheet.merged_cells.ranges:
        row_bounds.extend((merged.min_row, merged.max_row))
        column_bounds.extend((merged.min_col, merged.max_col))
    if not row_bounds or not column_bounds:
        return None
    start = f"{get_column_letter(min(column_bounds))}{min(row_bounds)}"
    end = f"{get_column_letter(max(column_bounds))}{max(row_bounds)}"
    return start if start == end else f"{start}:{end}"


def _row_heights(worksheet: Worksheet) -> dict[str, float]:
    return {
        str(index): float(dimension.height)
        for index, dimension in sorted(worksheet.row_dimensions.items())
        if dimension.height is not None
    }


def _column_widths(worksheet: Worksheet) -> dict[str, float]:
    entries: list[tuple[int, int, str, float]] = []
    for key, dimension in worksheet.column_dimensions.items():
        if dimension.width is None:
            continue
        minimum = int(dimension.min or 0)
        maximum = int(dimension.max or minimum)
        if minimum < 1 or maximum < minimum or maximum > 16_384:
            raise PatchValidationError(
                f"Invalid column dimension span for {worksheet.title!r}: {key!r}"
            )
        start = get_column_letter(minimum)
        end = get_column_letter(maximum)
        label = start if start == end else f"{start}:{end}"
        entries.append((minimum, maximum, label, float(dimension.width)))
    return {label: width for _, _, label, width in sorted(entries)}


def snapshot_from_workbook(workbook: Workbook, path: Path, source_sha256: str) -> WorkbookSnapshot:
    """Build a deterministic snapshot from an already-open, non-read-only workbook."""

    declared_dimensions = _declared_sheet_dimensions(path)
    sheets: list[SheetSnapshot] = []
    for index, worksheet in enumerate(workbook.worksheets):
        cells: dict[str, CellSnapshot] = {}
        # Iterating the sparse cell store avoids trusting a maliciously huge dimension range.
        sparse_cells = [cell for cell in worksheet._cells.values() if isinstance(cell, Cell)]
        for cell in sorted(
            sparse_cells,
            key=lambda item: (item.row, item.column),
        ):
            if cell.value is not None or cell.has_style:
                cells[cell.coordinate] = _snapshot_cell(cell, worksheet)
        validations: list[str] = []
        if worksheet.data_validations is not None:
            for validation in worksheet.data_validations.dataValidation:
                validations.append(
                    "|".join(
                        [
                            str(validation.sqref),
                            validation.type or "",
                            validation.operator or "",
                            validation.formula1 or "",
                            validation.formula2 or "",
                        ]
                    )
                )
        sheets.append(
            SheetSnapshot(
                name=worksheet.title,
                index=index,
                state=worksheet.sheet_state,
                max_row=max((cell.row for cell in sparse_cells), default=0),
                max_column=max((cell.column for cell in sparse_cells), default=0),
                declared_dimension=declared_dimensions.get(worksheet.title),
                content_dimension=_content_bounds(worksheet, sparse_cells),
                row_heights=_row_heights(worksheet),
                column_widths=_column_widths(worksheet),
                view_top_left_cell=worksheet.sheet_view.topLeftCell,
                view_zoom_scale=(
                    int(worksheet.sheet_view.zoomScale)
                    if worksheet.sheet_view.zoomScale is not None
                    else None
                ),
                cells=cells,
                merged_ranges=sorted(str(item) for item in worksheet.merged_cells.ranges),
                hidden_rows=sorted(
                    index
                    for index, dimension in worksheet.row_dimensions.items()
                    if dimension.hidden
                ),
                hidden_columns=hidden_column_labels(worksheet),
                data_validations=sorted(validations),
            )
        )
    calculation_mode = None
    if workbook.calculation is not None:
        calculation_mode = workbook.calculation.calcMode
    workbook_format = cast(Literal["xlsx", "xlsm"], path.suffix.lower().lstrip("."))
    return WorkbookSnapshot(
        source_name=path.name,
        source_sha256=source_sha256,
        format=workbook_format,
        sheets=sheets,
        defined_names=_defined_names(workbook),
        calculation_mode=calculation_mode,
    )


def load_for_analysis(path: Path, limits: PackageLimits | None = None) -> Workbook:
    """Safety-check and open a workbook without evaluating formulas or external content."""

    inspection = inspect_package(path, limits)
    return load_workbook(
        inspection.path,
        read_only=False,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )


def create_snapshot(path: Path, limits: PackageLimits | None = None) -> WorkbookSnapshot:
    """Inspect, open, snapshot, and close one workbook."""

    resolved = path.expanduser().resolve()
    workbook = load_for_analysis(resolved, limits)
    try:
        return snapshot_from_workbook(workbook, resolved, sha256_file(resolved))
    finally:
        workbook.close()
